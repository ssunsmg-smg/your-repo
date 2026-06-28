"""
==============================================================
  퀀트 주식 스크리너 v6.3 (단일 파일 완전통합판)
  ★ v35 핵심 스크리닝 엔진 전체를 본 파일에 직접 병합
    (더 이상 quant_screener_v41.py 별도 파일이 필요 없음)
  ★ 한국 KOSPI/KOSDAQ 전용 — 미국(S&P500) 별도 운영

  ┌─────────────────────────────────────────────────────────┐
  │  ① DART 재무 + 네이버 보조 + yfinance 가격/기술지표      │
  │  ② 7팩터 스코어링 + 100점 평가 + 매매시그널 7단계        │
  │  ③ 백테스트 엔진   (Walk-forward, IS/OOS 검증)          │
  │  ④ 팩터 모델 고도화 (Z-score, 섹터중립, 앙상블)        │
  │  ⑤ KIS API 자동매매 (한국투자증권 REST, 64bit OK)       │
  │  ⑥ 실시간 모니터링  (텔레그램 알림, 성과 추적)          │
  └─────────────────────────────────────────────────────────┘

  실행 방법 (이 파일 하나만 있으면 전체 기능 동작):
    python quant_screener_v36.py              # 대화형 메뉴
    python quant_screener_v36.py --auto       # 자동(스케줄러)
    python quant_screener_v36.py --backtest   # 백테스트 전용
    python quant_screener_v36.py --trade      # KIS 자동매매
    python quant_screener_v36.py --monitor    # 모니터링 전용

  KIS API 설정:
    - https://apiportal.koreainvestment.com 에서 앱키 발급 (무료)
    - kis_config.json 에 app_key / app_secret / account_no 입력
    - 64bit Python, VS Code 환경 그대로 사용 가능

  필요 패키지:
    pip install finance-datareader yfinance requests beautifulsoup4
                pandas numpy openpyxl tqdm dart-fss scikit-learn
==============================================================
"""

import os, sys, re, time, json, zipfile, io, warnings, argparse, traceback, hashlib
import threading
from datetime import datetime, timedelta
from concurrent.futures import ThreadPoolExecutor, as_completed
from typing import Optional

import requests
import pandas as pd
import numpy as np

try:
    from bs4 import BeautifulSoup
    HAS_BS4 = True
except ImportError:
    HAS_BS4 = False

try:
    import yfinance as yf
    HAS_YFINANCE = True
except ImportError:
    HAS_YFINANCE = False

try:
    import FinanceDataReader as fdr
    HAS_FDR = True
except ImportError:
    HAS_FDR = False

try:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side, GradientFill
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.merge import MergedCell
    from openpyxl.formatting.rule import ColorScaleRule
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    from tqdm import tqdm
    HAS_TQDM = True
except ImportError:
    HAS_TQDM = False

try:
    from sklearn.preprocessing import StandardScaler
    from sklearn.linear_model import Ridge
    HAS_SKLEARN = True
except ImportError:
    HAS_SKLEARN = False

warnings.filterwarnings("ignore")

VERSION   = "v6.3.0"
# [패치] 백테스트 현실화: 거래비용(수수료+세금+슬리피지) 모델 추가,
#        생존편향 경고 문구 자동 출력 (BacktestEngine.__init__ / run_walkforward /
#        run_basket_simulation 참고). 비용 없이 옛 결과와 비교하려면 --bt-no-cost.
BASE_DIR  = os.path.dirname(os.path.abspath(__file__))
CACHE_DIR = os.path.join(BASE_DIR, ".cache")
LOG_DIR   = os.path.join(BASE_DIR, "logs")
BT_DIR    = os.path.join(BASE_DIR, "backtest")
TRADE_DIR = os.path.join(BASE_DIR, "trades")

for _d in [CACHE_DIR, LOG_DIR, BT_DIR, TRADE_DIR]:
    os.makedirs(_d, exist_ok=True)

# ══════════════════════════════════════════════════════════
# 공통 스타일 상수 (v35 동일)
# ══════════════════════════════════════════════════════════
_HDR_FILL = PatternFill("solid", fgColor="1F3864") if HAS_OPENPYXL else None
_GLD_FILL = PatternFill("solid", fgColor="FFD700") if HAS_OPENPYXL else None
_SLV_FILL = PatternFill("solid", fgColor="C0C0C0") if HAS_OPENPYXL else None
_BRZ_FILL = PatternFill("solid", fgColor="CD7F32") if HAS_OPENPYXL else None
_GRN_FILL = PatternFill("solid", fgColor="E2EFDA") if HAS_OPENPYXL else None
_ALT_FILL = PatternFill("solid", fgColor="F5F5F5") if HAS_OPENPYXL else None
_THIN = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
) if HAS_OPENPYXL else None

NAVER_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/120.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "ko-KR,ko;q=0.9",
    "Referer": "https://finance.naver.com",
}


# ══════════════════════════════════════════════════════════
# 공통 유틸 (v35 동일)
# ══════════════════════════════════════════════════════════
def _safe(v, default=None):
    try:
        if v is None or (isinstance(v, float) and np.isnan(v)):
            return default
        return v
    except Exception:
        return default

def safe_div(a, b, default=np.nan):
    try:
        if b == 0 or pd.isna(b) or pd.isna(a):
            return default
        return a / b
    except Exception:
        return default

def grade_label(score):
    if score >= 85:   return "◆◆◆ A+ 탁월", "FF4444"
    elif score >= 75: return "◆◆  A  우수",  "FF9900"
    elif score >= 65: return "◆   B  양호",  "FFCC00"
    elif score >= 50: return "    C  보통",   "92D050"
    else:             return "    D  검토",   "BFBFBF"

def _rank_fill(rank):
    if not HAS_OPENPYXL: return None
    if rank == 1:   return _GLD_FILL
    elif rank == 2: return _SLV_FILL
    elif rank == 3: return _BRZ_FILL
    elif rank <= 5: return _GRN_FILL
    else:           return _ALT_FILL

def auto_col_width(ws, min_w=8, max_w=45):
    if not HAS_OPENPYXL: return
    col_widths = {}
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell, MergedCell) or cell.value is None:
                continue
            text = str(cell.value)
            w = sum(2 if ord(c) > 127 else 1 for c in text) + 2
            col_letter = get_column_letter(cell.column)
            col_widths[col_letter] = max(col_widths.get(col_letter, 0), w)
    for col_letter, w in col_widths.items():
        ws.column_dimensions[col_letter].width = min(max(w, min_w), max_w)


# ══════════════════════════════════════════════════════════
# 캐시 시스템 (v35 동일)
# ══════════════════════════════════════════════════════════
_TIER_TTL = {"A": 7, "B": 1}

def _cache_path(key: str, tier: str = "B") -> str:
    safe_key = hashlib.md5(key.encode()).hexdigest()[:12]
    tier_dir = os.path.join(CACHE_DIR, f"tier_{tier}")
    os.makedirs(tier_dir, exist_ok=True)
    return os.path.join(tier_dir, f"{safe_key}.json")

def cache_get(key: str, tier: str = "B"):
    ttl_days = _TIER_TTL.get(tier, 1)
    path = _cache_path(key, tier)
    if not os.path.exists(path):
        return None
    try:
        mtime = os.path.getmtime(path)
        if (time.time() - mtime) / 86400 > ttl_days:
            return None
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return None

def cache_set(key: str, data, tier: str = "B") -> None:
    path = _cache_path(key, tier)
    try:
        with open(path, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False)
    except Exception:
        pass

def cache_clear(days_old: int = 8) -> None:
    deleted = 0
    cutoff = time.time()
    try:
        for tier_name in ["tier_A", "tier_B", ""]:
            target_dir = os.path.join(CACHE_DIR, tier_name) if tier_name else CACHE_DIR
            if not os.path.isdir(target_dir):
                continue
            for fname in os.listdir(target_dir):
                if not fname.endswith(".json"):
                    continue
                fpath = os.path.join(target_dir, fname)
                try:
                    if (cutoff - os.path.getmtime(fpath)) / 86400 >= days_old:
                        os.remove(fpath)
                        deleted += 1
                except Exception:
                    pass
        if deleted:
            print(f"  [캐시] 오래된 캐시 {deleted}개 삭제")
    except Exception:
        pass

def cache_show_status() -> None:
    for tier in ["A", "B"]:
        tier_dir = os.path.join(CACHE_DIR, f"tier_{tier}")
        if not os.path.isdir(tier_dir):
            continue
        files = [f for f in os.listdir(tier_dir) if f.endswith(".json")]
        ttl = _TIER_TTL[tier]
        valid = sum(
            1 for f in files
            if (time.time() - os.path.getmtime(os.path.join(tier_dir, f))) / 86400 <= ttl
        )
        label = "DART 재무(7일)" if tier == "A" else "네이버 보조(1일)"
        print(f"  💾 캐시 TIER-{tier} [{label}]: {valid}/{len(files)}개 유효")


# ══════════════════════════════════════════════════════════
# NEW ① 백테스트 엔진 (Walk-forward + IS/OOS 검증)
# ══════════════════════════════════════════════════════════
class BacktestEngine:
    """
    Walk-forward 백테스트 엔진 (KOSPI/KOSDAQ 전용)
    - IS(학습) → OOS(검증) 슬라이딩 반복
    - 벤치마크: KOSPI 지수 (^KS11)
    - 과적합 판단: IS/OOS Sharpe 비율 비교
    """

    def __init__(self, universe_codes: list, start: str = "2020-01-01",
                 end: str = None, is_months: int = 12, oos_months: int = 3,
                 top_n: int = 20, initial_capital: float = 100_000_000,
                 markets: dict = None, cap_weighted_momentum: bool = False,
                 cap_weight_strength: float = 0.5,
                 dart_client=None,
                 fee_pct: float = 0.015, tax_pct: float = 0.18,
                 slippage_pct: float = 0.20):
        self.universe   = universe_codes
        self.start      = pd.Timestamp(start)
        self.end        = pd.Timestamp(end) if end else pd.Timestamp.today()
        self.is_months  = is_months
        self.oos_months = oos_months
        self.top_n      = top_n
        self.capital    = initial_capital
        self.results    = {}
        # {종목코드: "KOSPI"|"KOSDAQ"} — yfinance 접미사(.KS/.KQ) 정확히 매핑하기 위함
        # 안 넘기면 전부 KOSPI로 간주(.KS) — 이전 버전과 동일한 폴백
        self.markets    = markets or {}
        # KOSPI 지수(^KS11) 전체기간 종가 캐시 — 구간마다 재다운로드하지 않도록
        self._bench_series = None
        # ── 대형주 쏠림장 대응 옵션 ──
        # True면 모멘텀 스코어에 시가총액 가중을 곱해 반영
        # (2025년 반도체 슈퍼사이클처럼 소수 대형주가 지수를 끌 때
        #  중소형주 위주 동일가중 모멘텀이 그 흐름을 놓치는 문제 보완)
        self.cap_weighted_momentum = cap_weighted_momentum
        # 가중 강도: 최종 곱셈 범위 = [1-strength, 1+strength]
        #   0.5(기본) → 0.5~1.5배 (완만)
        #   0.7       → 0.3~1.7배 (강함)
        #   1.0       → 0.0~2.0배 (최강, 소형주 점수가 0에 가까워짐)
        self.cap_weight_strength    = max(0.0, min(1.0, cap_weight_strength))
        self._shares_outstanding    = None   # {code: 발행주식수} 캐시
        # DART 발행주식수 수집용 (스크리닝 단계와 동일한 DartClient 재사용 — 캐시도 7일 공유)
        self.dart_client = dart_client

        # ── 거래비용 모델 (수수료 + 매도세 + 슬리피지) ──
        # 백테스트가 실제 매매와 다르게 "비용 0"으로 계산되면 수익률이 비현실적으로
        # 부풀려진다. 편도(매수/매도 각각) 비용을 분리해서 적용한다.
        #   매수비용 = 수수료
        #   매도비용 = 수수료 + 거래세(코스피/코스닥 매도시에만 부과)
        #   슬리피지는 매수·매도 양쪽에 각각 적용 (체결가가 의도한 가격보다 불리하게 밀리는 효과 근사)
        self.fee_pct      = max(0.0, fee_pct)        # 편도 수수료(%), 기본 0.015%
        self.tax_pct      = max(0.0, tax_pct)        # 매도세(%), 기본 0.18% (2025년 기준 코스피/코스닥 공통)
        self.slippage_pct = max(0.0, slippage_pct)   # 편도 슬리피지(%), 기본 0.20%
        self.buy_cost_pct  = self.fee_pct + self.slippage_pct
        self.sell_cost_pct = self.fee_pct + self.tax_pct + self.slippage_pct
        self._cost_warning_printed = False

    def _print_cost_assumptions(self):
        """거래비용 가정과 생존편향 한계를 1회 출력 (결과를 그대로 믿지 않도록)."""
        if self._cost_warning_printed:
            return
        self._cost_warning_printed = True
        rt = self.buy_cost_pct + self.sell_cost_pct
        print(f"  [백테스트] 거래비용 적용: 매수 -{self.buy_cost_pct:.3f}% / "
              f"매도 -{self.sell_cost_pct:.3f}% (수수료{self.fee_pct:.3f}%+세금{self.tax_pct:.2f}%"
              f"+슬리피지{self.slippage_pct:.2f}%) → 왕복 -{rt:.3f}%/회 차감")
        print(f"  ⚠ [한계] 본 백테스트의 종목 유니버스는 '현재 시점' 시가총액 상위 종목으로 구성되어 "
              f"있어 과거 시점에 상장폐지·합병된 종목은 표본에서 빠집니다(생존편향). "
              f"실제 결과는 본 수치보다 낮을 수 있습니다.")

    def _load_shares_outstanding(self) -> dict:
        """
        시가총액 근사용 발행주식수를 DART API로 배치 수집.
        (시총 = 현재가 × 발행주식수)
        DartClient가 없으면 yfinance로 폴백 시도.
        실패한 종목은 dict에서 빠지며, 그 경우 동일가중으로 자동 폴백.
        """
        if self._shares_outstanding is not None:
            return self._shares_outstanding

        print(f"  [백테스트] 시가총액 가중용 발행주식수 수집 중... ({len(self.universe)}개)")
        result = {}

        if self.dart_client is not None:
            try:
                shares_map = self.dart_client.get_shares_bulk(self.universe)
                result = {code: float(v) for code, v in shares_map.items() if v}
            except Exception as e:
                print(f"  ⚠ DART 발행주식수 수집 실패: {e} → yfinance 폴백 시도")

        # DART로 못 채운 종목은 yfinance로 보완 시도 (실패해도 무방, 동일가중 폴백)
        missing = [c for c in self.universe if c not in result]
        if missing and HAS_YFINANCE:
            for code in missing:
                suffix = ".KQ" if "KOSDAQ" in str(self.markets.get(code, "KOSPI")).upper() else ".KS"
                yf_t = f"{code}{suffix}"
                try:
                    t = yf.Ticker(yf_t)
                    shares = t.fast_info.get("shares_outstanding") if hasattr(t, "fast_info") else None
                    if shares and shares > 0:
                        result[code] = float(shares)
                except Exception:
                    pass

        self._shares_outstanding = result
        ok = len(result)
        print(f"  [백테스트] 발행주식수 확보: {ok}/{len(self.universe)}개 "
              f"({'시총가중 적용' if ok > 0 else '데이터 부족 → 동일가중으로 폴백'})")
        return result

    def load_price_data(self) -> pd.DataFrame:
        """yfinance에서 KOSPI/KOSDAQ 종가 로드 (시장구분에 따라 .KS / .KQ 정확히 매핑)"""
        print(f"  [백테스트] 가격 로드 중... ({len(self.universe)}개 종목)")
        cache_key = f"bt_kr_{hashlib.md5(','.join(self.universe[:15]).encode()).hexdigest()[:8]}"
        cached = cache_get(cache_key, tier="A")
        if cached:
            df = pd.DataFrame(cached)
            df.index = pd.to_datetime(df.index)
            if not df.empty:
                print(f"  [백테스트] 캐시 로드: {df.shape}")
                return df
            print("  [백테스트] 캐시가 비어있음 → 다시 다운로드")

        if not HAS_YFINANCE:
            print("  ⚠ yfinance 미설치 → 백테스트 불가 (pip install yfinance)")
            return pd.DataFrame()

        # KOSDAQ은 .KQ, KOSPI(및 정보 없는 종목)는 .KS
        def _suffix(code):
            mkt = self.markets.get(code, "KOSPI")
            return ".KQ" if "KOSDAQ" in str(mkt).upper() else ".KS"

        yf_tickers = [f"{c}{_suffix(c)}" for c in self.universe]
        kospi_n  = sum(1 for c in self.universe if _suffix(c) == ".KS")
        kosdaq_n = len(self.universe) - kospi_n
        print(f"  [백테스트] 시장 구분: KOSPI(.KS) {kospi_n}개 / KOSDAQ(.KQ) {kosdaq_n}개")

        try:
            raw = yf.download(
                yf_tickers,
                start=self.start.strftime("%Y-%m-%d"),
                end=self.end.strftime("%Y-%m-%d"),
                progress=False, group_by="ticker",
                auto_adjust=True, threads=True
            )

            if raw is None or raw.empty:
                print("  ⚠ yfinance가 빈 데이터를 반환했습니다.")
                print("     가능한 원인: (1) 네트워크/방화벽이 finance.yahoo.com을 차단")
                print("                  (2) yfinance 버전이 오래되어 Yahoo 쪽 변경에 대응 못 함")
                print("     확인: pip install --upgrade yfinance 후 재시도")
                print("     단독 테스트: python -c \"import yfinance as yf; "
                      "print(yf.download('005930.KS', period='1mo'))\"")
                return pd.DataFrame()

            close_df = pd.DataFrame()
            fail_examples = []

            def _extract_close(raw_df, code, yf_t):
                """
                yfinance group_by='ticker' 반환 구조는 버전에 따라
                (Ticker, Price) 또는 (Price, Ticker) 순서가 다를 수 있어
                둘 다 시도. 단일 종목이면 MultiIndex가 아예 없을 수도 있음.
                """
                if not isinstance(raw_df.columns, pd.MultiIndex):
                    # 단일 티커 다운로드 시 평범한 컬럼 구조
                    return raw_df["Close"]
                # 진단 결과 기준 정상 순서: (Ticker, Price) → raw[yf_t]["Close"]
                try:
                    return raw_df[yf_t]["Close"]
                except Exception:
                    pass
                # 혹시 반대 순서(Price, Ticker)인 yfinance 버전 대비 폴백
                return raw_df["Close"][yf_t]

            for code, yf_t in zip(self.universe, yf_tickers):
                try:
                    c_s = _extract_close(raw, code, yf_t).dropna()
                    if len(c_s) > 50:
                        close_df[code] = c_s
                    elif len(fail_examples) < 5:
                        fail_examples.append(f"{code}({yf_t}): {len(c_s)}일치만 수집됨")
                except Exception as e:
                    if len(fail_examples) < 5:
                        fail_examples.append(f"{code}({yf_t}): {e}")

            if close_df.empty and fail_examples:
                print("  ⚠ 가격 데이터 없음 — 실패 예시:")
                for ex in fail_examples:
                    print(f"     - {ex}")

            cache_set(cache_key, close_df.to_dict(), tier="A")
            print(f"  [백테스트] 가격 로드 완료: {close_df.shape}")
            return close_df
        except Exception as e:
            print(f"  ⚠ 가격 로드 실패: {e}")
            return pd.DataFrame()

    def compute_factor_scores(self, price_df: pd.DataFrame,
                               as_of: pd.Timestamp,
                               fundamentals: pd.DataFrame = None,
                               enable_pullback: bool = True) -> pd.Series:
        """
        포인트-인-타임 팩터 스코어 (모멘텀 50% + 저변동성 30% + 52주위치 20%)
        cap_weighted_momentum=True 이면 모멘텀 항에 시가총액 가중을 곱해서
        대형주 쏠림장(소수 대형주가 지수를 끄는 구간)에서의 추종력을 보강

        ── 합성 진입설계 (실익 보강) ──
        ① 추세 필터: 20일선 > 60일선 > 120일선 정배열이 아니면 점수 50% 감점
           (가격만 보는 모멘텀의 가짜신호 — 일시적 반등 — 를 줄이기 위함)
        ② 펀더멘털 필터: fundamentals가 주어지면 PER·ROE·부채비율 기준 미달 종목 제외
           (오르고 있지만 망해가는 회사를 걸러내기 위함, compute_scores와 동일 철학)
        ③ 눌림목(Pull-back) 보너스: 정배열 유지 + 20일 고점대비 5~15% 하락한 종목에
           최대 +15점 가산 (추세는 살아있는데 일시적으로 쉬어가는 구간을 매수 우대)
           enable_pullback=False 로 끄면 ①②만 적용된 상태로 A/B 비교 가능
        """
        shares = self._load_shares_outstanding() if self.cap_weighted_momentum else {}

        scores = {}
        raw_mcaps = {}
        shares_used = {}
        pullback_flags = {}   # {code: 고점대비 하락폭(%)} — 눌림목 보너스 받은 종목 추적
        for code in price_df.columns:
            try:
                hist = price_df[code][:as_of].dropna()
                if len(hist) < 120:   # 120일선 계산을 위해 최소 길이 상향
                    continue

                # ── 펀더멘털 필터 (있으면 적용) ──
                if fundamentals is not None and code in fundamentals.index:
                    f = fundamentals.loc[code]
                    per = f.get("PER")
                    roe = f.get("ROE")
                    debt = f.get("부채비율")
                    if pd.notna(per) and per > 0 and per > 30:
                        continue   # 고평가(PER 30배 초과) 제외
                    if pd.notna(roe) and roe < 5:
                        continue   # 저수익(ROE 5% 미만) 제외
                    if pd.notna(debt) and debt > 200:
                        continue   # 과다부채(부채비율 200% 초과) 제외

                p_now = hist.iloc[-1]
                p_1m  = hist.iloc[-21] if len(hist) >= 21 else hist.iloc[0]
                p_12m = hist.iloc[-252] if len(hist) >= 252 else hist.iloc[0]
                mom   = (p_1m / p_12m - 1) * 100 if p_12m > 0 else 0
                vol   = hist.pct_change().dropna().tail(60).std() * np.sqrt(252) * 100
                hi52  = hist.tail(252).max()
                lo52  = hist.tail(252).min()
                pos52 = (p_now - lo52) / (hi52 - lo52) * 100 if hi52 > lo52 else 50

                base_score = mom * 0.5 + (100 - min(vol, 100)) * 0.3 + pos52 * 0.2

                # ── 추세(이동평균 정배열) 필터 ──
                ma20  = hist.tail(20).mean()
                ma60  = hist.tail(60).mean()
                ma120 = hist.tail(120).mean()
                is_golden_aligned = ma20 > ma60 > ma120
                if not is_golden_aligned:
                    base_score *= 0.5   # 정배열 아니면 50% 감점 (제외는 아님, 비중 약화)

                # ── 눌림목(Pull-back) 매수 보너스 ──
                # 정배열(추세 유효)인 종목이 최근 20일 고점에서 5~15% 빠진 상태면
                # "추세는 살아있는데 일시적으로 쉬어가는 구간"으로 보고 가산점.
                # 너무 적게 빠지면(<5%) 아직 고점이라 보너스 없음.
                # 너무 많이 빠지면(>15%) 추세 붕괴 가능성으로 보고 보너스 없음(페널티는 안 줌).
                # 거래량 데이터가 없는 가격 시리즈 한계상, 가격 되돌림 폭만으로 판단.
                if enable_pullback and is_golden_aligned:
                    recent_high_20d = hist.tail(20).max()
                    if recent_high_20d > 0:
                        pullback_pct = (recent_high_20d - p_now) / recent_high_20d * 100
                        if 5.0 <= pullback_pct <= 15.0:
                            # 가장 이상적인 눌림목 구간(고점대비 -5~-15%)일수록 보너스 최대 +15점
                            # 5%에서 0, 10%에서 최대(+15), 15%에서 다시 0이 되는 삼각함수형 보너스
                            depth_score = 1 - abs(pullback_pct - 10.0) / 5.0  # 10%일 때 1.0, 양끝일 때 0
                            base_score += depth_score * 15
                            pullback_flags[code] = round(pullback_pct, 1)

                scores[code] = base_score

                if shares.get(code):
                    raw_mcaps[code] = float(p_now) * shares[code]
                    shares_used[code] = shares[code]
            except Exception:
                pass

        # 이 리밸런싱 시점에서 눌림목 보너스를 받은 종목 기록 (로그/엑셀 참고용)
        if pullback_flags:
            if not hasattr(self, "_pullback_log"):
                self._pullback_log = {}
            self._pullback_log[str(as_of.date())] = dict(pullback_flags)

        # 이 리밸런싱 시점의 현재가·발행주식수·시총을 기록 (엑셀 "시총현황" 시트용)
        # 키: as_of 날짜 문자열 → {code: {price, shares, mcap}}
        if raw_mcaps:
            if not hasattr(self, "_mcap_snapshots"):
                self._mcap_snapshots = {}
            self._mcap_snapshots[str(as_of.date())] = {
                code: {
                    "price":  float(price_df[code][:as_of].dropna().iloc[-1]),
                    "shares": shares_used.get(code),
                    "mcap":   raw_mcaps.get(code),
                }
                for code in scores.keys()
                if code in raw_mcaps
            }

        if not self.cap_weighted_momentum or not raw_mcaps:
            return pd.Series(scores)

        # 시총 가중치: 로그 스케일로 압축 후 0~1 정규화
        # (시총 차이가 수백 배라 선형 가중하면 초대형주가 모든 점수를 압도하므로 완화)
        mcap_series = pd.Series(raw_mcaps)
        log_mcap    = np.log(mcap_series.clip(lower=1))
        norm_weight = (log_mcap - log_mcap.min()) / (log_mcap.max() - log_mcap.min() + 1e-9)

        adjusted = {}
        s = self.cap_weight_strength   # 0.0~1.0
        for code, score in scores.items():
            w = norm_weight.get(code, 0.5)   # 시총 정보 없으면 중간값으로 폴백
            # 가중치 범위: [1-strength, 1+strength]
            #   strength=0.5(기본) → 0.5~1.5배
            #   strength=0.7       → 0.3~1.7배
            #   strength=1.0       → 0.0~2.0배
            mult = (1 - s) + (2 * s * w)
            adjusted[code] = score * mult

        return pd.Series(adjusted)

    def run_pullback_ab_test(self, price_df: pd.DataFrame) -> dict:
        """
        눌림목(Pull-back) 보너스의 효과를 정확히 측정하는 A/B 비교.

        ── run_topn_comparison과의 차이 ──
        run_topn_comparison: top_n만 바꿔서 비교 (다른 실행끼리 비교하면
            KRX 장애 등으로 종목 표본이 미세하게 달라질 수 있어 완전한 비교가 아님)
        run_pullback_ab_test: 완전히 동일한 price_df, top_n, 기간으로
            눌림목 ON/OFF만 다르게 실행 → 표본 차이 없이 순수한 효과만 측정

        반환: {"with_pullback": {...}, "without_pullback": {...}, "diff": {...}}
        """
        print("\n" + "=" * 65)
        print(f"  📊 눌림목(Pull-back) A/B 비교 — top_n={self.top_n} (동일 데이터)")
        print("=" * 65)

        print("\n  ── A: 눌림목 OFF (추세+레짐+펀더멘털만) ──")
        self._enable_pullback = False
        result_off = self.run_walkforward(price_df)

        print("\n  ── B: 눌림목 ON (전체 합성설계) ──")
        self._enable_pullback = True
        result_on = self.run_walkforward(price_df)

        self._enable_pullback = True   # 기본값으로 복원

        s_off = result_off.get("summary", {}) if result_off else {}
        s_on  = result_on.get("summary", {}) if result_on else {}

        diff = {}
        for k in ["CAGR(%)", "Sharpe_OOS", "MDD(%)", "승률(%)", "총수익률(%)"]:
            v_off = s_off.get(k, 0)
            v_on  = s_on.get(k, 0)
            try:
                diff[k] = round(float(v_on) - float(v_off), 2)
            except (TypeError, ValueError):
                diff[k] = None

        print("\n" + "=" * 65)
        print(f"  📊 눌림목 A/B 비교 결과 (동일 데이터·동일 top_n={self.top_n})")
        print("=" * 65)
        print(f"  {'지표':<14} | {'OFF (A)':>10} | {'ON (B)':>10} | {'차이(B-A)':>10}")
        print("  " + "-" * 54)
        for k in ["CAGR(%)", "Sharpe_OOS", "MDD(%)", "승률(%)", "총수익률(%)"]:
            v_off = s_off.get(k, 0)
            v_on  = s_on.get(k, 0)
            d     = diff.get(k)
            d_str = f"{d:+.2f}" if d is not None else "N/A"
            print(f"  {k:<14} | {v_off:>10} | {v_on:>10} | {d_str:>10}")
        print("=" * 65)
        verdict = "✅ 눌림목 도입이 더 좋음" if diff.get("CAGR(%)", 0) > 0 else \
                  "⚠ 눌림목 도입이 더 나쁨" if diff.get("CAGR(%)", 0) < 0 else "─ 차이 없음"
        print(f"  결론(CAGR 기준): {verdict}")
        print("=" * 65)

        return {"with_pullback": result_on, "without_pullback": result_off, "diff": diff}

    def save_pullback_ab_to_excel(self, ab_result: dict) -> str:
        """눌림목 A/B 비교 결과를 엑셀로 저장"""
        if not HAS_OPENPYXL or not ab_result.get("with_pullback"):
            return ""
        fname = datetime.today().strftime("backtest_pullback_ab_%Y%m%d_%H%M%S.xlsx")
        fpath = os.path.join(BT_DIR, fname)
        wb = Workbook()
        KR = "맑은 고딕"

        ws1 = wb.active
        ws1.title = "AB비교요약"
        ws1.merge_cells("A1:D1")
        h = ws1.cell(1, 1, f"눌림목(Pull-back) A/B 비교  {VERSION}")
        h.font = Font(name=KR, bold=True, size=14, color="FFFFFF")
        h.fill = PatternFill("solid", fgColor="1F3864")
        h.alignment = Alignment(horizontal="center")
        ws1.row_dimensions[1].height = 28

        hdrs = ["지표", "OFF(눌림목없음)", "ON(눌림목적용)", "차이"]
        for j, hh in enumerate(hdrs, 1):
            c = ws1.cell(2, j, hh)
            c.font = Font(name=KR, bold=True, color="FFFFFF", size=10)
            c.fill = PatternFill("solid", fgColor="2E75B6")
            c.border = _THIN

        s_off = ab_result["without_pullback"].get("summary", {})
        s_on  = ab_result["with_pullback"].get("summary", {})
        diff  = ab_result.get("diff", {})

        ri = 3
        for k in ["CAGR(%)", "Sharpe_OOS", "MDD(%)", "승률(%)", "총수익률(%)"]:
            ws1.cell(ri, 1, k).font = Font(name=KR, bold=True, size=10)
            ws1.cell(ri, 2, s_off.get(k, 0)).border = _THIN
            ws1.cell(ri, 3, s_on.get(k, 0)).border = _THIN
            d = diff.get(k)
            c4 = ws1.cell(ri, 4, d)
            c4.border = _THIN
            if isinstance(d, (int, float)):
                c4.font = Font(name=KR, bold=True,
                               color="0070C0" if d > 0 else ("C00000" if d < 0 else "808080"))
            ri += 1

        # 구간별 상세 비교 시트
        ws2 = wb.create_sheet("구간별비교")
        plog_off = ab_result["without_pullback"].get("period_log", [])
        plog_on  = ab_result["with_pullback"].get("period_log", [])
        ws2.cell(1, 1, "기간").font = Font(name=KR, bold=True)
        ws2.cell(1, 2, "포트수익_OFF(%)").font = Font(name=KR, bold=True)
        ws2.cell(1, 3, "포트수익_ON(%)").font = Font(name=KR, bold=True)
        ws2.cell(1, 4, "차이").font = Font(name=KR, bold=True)
        for ri2, (p_off, p_on) in enumerate(zip(plog_off, plog_on), 2):
            ws2.cell(ri2, 1, p_off.get("기간"))
            ws2.cell(ri2, 2, p_off.get("포트수익(%)"))
            ws2.cell(ri2, 3, p_on.get("포트수익(%)"))
            try:
                d = round(p_on.get("포트수익(%)", 0) - p_off.get("포트수익(%)", 0), 2)
            except TypeError:
                d = None
            c = ws2.cell(ri2, 4, d)
            if isinstance(d, (int, float)):
                c.font = Font(name=KR, color="0070C0" if d > 0 else ("C00000" if d < 0 else "808080"))

        auto_col_width(ws1)
        auto_col_width(ws2)
        wb.save(fpath)
        print(f"  ✅ 눌림목 A/B 비교 엑셀: {fpath}")
        return fpath

    def run_topn_comparison(self, price_df: pd.DataFrame,
                            top_n_list: list = None) -> dict:
        """
        여러 포트폴리오 집중도(top_n)로 동일 가격데이터에 대해
        Walk-forward를 반복 실행하고 성과를 비교.
        예: top_n_list=[5, 10, 20] → 5종목 집중 vs 10종목 vs 20종목 분산 비교

        반환: {
            "comparison": [{"top_n": 5, "CAGR(%)": ..., "Sharpe_OOS": ..., ...}, ...],
            "details": {5: {...전체 result...}, 10: {...}, 20: {...}}
        }
        """
        top_n_list = top_n_list or [5, 10, 20]
        original_top_n = self.top_n
        comparison = []
        details = {}

        print("\n" + "=" * 65)
        print(f"  📊 포트폴리오 집중도 비교: top_n = {top_n_list}")
        print("=" * 65)

        for n in top_n_list:
            print(f"\n  ── top_n={n} 실행 중 ──")
            self.top_n = n
            result = self.run_walkforward(price_df)
            if result:
                s = result.get("summary", {})
                comparison.append({
                    "top_n":          n,
                    "총수익률(%)":    s.get("총수익률(%)", 0),
                    "CAGR(%)":        s.get("CAGR(%)", 0),
                    "연변동성(%)":    s.get("연변동성(%)", 0),
                    "Sharpe_OOS":     s.get("Sharpe_OOS", 0),
                    "MDD(%)":         s.get("MDD(%)", 0),
                    "승률(%)":        s.get("승률(%)", 0),
                    "과적합판단":     s.get("과적합판단", ""),
                })
                details[n] = result

        self.top_n = original_top_n   # 원래 값 복원

        print("\n" + "=" * 65)
        print(f"  📊 집중도별 비교 요약")
        print("=" * 65)
        if comparison:
            hdr = f"  {'top_n':>6} | {'CAGR(%)':>8} | {'Sharpe':>7} | {'MDD(%)':>7} | {'승률(%)':>7} | 과적합"
            print(hdr)
            print("  " + "-" * (len(hdr) - 2))
            for c in comparison:
                print(f"  {c['top_n']:>6} | {c['CAGR(%)']:>+8.2f} | "
                      f"{c['Sharpe_OOS']:>7.2f} | {c['MDD(%)']:>7.2f} | "
                      f"{c['승률(%)']:>7.1f} | {c['과적합판단']}")
        print("=" * 65)

        return {"comparison": comparison, "details": details}

    def save_comparison_to_excel(self, comparison_result: dict) -> str:
        """top_n 비교 결과를 엑셀로 저장 (시트1: 비교표, 시트2~N: 각 top_n 상세)"""
        if not HAS_OPENPYXL or not comparison_result.get("comparison"):
            return ""

        fname = datetime.today().strftime("backtest_topn_compare_%Y%m%d_%H%M%S.xlsx")
        fpath = os.path.join(BT_DIR, fname)
        wb = Workbook()
        KR = "맑은 고딕"

        ws1 = wb.active
        ws1.title = "집중도비교"
        comp = comparison_result["comparison"]
        hdrs = list(comp[0].keys())
        for j, h in enumerate(hdrs, 1):
            c = ws1.cell(1, j, h)
            c.font   = Font(name=KR, bold=True, color="FFFFFF", size=10)
            c.fill   = PatternFill("solid", fgColor="1F3864")
            c.border = _THIN
            c.alignment = Alignment(horizontal="center")
        for ri, row in enumerate(comp, 2):
            for j, h in enumerate(hdrs, 1):
                v = row.get(h)
                c = ws1.cell(ri, j, v)
                c.font   = Font(name=KR, size=10)
                c.border = _THIN
                c.alignment = Alignment(horizontal="center")
                if h in ("CAGR(%)", "Sharpe_OOS") and isinstance(v, (int, float)):
                    c.font = Font(name=KR, size=10, bold=True,
                                  color="0070C0" if v > 0 else "C00000")

        # 각 top_n 별 구간상세 시트
        for n, detail in comparison_result.get("details", {}).items():
            ws = wb.create_sheet(f"top{n}_상세")
            plog = detail.get("period_log", [])
            if not plog:
                continue
            hdrs2 = list(plog[0].keys())
            for j, h in enumerate(hdrs2, 1):
                c = ws.cell(1, j, h)
                c.font = Font(name=KR, bold=True, color="FFFFFF", size=9)
                c.fill = PatternFill("solid", fgColor="2E75B6")
                c.border = _THIN
            for ri, p in enumerate(plog, 2):
                for j, h in enumerate(hdrs2, 1):
                    ws.cell(ri, j, p.get(h)).border = _THIN

        # 종목별 시총현황 (top_n 비교 실행 동안 누적된 모든 스냅샷)
        self._add_mcap_sheet(wb, KR)

        wb.save(fpath)
        print(f"  ✅ 집중도 비교 엑셀: {fpath}")
        return fpath

    def run_walkforward(self, price_df: pd.DataFrame) -> dict:
        """Walk-forward 검증 실행"""
        self._print_cost_assumptions()
        print(f"\n  [백테스트] Walk-forward 시작")
        print(f"     기간: {self.start.date()} ~ {self.end.date()}")
        print(f"     IS: {self.is_months}개월 / OOS: {self.oos_months}개월 / 상위: {self.top_n}종목")

        rebal_dates = pd.date_range(
            start=self.start + pd.DateOffset(months=self.is_months),
            end=self.end, freq=f"{self.oos_months}MS"
        )

        portfolio_value = self.capital
        equity_curve    = [{"date": str(self.start.date()), "value": portfolio_value}]
        oos_returns, is_sharpes, period_log = [], [], []

        for rebal_date in rebal_dates:
            is_end   = rebal_date - pd.DateOffset(days=1)
            is_start = is_end - pd.DateOffset(months=self.is_months)
            oos_end  = min(rebal_date + pd.DateOffset(months=self.oos_months) - pd.DateOffset(days=1),
                           self.end)

            scores   = self.compute_factor_scores(price_df, is_end, fundamentals=getattr(self, "_fundamentals_cache", None), enable_pullback=getattr(self, "_enable_pullback", True))
            if scores.empty:
                continue

            # ── 시장 레짐 필터 (run_basket_simulation과 동일 규칙) ──
            market_ok = self._is_market_healthy(is_end)
            effective_top_n = self.top_n if market_ok else max(3, self.top_n // 2)
            selected = scores.nlargest(effective_top_n).index.tolist()

            oos_rets = self._period_returns(price_df, selected, rebal_date, oos_end)
            if oos_rets is None or len(oos_rets) < 2:
                continue

            port_ret  = oos_rets.mean(axis=1)
            gross_period_r = float((1 + port_ret).cumprod().iloc[-1] - 1) * 100
            # 매 리밸런싱마다 보유종목을 전량 교체(매도+매수)한다고 가정 → 왕복비용 1회 차감
            period_r  = gross_period_r - (self.buy_cost_pct + self.sell_cost_pct)
            bench_r   = self._benchmark_return(rebal_date, oos_end)
            oos_sharpe= self._sharpe(port_ret)

            is_rets   = self._period_returns(price_df, selected, is_start, is_end)
            is_sharpe = self._sharpe(is_rets.mean(axis=1)) if is_rets is not None else 0

            portfolio_value *= (1 + period_r / 100)
            equity_curve.append({"date": str(oos_end.date()), "value": round(portfolio_value)})
            oos_returns.append(period_r)
            is_sharpes.append(is_sharpe)

            # 이 리밸런싱 시점(is_end)의 선정종목 시총 정보 (있으면)
            snap = getattr(self, "_mcap_snapshots", {}).get(str(is_end.date()), {})
            mcap_lines = []
            for code in selected[:5]:
                info = snap.get(code)
                if info and info.get("mcap"):
                    mcap_eok = info["mcap"] / 1e8   # 원 → 억원
                    mcap_lines.append(f"{code}:{mcap_eok:,.0f}억")
            mcap_str = ", ".join(mcap_lines) if mcap_lines else ""

            period_log.append({
                "기간":       f"{rebal_date.date()}~{oos_end.date()}",
                "시장레짐":   "정상장" if market_ok else "위험장",
                "선정종목":   ", ".join(selected[:5]),
                "선정종목_시총": mcap_str,
                "포트수익(%)":round(period_r, 2),
                "벤치수익(%)":round(bench_r, 2),
                "Alpha(%)":   round(period_r - bench_r, 2),
                "OOS_Sharpe": round(oos_sharpe, 2),
                "IS_Sharpe":  round(is_sharpe, 2),
            })
            print(f"  {rebal_date.date()}~{oos_end.date()} [{'정상장' if market_ok else '⚠위험장'}]: "
                  f"포트 {period_r:+.1f}% / 벤치 {bench_r:+.1f}% / "
                  f"Alpha {period_r-bench_r:+.1f}% / Sharpe {oos_sharpe:.2f}")

        if not oos_returns:
            print("  ⚠ 유효한 OOS 구간 없음")
            return {}

        years = (self.end - self.start).days / 365.25
        cagr  = ((portfolio_value / self.capital) ** (1 / max(years, 0.1)) - 1) * 100
        all_r = np.array(oos_returns) / 100
        ann_vol = np.std(all_r) * np.sqrt(12 / self.oos_months) * 100
        sharpe  = (np.mean(all_r) * 12/self.oos_months) / max(
                   np.std(all_r) * np.sqrt(12/self.oos_months), 0.001)

        equity_vals = [e["value"] for e in equity_curve]
        peak, mdd = equity_vals[0], 0
        for v in equity_vals:
            peak = max(peak, v)
            mdd  = min(mdd, (v - peak) / peak * 100)

        avg_is  = np.mean(is_sharpes)
        overfit = "⚠ 과적합 의심" if sharpe / max(abs(avg_is), 0.01) < 0.5 else "✅ 정상"

        result = {
            "summary": {
                "기간":        f"{self.start.date()} ~ {self.end.date()}",
                "총수익률(%)": round((portfolio_value / self.capital - 1) * 100, 2),
                "CAGR(%)":     round(cagr, 2),
                "연변동성(%)": round(ann_vol, 2),
                "Sharpe_OOS":  round(sharpe, 2),
                "IS_Sharpe":   round(avg_is, 2),
                "MDD(%)":      round(mdd, 2),
                "승률(%)":     round(sum(1 for r in oos_returns if r > 0) / len(oos_returns) * 100, 1),
                "구간수":      len(oos_returns),
                "과적합판단":  overfit,
                "초기자본(원)":self.capital,
                "최종자본(원)":round(portfolio_value),
                "왕복거래비용(%)": round(self.buy_cost_pct + self.sell_cost_pct, 3),
                "주의":        "유니버스가 현재 시총 상위 종목 기준 → 생존편향 있음, 실제는 더 낮을 수 있음",
            },
            "equity_curve": equity_curve,
            "period_log":   period_log,
        }
        self.results = result
        self._print_summary(result)
        return result

    def _is_market_healthy(self, as_of: pd.Timestamp) -> bool:
        """
        시장 레짐 필터: KOSPI 종가가 200일 이동평균선 위면 "정상장", 아래면 "위험장".
        위험장이면 run_basket_simulation/run_walkforward에서 신규 매수를 보류하거나
        종목 수를 줄여서, 시장 전체가 하락하는 구간의 손실을 줄이는 목적.
        """
        bench = self._load_benchmark_series()
        if bench.empty:
            return True   # 벤치마크 없으면 필터 없이 정상장으로 간주(폴백)
        try:
            hist = bench.loc[:as_of].dropna()
            if len(hist) < 200:
                return True   # 200일치가 안 모이면 판단 보류, 정상장으로 간주
            ma200 = hist.tail(200).mean()
            return float(hist.iloc[-1]) >= float(ma200)
        except Exception:
            return True

    def run_basket_simulation(self, price_df: pd.DataFrame,
                              stop_loss_pct: float = 7.0,
                              take_profit_min_pct: float = 5.0,
                              take_profit_max_pct: float = 30.0,
                              rescore_freq_days: int = 21) -> dict:
        """
        실제 KIS 자동매매와 동일한 "포트폴리오(바스켓) 단위 판단"을 백테스트로 재현.

        ── run_walkforward와의 차이 ──
        run_walkforward: 고정된 분기 캘린더로 강제 리밸런싱 (3개월마다 무조건 교체)
        run_basket_simulation: 매일 평가해서, 보유종목 합산수익률이
            손절(-stop_loss_pct%) 또는 익절(+tp_min~+tp_max%) 범위에 들 때만 매도.
            그 사이면 계속 보유 — 실제 KISAutoTrader.execute_signals와 동일한 규칙.

        ── 동작 ──
        1) IS(학습)기간이 끝난 첫날부터 시뮬레이션 시작
        2) 보유 종목이 없는 날: rescore_freq_days(기본 21영업일≈1개월)마다 한 번씩
           그 시점 팩터스코어로 top_n 선정 → 가용 자본을 N분할 균등매수
        3) 보유 종목이 있는 날: 매일 (현재가합-매수가합)/매수가합 으로 합산수익률 계산
           → 손절/익절 범위면 전량매도(다음날 재평가 대상), 범위 밖이면 보유 유지
        4) 매도로 회수된 금액 전체가 그대로 다음 매수의 가용 자본이 됨(복리, 재투자)
        """
        # 입력값 정규화: 손절은 항상 양수(절댓값)로, 익절도 항상 양수로 처리
        # (CLI/대화형 메뉴에서 -3처럼 음수로 들어와도 "--3.0%" 같은 표시 오류 방지)
        stop_loss_pct       = abs(stop_loss_pct)
        take_profit_min_pct = abs(take_profit_min_pct)
        take_profit_max_pct = abs(take_profit_max_pct)

        print(f"\n  [백테스트-바스켓] 포트폴리오 단위 시뮬레이션 시작")
        self._print_cost_assumptions()
        print(f"     기간: {self.start.date()} ~ {self.end.date()}")
        print(f"     손절: -{stop_loss_pct:.1f}% | 익절: +{take_profit_min_pct:.1f}%~+{take_profit_max_pct:.1f}%")
        print(f"     재평가 주기(보유없을때): {rescore_freq_days}영업일 | 상위: {self.top_n}종목")

        stop   = stop_loss_pct / 100
        tp_min = take_profit_min_pct / 100
        tp_max = take_profit_max_pct / 100

        sim_start = self.start + pd.DateOffset(months=self.is_months)
        trading_days = price_df.index[(price_df.index >= sim_start) & (price_df.index <= self.end)]
        if len(trading_days) < 2:
            print("  ⚠ 시뮬레이션 가능한 거래일이 부족합니다.")
            return {}

        capital = self.capital          # 현재 가용 자본(재투자 풀과 동일한 개념)
        positions = {}                  # {code: {"qty": 가상수량비율, "avg_price": 매수가}}
        equity_curve = [{"date": str(sim_start.date()), "value": capital}]
        trade_log = []                   # 매수/매도 이벤트 로그
        last_rescore_idx = -10**9        # 마지막으로 신규 매수했던 거래일 인덱스

        for i, today in enumerate(trading_days):
            if positions:
                # ── 보유 중: 합산 수익률 평가 ──
                cost_sum = sum(p["avg_price"] * p["qty"] for p in positions.values())
                value_sum = 0.0
                missing = False
                for code, p in positions.items():
                    if code in price_df.columns and today in price_df.index:
                        cur = price_df.at[today, code]
                        if pd.isna(cur):
                            missing = True
                            cur = p["avg_price"]
                    else:
                        missing = True
                        cur = p["avg_price"]
                    value_sum += cur * p["qty"]

                basket_ret = (value_sum - cost_sum) / cost_sum if cost_sum > 0 else 0.0

                if basket_ret <= -stop or (tp_min <= basket_ret <= tp_max):
                    action = "STOP_LOSS" if basket_ret <= -stop else "TAKE_PROFIT"
                    capital = value_sum * (1 - self.sell_cost_pct / 100)   # 매도비용 차감 후 재투자
                    trade_log.append({
                        "date": str(today.date()), "action": action,
                        "codes": list(positions.keys()),
                        "ret_pct": round(basket_ret * 100, 2),
                        "capital_after": round(capital),
                    })
                    positions = {}
                    last_rescore_idx = i   # 매도된 날 즉시 재평가 가능하게

                equity_curve.append({"date": str(today.date()), "value": round(capital)})
                continue

            # ── 보유 없음: 재평가 주기 도달 시 신규 매수 ──
            if i - last_rescore_idx < rescore_freq_days and i != 0:
                equity_curve.append({"date": str(today.date()), "value": round(capital)})
                continue

            scores = self.compute_factor_scores(price_df, today, fundamentals=getattr(self, "_fundamentals_cache", None), enable_pullback=getattr(self, "_enable_pullback", True))
            if scores.empty:
                equity_curve.append({"date": str(today.date()), "value": round(capital)})
                continue

            # ── 시장 레짐 필터 ──
            # KOSPI가 200일선 아래(위험장)면 매수 종목 수를 절반으로 줄여
            # 하락장에서의 노출(포지션 크기)을 줄임. 완전 매수중단은 아님(분산은 유지).
            market_ok = self._is_market_healthy(today)
            effective_top_n = self.top_n if market_ok else max(3, self.top_n // 2)

            selected = scores.nlargest(effective_top_n).index.tolist()
            per_stock_budget = capital / max(len(selected), 1)
            effective_budget = per_stock_budget * (1 - self.buy_cost_pct / 100)  # 매수비용 차감 후 실제 매수액

            new_positions = {}
            for code in selected:
                if code not in price_df.columns or today not in price_df.index:
                    continue
                px = price_df.at[today, code]
                if pd.isna(px) or px <= 0:
                    continue
                qty = effective_budget / px   # 백테스트는 가상 비율 수량(소수 허용), 매수비용 차감 반영
                new_positions[code] = {"avg_price": float(px), "qty": qty}

            if new_positions:
                positions = new_positions
                trade_log.append({
                    "date": str(today.date()), "action": "BUY",
                    "codes": list(positions.keys()),
                    "budget": round(capital),
                    "market_regime": "정상장" if market_ok else "위험장(종목수축소)",
                })
                last_rescore_idx = i

            equity_curve.append({"date": str(today.date()), "value": round(capital)})

        # ── 성과 집계 ──
        sell_events = [t for t in trade_log if t["action"] in ("STOP_LOSS", "TAKE_PROFIT")]
        if not sell_events:
            print("  ⚠ 매도(손절/익절) 이벤트가 한 번도 발생하지 않았습니다.")
            print("     stop_loss/take_profit 범위를 넓히거나 기간을 늘려보세요.")

        final_capital = capital if not positions else sum(
            (price_df[code].reindex(trading_days).ffill().iloc[-1] or p["avg_price"]) * p["qty"]
            for code, p in positions.items()
        )

        years = (self.end - sim_start).days / 365.25
        total_ret = (final_capital / self.capital - 1) * 100
        cagr = ((final_capital / self.capital) ** (1 / max(years, 0.1)) - 1) * 100 if final_capital > 0 else -100

        equity_vals = [e["value"] for e in equity_curve]
        peak, mdd = equity_vals[0], 0
        for v in equity_vals:
            peak = max(peak, v)
            mdd = min(mdd, (v - peak) / peak * 100) if peak > 0 else mdd

        win = sum(1 for t in sell_events if t["ret_pct"] > 0)
        win_rate = (win / len(sell_events) * 100) if sell_events else 0.0
        avg_ret_per_trade = np.mean([t["ret_pct"] for t in sell_events]) if sell_events else 0.0

        result = {
            "summary": {
                "기간":          f"{sim_start.date()} ~ {self.end.date()}",
                "총수익률(%)":   round(total_ret, 2),
                "CAGR(%)":       round(cagr, 2),
                "MDD(%)":        round(mdd, 2),
                "매도건수":      len(sell_events),
                "손절건수":      sum(1 for t in sell_events if t["action"] == "STOP_LOSS"),
                "익절건수":      sum(1 for t in sell_events if t["action"] == "TAKE_PROFIT"),
                "승률(%)":       round(win_rate, 1),
                "평균거래수익(%)": round(avg_ret_per_trade, 2),
                "초기자본(원)":  self.capital,
                "최종자본(원)":  round(final_capital),
                "왕복거래비용(%)": round(self.buy_cost_pct + self.sell_cost_pct, 3),
                "주의":          "유니버스가 현재 시총 상위 종목 기준 → 생존편향 있음, 실제는 더 낮을 수 있음",
            },
            "equity_curve": equity_curve,
            "trade_log":    trade_log,
        }

        print("\n" + "=" * 65)
        print(f"  📊 바스켓 시뮬레이션 결과 [{result['summary']['기간']}]")
        print("=" * 65)
        for k, v in result["summary"].items():
            if k in ("초기자본(원)", "최종자본(원)"):
                print(f"  {k:<14}: {int(v):>14,}원")
            else:
                print(f"  {k:<14}: {v}")
        print("=" * 65)

        return result

    def run_grid_search(self, price_df: pd.DataFrame,
                        top_n_list: list = None,
                        stop_loss_list: list = None,
                        take_profit_ranges: list = None) -> dict:
        """
        여러 (top_n, 손절%, 익절범위) 조합을 한 번에 돌려 표로 비교하는 그리드서치.
        run_basket_simulation을 조합 수만큼 반복 실행.

        예) top_n_list=[5,10,20], stop_loss_list=[5,7,10],
            take_profit_ranges=[(2,25),(5,30)]
            → 3×3×2 = 18가지 조합을 전부 실행

        반환: {"results": [{"top_n":.., "stop_loss":.., "tp_min":.., "tp_max":..,
                            "CAGR(%)":.., "Sharpe":.., "MDD(%)":.., "승률(%)":..}, ...]}
        """
        top_n_list         = top_n_list or [5, 10, 20]
        stop_loss_list      = stop_loss_list or [5.0, 7.0, 10.0]
        take_profit_ranges = take_profit_ranges or [(2.0, 25.0), (5.0, 30.0)]

        total_combos = len(top_n_list) * len(stop_loss_list) * len(take_profit_ranges)
        print("\n" + "=" * 70)
        print(f"  📊 그리드서치: top_n×{len(top_n_list)} × 손절×{len(stop_loss_list)} × "
              f"익절범위×{len(take_profit_ranges)} = {total_combos}가지 조합")
        print("=" * 70)

        original_top_n = self.top_n
        results = []
        combo_idx = 0

        for n in top_n_list:
            self.top_n = n
            for sl in stop_loss_list:
                for tp_min, tp_max in take_profit_ranges:
                    combo_idx += 1
                    print(f"\n  ── [{combo_idx}/{total_combos}] top_n={n}, "
                          f"손절=-{sl}%, 익절=+{tp_min}~+{tp_max}% ──")
                    result = self.run_basket_simulation(
                        price_df, stop_loss_pct=sl,
                        take_profit_min_pct=tp_min, take_profit_max_pct=tp_max
                    )
                    if result:
                        s = result.get("summary", {})
                        results.append({
                            "top_n":       n,
                            "손절(%)":     sl,
                            "익절범위":    f"{tp_min}~{tp_max}",
                            "CAGR(%)":     s.get("CAGR(%)", 0),
                            "MDD(%)":      s.get("MDD(%)", 0),
                            "승률(%)":     s.get("승률(%)", 0),
                            "매도건수":    s.get("매도건수", 0),
                            "평균거래수익(%)": s.get("평균거래수익(%)", 0),
                        })

        self.top_n = original_top_n   # 원래 값 복원

        # CAGR 내림차순 정렬해서 가장 좋은 조합이 위로 오게
        results_sorted = sorted(results, key=lambda r: r.get("CAGR(%)", -999), reverse=True)

        print("\n" + "=" * 70)
        print(f"  📊 그리드서치 결과 (CAGR 내림차순, 상위 10개)")
        print("=" * 70)
        hdr = (f"  {'top_n':>5} | {'손절':>6} | {'익절범위':>10} | "
               f"{'CAGR(%)':>8} | {'MDD(%)':>7} | {'승률(%)':>7} | 매도건수")
        print(hdr)
        print("  " + "-" * (len(hdr) - 2))
        for r in results_sorted[:10]:
            print(f"  {r['top_n']:>5} | -{r['손절(%)']:>5} | {r['익절범위']:>10} | "
                  f"{r['CAGR(%)']:>+8.2f} | {r['MDD(%)']:>7.2f} | "
                  f"{r['승률(%)']:>7.1f} | {r['매도건수']:>6}")
        print("=" * 70)
        if results_sorted:
            best = results_sorted[0]
            print(f"  🏆 최고 조합: top_n={best['top_n']}, 손절=-{best['손절(%)']}%, "
                  f"익절={best['익절범위']}% → CAGR {best['CAGR(%)']:+.2f}%")
        print("=" * 70)

        return {"results": results_sorted}

    def save_grid_search_to_excel(self, grid_result: dict) -> str:
        """그리드서치 결과를 엑셀로 저장"""
        if not HAS_OPENPYXL or not grid_result.get("results"):
            return ""
        fname = datetime.today().strftime("backtest_gridsearch_%Y%m%d_%H%M%S.xlsx")
        fpath = os.path.join(BT_DIR, fname)
        wb = Workbook()
        KR = "맑은 고딕"

        ws = wb.active
        ws.title = "그리드서치"
        results = grid_result["results"]
        hdrs = list(results[0].keys())
        for j, h in enumerate(hdrs, 1):
            c = ws.cell(1, j, h)
            c.font   = Font(name=KR, bold=True, color="FFFFFF", size=10)
            c.fill   = PatternFill("solid", fgColor="1F3864")
            c.border = _THIN
        for ri, r in enumerate(results, 2):
            for j, h in enumerate(hdrs, 1):
                v = r.get(h)
                c = ws.cell(ri, j, v)
                c.border = _THIN
                if h == "CAGR(%)" and isinstance(v, (int, float)):
                    c.font = Font(name=KR, bold=True, color="0070C0" if v > 0 else "C00000")

        auto_col_width(ws)
        wb.save(fpath)
        print(f"  ✅ 그리드서치 엑셀: {fpath}")
        return fpath

    def _period_returns(self, price_df, codes, start, end):
        try:
            avail = [c for c in codes if c in price_df.columns]
            if not avail:
                return None
            sub = price_df[avail].loc[start:end].dropna(how="all")
            if len(sub) < 2:
                return None
            return sub.pct_change().dropna(how="all")
        except Exception:
            return None

    def _load_benchmark_series(self) -> pd.Series:
        """KOSPI 지수(^KS11) 전체 기간 종가를 한 번만 다운로드해서 캐시"""
        if self._bench_series is not None:
            return self._bench_series

        try:
            h = yf.download("^KS11",
                            start=self.start.strftime("%Y-%m-%d"),
                            end=(self.end + timedelta(days=5)).strftime("%Y-%m-%d"),
                            progress=False, auto_adjust=True)
            if h is None or h.empty:
                print("  ⚠ 벤치마크(^KS11) 다운로드 결과가 비어있음")
                self._bench_series = pd.Series(dtype=float)
                return self._bench_series

            if isinstance(h.columns, pd.MultiIndex):
                try:
                    c = h["Close"]
                    if isinstance(c, pd.DataFrame):
                        c = c.iloc[:, 0]
                except Exception:
                    c = h.xs("Close", level=-1, axis=1).iloc[:, 0]
            else:
                c = h["Close"]

            self._bench_series = c.dropna()
            print(f"  [백테스트] 벤치마크(KOSPI) 로드 완료: {len(self._bench_series)}일치")
        except Exception as e:
            print(f"  ⚠ 벤치마크 로드 실패: {e} → 모든 구간 벤치마크 0%로 처리됩니다")
            self._bench_series = pd.Series(dtype=float)

        return self._bench_series

    def _benchmark_return(self, start, end) -> float:
        bench = self._load_benchmark_series()
        if bench.empty:
            return 0.0
        try:
            sub = bench.loc[start:end].dropna()
            if len(sub) < 2:
                return 0.0
            first, last = float(sub.iloc[0]), float(sub.iloc[-1])
            if first == 0:
                return 0.0
            return (last / first - 1) * 100
        except Exception as e:
            print(f"  ⚠ 벤치마크 구간 계산 오류({start.date()}~{end.date()}): {e}")
            return 0.0

    def _sharpe(self, returns_series, rf=0.03) -> float:
        try:
            if returns_series is None or len(returns_series) < 2:
                return 0.0
            r  = np.array(returns_series, dtype=float)
            ex = r - rf / 252
            return float(np.mean(ex) / np.std(ex) * np.sqrt(252)) if np.std(ex) > 1e-10 else 0.0
        except Exception:
            return 0.0

    def _print_summary(self, result: dict):
        s = result.get("summary", {})
        print("\n" + "=" * 65)
        print(f"  📊 백테스트 결과 [{s.get('기간','')}]")
        print("=" * 65)
        for k, v in s.items():
            if k in ("초기자본(원)", "최종자본(원)"):
                print(f"  {k:<14}: {int(v):>14,}원")
            elif isinstance(v, float):
                color = "\033[34m" if v > 0 else "\033[31m"
                print(f"  {k:<14}: {color}{v:>+.2f}\033[0m")
            else:
                print(f"  {k:<14}: {v}")
        print("=" * 65)

    def save_to_excel(self, result: dict) -> str:
        if not HAS_OPENPYXL or not result:
            return ""
        fname = datetime.today().strftime("backtest_KR_%Y%m%d_%H%M%S.xlsx")
        fpath = os.path.join(BT_DIR, fname)
        wb = Workbook()
        KR = "맑은 고딕"

        # 시트1: 요약
        ws1 = wb.active
        ws1.title = "백테스트요약"
        ws1.merge_cells("A1:C1")
        h = ws1.cell(1, 1, f"Walk-forward 백테스트 결과  {VERSION}")
        h.font = Font(name=KR, bold=True, size=14, color="FFFFFF")
        h.fill = PatternFill("solid", fgColor="1F3864")
        h.alignment = Alignment(horizontal="center")
        ws1.row_dimensions[1].height = 28

        for ri, (k, v) in enumerate(result.get("summary", {}).items(), 2):
            c1 = ws1.cell(ri, 1, k)
            c1.font = Font(name=KR, bold=True, size=10)
            c2 = ws1.cell(ri, 2, v)
            c2.font = Font(name=KR, size=10)
            c2.border = _THIN
            if any(x in k for x in ("수익률", "CAGR", "Sharpe")):
                try:
                    fv = float(v)
                    c2.font = Font(name=KR, size=10, bold=True,
                                   color="0070C0" if fv > 0 else "C00000")
                except Exception:
                    pass

        # 시트2: 구간별 성과
        ws2 = wb.create_sheet("구간별성과")
        plog = result.get("period_log", [])
        if plog:
            hdrs = list(plog[0].keys())
            for j, h in enumerate(hdrs, 1):
                c = ws2.cell(1, j, h)
                c.font  = Font(name=KR, bold=True, color="FFFFFF", size=9)
                c.fill  = PatternFill("solid", fgColor="1F3864")
                c.border= _THIN
                c.alignment = Alignment(horizontal="center")
            for ri, p in enumerate(plog, 2):
                for j, k in enumerate(hdrs, 1):
                    v = p.get(k)
                    c = ws2.cell(ri, j, v)
                    c.font   = Font(name=KR, size=9)
                    c.border = _THIN
                    c.alignment = Alignment(horizontal="center")
                    if k == "Alpha(%)":
                        try:
                            fv = float(v)
                            c.font = Font(name=KR, size=9, bold=True,
                                          color="0070C0" if fv > 0 else "C00000")
                        except Exception:
                            pass

        # 시트3: 수익곡선
        ws3 = wb.create_sheet("수익곡선")
        ws3.cell(1, 1, "날짜").font = Font(name=KR, bold=True)
        ws3.cell(1, 2, "포트폴리오(원)").font = Font(name=KR, bold=True)
        for ri, e in enumerate(result.get("equity_curve", []), 2):
            ws3.cell(ri, 1, e["date"])
            c = ws3.cell(ri, 2, e["value"])
            c.number_format = "#,##0"

        # 시트4: 종목별 시총현황 (시총가중 모드일 때만)
        self._add_mcap_sheet(wb, KR)

        wb.save(fpath)
        print(f"  ✅ 백테스트 엑셀: {fpath}")
        return fpath

    def save_basket_to_excel(self, result: dict) -> str:
        """바스켓(포트폴리오 단위) 시뮬레이션 결과를 엑셀로 저장"""
        if not HAS_OPENPYXL or not result:
            return ""
        fname = datetime.today().strftime("backtest_basket_KR_%Y%m%d_%H%M%S.xlsx")
        fpath = os.path.join(BT_DIR, fname)
        wb = Workbook()
        KR = "맑은 고딕"

        # 시트1: 요약
        ws1 = wb.active
        ws1.title = "바스켓요약"
        ws1.merge_cells("A1:C1")
        h = ws1.cell(1, 1, f"바스켓(포트폴리오 단위) 백테스트  {VERSION}")
        h.font = Font(name=KR, bold=True, size=14, color="FFFFFF")
        h.fill = PatternFill("solid", fgColor="1F3864")
        h.alignment = Alignment(horizontal="center")
        ws1.row_dimensions[1].height = 28

        for ri, (k, v) in enumerate(result.get("summary", {}).items(), 2):
            c1 = ws1.cell(ri, 1, k)
            c1.font = Font(name=KR, bold=True, size=10)
            c2 = ws1.cell(ri, 2, v)
            c2.font = Font(name=KR, size=10)
            c2.border = _THIN
            if any(x in k for x in ("수익률", "CAGR", "수익")):
                try:
                    fv = float(v)
                    c2.font = Font(name=KR, size=10, bold=True,
                                   color="0070C0" if fv > 0 else "C00000")
                except Exception:
                    pass

        # 시트2: 매매로그 (매수/손절/익절 전체)
        ws2 = wb.create_sheet("매매로그")
        tlog = result.get("trade_log", [])
        if tlog:
            hdrs = ["date", "action", "codes", "ret_pct", "budget", "capital_after"]
            for j, h2 in enumerate(hdrs, 1):
                c = ws2.cell(1, j, h2)
                c.font   = Font(name=KR, bold=True, color="FFFFFF", size=9)
                c.fill   = PatternFill("solid", fgColor="1F3864")
                c.border = _THIN
            for ri, t in enumerate(tlog, 2):
                for j, h2 in enumerate(hdrs, 1):
                    v = t.get(h2)
                    if isinstance(v, list):
                        v = ", ".join(str(x) for x in v)
                    c = ws2.cell(ri, j, v)
                    c.border = _THIN
                    if h2 == "ret_pct" and isinstance(v, (int, float)):
                        c.font = Font(name=KR, size=9, bold=True,
                                      color="0070C0" if v > 0 else "C00000")
            auto_col_width(ws2)

        # 시트3: 자본 변화 곡선
        ws3 = wb.create_sheet("자본변화")
        ws3.cell(1, 1, "날짜").font = Font(name=KR, bold=True)
        ws3.cell(1, 2, "자본(원)").font = Font(name=KR, bold=True)
        for ri, e in enumerate(result.get("equity_curve", []), 2):
            ws3.cell(ri, 1, e["date"])
            c = ws3.cell(ri, 2, e["value"])
            c.number_format = "#,##0"

        wb.save(fpath)
        print(f"  ✅ 바스켓 백테스트 엑셀: {fpath}")
        return fpath

    def _add_mcap_sheet(self, wb, kr_font: str):
        """
        시가총액 가중 모드에서 수집된 종목별 현재가·발행주식수·시총 스냅샷을
        "시총현황" 시트로 정리. cap_weighted_momentum=False면 아무것도 안 함.
        """
        snapshots = getattr(self, "_mcap_snapshots", None)
        if not snapshots:
            return

        ws = wb.create_sheet("시총현황")
        hdrs = ["리밸런싱일", "종목코드", "현재가(원)", "발행주식수", "시가총액(억원)"]
        for j, h in enumerate(hdrs, 1):
            c = ws.cell(1, j, h)
            c.font   = Font(name=kr_font, bold=True, color="FFFFFF", size=10)
            c.fill   = PatternFill("solid", fgColor="1F3864")
            c.border = _THIN
            c.alignment = Alignment(horizontal="center")

        ri = 2
        for date_str in sorted(snapshots.keys()):
            stocks = snapshots[date_str]
            # 시총 내림차순 정렬해서 대형주가 위로 오게
            for code, info in sorted(stocks.items(),
                                     key=lambda kv: kv[1].get("mcap") or 0,
                                     reverse=True):
                ws.cell(ri, 1, date_str).border = _THIN
                ws.cell(ri, 2, code).border = _THIN

                c_price = ws.cell(ri, 3, info.get("price"))
                c_price.number_format = "#,##0"
                c_price.border = _THIN

                c_shares = ws.cell(ri, 4, info.get("shares"))
                c_shares.number_format = "#,##0"
                c_shares.border = _THIN

                mcap = info.get("mcap")
                mcap_eok = round(mcap / 1e8, 1) if mcap else None
                c_mcap = ws.cell(ri, 5, mcap_eok)
                c_mcap.number_format = "#,##0.0"
                c_mcap.border = _THIN

                ri += 1

        auto_col_width(ws)


# ══════════════════════════════════════════════════════════
# NEW ② 팩터 모델 고도화 (Z-score 섹터중립 + 앙상블)
# ══════════════════════════════════════════════════════════
class FactorModel:
    """
    Z-score 표준화 기반 팩터 모델 (한국 주식 전용)
    - 섹터(업종) 내 Z-score → 업종 편향 제거
    - Winsorize ±3σ → 아웃라이어 클리핑
    - 강화복합점수 = 기존점수 70% + Z팩터 30%
    - 선택적 Ridge ML 가중치 학습
    """

    # 팩터 → 컬럼 매핑 (v35 컬럼명 기준)
    FACTOR_COLS = {
        "value":    ["PBR",   "ROE",    "DIV",    "영업이익률"],
        "momentum": ["52주위치", "6개월수익률", "거래량비율VR(%)"],
        "quality":  ["ROA",   "부채비율", "interest_coverage", "altman_z"],
        "growth":   ["매출성장률", "fcf_margin"],
        "cashflow": ["cfo",   "fcf",    "fcf_margin"],
    }
    # 낮을수록 좋은 팩터 (Z-score 반전 대상)
    INVERT_COLS = {"PBR", "부채비율"}

    def __init__(self, sector_neutral: bool = True, winsorize_sigma: float = 3.0,
                 cap_weighted_momentum: bool = False, cap_weight_strength: float = 0.7,
                 enable_pullback: bool = True):
        self.sector_neutral  = sector_neutral
        self.winsorize_sigma = winsorize_sigma
        # ── 실전 확정: 분기리밸런싱 top20 + 시총가중 강도 0.7 + 눌림목 보너스 ──
        # 백테스트(BacktestEngine.compute_factor_scores)와 동일한 가중 공식을
        # 스크리닝 모멘텀 팩터(factor_momentum)에도 적용
        self.cap_weighted_momentum = cap_weighted_momentum
        self.cap_weight_strength   = max(0.0, min(1.0, cap_weight_strength))
        # 눌림목(Pull-back) 보너스: 백테스트는 일별가격으로 20일 고점 대비 하락폭을
        # 계산하지만, 실전 스크리닝은 시계열이 없어 52주고가 대비 하락폭으로 근사.
        # (52주 단위라 백테스트의 20일 단위보다 범위를 5~20%로 약간 넓혀 보정)
        self.enable_pullback = enable_pullback

    def _winsorize(self, s: pd.Series) -> pd.Series:
        mu, sigma = s.mean(), s.std()
        if sigma < 1e-10:
            return s
        return s.clip(mu - self.winsorize_sigma * sigma,
                      mu + self.winsorize_sigma * sigma)

    def _zscore(self, s: pd.Series) -> pd.Series:
        mu, sigma = s.mean(), s.std()
        if sigma < 1e-10:
            return pd.Series(0.0, index=s.index)
        return (s - mu) / sigma

    def _sector_zscore(self, df: pd.DataFrame, col: str) -> pd.Series:
        """업종 내 Z-score. 업종 종목 수 < 3이면 전체 Z-score 사용"""
        result = pd.Series(np.nan, index=df.index)
        has_sector = "업종" in df.columns and self.sector_neutral

        if has_sector:
            for sector, grp in df.groupby("업종"):
                vals = pd.to_numeric(grp[col], errors="coerce").dropna()
                if len(vals) < 3:
                    ref  = pd.to_numeric(df[col], errors="coerce").dropna()
                    z    = self._zscore(self._winsorize(ref))
                    result.loc[grp.index] = z.reindex(grp.index)
                else:
                    z = self._zscore(self._winsorize(vals))
                    result.loc[vals.index] = z.values
        else:
            vals = pd.to_numeric(df[col], errors="coerce").dropna()
            z    = self._zscore(self._winsorize(vals))
            result.loc[vals.index] = z.values

        return result

    def compute_factor_zscores(self, df: pd.DataFrame) -> pd.DataFrame:
        out = df.copy()
        group_scores = {}

        for group, cols in self.FACTOR_COLS.items():
            group_z = []
            for col in cols:
                if col not in df.columns:
                    continue
                z = self._sector_zscore(df, col)
                z_col = f"z_{col}"
                out[z_col] = -z if col in self.INVERT_COLS else z
                group_z.append(out[z_col])

            if group_z:
                out[f"factor_{group}"] = pd.concat(group_z, axis=1).mean(axis=1)
                group_scores[group] = f"factor_{group}"

        # ── 시가총액 가중 모멘텀 (실전 확정: 강도 0.7) ──
        # 대형주가 지수를 끄는 쏠림장(2025년 반도체 슈퍼사이클 등)에서
        # 모멘텀 팩터가 그 흐름을 더 잘 따라가도록 보강 (백테스트와 동일 공식)
        if self.cap_weighted_momentum and "factor_momentum" in out.columns:
            mcap_col = next((c for c in ["시가총액(억)", "시총"] if c in df.columns), None)
            if mcap_col:
                mcap = pd.to_numeric(df[mcap_col], errors="coerce")
                valid = mcap.dropna()
                valid = valid[valid > 0]
                if len(valid) >= 3:
                    log_mcap = np.log(valid)
                    norm_w = (log_mcap - log_mcap.min()) / (log_mcap.max() - log_mcap.min() + 1e-9)
                    s = self.cap_weight_strength
                    # 가중치 범위: [1-s, 1+s] — strength=0.7 → 0.3~1.7배
                    mult = (1 - s) + (2 * s * norm_w)
                    mult = mult.reindex(out.index).fillna(1.0)   # 시총 정보 없는 종목은 가중 없음(1배)
                    out["factor_momentum"] = out["factor_momentum"] * mult
                    print(f"  [팩터모델] 시총가중 모멘텀 적용 (강도 {s}, "
                          f"{len(valid)}/{len(out)}개 종목 시총 확인됨)")
                else:
                    print(f"  ⚠ 시총가중 모멘텀: 유효 시총 데이터 부족({len(valid)}개) → 미적용")
            else:
                print(f"  ⚠ 시총가중 모멘텀: '시가총액(억)' 컬럼 없음 → 미적용")

        avail = [v for v in group_scores.values() if v in out.columns]
        if avail:
            out["factor_composite"]  = out[avail].mean(axis=1)
            out["factor_score_pct"]  = (out["factor_composite"].rank(pct=True) * 100).round(1)

        return out

    def ml_signal(self, df: pd.DataFrame, target_col: str = "6개월수익률") -> pd.Series:
        """Ridge 회귀 팩터 가중치 학습 (백테스트 전용 — 미래 데이터 주의)"""
        if not HAS_SKLEARN:
            print("  ⚠ scikit-learn 미설치 → ML 신호 스킵")
            return pd.Series(dtype=float)
        f_cols = [c for c in df.columns
                  if c.startswith("factor_") and c != "factor_composite"]
        if not f_cols or target_col not in df.columns:
            return pd.Series(dtype=float)
        X  = df[f_cols].fillna(0)
        y  = pd.to_numeric(df[target_col], errors="coerce").fillna(0)
        ok = X.notna().all(axis=1) & y.notna()
        if ok.sum() < 30:
            return pd.Series(dtype=float)
        mdl = Ridge(alpha=1.0).fit(X[ok], y[ok])
        print(f"  [팩터모델] Ridge 학습 ({ok.sum()}개 샘플)")
        for c, w in zip(f_cols, mdl.coef_):
            print(f"     {c.replace('factor_',''):>12}: {w:+.3f}")
        return (pd.Series(mdl.predict(X), index=df.index).rank(pct=True) * 100).round(1)

    def enhance_scores(self, df: pd.DataFrame) -> pd.DataFrame:
        """복합점수 강화: 기존 70% + Z팩터 30% (+ 눌림목 보너스, 실전 확정 적용)"""
        out  = self.compute_factor_zscores(df)
        comp = pd.to_numeric(out.get("복합점수", pd.Series(dtype=float)), errors="coerce").fillna(0)
        zfac = pd.to_numeric(out.get("factor_score_pct", pd.Series(50, index=out.index)),
                             errors="coerce").fillna(50)
        enhanced = comp * 0.70 + zfac * 0.30

        # ── 눌림목(Pull-back) 보너스 (실전 확정: top20+시총가중과 함께 사용) ──
        # 백테스트는 일별가격의 20일 고점 대비 하락폭을 쓰지만, 실전 스크리닝엔
        # 시계열이 없어 52주고가 대비 하락폭으로 근사 (52주 단위라 범위를 5~20%로 보정)
        if self.enable_pullback and "52주고가" in out.columns and "현재가" in out.columns:
            hi52 = pd.to_numeric(out["52주고가"], errors="coerce")
            cur  = pd.to_numeric(out["현재가"], errors="coerce")
            valid = (hi52 > 0) & cur.notna()
            pullback_pct = pd.Series(np.nan, index=out.index)
            pullback_pct.loc[valid] = (hi52[valid] - cur[valid]) / hi52[valid] * 100

            # 5~20% 구간에서 보너스, 12.5%(중간)에서 최대 +15점인 삼각형 보너스
            in_range = (pullback_pct >= 5.0) & (pullback_pct <= 20.0)
            depth_score = (1 - (pullback_pct - 12.5).abs() / 7.5).clip(lower=0)
            bonus = pd.Series(0.0, index=out.index)
            bonus.loc[in_range] = depth_score.loc[in_range] * 15

            out["눌림목_하락폭(%)"] = pullback_pct.round(1)
            out["눌림목_보너스"] = bonus.round(1)
            enhanced = enhanced + bonus
            n_flagged = int(in_range.sum())
            print(f"  [팩터모델] 눌림목 보너스 적용: {n_flagged}/{len(out)}개 종목 "
                  f"(52주고가 대비 5~20% 하락 구간)")

        out["강화복합점수"] = enhanced.round(1)
        print(f"  [팩터모델] 강화복합점수 계산 완료 (기존 70% + Z팩터 30%"
              f"{' + 눌림목보너스' if self.enable_pullback else ''})")
        return out


# ══════════════════════════════════════════════════════════
# NEW ③ KIS API 자동매매 엔진
# (한국투자증권 KIS Developers REST API)
#
#  ★ 64bit Python / VS Code 환경 그대로 사용 가능
#  ★ 별도 설치 없음 — requests 라이브러리만 사용
#
#  설정파일: kis_config.json
#  {
#      "app_key":             "발급받은 앱키",
#      "app_secret":          "발급받은 앱시크릿",
#      "account_no":          "12345678-01",  ← 계좌번호-상품코드
#      "is_real":             false,          ← true=실전, false=모의
#      "max_daily_trades":    10,             ← 일 최대 거래횟수
#      "stop_loss_pct":       7.0,            ← 손절 기준(%) — 보유종목 합산수익률 이하시 전량매도
#      "take_profit_min_pct": 2.0,            ← 익절 시작 기준(%)
#      "take_profit_max_pct": 25.0,           ← 익절 종료 기준(%) — 합산수익률 2~25% 범위면 전량매도 (실전 확정값)
#      "base_invest_amount":  10000000,       ← 최초 매수 예산(원), 재투자 풀이 없을 때 사용
#      "reinvest_mode":       true            ← true면 매도회수금을 다음 매수예산으로 누적 재투자
#  }
#
#  ── 매매 판단 단위: 포트폴리오(바스켓) 전체 ──
#  개별 종목이 아니라 "보유 중인 모든 종목을 하나의 묶음"으로 보고 판단합니다.
#    합산수익률 = (보유종목 현재가×수량의 합 − 매수가×수량의 합) ÷ 매수가×수량의 합
#    합산수익률이 손절/익절 기준에 들면 → 보유 종목 전부를 같은 날 동시에 매도
#    범위 밖이면 → 전부 그대로 보유 (일부만 매도하지 않음)
#  보유 종목이 하나라도 남아있으면 신규 매수를 하지 않고, 바스켓이 통째로
#  비워진 날에만 새로운 시그널로 N종목을 동일 비중으로 매수합니다.
#
#  ── 재투자(복리) 동작 ──
#  reinvest_mode=true 이면 바스켓 전체 매도로 회수된 금액을 trades/reinvest_pool.json 에
#  저장해서, 다음 매수 시점에 그 회수금 전체를 그날의 매수 예산으로 사용합니다.
#  예) 1,000만원 5종목(종목당 200만원) → 합산 +10% 매도 → 회수 1,100만원
#      → 다음 매수 시그널이 4종목이면 1,100만원÷4 = 종목당 275만원으로 매수
#
#  앱키 발급: https://apiportal.koreainvestment.com
# ══════════════════════════════════════════════════════════
class KISAutoTrader:
    """
    한국투자증권 KIS Developers API 자동매매 클라이언트
    - REST API (HTTP) 방식 → 64bit Python 완전 지원
    - 실전투자 / 모의투자 구분
    - 매수·매도·잔고·현재가 조회
    - 포지션 관리, 손절·익절 자동 실행
    - 일일 거래한도 안전장치
    """

    BASE_REAL = "https://openapi.koreainvestment.com:9443"
    BASE_MOCK = "https://openapivts.koreainvestment.com:29443"

    def __init__(self, config_path: str = None):
        self.config_path  = config_path or os.path.join(BASE_DIR, "kis_config.json")
        self.cfg          = {}
        self.token        = None
        self.token_exp    = None
        self.is_real      = False
        self.daily_trades = 0
        self.positions    = {}   # {code: {"qty": N, "avg_price": P, "buy_date": D}}
        self._lock        = threading.Lock()
        # 재투자 자금 풀(원) — "오늘 매수에 쓸 총 자금" (계좌 전체 단위 복리)
        # None이면 아직 초기화 안 됨 → execute_signals에서 base_invest_amount로 첫 세팅
        self._reinvest_pool   = None
        self._reinvest_path   = os.path.join(TRADE_DIR, "reinvest_pool.json")
        # 보유 포지션도 reinvest_pool과 동일하게 파일로 영속화.
        # (이게 없으면 매 실행 프로세스가 "보유 0종목"으로 리셋돼서, 어제 산 종목을
        #  오늘 또 신규매수로 잘못 판단하고, 손절/익절 판단도 그 종목엔 작동하지 않음)
        self._positions_path  = os.path.join(TRADE_DIR, "positions.json")
        self._load_config()
        self._load_reinvest_pool()
        self._load_positions()

    # ── 설정 로드 ──
    def _load_config(self):
        if not os.path.exists(self.config_path):
            self._create_default_config()
            return
        try:
            with open(self.config_path, "r", encoding="utf-8") as f:
                self.cfg = json.load(f)
            self.is_real = self.cfg.get("is_real", False)
            mode = "🔴 실전투자" if self.is_real else "🟡 모의투자"
            print(f"  [KIS] 설정 로드 완료 ({mode})")
            print(f"  [KIS] 계좌: {self.cfg.get('account_no', '')}")
        except Exception as e:
            print(f"  ⚠ KIS 설정 로드 오류: {e}")

    def _create_default_config(self):
        default = {
            "app_key":          "YOUR_APP_KEY",
            "app_secret":       "YOUR_APP_SECRET",
            "account_no":       "12345678-01",
            "is_real":          False,
            "max_daily_trades": 10,
            "stop_loss_pct":    7.0,
            "take_profit_min_pct": 2.0,
            "take_profit_max_pct": 25.0,
            "base_invest_amount": 10_000_000,
            "reinvest_mode":    True,
            "buy_top_n":        20,
        }
        with open(self.config_path, "w", encoding="utf-8") as f:
            json.dump(default, f, ensure_ascii=False, indent=2)
        print(f"  [KIS] 기본 설정파일 생성: {self.config_path}")
        print(f"  → https://apiportal.koreainvestment.com 에서 앱키 발급 후 입력")

    # ── 재투자 자금 풀 영속 저장/로드 ──
    def _load_reinvest_pool(self):
        """
        전날 매도(익절/손절)로 회수된 자금 풀을 파일에서 로드.
        파일이 없으면(첫 실행) None으로 두고, execute_signals에서
        base_invest_amount로 최초 세팅.
        """
        try:
            if os.path.exists(self._reinvest_path):
                with open(self._reinvest_path, "r", encoding="utf-8") as f:
                    data = json.load(f)
                self._reinvest_pool = data.get("pool_won")
                print(f"  [재투자] 자금 풀 로드: {self._reinvest_pool:,.0f}원"
                      if self._reinvest_pool else "  [재투자] 저장된 풀 없음 (최초 실행)")
        except Exception as e:
            print(f"  ⚠ 재투자 풀 로드 오류: {e}")
            self._reinvest_pool = None

    def _save_reinvest_pool(self, pool_won: float, note: str = ""):
        """오늘 매도 후 갱신된 자금 풀을 파일에 저장 (다음 실행에서 이어서 사용)"""
        self._reinvest_pool = pool_won
        try:
            os.makedirs(TRADE_DIR, exist_ok=True)
            with open(self._reinvest_path, "w", encoding="utf-8") as f:
                json.dump({
                    "pool_won": pool_won,
                    "updated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    "note": note,
                }, f, ensure_ascii=False, indent=2)
        except Exception as e:
            print(f"  ⚠ 재투자 풀 저장 오류: {e}")

    # ── 보유 포지션 영속 저장/로드 ──
    def _load_positions(self):
        """직전 실행에서 저장된 보유 포지션을 로드 (프로세스가 바뀌어도 보유내역 유지)."""
        try:
            if os.path.exists(self._positions_path):
                with open(self._positions_path, "r", encoding="utf-8") as f:
                    data = json.load(f)
                self.positions = data.get("positions", {})
                if self.positions:
                    print(f"  [포지션] 로컬 저장분 로드: {len(self.positions)}종목 "
                          f"({', '.join(self.positions.keys())})")
        except Exception as e:
            print(f"  ⚠ 포지션 로드 오류: {e}")

    def _save_positions(self):
        """현재 self.positions를 파일에 저장 (다음 프로세스 실행에서도 이어서 사용)."""
        try:
            os.makedirs(TRADE_DIR, exist_ok=True)
            with open(self._positions_path, "w", encoding="utf-8") as f:
                json.dump({
                    "positions": self.positions,
                    "updated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                }, f, ensure_ascii=False, indent=2)
        except Exception as e:
            print(f"  ⚠ 포지션 저장 오류: {e}")

    def sync_positions_from_broker(self):
        """
        실행 시작 시 KIS 실제 잔고(get_balance의 holdings)를 조회해서
        self.positions를 '진짜 계좌 상태'로 덮어씀.

        로컬 positions.json만 믿으면, 그 파일이 유실되거나(예: 다른 머신/컨테이너에서
        실행) 수동으로 따로 매매한 경우 실제 보유와 어긋날 수 있다. 브로커 응답을
        항상 최종 진실로 삼고, 브로커 조회가 실패했을 때만 로컬 파일 값을 그대로 둔다.
        """
        if not self._is_configured():
            return
        bal = self.get_balance()
        holdings = bal.get("holdings")
        if holdings is None:
            print("  ⚠ 잔고 조회 실패 → 로컬 positions.json 값을 그대로 사용")
            return

        broker_positions = {}
        for h in holdings:
            code = str(h.get("pdno", "")).strip()
            qty  = int(float(h.get("hldg_qty", 0) or 0))
            if not code or qty <= 0:
                continue
            avg_price = float(h.get("pchs_avg_pric", 0) or 0)
            # 매수일자는 브로커가 안 주므로, 기존 로컬 기록이 있으면 그대로 유지
            buy_date = self.positions.get(code, {}).get("buy_date",
                        datetime.today().strftime("%Y-%m-%d"))
            broker_positions[code] = {"qty": qty, "avg_price": round(avg_price),
                                       "buy_date": buy_date}

        if broker_positions != self.positions:
            print(f"  [포지션] 브로커 실잔고로 동기화: {len(broker_positions)}종목 "
                  f"(로컬 기록 {len(self.positions)}종목 → 교체)")
        self.positions = broker_positions
        self._save_positions()

    @property
    def base_url(self):
        return self.BASE_REAL if self.is_real else self.BASE_MOCK

    def _is_configured(self) -> bool:
        key = self.cfg.get("app_key", "")
        return bool(key and key != "YOUR_APP_KEY")

    # ── OAuth2 토큰 발급 ──
    def get_access_token(self) -> Optional[str]:
        """액세스 토큰 발급 (유효기간 24시간, 자동 캐시)"""
        if self.token and self.token_exp and datetime.now() < self.token_exp:
            return self.token
        if not self._is_configured():
            print("  ⚠ KIS 앱키 미설정 → kis_config.json 확인")
            return None
        try:
            resp = requests.post(
                f"{self.base_url}/oauth2/tokenP",
                json={
                    "grant_type": "client_credentials",
                    "appkey":     self.cfg["app_key"],
                    "appsecret":  self.cfg["app_secret"],
                },
                timeout=10
            )
            data = resp.json()
            if "access_token" in data:
                self.token     = data["access_token"]
                self.token_exp = datetime.now() + timedelta(hours=23)
                print(f"  [KIS] 토큰 발급 완료 (유효: {self.token_exp.strftime('%H:%M')}까지)")
                return self.token
            else:
                print(f"  ⚠ KIS 토큰 발급 실패: {data.get('msg1', '')}")
                return None
        except Exception as e:
            print(f"  ⚠ KIS 토큰 요청 오류: {e}")
            return None

    def _headers(self, tr_id: str) -> dict:
        """공통 API 요청 헤더"""
        token = self.get_access_token()
        return {
            "Content-Type":  "application/json; charset=utf-8",
            "authorization": f"Bearer {token}",
            "appkey":        self.cfg.get("app_key", ""),
            "appsecret":     self.cfg.get("app_secret", ""),
            "tr_id":         tr_id,
            "custtype":      "P",
        }

    # ── 잔고 조회 ──
    def get_balance(self) -> dict:
        """주식 잔고 조회 (총평가금액·예수금·순자산)"""
        if not self._is_configured():
            return {"총평가금액": 0, "예수금총액": 0, "순자산": 10_000_000, "holdings": []}

        tr_id   = "TTTC8434R" if self.is_real else "VTTC8434R"
        acc     = self.cfg.get("account_no", "")
        acc_no  = acc[:8] if len(acc) >= 8 else acc
        acc_prd = acc[9:] if len(acc) > 9 else "01"

        try:
            resp = requests.get(
                f"{self.base_url}/uapi/domestic-stock/v1/trading/inquire-balance",
                headers=self._headers(tr_id),
                params={
                    "CANO": acc_no, "ACNT_PRDT_CD": acc_prd,
                    "AFHR_FLPR_YN": "N", "OFL_YN": "",
                    "INQR_DVSN": "02", "UNPR_DVSN": "01",
                    "FUND_STTL_ICLD_YN": "N", "FNCG_AMT_AUTO_RDPT_YN": "N",
                    "PRCS_DVSN": "01", "CTX_AREA_FK100": "", "CTX_AREA_NK100": "",
                },
                timeout=10
            )
            data = resp.json()
            if data.get("rt_cd") == "0":
                out2    = data.get("output2", [{}])
                bal     = out2[0] if out2 else {}
                result  = {
                    "총평가금액": int(bal.get("tot_evlu_amt", 0)),
                    "예수금총액": int(bal.get("dnca_tot_amt", 0)),
                    "순자산":     int(bal.get("nass_amt", 0)),
                    "holdings":   data.get("output1", []),
                }
                print(f"  [KIS] 잔고: 총평가 {result['총평가금액']:,}원 / "
                      f"예수금 {result['예수금총액']:,}원")
                return result
            else:
                print(f"  ⚠ KIS 잔고 조회 실패: {data.get('msg1', '')}")
                return {"총평가금액": 0, "예수금총액": 0, "순자산": 10_000_000, "holdings": []}
        except Exception as e:
            print(f"  ⚠ KIS 잔고 조회 오류: {e}")
            return {"총평가금액": 0, "예수금총액": 0, "순자산": 10_000_000, "holdings": []}

    # ── 현재가 조회 ──
    def get_current_price(self, stock_code: str) -> int:
        """주식 현재가 조회"""
        try:
            resp = requests.get(
                f"{self.base_url}/uapi/domestic-stock/v1/quotations/inquire-price",
                headers=self._headers("FHKST01010100"),
                params={"FID_COND_MRKT_DIV_CODE": "J", "FID_INPUT_ISCD": stock_code},
                timeout=5
            )
            data = resp.json()
            if data.get("rt_cd") == "0":
                return int(data["output"].get("stck_prpr", 0))
        except Exception:
            pass
        # 폴백: 네이버 금융
        return self._price_naver(stock_code)

    def _price_naver(self, stock_code: str) -> int:
        """네이버 금융 현재가 (KIS API 폴백용)"""
        try:
            url  = f"https://finance.naver.com/item/sise.naver?code={stock_code}"
            resp = requests.get(url, headers=NAVER_HEADERS, timeout=5)
            m    = re.search(r'id="_nowVal"[^>]*>([\d,]+)', resp.text)
            if m:
                return int(m.group(1).replace(",", ""))
        except Exception:
            pass
        return 0

    # ── 주문 실행 ──
    def place_order(self, stock_code: str, order_type: str,
                    qty: int, price: int = 0, reason: str = "") -> dict:
        """
        주문 실행
        order_type : "BUY" | "SELL"
        price      : 0 = 시장가, 양수 = 지정가
        """
        with self._lock:
            max_trades = self.cfg.get("max_daily_trades", 10)
            if self.daily_trades >= max_trades:
                msg = f"  ⚠ 일일 거래한도({max_trades}회) 초과 → 주문 차단: {stock_code}"
                print(msg)
                return {"success": False, "msg": msg}

            if not self._is_configured():
                print("  ⚠ KIS 앱키 미설정 → 주문 불가")
                return {"success": False, "msg": "앱키 미설정"}

            # TR ID (실전 / 모의 구분)
            if self.is_real:
                tr_id = "TTTC0802U" if order_type == "BUY" else "TTTC0801U"
            else:
                tr_id = "VTTC0802U" if order_type == "BUY" else "VTTC0801U"

            acc    = self.cfg.get("account_no", "")
            acc_no = acc[:8] if len(acc) >= 8 else acc
            acc_prd= acc[9:] if len(acc) > 9 else "01"

            # 시장가 / 지정가 구분
            ord_dvsn = "01" if price == 0 else "00"   # 01=시장가, 00=지정가

            try:
                resp = requests.post(
                    f"{self.base_url}/uapi/domestic-stock/v1/trading/order-cash",
                    headers=self._headers(tr_id),
                    json={
                        "CANO":         acc_no,
                        "ACNT_PRDT_CD": acc_prd,
                        "PDNO":         stock_code,
                        "ORD_DVSN":     ord_dvsn,
                        "ORD_QTY":      str(qty),
                        "ORD_UNPR":     str(price),
                    },
                    timeout=10
                )
                data = resp.json()

                if data.get("rt_cd") == "0":
                    self.daily_trades += 1
                    order_no  = data.get("output", {}).get("ODNO", "")
                    cur_price = price or self.get_current_price(stock_code)

                    # 포지션 추적
                    self._update_position(stock_code, order_type, qty, cur_price)

                    # 거래 로그 저장
                    self._save_log(stock_code, order_type, qty, cur_price, order_no, reason)

                    mode_str = "🔴실전" if self.is_real else "🟡모의"
                    emoji    = "📈" if order_type == "BUY" else "📉"
                    print(f"  {emoji} [{mode_str}] {order_type} {stock_code} "
                          f"{qty:,}주 @{cur_price:,}원 → 주문번호: {order_no}")
                    return {"success": True, "order_no": order_no, "price": cur_price}
                else:
                    print(f"  ⚠ KIS 주문 실패: {data.get('msg1', '')}")
                    return {"success": False, "msg": data.get("msg1", "")}

            except Exception as e:
                print(f"  ⚠ KIS 주문 오류: {e}")
                return {"success": False, "msg": str(e)}

    def _update_position(self, code, order_type, qty, price):
        if order_type == "BUY":
            if code in self.positions:
                old   = self.positions[code]
                total = old["qty"] + qty
                avg   = (old["avg_price"] * old["qty"] + price * qty) / total
                self.positions[code] = {
                    "qty": total, "avg_price": round(avg),
                    "buy_date": old["buy_date"]
                }
            else:
                self.positions[code] = {
                    "qty": qty, "avg_price": price,
                    "buy_date": datetime.today().strftime("%Y-%m-%d")
                }
        elif order_type == "SELL" and code in self.positions:
            new_qty = self.positions[code]["qty"] - qty
            if new_qty <= 0:
                del self.positions[code]
            else:
                self.positions[code]["qty"] = new_qty
        self._save_positions()   # 주문 직후 즉시 영속화 — 프로세스가 끝나도 보유내역 유지

    def _save_log(self, code, order_type, qty, price, order_no, reason):
        """거래 로그를 trades_KR_YYYYMMDD.json 에 누적 저장"""
        fname = datetime.today().strftime("trades_KR_%Y%m%d.json")
        fpath = os.path.join(TRADE_DIR, fname)
        logs  = []
        if os.path.exists(fpath):
            try:
                with open(fpath, "r", encoding="utf-8") as f:
                    logs = json.load(f)
            except Exception:
                pass
        logs.append({
            "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "type":      order_type,
            "code":      code,
            "qty":       qty,
            "price":     price,
            "amount":    qty * price,
            "order_no":  order_no,
            "reason":    reason,
            "mode":      "REAL" if self.is_real else "MOCK",
        })
        with open(fpath, "w", encoding="utf-8") as f:
            json.dump(logs, f, ensure_ascii=False, indent=2)

    def _check_market_regime(self) -> bool:
        """
        시장 레짐 필터: KOSPI 현재가가 200일 이동평균선 위인지 확인.
        True=정상장, False=위험장 (위험장이면 execute_signals에서 매수 종목 수를 줄임).
        yfinance 호출 실패 시 안전하게 True(정상장)로 폴백 — 매매 자체를 막지 않음.
        """
        if not HAS_YFINANCE:
            return True
        try:
            hist = yf.download("^KS11", period="300d", progress=False, auto_adjust=True)
            if hist is None or hist.empty:
                return True
            if isinstance(hist.columns, pd.MultiIndex):
                try:
                    c = hist["Close"]
                    if isinstance(c, pd.DataFrame):
                        c = c.iloc[:, 0]
                except Exception:
                    c = hist.xs("Close", level=-1, axis=1).iloc[:, 0]
            else:
                c = hist["Close"]
            c = c.dropna()
            if len(c) < 200:
                return True
            ma200 = c.tail(200).mean()
            is_healthy = float(c.iloc[-1]) >= float(ma200)
            print(f"  [시장레짐] KOSPI {'정상장(200일선 위)' if is_healthy else '⚠위험장(200일선 아래)'}")
            return is_healthy
        except Exception as e:
            print(f"  ⚠ 시장레짐 확인 실패({e}) → 정상장으로 간주")
            return True

    # ── 시그널 실행 ──
    def execute_signals(self, df_signals: pd.DataFrame, total_capital: int) -> list:
        """
        스크리닝 매매 시그널 자동 실행 — 종목별 개별 판단 매매 (실전 확정 방식)

        ── 핵심 로직: 각 종목을 독립적으로 판단 ──
        ① 보유 중인 종목 각각에 대해 (현재가-매수가)/매수가 수익률을 계산
        ② 종목별로
             -stop_loss_pct 이하        → 그 종목만 매도(손절)
             +tp_min ~ +tp_max 범위     → 그 종목만 매도(익절)
             그 사이                    → 그 종목만 계속 보유
           (다른 종목의 매도/보유 상태와 무관하게 독립적으로 결정)
        ② 매도된 종목의 회수금은 즉시 재투자 풀에 더해짐
        ③ 매수는 매매시그널 상위 top_n(기본 10)종목 중, 아직 보유하지 않은 종목만 신규 매수
           (스크리닝 단계에서 이미 top_n=10 + 시총가중 강도 0.7이 반영된 결과를 그대로 사용)

        ── 재투자 풀(계좌 단위 복리) ──
        예) base_invest_amount = 1,000만원, 매수 시그널 10종목 중 보유 0개
            1일차: 1,000만원 ÷ 10 = 종목당 100만원 매수
            n일차: A종목이 +10% 익절 → A만 매도, 회수금 110만원 → 풀에 더해짐
                   (B~J 종목은 그대로 보유, 매도 안 됨)
            다음날: 매수 시그널에 새 종목이 들어오면 그 풀(이번엔 110만원)이 매수 예산에 합산됨
        """
        if not self._is_configured():
            print("  ⚠ KIS 앱키 미설정 → 자동매매 중단")
            return []

        # 실행마다 새 프로세스로 떠도(예: GitHub Actions) 실제 보유내역을 알 수 있도록
        # 브로커 잔고를 먼저 조회해 self.positions를 진짜 계좌 상태로 맞춘다.
        self.sync_positions_from_broker()

        # kis_config.json에 실수로 음수가 입력돼도(예: -3.0) 항상 절댓값으로 처리
        stop      = abs(self.cfg.get("stop_loss_pct", 7.0)) / 100
        tp_min    = abs(self.cfg.get("take_profit_min_pct", 2.0)) / 100
        tp_max    = abs(self.cfg.get("take_profit_max_pct", 25.0)) / 100
        base_amt  = self.cfg.get("base_invest_amount", 10_000_000)
        reinvest_on = self.cfg.get("reinvest_mode", True)
        buy_top_n = self.cfg.get("buy_top_n", 20)   # 실전 확정: top20 (시총가중+눌림목 적용)
        executed  = []

        # ── KIS Rate Limit: 모의투자 초당 5건 / 실전 초당 20건 (운영사 공지 기준) ──
        # 호출당(현재가조회 + 주문) 2건이 나갈 수 있으므로 넉넉하게 종목당 1.2초 간격
        api_delay = 0.4 if self.is_real else 1.2

        # ① 종목별 개별 손절·익절 판단
        recovered_today = 0.0

        if self.positions:
            print(f"\n  [KIS] 포지션 점검 (보유 {len(self.positions)}종목, 종목별 개별 판단)...")
            print(f"     손절: -{stop*100:.1f}% | 익절: +{tp_min*100:.1f}%~+{tp_max*100:.1f}% (종목별 독립 판단)")

            for code, pos in list(self.positions.items()):
                cur = self.get_current_price(code)
                time.sleep(api_delay)
                if cur <= 0:
                    continue
                ret = (cur - pos["avg_price"]) / pos["avg_price"] if pos["avg_price"] > 0 else 0.0

                sell_reason = None
                if ret <= -stop:
                    sell_reason = f"손절 {ret*100:+.2f}%"
                elif tp_min <= ret <= tp_max:
                    sell_reason = f"익절 {ret*100:+.2f}%"

                if sell_reason:
                    print(f"  {'🔴' if ret <= -stop else '🟢'} {code} {sell_reason} → 매도")
                    r = self.place_order(code, "SELL", pos["qty"], reason=sell_reason)
                    if r.get("success"):
                        sell_value = pos["qty"] * cur
                        recovered_today += sell_value
                        action = "STOP_LOSS" if ret <= -stop else "TAKE_PROFIT"
                        executed.append({"action": action, "code": code,
                                         "ret_pct": round(ret * 100, 1),
                                         "recovered_won": round(sell_value)})
                    time.sleep(api_delay)

            if not executed:
                print(f"     판단: 전 종목 범위 밖 → 보유 유지 (매도 없음)")

        # ── 재투자 풀 갱신 ──
        # 오늘 매도로 회수된 금액이 있으면 풀에 더함. 풀이 아예 None(최초 실행)이면 base_amt로 시작.
        if recovered_today > 0:
            self._reinvest_pool = (self._reinvest_pool or 0.0) + recovered_today
            self._save_reinvest_pool(
                self._reinvest_pool,
                note=f"종목별 매도 회수 {recovered_today:,.0f}원 누적"
            )
            print(f"  💰 매도 회수: {recovered_today:,.0f}원 → 재투자 풀: {self._reinvest_pool:,.0f}원")
        elif self._reinvest_pool is None:
            self._reinvest_pool = base_amt

        # ② 신규 매수 — 매매시그널 상위 top_n종목 중 미보유 종목만 매수
        if "매매시그널" not in df_signals.columns or df_signals.empty:
            print(f"  [KIS] 자동매매 완료: {len(executed)}건")
            return executed

        # ── 시장 레짐 필터: 위험장(KOSPI 200일선 아래)이면 매수 종목 수를 절반으로 축소 ──
        market_ok = self._check_market_regime()
        effective_top_n = buy_top_n if market_ok else max(3, buy_top_n // 2)

        # 빈 DataFrame이거나 컬럼이 float dtype(전부 NaN)인 경우를 방어
        sig_series = df_signals["매매시그널"].astype(str)
        buy_df = df_signals[sig_series.str.contains("매수", na=False)].head(effective_top_n)
        # 이미 보유 중인 종목은 매수 후보에서 제외하고 카운트
        new_targets = [c for c in buy_df.index if str(c) not in self.positions]
        print(f"  [KIS] 매수 시그널 top{effective_top_n}"
              f"{' (위험장 축소)' if not market_ok else ''} 중 신규 매수 대상 {len(new_targets)}개 처리 중...")

        if not new_targets:
            print(f"  [KIS] 자동매매 완료: {len(executed)}건")
            return executed

        # 오늘 매수에 쓸 총 예산 결정 (신규 매수 대상 종목 수로 분할)
        if reinvest_on:
            invest_budget = self._reinvest_pool if self._reinvest_pool else base_amt
            print(f"  [재투자] 오늘 매수 예산: {invest_budget:,.0f}원 "
                  f"({'재투자 풀' if self._reinvest_pool else '기본금액(최초실행)'})")
        else:
            invest_budget = base_amt

        per_stock_budget = invest_budget / len(new_targets)
        print(f"  [재투자] 종목당 매수 배분: {per_stock_budget:,.0f}원 "
              f"({invest_budget:,.0f}원 ÷ {len(new_targets)}종목)")

        spent_today = 0.0
        for code in new_targets:
            str_code = str(code)
            row = df_signals.loc[code]

            cur = self.get_current_price(str_code)
            time.sleep(api_delay)   # 현재가조회 직후 대기 — 바로 이어지는 주문 호출과의 간격 확보
            if cur <= 0:
                continue

            qty = max(1, int(per_stock_budget / cur))
            sig = str(row.get("매매시그널", ""))
            rsn = (f"{sig} | "
                   f"복합:{row.get('복합점수', 0):.1f} | "
                   f"100점:{row.get('100점_합계', 0):.0f} | "
                   f"괴리율:{row.get('괴리율(%)', 'N/A')}")

            r = self.place_order(str_code, "BUY", qty, reason=rsn)
            if r.get("success"):
                spent_today += qty * r.get("price", cur)
                executed.append({"action": "BUY", "code": str_code,
                                  "qty": qty, "price": r.get("price")})
            time.sleep(api_delay)   # 현재가조회+주문 2건 소진했으므로 더 길게 대기

        # 오늘 매수에 실제로 쓴 금액만큼 풀에서 차감
        # (매수에 안 쓴 잔액은 풀에 남아 다음날에도 계속 누적됨)
        if reinvest_on and spent_today > 0:
            self._reinvest_pool = max(0.0, self._reinvest_pool - spent_today)
            self._save_reinvest_pool(
                self._reinvest_pool,
                note=f"매수 집행 {spent_today:,.0f}원 차감"
            )

        print(f"  [KIS] 자동매매 완료: {len(executed)}건")
        return executed

    def reset_daily(self):
        """일일 거래 카운터 리셋 (매일 장 시작 전 호출)"""
        self.daily_trades = 0
        print(f"  [KIS] 일일 거래카운터 리셋 ({datetime.now().strftime('%Y-%m-%d')})")

    def status_summary(self) -> str:
        mode = "🔴실전" if self.is_real else "🟡모의"
        cfg_ok = "✅설정됨" if self._is_configured() else "❌미설정(kis_config.json 확인)"
        return (f"KIS API {mode} {cfg_ok} | "
                f"계좌:{self.cfg.get('account_no', '미설정')} | "
                f"오늘 거래:{self.daily_trades}회")


# ══════════════════════════════════════════════════════════
# NEW ④ 실시간 모니터링 & 텔레그램 알림
# ══════════════════════════════════════════════════════════
class MonitorEngine:
    """
    텔레그램 알림 + 성과 추적 엔진

    설정파일: telegram_config.json
    {
        "bot_token": "YOUR_BOT_TOKEN",
        "chat_id":   "YOUR_CHAT_ID"
    }

    봇 설정 방법:
      1. 텔레그램에서 @BotFather → /newbot → 토큰 복사
      2. 봇에게 메시지를 먼저 보낸 뒤
         https://api.telegram.org/bot<TOKEN>/getUpdates 에서 chat_id 확인
    """

    def __init__(self, config_path: str = None):
        self.config_path  = config_path or os.path.join(BASE_DIR, "telegram_config.json")
        self.bot_token    = ""
        self.chat_id      = ""
        self.enabled      = False
        self.perf_history = []
        self._load_config()
        self._load_perf_history()

    def _load_config(self):
        if not os.path.exists(self.config_path):
            self._create_default_config()
            return
        try:
            with open(self.config_path, "r", encoding="utf-8") as f:
                cfg = json.load(f)
            self.bot_token = cfg.get("bot_token", "")
            self.chat_id   = cfg.get("chat_id",   "")
            self.enabled   = bool(self.bot_token and
                                  self.bot_token != "YOUR_BOT_TOKEN")
            print(f"  [텔레그램] {'✅ 활성' if self.enabled else '❌ 비활성 (설정 필요)'}")
        except Exception as e:
            print(f"  ⚠ 텔레그램 설정 오류: {e}")

    def _create_default_config(self):
        default = {"bot_token": "YOUR_BOT_TOKEN", "chat_id": "YOUR_CHAT_ID"}
        with open(self.config_path, "w", encoding="utf-8") as f:
            json.dump(default, f, ensure_ascii=False, indent=2)
        print(f"  [텔레그램] 설정파일 생성: {self.config_path}")

    def _load_perf_history(self):
        path = os.path.join(LOG_DIR, "perf_history.json")
        if os.path.exists(path):
            try:
                with open(path, "r", encoding="utf-8") as f:
                    self.perf_history = json.load(f)
            except Exception:
                self.perf_history = []

    def _save_perf_history(self):
        path = os.path.join(LOG_DIR, "perf_history.json")
        with open(path, "w", encoding="utf-8") as f:
            json.dump(self.perf_history[-365:], f, ensure_ascii=False, indent=2)

    def send(self, text: str) -> bool:
        """텔레그램 메시지 전송"""
        if not self.enabled:
            print(f"  [텔레그램 비활성] {text[:80]}...")
            return False
        url = f"https://api.telegram.org/bot{self.bot_token}/sendMessage"
        try:
            r = requests.post(url, json={
                "chat_id": self.chat_id, "text": text, "parse_mode": "HTML"
            }, timeout=10)
            return r.status_code == 200
        except Exception as e:
            print(f"  ⚠ 텔레그램 전송 실패: {e}")
            return False

    def notify_screening(self, df_top: pd.DataFrame) -> None:
        """스크리닝 결과 알림 (상위 10종목)"""
        today = datetime.today().strftime("%Y-%m-%d")
        lines = [
            f"📊 <b>퀀트 스크리너 {VERSION} [KR]</b>",
            f"🗓 {today}",
            "━━━━━━━━━━━━━━━━━━━━",
        ]
        for rank, (code, row) in enumerate(df_top.head(10).iterrows(), 1):
            name   = str(row.get("종목명", ""))[:8]
            score  = _safe(row.get("강화복합점수") or row.get("복합점수"), 0)
            pt100  = _safe(row.get("100점_합계"), 0)
            sig    = str(row.get("매매시그널", "─"))
            upside = _safe(row.get("괴리율(%)"))

            if "강력매수" in sig:   em = "🔵"
            elif "■■ 매수" in sig:  em = "🟢"
            elif "관심" in sig:     em = "🟡"
            elif "매도" in sig:     em = "🔴"
            else:                   em = "⬜"

            up_str = f"({upside:+.0f}%)" if upside is not None else ""
            lines.append(
                f"{rank:2d}. {em} <b>{name}</b>({code}) "
                f"복합:{score:.0f} 100점:{pt100:.0f} {up_str}"
            )

        lines.append("━━━━━━━━━━━━━━━━━━━━")

        # 강력매수 강조
        if "매매시그널" in df_top.columns:
            sb = df_top[df_top["매매시그널"].str.contains("강력매수", na=False)]
            if not sb.empty:
                lines.append(f"🔵 <b>강력매수 {len(sb)}종목</b>")
                for code, row in sb.iterrows():
                    name = str(row.get("종목명", ""))[:8]
                    tgt  = _safe(row.get("1차목표가"))
                    stp  = _safe(row.get("손절선"))
                    if tgt and stp:
                        lines.append(f"   {name}({code}) → "
                                     f"목표:{int(tgt):,}원 / 손절:{int(stp):,}원")
                    else:
                        lines.append(f"   {name}({code})")

        self.send("\n".join(lines))

    def notify_positions(self, trader: "KISAutoTrader") -> None:
        """포지션 현황 알림"""
        if not trader.positions:
            self.send("📋 현재 보유 포지션 없음")
            return
        lines = [f"📋 <b>포지션 현황</b> ({datetime.now().strftime('%H:%M')})",
                 "━━━━━━━━━━━━━━━━━━━━"]
        total_pnl = 0
        for code, pos in trader.positions.items():
            cur = trader.get_current_price(code)
            if cur > 0:
                ret     = (cur - pos["avg_price"]) / pos["avg_price"] * 100
                pnl_won = (cur - pos["avg_price"]) * pos["qty"]
                total_pnl += pnl_won
                em = "📈" if ret >= 0 else "📉"
                lines.append(f"{em} {code}: {pos['qty']}주 "
                             f"@{pos['avg_price']:,}→{cur:,}원 "
                             f"({ret:+.1f}%, {pnl_won:+,.0f}원)")
        lines.append("━━━━━━━━━━━━━━━━━━━━")
        lines.append(f"💰 총 평가손익: {total_pnl:+,.0f}원")
        self.send("\n".join(lines))

    def notify_trade(self, executed: list, reinvest_pool: float = None) -> None:
        """자동매매 실행 결과 알림"""
        if not executed:
            return
        lines = ["🤖 <b>자동매매 실행 결과</b>"]
        for e in executed:
            action = e.get("action", "")
            code   = e.get("code", "")
            if action == "BUY":
                lines.append(f"  📈 매수: {code} "
                             f"{e.get('qty',0)}주 @{e.get('price',0):,}원")
            elif action == "STOP_LOSS":
                rec = e.get("recovered_won")
                rec_str = f" → 회수 {rec:,.0f}원" if rec else ""
                lines.append(f"  🔴 손절: {code} ({e.get('ret_pct',0):+.1f}%){rec_str}")
            elif action == "TAKE_PROFIT":
                rec = e.get("recovered_won")
                rec_str = f" → 회수 {rec:,.0f}원" if rec else ""
                lines.append(f"  🟢 익절: {code} ({e.get('ret_pct',0):+.1f}%){rec_str}")
        if reinvest_pool is not None:
            lines.append(f"\n  💰 다음 매수 예산(재투자 풀): {reinvest_pool:,.0f}원")
        self.send("\n".join(lines))

    def notify_signal_change(self, df_today: pd.DataFrame,
                              df_prev: Optional[pd.DataFrame]) -> None:
        """전일 대비 시그널 변경 감지"""
        if df_prev is None or df_prev.empty:
            return
        changes = []
        for code in df_today.index:
            if code not in df_prev.index:
                continue
            s_now  = str(df_today.loc[code, "매매시그널"]
                         if "매매시그널" in df_today.columns else "")
            s_prev = str(df_prev.loc[code, "매매시그널"]
                         if "매매시그널" in df_prev.columns else "")
            if s_now and s_now != s_prev:
                name = str(df_today.loc[code, "종목명"]
                           if "종목명" in df_today.columns else code)
                changes.append(f"  {name}({code}): {s_prev} → <b>{s_now}</b>")
        if changes:
            self.send("⚡ <b>시그널 변경</b>\n" + "\n".join(changes[:10]))

    def track(self, df_top: pd.DataFrame) -> None:
        """성과 히스토리 기록 + 주간 리포트 (월요일)"""
        top10 = df_top.index[:10].tolist()
        scores = [_safe(df_top.loc[c, "강화복합점수"]
                        if "강화복합점수" in df_top.columns
                        else df_top.loc[c, "복합점수"], 0)
                  for c in top10 if c in df_top.index]
        entry = {
            "date":     datetime.today().strftime("%Y-%m-%d"),
            "top10":    [str(c) for c in top10],
            "avg_score":round(float(np.mean(scores)), 1) if scores else 0,
            "강력매수": int(df_top["매매시그널"].str.contains("강력매수", na=False).sum())
                        if "매매시그널" in df_top.columns else 0,
        }
        self.perf_history.append(entry)
        self._save_perf_history()

        if datetime.today().weekday() == 0:   # 월요일 → 주간 리포트
            self._weekly_report()

    def _weekly_report(self):
        week_ago = (datetime.today() - timedelta(days=7)).strftime("%Y-%m-%d")
        recent   = [e for e in self.perf_history if e["date"] >= week_ago]
        if not recent:
            return
        avg_score  = np.mean([e.get("avg_score", 0) for e in recent])
        avg_strong = np.mean([e.get("강력매수", 0) for e in recent])
        self.send(
            f"📅 <b>주간 리포트 [KR]</b>\n"
            f"기간: {week_ago} ~ {datetime.today().strftime('%Y-%m-%d')}\n"
            f"━━━━━━━━━━━━━━━━━━━━\n"
            f"평균 복합점수: {avg_score:.1f}\n"
            f"일평균 강력매수: {avg_strong:.1f}개\n"
            f"스크리닝 실행: {len(recent)}회"
        )

    def load_yesterday(self) -> Optional[pd.DataFrame]:
        """전일 매매신호 JSON 로드"""
        yesterday = (datetime.today() - timedelta(days=1)).strftime("%Y%m%d")
        fpath = os.path.join(BASE_DIR, f"매매신호_KR_{yesterday}.json")
        if not os.path.exists(fpath):
            return None
        try:
            with open(fpath, "r", encoding="utf-8") as f:
                data = json.load(f)
            rows = {s["code"]: {
                "종목명":     s.get("name", ""),
                "매매시그널": s.get("trading", {}).get("signal", ""),
                "복합점수":   s.get("composite_score", 0),
            } for s in data.get("top_stocks", [])}
            return pd.DataFrame.from_dict(rows, orient="index") if rows else None
        except Exception:
            return None



# ══════════════════════════════════════════════════════════
# v35 핵심 스크리닝 엔진 (단일 파일로 통합됨)
#
#  ★ 원래는 별도 파일(quant_screener_v35.py)을 런타임에 동적
#    임포트하는 구조였으나, "v35 파일을 빠뜨리면 조용히 빈
#    결과로 끝나는" 운영 사고를 막기 위해 v36 본체에 직접 병합함.
#
#  ★ 이 블록에 포함된 것: DART 재무수집(DartClient), 네이버
#    보조지표 크롤링, yfinance 가격/기술지표, 7팩터 점수계산,
#    100점 평가, 매매시그널, 하드필터, 엑셀 5시트 출력, JSON
#    매매신호 저장 등 — 미국(S&P500) 관련 함수만 제외됨.
# ══════════════════════════════════════════════════════════
HAS_V35 = True   # 병합되었으므로 항상 사용 가능

def _print_pkg_status():
    missing = []
    if not HAS_BS4:      missing.append("beautifulsoup4  (네이버 스크래핑)")
    if not HAS_YFINANCE: missing.append("yfinance        (가격/모멘텀)")
    if not HAS_FDR:      missing.append("finance-datareader (유니버스 수집 1순위)")
    if not HAS_OPENPYXL: missing.append("openpyxl        (엑셀 저장)")
    if not HAS_TQDM:     missing.append("tqdm            (진행률 바)")
    if missing:
        print("\n  ⚠ 미설치 패키지 감지 → 해당 기능은 중립값/폴백으로 대체됩니다.")
        print("    pip install " + " ".join([m.split()[0] for m in missing]))
        for m in missing:
            print(f"    - {m}")
        print()

# ══════════════════════════════════════════════════════════
# 0. 상수 / 스타일 / 설정
DART_ACCOUNTS = {
    "영업활동현금흐름": [
        "ifrs-full_CashFlowsFromUsedInOperatingActivities",
        "dart_CashFlowsFromOperatingActivities",
        "영업활동으로인한현금흐름", "영업활동현금흐름", "영업활동으로인현금흐름",
    ],
    "투자활동현금흐름": [
        "ifrs-full_CashFlowsFromUsedInInvestingActivities",
        "dart_CashFlowsFromInvestingActivities",
        "투자활동으로인한현금흐름", "투자활동현금흐름", "투자활동으로인현금흐름",
    ],
    "당기순이익": [
        "ifrs-full_ProfitLoss", "dart_ProfitLoss",
        "당기순이익", "당기순손익", "연결당기순이익", "당기순이익(손실)",
    ],
    "매출총이익": [
        "ifrs-full_GrossProfit", "dart_GrossProfit",
        "매출총이익", "매출총손익",
    ],
    "영업이익": [
        "ifrs-full_ProfitLossFromOperatingActivities",
        "dart_OperatingIncomeLoss",
        "영업이익", "영업손익", "영업이익(손실)",
    ],
    "매출액": [
        "ifrs-full_Revenue", "dart_Revenue",
        "매출액", "수익(매출액)", "영업수익", "매출",
    ],
    "이자비용": [
        "ifrs-full_FinanceCosts", "dart_FinanceCosts",
        "이자비용", "금융비용", "이자비용합계",
    ],
    "감가상각비": [
        "ifrs-full_DepreciationAndAmortisationExpense",
        "dart_DepreciationAndAmortisation",
        "감가상각비", "감가상각비및상각비", "유형자산감가상각비",
    ],
    "총자산": [
        "ifrs-full_Assets", "dart_Assets",
        "자산총계", "총자산",
    ],
    "총부채": [
        "ifrs-full_Liabilities", "dart_Liabilities",
        "부채총계", "총부채",
    ],
    "자기자본": [
        "ifrs-full_Equity", "dart_Equity",
        "자본총계", "자기자본", "지배기업소유주지분",
    ],
    "현금및현금성자산": [
        "ifrs-full_CashAndCashEquivalents",
        "dart_CashAndCashEquivalents",
        "현금및현금성자산", "현금및현금등가물",
    ],
    "유동자산": [
        "ifrs-full_CurrentAssets", "dart_CurrentAssets",
        "유동자산", "유동자산합계",
    ],
    "유동부채": [
        "ifrs-full_CurrentLiabilities", "dart_CurrentLiabilities",
        "유동부채", "유동부채합계",
    ],
    "배당금지급": [
        "ifrs-full_DividendsPaid", "dart_DividendsPaid",
        "배당금지급", "배당금의지급", "현금배당금지급",
        "배당금지급(현금배당)", "주주에대한배당금지급",
    ],
    # ── v30: 주당 지표 (PBR/PER 현재가 기준 계산에 직접 사용) ──
    "주당순이익": [
        "ifrs-full_BasicEarningsLossPerShare",
        "ifrs-full_EarningsPerShare",
        "dart_EarningsPerShare",
        "기본주당이익(손실)", "기본주당순이익", "주당순이익(기본)",
        "주당순이익", "기본EPS",
    ],
    "주당순자산": [
        "dart_BookValuePerShare",
        "ifrs-full_NetAssetsPerShare",
        "주당순자산", "주당장부금액", "주당순자산가치",
        "주당자산가치", "기본BPS",
    ],
}

# 계정명 → 표준키 역방향 매핑 (빠른 검색용)
_ACNT_NM_MAP = {}
for _k, _ids in DART_ACCOUNTS.items():
    for _id in _ids:
        _ACNT_NM_MAP[_id] = _k
        _ACNT_NM_MAP[_id.replace(" ", "")] = _k

# ══════════════════════════════════════════════════════════
# 1. 공통 유틸
# ══════════════════════════════════════════════════════════

def _fmt(v, digits=2, suffix=""):
    try:
        if v is None or (isinstance(v, float) and np.isnan(v)):
            return "N/A"
        return f"{v:.{digits}f}{suffix}"
    except Exception:
        return "N/A"

def _clamp(v, lo, hi):
    """네이버 파싱값 이상치 필터: lo~hi 범위 밖이면 None 반환"""
    if v is None:
        return None
    try:
        fv = float(v)
        if np.isnan(fv):
            return None
        return fv if lo <= fv <= hi else None
    except Exception:
        return None

def percentile_rank(series: pd.Series) -> pd.Series:
    return series.rank(pct=True) * 100

def load_dart_key() -> str:
    # 항상 스크립트 폴더에 저장 (실행 위치 무관)
    key_file = os.path.join(BASE_DIR, "dart_api.txt")
    if os.path.exists(key_file):
        with open(key_file) as f:
            key = f.read().strip()
        if key:
            print(f"  ✅ DART API 키 자동 로드 완료")
            return key
    print("\n  DART API 키를 입력하세요 (opendart.fss.or.kr 에서 발급, 무료):")
    print(f"  (한 번 입력하면 {key_file} 에 저장되어 다음부터 자동 로드됩니다)")
    key = input("  키 입력: ").strip()
    if key:
        with open(key_file, "w") as f:
            f.write(key)
        print(f"  💾 저장 완료 → 앞으로 자동 로드됩니다")
    return key

class DartClient:
    BASE = "https://opendart.fss.or.kr/api"

    def __init__(self, api_key: str):
        self.api_key  = api_key
        self._corp_map = {}

    def load_corp_codes(self) -> dict:
        if self._corp_map:
            return self._corp_map

        # 캐시 파일 확인 (당일 다운로드한 파일 재사용)
        cache_path = os.path.join(BASE_DIR, ".cache", "dart_corp_codes.json")
        if os.path.exists(cache_path):
            try:
                mtime = os.path.getmtime(cache_path)
                # 7일 이내 캐시 재사용
                if time.time() - mtime < 7 * 86400:
                    with open(cache_path, "r", encoding="utf-8") as f:
                        self._corp_map = json.load(f)
                    print(f"  [DART] 기업 코드 캐시 로드: {len(self._corp_map)}개")
                    return self._corp_map
            except Exception:
                pass

        print("  [DART] 기업 고유번호 목록 다운로드 중...")
        last_err = None
        for attempt in range(3):
            try:
                resp = requests.get(
                    f"{self.BASE}/corpCode.xml",
                    params={"crtfc_key": self.api_key},
                    timeout=30
                )
                # 응답이 ZIP인지 확인
                if resp.content[:2] != b'PK':
                    raise ValueError(f"ZIP 응답 아님 (HTTP {resp.status_code}): {resp.text[:200]}")

                with zipfile.ZipFile(io.BytesIO(resp.content)) as z:
                    xml_data = z.read("CORPCODE.xml").decode("utf-8")
                import xml.etree.ElementTree as ET
                root = ET.fromstring(xml_data)
                for item in root.findall("list"):
                    stock_code = item.findtext("stock_code","").strip()
                    corp_code  = item.findtext("corp_code","").strip()
                    if stock_code and len(stock_code) == 6:
                        self._corp_map[stock_code] = corp_code
                print(f"  [DART] 기업 코드 로드: {len(self._corp_map)}개")

                # 캐시 저장
                try:
                    os.makedirs(os.path.dirname(cache_path), exist_ok=True)
                    with open(cache_path, "w", encoding="utf-8") as f:
                        json.dump(self._corp_map, f, ensure_ascii=False)
                except Exception:
                    pass
                return self._corp_map

            except Exception as e:
                last_err = e
                print(f"  [DART] 기업코드 다운로드 실패 (시도 {attempt+1}/3): {e}")
                time.sleep(3)

        print(f"  ⚠ DART 기업코드 로드 실패 — 재무 데이터 수집 불가: {last_err}")
        return self._corp_map

    def get_financial(self, corp_code: str, year: str, reprt_code="11011") -> dict:
        url = f"{self.BASE}/fnlttSinglAcntAll.json"
        for fs_div in ("CFS", "OFS"):
            params = {"crtfc_key": self.api_key, "corp_code": corp_code,
                      "bsns_year": year, "reprt_code": reprt_code, "fs_div": fs_div}
            try:
                resp = requests.get(url, params=params, timeout=15)
                data = resp.json()
                if data.get("status") == "000" and data.get("list"):
                    result = {}
                    for row in data["list"]:
                        acnt_id = row.get("account_id", "").strip()
                        acnt_nm = row.get("account_nm", "").strip()
                        thstrm  = row.get("thstrm_amount", "").replace(",", "").strip()
                        if not thstrm:
                            continue
                        try:
                            val = float(thstrm)
                        except ValueError:
                            continue

                        # ① account_id 역방향 매핑 (가장 정확)
                        std_key = _ACNT_NM_MAP.get(acnt_id) or _ACNT_NM_MAP.get(acnt_id.replace(" ",""))
                        if std_key and std_key not in result:
                            result[std_key] = val
                            continue

                        # ② account_nm 역방향 매핑
                        std_key = _ACNT_NM_MAP.get(acnt_nm) or _ACNT_NM_MAP.get(acnt_nm.replace(" ",""))
                        if std_key and std_key not in result:
                            result[std_key] = val
                            continue

                        # ③ account_nm 부분 포함 매칭 (회사별 자유 명칭 대응)
                        for kor_name, ids in DART_ACCOUNTS.items():
                            if kor_name not in result:
                                for alias in ids:
                                    if len(alias) > 4 and alias in acnt_nm:
                                        result[kor_name] = val
                                        break

                    if result:
                        return result
            except Exception:
                pass
        return {}

    def get_latest_year(self) -> str:
        now = datetime.now()
        return str(now.year - 1) if now.month >= 4 else str(now.year - 2)

    def get_shares_outstanding(self, corp_code: str) -> int | None:
        """
        DART 발행주식수 조회 — 정확한 엔드포인트: /api/stockTotqySttus.json
        (※ /api/company.json 은 회사 개황만 제공하며 발행주식수 필드가 없음 — 과거 버그)

        필수 파라미터: bsns_year(사업연도), reprt_code(보고서코드)
        응답 필드: istc_totqy = 발행주식의 총수 (Ⅳ. = Ⅱ-Ⅲ)

        최신 사업보고서(11011) 기준으로 최근 3개 사업연도를 순서대로 시도.
        사업보고서가 아직 안 나온 연도는 가장 최근 분기보고서(11013)로 폴백.
        """
        url = f"{self.BASE}/stockTotqySttus.json"
        this_year = datetime.now().year

        # 사업연도 후보: 작년 → 재작년 (올해 사업보고서는 보통 다음해 3월에 나옴)
        # 보고서 코드: 11011=사업보고서, 11014=3분기, 11012=반기, 11013=1분기
        attempts = [
            (this_year - 1, "11011"),
            (this_year - 2, "11011"),
            (this_year,     "11013"),   # 올해 1분기 보고서로 폴백
        ]

        for year, reprt_code in attempts:
            try:
                resp = requests.get(
                    url,
                    params={
                        "crtfc_key":  self.api_key,
                        "corp_code":  corp_code,
                        "bsns_year":  str(year),
                        "reprt_code": reprt_code,
                    },
                    timeout=10
                )
                data = resp.json()
                if data.get("status") != "000":
                    continue   # 013(데이터없음) 등이면 다음 후보 시도

                rows = data.get("list", [])
                # "합계" 행 우선, 없으면 첫 보통주 행 사용
                target = None
                for row in rows:
                    se = str(row.get("se", ""))
                    if "합계" in se:
                        target = row
                        break
                if target is None and rows:
                    target = rows[0]
                if not target:
                    continue

                cnt_raw = target.get("istc_totqy") or ""
                cnt = "".join(c for c in str(cnt_raw) if c.isdigit())
                if cnt:
                    val = int(cnt)
                    if 1_000 <= val <= 1_000_000_000_000:
                        return val
            except Exception:
                continue

        return None

    def get_shares_bulk(self, stock_codes: list, max_workers: int = 30) -> dict:
        """발행주식수 배치 수집 — 7일 캐시 적용, 병렬처리
        반환: {stock_code: shares(int)} 딕셔너리
        """
        cache_path = os.path.join(BASE_DIR, ".cache", "dart_shares.json")

        # ── 캐시 로드 (7일 이내) ──
        cached = {}
        if os.path.exists(cache_path):
            try:
                mtime = os.path.getmtime(cache_path)
                if time.time() - mtime < 7 * 86400:
                    with open(cache_path, "r", encoding="utf-8") as f:
                        cached = json.load(f)
                    # 정수 타입 보정
                    cached = {k: int(v) for k, v in cached.items() if v}
            except Exception:
                cached = {}

        corp_map = self.load_corp_codes()
        need = [c for c in stock_codes if c not in cached]

        if need:
            print(f"  [주식수] DART 배치 수집: {len(need)}개 종목 (미캐시)")
            from concurrent.futures import ThreadPoolExecutor, as_completed

            def _fetch_one(code):
                corp_code = corp_map.get(code)
                if not corp_code:
                    return code, None
                shares = self.get_shares_outstanding(corp_code)
                return code, shares

            results = {}
            with ThreadPoolExecutor(max_workers=max_workers) as ex:
                futures = {ex.submit(_fetch_one, c): c for c in need}
                done = 0
                for fut in as_completed(futures):
                    code, shares = fut.result()
                    if shares:
                        results[code] = shares
                    done += 1
                    if done % 200 == 0:
                        print(f"    {done}/{len(need)} 완료...")

            cached.update(results)
            ok = sum(1 for v in results.values() if v)
            print(f"  [주식수] 수집 완료: {ok}/{len(need)}개 성공")

            # 캐시 저장
            try:
                os.makedirs(os.path.dirname(cache_path), exist_ok=True)
                with open(cache_path, "w", encoding="utf-8") as f:
                    json.dump({k: str(v) for k, v in cached.items()}, f, ensure_ascii=False)
            except Exception:
                pass
        else:
            print(f"  [주식수] 캐시에서 로드: {len(cached)}개 종목")

        return cached

    def fetch_financials_dart(self, stock_code: str) -> dict:
        keys = [
            "cfo","cfi","net_income","gross_profit","total_assets",
            "total_liabilities","equity","op_income","interest_expense",
            "depreciation","cash","current_assets","current_liabilities",
            "dividends_paid",
            "eps_dart","bps_dart",           # v30: 주당순이익/주당순자산
            # 파생 지표
            "debt_ratio","net_debt_ratio","current_ratio","interest_coverage",
            "roe","gpa","fcf","fcf_margin","ebitda","altman_z",
            "dividend_payout","vol_ratio",
        ]
        empty = {k: np.nan for k in keys}
        corp_map  = self.load_corp_codes()
        corp_code = corp_map.get(stock_code)
        if not corp_code:
            return empty
        year = self.get_latest_year()
        fs = self.get_financial(corp_code, year)
        if not fs:
            fs = self.get_financial(corp_code, str(int(year)-1))
        if not fs:
            return empty
        r = empty.copy()

        # ── 원본 항목 ──
        r["cfo"]                = fs.get("영업활동현금흐름", np.nan)
        r["cfi"]                = fs.get("투자활동현금흐름", np.nan)
        r["net_income"]         = fs.get("당기순이익",       np.nan)
        r["gross_profit"]       = fs.get("매출총이익",       np.nan)
        r["total_assets"]       = fs.get("총자산",           np.nan)
        r["total_liabilities"]  = fs.get("총부채",           np.nan)
        r["equity"]             = fs.get("자기자본",         np.nan)
        r["op_income"]          = fs.get("영업이익",         np.nan)
        r["interest_expense"]   = fs.get("이자비용",         np.nan)
        r["depreciation"]       = fs.get("감가상각비",       np.nan)
        r["cash"]               = fs.get("현금및현금성자산", np.nan)
        r["current_assets"]     = fs.get("유동자산",         np.nan)
        r["current_liabilities"]= fs.get("유동부채",         np.nan)
        r["dividends_paid"]     = fs.get("배당금지급",       np.nan)
        # v30: 주당 지표 (원/주) → PBR/PER 현재가 기준 계산의 핵심
        r["eps_dart"]           = fs.get("주당순이익",       np.nan)
        r["bps_dart"]           = fs.get("주당순자산",       np.nan)

        # ── 파생 지표 계산 ──
        eq  = r["equity"]
        tl  = r["total_liabilities"]
        ta  = r["total_assets"]
        ni  = r["net_income"]
        cfo = r["cfo"]
        cfi = r["cfi"]
        opm = r["op_income"]
        ie  = r["interest_expense"]
        ca  = r["current_assets"]
        cl  = r["current_liabilities"]
        dep = r["depreciation"]
        cash= r["cash"]

        def _ok(*vals):
            return all(v is not None and not np.isnan(v) for v in vals)

        # 부채비율
        if _ok(tl, eq) and eq > 0:
            r["debt_ratio"] = (tl / eq) * 100

        # 순부채비율 = (총부채 - 현금) / 자기자본
        if _ok(tl, cash, eq) and eq > 0:
            r["net_debt_ratio"] = ((tl - cash) / eq) * 100

        # 유동비율
        if _ok(ca, cl) and cl > 0:
            r["current_ratio"] = (ca / cl) * 100

        # 이자보상배율
        if _ok(opm, ie) and ie > 0:
            r["interest_coverage"] = opm / abs(ie)

        # ROE
        if _ok(ni, eq) and eq > 0:
            r["roe"] = (ni / eq) * 100

        # GP/A
        if _ok(r["gross_profit"], ta) and ta > 0:
            r["gpa"] = r["gross_profit"] / ta

        # FCF = CFO + CFI (투자CF는 음수가 정상 → CAPEX 근사)
        # 더 정확: FCF = CFO - |CAPEX|, CAPEX ≈ |CFI| 근사 사용
        if _ok(cfo, cfi):
            capex_est = abs(cfi)  # 투자CF 절댓값 = 대략적 CAPEX
            r["fcf"] = cfo - capex_est

        # FCF 마진 = FCF / 매출총이익(근사)
        gp = r["gross_profit"]
        if _ok(r["fcf"], gp) and gp > 0:
            r["fcf_margin"] = (r["fcf"] / gp) * 100

        # EBITDA = 영업이익 + 감가상각비
        if _ok(opm, dep):
            r["ebitda"] = opm + abs(dep)
        elif _ok(opm):
            r["ebitda"] = opm  # 감가상각 없으면 영업이익으로 대체

        # Altman Z-Score (비금융업 기준)
        # Z = 1.2*X1 + 1.4*X2 + 3.3*X3 + 0.6*X4 + 1.0*X5
        # X1=운전자본/총자산, X2=이익잉여금/총자산(ROE 근사),
        # X3=EBIT/총자산, X4=자기자본/총부채, X5=매출/총자산(GP 근사)
        if _ok(ca, cl, ta, ni, opm, eq, tl, gp) and ta > 0 and tl > 0:
            x1 = (ca - cl) / ta
            x2 = (ni / ta) if ta > 0 else 0          # 순이익/총자산 (이익잉여금 근사)
            x3 = opm / ta
            x4 = eq / tl
            x5 = gp / ta                              # 매출총이익/총자산 (매출 근사)
            r["altman_z"] = round(1.2*x1 + 1.4*x2 + 3.3*x3 + 0.6*x4 + 1.0*x5, 2)

        # 배당성향 = 배당금지급 / 순이익
        div_paid = r["dividends_paid"]
        if _ok(div_paid, ni) and ni > 0:
            r["dividend_payout"] = abs(div_paid) / ni * 100

        return r

# ══════════════════════════════════════════════════════════
# 3. 네이버 금융 스크래핑 (배당/업종/PER/PBR/영업이익률 등 보조지표)
# ══════════════════════════════════════════════════════════


def _decode_html(content: bytes) -> str:
    """네이버 금융 HTML 바이트 → 문자열 (cp949 완전 지원)"""
    for enc in ("cp949", "euc-kr", "utf-8"):
        try:
            return content.decode(enc)
        except (UnicodeDecodeError, LookupError):
            continue
    return content.decode("utf-8", errors="ignore")


def _clean_str(s: str) -> str:
    """인코딩 깨짐 문자(□, ?, 특수문자 덩어리) 감지 후 빈문자열 반환"""
    if not s:
        return s
    # 유니코드 대체문자(U+FFFD), 알 수 없는 문자 비율이 높으면 빈값
    bad = sum(1 for c in s if ord(c) in (0xFFFD, 0x25A1, 0x3013))
    if bad / max(len(s), 1) > 0.3:
        return ""
    return s.strip()

def _safe_num(raw_text):
    if not raw_text:
        return None
    txt = raw_text.replace(",","").replace("%","").replace("▲","").replace("▼","-").strip()
    txt = re.sub(r"[^\d.\-]","", txt)
    try:
        return float(txt)
    except ValueError:
        return None

def _safe_price(raw_text):
    if not raw_text:
        return None
    nums = re.findall(r"[\d,]+", raw_text.replace(" ",""))
    if not nums:
        return None
    try:
        candidates = [int(n.replace(",","")) for n in nums]
        valid = [v for v in candidates if 100 <= v <= 20_000_000]
        return valid[0] if valid else None
    except (ValueError, IndexError):
        return None

def _parse_naver_current_price(code: str):
    try:
        url  = f"https://finance.naver.com/item/sise.naver?code={code}"
        resp = requests.get(url, headers=NAVER_HEADERS, timeout=5)
        html = _decode_html(resp.content)
        m = re.search(r'<strong[^>]*id="_nowVal"[^>]*>([\d,]+)</strong>', html)
        if m:
            return int(m.group(1).replace(",",""))
        soup = BeautifulSoup(html, "html.parser") if HAS_BS4 else None
        if soup:
            tag = soup.find(id="_nowVal") or soup.find("strong", class_="tah p11")
            if tag:
                v = _safe_price(tag.get_text(strip=True))
                if v:
                    return v
    except Exception:
        pass
    return None

def _get_latest_annual_idx(soup):
    try:
        th_eps = soup.find("th", class_="th_cop_anal17")
        if not th_eps:
            return 0
        table = th_eps.find_parent("table")
        if not table:
            return 0
        thead = table.find("thead")
        if not thead:
            return 0
        all_trs = thead.find_all("tr")
        year_pat = re.compile(r'^\d{4}\.\d{2}')
        annual_cols = 4
        for tr in all_trs:
            for cell in tr.find_all(["th","td"]):
                txt = cell.get_text(strip=True)
                if "연간" in txt:
                    try:
                        annual_cols = int(cell.get("colspan", 4))
                    except (ValueError, TypeError):
                        annual_cols = 4
                    break
        for tr in all_trs:
            cells     = tr.find_all(["th","td"])
            year_texts = [c.get_text(strip=True) for c in cells
                          if year_pat.match(c.get_text(strip=True))]
            if len(year_texts) < 2:
                continue
            ann_years = year_texts[:annual_cols]
            confirmed_idx = -1
            for i, t in enumerate(ann_years):
                if "(E)" not in t:
                    confirmed_idx = i
            return confirmed_idx if confirmed_idx >= 0 else len(ann_years) - 1
    except Exception:
        pass
    return 0

def _parse_naver_stock(code: str) -> dict:
    """네이버 금융에서 보조 재무지표 수집 (배당/업종/PER/영업이익률/ROA 등)"""
    result = {"code": code}
    if not HAS_BS4:
        return result
    for attempt in range(3):
        try:
            # 현재가
            url  = f"https://finance.naver.com/item/main.naver?code={code}"
            resp = requests.get(url, headers=NAVER_HEADERS, timeout=6)
            html = _decode_html(resp.content)
            soup = BeautifulSoup(html, "html.parser")

            # 현재가
            cur_price = None
            for sel in [("strong","tah p11"), ("p","no_today")]:
                tag = soup.find(sel[0], class_=sel[1])
                if tag:
                    v = _safe_price(tag.get_text(strip=True))
                    if v:
                        cur_price = v
                        break
            if not cur_price:
                cur_price = _parse_naver_current_price(code)
            if cur_price:
                result["current_price"] = cur_price

            # 시가총액 (현재가 정합성 검증에 사용)
            marcap_eok_nv = None
            for th in soup.find_all("th"):
                if th.get_text(strip=True) == "시가총액":
                    td = th.find_next_sibling("td")
                    if td:
                        raw = td.get_text(strip=True)
                        if "조" in raw:
                            m = re.search(r"([\d.]+)조", raw)
                            if m:
                                marcap_eok_nv = int(float(m.group(1)) * 10000)
                        elif "억" in raw:
                            m = re.search(r"([\d,]+)억", raw)
                            if m:
                                marcap_eok_nv = int(m.group(1).replace(",",""))
                    break

            # ── 현재가 정합성 검증 (시가총액 기반) ──
            if cur_price and marcap_eok_nv and marcap_eok_nv > 0:
                implied_shares = marcap_eok_nv * 1e8 / cur_price
                if implied_shares < 100_000 or implied_shares > 5e10:
                    # 비현실적 주식수 → 현재가 오류
                    result["current_price"] = None
                    result["price_error"] = f"현재가오류({cur_price:,}원,추정주식수{implied_shares:.0f}주)"
            if marcap_eok_nv:
                result["market_cap_억"] = marcap_eok_nv

            # 업종
            for em in soup.find_all("em", class_="coinfo01"):
                result["sector"] = _clean_str(em.get_text(strip=True))
                break
            if "sector" not in result:
                a_sec = soup.find("a", href=re.compile(r"sise_group_detail"))
                if a_sec:
                    result["sector"] = _clean_str(a_sec.get_text(strip=True))

            # 재무 분석 지표 (CLASS_MAP 방식)
            # ★ EPS/BPS(th_cop_anal17/18)는 네이버에서 억원 단위로 표시되어
            #   자릿수 오류가 생기므로 수집하지 않음 (결산 PER/PBR에서 역산)
            annual_idx = _get_latest_annual_idx(soup)
            CLASS_MAP = {
                "th_cop_anal11": "debt_ratio_nv",   # 부채비율 (네이버)
                "th_cop_anal13": "op_margin",        # 영업이익률
                "th_cop_anal14": "net_margin",       # 순이익률
                "th_cop_anal15": "roe_nv",           # ROE (네이버)
                "th_cop_anal16": "roa",              # ROA
                # th_cop_anal17(EPS), th_cop_anal18(BPS) 제거 — 단위 오류
                "th_cop_anal19": "dps",              # DPS
                "th_cop_anal20": "per_결산",         # PER(결산)
                "th_cop_anal21": "pbr_결산",         # PBR(결산)
                "th_cop_anal22": "div",              # 배당수익률
            }
            for cls, field in CLASS_MAP.items():
                th = soup.find("th", class_=cls)
                if not th:
                    continue
                tds = []
                sib = th.find_next_sibling("td")
                while sib is not None and sib.name == "td":
                    tds.append(sib)
                    sib = sib.find_next_sibling("td")
                if not tds:
                    continue
                idx = annual_idx if annual_idx < len(tds) else (len(tds)-1)
                v   = _safe_num(tds[idx].get_text(strip=True))
                if v is None:
                    continue
                # ── 필드별 유효성 검사 ──
                if field == "roa":
                    # ROA는 -50% ~ +50% 범위 (단위가 % 아닌 소수면 *100)
                    if abs(v) > 100:
                        continue   # 비정상값 제거
                    result[field] = v
                elif field == "roe_nv":
                    if abs(v) > 200:
                        continue
                    result[field] = v
                elif field == "debt_ratio_nv":
                    if v < 0 or v > 10000:
                        continue
                    result[field] = v
                elif field in ("per_결산", "pbr_결산"):
                    if v <= 0 or v > 500:
                        continue
                    result[field] = v
                elif field == "dps":
                    result[field] = int(v) if v > 0 else 0
                else:
                    result[field] = v

            # 매출액·영업이익 (텍스트 매칭)
            TEXT_MAP = {"매출액": "revenue", "영업이익": "op_profit"}
            for label, field in TEXT_MAP.items():
                for th in soup.find_all("th"):
                    if th.get_text(strip=True) == label:
                        tds = []
                        sib = th.find_next_sibling("td")
                        while sib and sib.name == "td":
                            tds.append(sib)
                            sib = sib.find_next_sibling("td")
                        if tds:
                            idx = annual_idx if annual_idx < len(tds) else (len(tds)-1)
                            v = _safe_num(tds[idx].get_text(strip=True))
                            if v is not None:
                                result[field] = v
                            if field == "revenue" and idx > 0:
                                pv = _safe_num(tds[idx-1].get_text(strip=True))
                                if pv and pv > 0:
                                    result["revenue_prev"] = pv
                        break

            # 매출 성장률
            rev   = result.get("revenue")
            rev_p = result.get("revenue_prev")
            if rev and rev_p and rev_p > 0:
                result["revenue_growth"] = round((rev - rev_p) / rev_p * 100, 1)

            # PBR/PER: 결산 기준 사용 (EPS/BPS 파싱 제거로 현재가 기준 재계산 불가)
            # 결산 PER/PBR이 가장 신뢰할 수 있는 값
            result["per"] = result.get("per_결산")
            result["pbr"] = result.get("pbr_결산")

            # EPS/BPS 컬럼도 제거 (오염된 값 전파 방지)
            result.pop("eps", None)
            result.pop("bps", None)

            return result

        except requests.exceptions.Timeout:
            time.sleep(1)
        except Exception as e:
            if attempt == 2:
                return {"code": code, "error": str(e)}
            time.sleep(0.5)
    return {"code": code, "error": "max retries"}

def fetch_naver_bulk(tickers, names, max_workers=20, no_cache=False) -> pd.DataFrame:
    """네이버 금융 병렬 수집 (캐시 적용)
    v33 변경:
      - 업종/배당/결산PER·PBR 등 보조지표 → TIER-B 캐시 (1일 유효)
      - 현재가(current_price) → 캐시 우회, 매 실행마다 항상 새로 수집 (TIER-C)
    """
    total_nv = len(tickers)

    # ── TIER-B 캐시 확인 (현재가 제외 보조지표) ──
    cache_key_nv = f"naver_bulk_{','.join(sorted(tickers[:20]))}_n{total_nv}"
    cached_meta = None  # 현재가 제외 보조지표 캐시
    if not no_cache:
        cached_meta = cache_get(cache_key_nv, tier="B")

    # ── TIER-C: 현재가는 항상 새로 수집 ──
    print(f"  [네이버] 현재가 실시간 수집 중 ({total_nv}개, 병렬처리)...")

    def _fetch_price_only(code):
        """현재가만 빠르게 수집 (TIER-C — 항상 최신)"""
        price = _parse_naver_current_price(code)
        time.sleep(0.02)
        return code, price

    price_map = {}
    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        futs = {executor.submit(_fetch_price_only, c): c for c in tickers}
        done = 0
        for fut in as_completed(futs):
            code, price = fut.result()
            if price:
                price_map[code] = price
            done += 1
            if done % 500 == 0:
                print(f"  [현재가] {done}/{total_nv} 수집 중...")
    ok_cnt = sum(1 for v in price_map.values() if v)
    print(f"  ✓ 현재가 수집 완료: {ok_cnt}/{total_nv}개")

    # 보조지표: 캐시 있으면 재사용, 없으면 새로 수집
    if cached_meta is not None:
        print(f"  [캐시] 네이버 보조지표 TIER-B 캐시 적중 → 재수집 생략 (업종/배당/PER/PBR 등)")
        try:
            df = pd.DataFrame(cached_meta).set_index("코드")
            # 현재가만 오늘 수집값으로 교체
            for code, price in price_map.items():
                if code in df.index:
                    df.loc[code, "현재가_NV"] = price
            return df
        except Exception:
            pass  # 캐시 파싱 실패 → 전체 재수집

    # 전체 재수집 (보조지표 캐시 없을 때)
    print(f"  [네이버] 보조지표 전체 수집 중 ({total_nv}개)...")
    code_to_name = dict(zip(tickers, names))
    results = {}
    _nv_done = [0]

    if HAS_TQDM:
        pbar = tqdm(total=total_nv, desc="  네이버 수집", unit="종목",
                    bar_format="{l_bar}{bar}| {n_fmt}/{total_fmt} [{elapsed}<{remaining}]")

    def _fetch_one(code):
        r = _parse_naver_stock(code)
        time.sleep(0.03)
        return code, r

    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        futures = {executor.submit(_fetch_one, c): c for c in tickers}
        for future in as_completed(futures):
            code, data = future.result()
            results[code] = data
            if HAS_TQDM:
                pbar.update(1)
            else:
                _nv_done[0] += 1
                if _nv_done[0] % 200 == 0:
                    print(f"  [네이버] {_nv_done[0]}/{total_nv} ({_nv_done[0]/total_nv*100:.0f}%) 수집 중...")

    if HAS_TQDM:
        pbar.close()

    rows = []
    for code in tickers:
        d = results.get(code, {})
        name = code_to_name.get(code) or d.get("name", code)
        rows.append({
            "코드":          code,
            "종목명":        _clean_str(str(name)) if name else "",
            "시가총액(억)":  _safe(d.get("market_cap_억")),
            "업종":          d.get("sector",""),
            "현재가_NV":     _safe(price_map.get(code) or d.get("current_price")),
            "PBR_결산":      _safe(d.get("pbr_결산")),
            "PBR":           _safe(d.get("pbr")),
            "PER_결산":      _safe(d.get("per_결산")),
            "PER":           _safe(d.get("per")),
            # EPS/BPS 제거 — 네이버 억원 단위 파싱 오류로 신뢰 불가
            "ROE_NV":        _clamp(d.get("roe_nv"),  -200, 200),
            "ROA":           _clamp(d.get("roa"),     -50,  50),   # ROA: -50%~+50% 범위
            "DIV":           _safe(d.get("div"), 0.0),
            "영업이익률":    _clamp(d.get("op_margin"),  -100, 100),
            "순이익률":      _clamp(d.get("net_margin"), -200, 200),
            "부채비율_NV":   _clamp(d.get("debt_ratio_nv"), 0, 5000),
            "매출액":        _safe(d.get("revenue")),
            "매출성장률":    _safe(d.get("revenue_growth")),
            "영업이익_NV":   _safe(d.get("op_profit")),
        })

    df = pd.DataFrame(rows).set_index("코드")
    print(f"  ✓ 네이버 수집 완료: {df['PBR'].notna().sum()}개 PBR 수집")

    # TIER-B 캐시 저장 (현재가 제외 보조지표만 저장)
    try:
        df_cache = df.copy()
        df_cache["현재가_NV"] = None  # 현재가는 캐시에 저장 안 함
        cache_set(cache_key_nv, df_cache.reset_index().to_dict(orient="records"), tier="B")
        print(f"  [캐시] 네이버 보조지표 TIER-B 저장 (1일 유효, 현재가 제외)")
    except Exception:
        pass

    return df

# ══════════════════════════════════════════════════════════
# 4. yfinance 가격/모멘텀 (배치 다운로드)
# ══════════════════════════════════════════════════════════

def _safe_get_ohlc(hist, yf_ticker):
    """Close/High/Low/Volume 시리즈를 안전하게 추출"""
    t = yf_ticker

    def _extract(sub):
        c = sub["Close"].dropna()
        h = sub["High"].dropna()
        l = sub["Low"].dropna()
        v = sub["Volume"].dropna() if "Volume" in sub.columns else pd.Series(dtype=float)
        return c, h, l, v

    if not isinstance(hist.columns, pd.MultiIndex):
        return _extract(hist)
    lv0 = hist.columns.get_level_values(0).unique()
    lv1 = hist.columns.get_level_values(1).unique()
    if t in lv0:
        sub = hist[t]
        if isinstance(sub, pd.DataFrame):
            return _extract(sub)
    if t in lv1:
        c = hist["Close"][t].dropna()
        h = hist["High"][t].dropna()
        l = hist["Low"][t].dropna()
        v = hist["Volume"][t].dropna() if "Volume" in hist.columns.get_level_values(0) else pd.Series(dtype=float)
        return c, h, l, v
    for lv in (0, 1):
        try:
            sub = hist.xs(t, level=lv, axis=1)
            if isinstance(sub, pd.DataFrame) and "Close" in sub.columns:
                return _extract(sub)
        except Exception:
            pass
    raise KeyError(f"{t} not found in columns")


# ── 기술적 지표 계산 함수 (외부 라이브러리 불필요 — 순수 pandas/numpy) ──

def _ta_macd(close):
    """MACD(12,26,9) — 히스토그램 부호 교차로 매수/매도 판단"""
    if len(close) < 30:
        return None, None, None, "중립"
    ema12 = close.ewm(span=12, adjust=False).mean()
    ema26 = close.ewm(span=26, adjust=False).mean()
    macd  = ema12 - ema26
    sig   = macd.ewm(span=9, adjust=False).mean()
    hist  = macd - sig
    cur_h, prev_h = float(hist.iloc[-1]), float(hist.iloc[-2])
    # 히스토그램 부호 교차
    if prev_h <= 0 and cur_h > 0:   sig_txt = "매수"       # 골든크로스
    elif prev_h >= 0 and cur_h < 0: sig_txt = "매도"       # 데드크로스
    elif cur_h > 0:                  sig_txt = "매수유지"
    elif cur_h < 0:                  sig_txt = "매도유지"
    else:                            sig_txt = "중립"
    return round(float(macd.iloc[-1]),2), round(float(sig.iloc[-1]),2), round(cur_h,2), sig_txt


def _ta_rsi(close, period=14):
    """RSI(14) — 30 이하 과매도=매수, 70 이상 과매수=매도"""
    if len(close) < period + 1:
        return None, "중립"
    delta = close.diff()
    gain  = delta.clip(lower=0)
    loss  = (-delta).clip(lower=0)
    avg_g = gain.ewm(alpha=1/period, adjust=False).mean()
    avg_l = loss.ewm(alpha=1/period, adjust=False).mean()
    rs    = avg_g / avg_l.replace(0, np.nan)
    rsi   = 100 - (100 / (1 + rs))
    val   = float(rsi.iloc[-1])
    if   val < 30: sig = "매수"       # 과매도
    elif val > 70: sig = "매도"       # 과매수
    elif val < 45: sig = "매수유지"
    elif val > 55: sig = "매도유지"
    else:          sig = "중립"
    return round(val, 1), sig


def _ta_obv(close, volume):
    """OBV — 5일/20일 OBV 이동평균 비교로 추세 판단"""
    if len(close) < 20 or volume is None or len(volume) < 20:
        return "중립", "중립"
    try:
        # 인덱스 정렬
        common = close.index.intersection(volume.index)
        c = close.reindex(common)
        v = volume.reindex(common)
        direction = np.sign(c.diff().fillna(0))
        obv  = (direction * v).cumsum()
        obv5  = float(obv.rolling(5).mean().iloc[-1])
        obv20 = float(obv.rolling(20).mean().iloc[-1])
        trend = "상승" if float(obv.iloc[-1]) > float(obv.iloc[-5]) else "하락"
        if   obv5 > obv20 * 1.02: sig = "매수"
        elif obv5 < obv20 * 0.98: sig = "매도"
        else:                      sig = "중립"
        return trend, sig
    except Exception:
        return "중립", "중립"


def _ta_bb(close, period=20, std_mult=2.0):
    """Bollinger Bands(20,2) — %B 기준 매수/매도"""
    if len(close) < period:
        return None, None, None, None, "중립"
    ma    = close.rolling(period).mean()
    std   = close.rolling(period).std()
    upper = ma + std_mult * std
    lower = ma - std_mult * std
    cur   = float(close.iloc[-1])
    u     = float(upper.iloc[-1])
    l     = float(lower.iloc[-1])
    pct_b = (cur - l) / (u - l) if (u - l) > 0 else 0.5
    # 밴드폭 = (상단-하단)/중간×100 (%)  — 수치가 낮으면 스퀴즈
    bw    = (u - l) / float(ma.iloc[-1]) * 100 if float(ma.iloc[-1]) > 0 else 0
    if   pct_b < 0.0: sig = "매수"       # 하단 이탈
    elif pct_b > 1.0: sig = "매도"       # 상단 이탈
    elif pct_b < 0.2: sig = "매수유지"
    elif pct_b > 0.8: sig = "매도유지"
    else:             sig = "중립"
    return round(u), round(l), round(pct_b*100,1), round(bw,1), sig


def _detect_box_range(close_s, high_s, low_s, vol_s=None, recent_excl=3):
    """
    박스권(장기 횡보 구간) 탐지 (v41g 신규)
    ─────────────────────────────────────────────
    1개월 내 단기 등락이 아니라, 최소 약 3개월(60거래일) 이상
    일정 가격대 안에서 횡보해온 구간을 찾는다.
    박스 상단을 거래량 급증과 함께 돌파하면
    '뉴스·재료성 급등 후보'로 표시한다.

    판정 절차:
      ① '박스 구간'은 최근 recent_excl(기본 3)일을 제외한 과거 데이터로만
         산정한다 — 오늘의 급등이 박스 상단 계산에 섞여
         '자기 자신을 못 뚫는' 모순을 막기 위함.
      ② 120일(≈6개월) → 90일(≈4.5개월) → 60일(≈3개월) 순으로
         가장 긴 구간부터 박스권 조건을 검사해 채택한다
         (긴 박스일수록 돌파 시 추세 신뢰도가 높다고 보기 때문).
      ③ 박스 조건: 구간 내 가격폭 35% 이내
                    + 전반부·후반부 평균가 차이(추세/드리프트) 12% 이내
                    + 종가의 90% 이상이 밴드(상단×1.02~하단×0.98) 안에 머묾
         → 추세(상승·하락) 구간은 제외하고 '진짜 횡보'만 박스권으로 인정.
      ④ 산정된 박스를 기준으로 '오늘 종가'가 상단을 넘었는지 비교한다.
         거래량 급증(최근5일 거래대금 ÷ 20일 거래대금 ≥ 1.5배) 상태에서
         박스 상단을 돌파하면 → "상단돌파+거래량급증(뉴스성모멘텀후보)"
    """
    n = len(close_s)
    if n < 60 + recent_excl:
        return {
            "박스권여부": "N", "박스기간(일)": None,
            "박스상단": None, "박스하단": None,
            "박스폭(%)": None, "박스내위치(%)": None,
            "박스이탈신호": "데이터부족",
        }

    cur = float(close_s.iloc[-1])
    candidates = [w for w in (120, 90, 60) if n >= w + recent_excl]
    best = None

    for w in candidates:
        try:
            # 박스는 '최근 recent_excl일을 제외한' 과거 구간으로만 산정
            c_win = close_s.iloc[-(w + recent_excl):-recent_excl]
            h_win = high_s.iloc[-(w + recent_excl):-recent_excl]
            l_win = low_s.iloc[-(w + recent_excl):-recent_excl]
            top = float(h_win.max())
            bot = float(l_win.min())
            if bot <= 0:
                continue
            width_pct = (top - bot) / bot * 100

            half = w // 2
            avg_front = float(c_win.iloc[:half].mean())
            avg_back  = float(c_win.iloc[half:].mean())
            drift_pct = abs(avg_back - avg_front) / avg_front * 100 if avg_front > 0 else 999

            tol_top = top * 1.02
            tol_bot = bot * 0.98
            inside_ratio = float(((c_win >= tol_bot) & (c_win <= tol_top)).mean())

            if width_pct <= 35 and drift_pct <= 12 and inside_ratio >= 0.90:
                best = {"window": w, "top": top, "bot": bot, "width_pct": width_pct}
                break   # 가장 긴 구간부터 검사하므로 처음 통과하는 것을 채택
        except Exception:
            continue

    if best is None:
        return {
            "박스권여부": "N", "박스기간(일)": None,
            "박스상단": None, "박스하단": None,
            "박스폭(%)": None, "박스내위치(%)": None,
            "박스이탈신호": "박스아님(추세/등락)",
        }

    top, bot, w = best["top"], best["bot"], best["window"]
    pos_in_box = (cur - bot) / (top - bot) * 100 if top > bot else 50.0
    pos_in_box = max(-20.0, min(120.0, pos_in_box))

    # 거래량 급증 여부 (최근5일 거래대금 ÷ 20일 거래대금)
    vol_surge = False
    if vol_s is not None and len(vol_s) >= 20:
        try:
            common = close_s.index.intersection(vol_s.index)
            c_aln = close_s.reindex(common)
            v_aln = vol_s.reindex(common)
            amt = c_aln * v_aln
            a5  = float(amt.iloc[-5:].mean())
            a20 = float(amt.iloc[-20:].mean())
            vol_surge = a20 > 0 and (a5 / a20) >= 1.5
        except Exception:
            pass

    if cur > top:
        box_sig = "상단돌파+거래량급증(뉴스성모멘텀후보)" if vol_surge else "상단돌파(확인필요)"
    elif cur < bot:
        box_sig = "하단이탈(주의)"
    elif pos_in_box >= 80:
        box_sig = "상단근접(돌파대기)"
    elif pos_in_box <= 20:
        box_sig = "하단근접(지지테스트)"
    else:
        box_sig = "박스권유지"

    return {
        "박스권여부":    "Y",
        "박스기간(일)":  w,
        "박스상단":      int(round(top)),
        "박스하단":      int(round(bot)),
        "박스폭(%)":     round(best["width_pct"], 1),
        "박스내위치(%)": round(pos_in_box, 1),
        "박스이탈신호":  box_sig,
    }


def _ta_summary(macd_sig, rsi_sig, obv_sig, bb_sig):
    """4개 지표 종합 → 기술신호 + 매수/매도 카운트"""
    sigs = [macd_sig, rsi_sig, obv_sig, bb_sig]
    buy  = sum(1 for s in sigs if "매수" in s)
    sell = sum(1 for s in sigs if "매도" in s)
    if   buy >= 3:  overall = "⚡ 강력매수"
    elif buy == 2:  overall = "▲ 매수"
    elif sell >= 3: overall = "🔻 강력매도"
    elif sell == 2: overall = "▼ 매도"
    else:           overall = "─ 중립"
    summary = f"MACD:{macd_sig} RSI:{rsi_sig} OBV:{obv_sig} BB:{bb_sig}"
    return buy, sell, overall, summary


def _compute_price_row(close_s, high_s, low_s, vol_s=None):
    """가격/모멘텀/거래량/기술지표(MACD·RSI·OBV·BB) 통합 계산 (v32)"""
    cur   = float(close_s.iloc[-1])
    hi52  = float(high_s.max())
    lo52  = float(low_s.min())
    pos   = (cur - lo52) / (hi52 - lo52) * 100 if hi52 > lo52 else 50.0
    first = float(close_s.iloc[0])
    ret52 = (cur / first - 1) * 100 if first > 0 else 0.0
    idx_6m = max(0, len(close_s) - 126)
    ret6m = (cur / float(close_s.iloc[idx_6m]) - 1) * 100 if float(close_s.iloc[idx_6m]) > 0 else 0.0

    # ── 거래량 지표 ──
    amt5 = amt20 = vr = None
    if vol_s is not None and len(vol_s) >= 5:
        try:
            common_idx = close_s.index.intersection(vol_s.index)
            if len(common_idx) >= 5:
                c_aln = close_s.reindex(common_idx)
                v_aln = vol_s.reindex(common_idx)
                amt_s = c_aln * v_aln
                n = len(amt_s)
                amt5  = round(float(amt_s.iloc[-5:].mean()) / 1e8, 1)
                amt20 = round(float(amt_s.iloc[-min(20,n):].mean()) / 1e8, 1)
                if amt20 and amt20 > 0:
                    vr = round(amt5 / amt20 * 100, 1)
        except Exception:
            pass

    # ── 4대 기술지표 계산 ──
    macd_line = macd_sig_line = macd_hist = None
    macd_sig = rsi_val = "중립"
    obv_trend = obv_sig = "중립"
    bb_upper = bb_lower = bb_pct_b = bb_width = None
    bb_sig = "중립"
    ta_buy = ta_sell = 0
    ta_signal = "─ 중립"
    ta_summary_str = ""

    try:
        if len(close_s) >= 30:
            macd_line, macd_sig_line, macd_hist, macd_sig = _ta_macd(close_s)
            rsi_val, rsi_sig = _ta_rsi(close_s)
            obv_trend, obv_sig = _ta_obv(close_s, vol_s)
            bb_upper, bb_lower, bb_pct_b, bb_width, bb_sig = _ta_bb(close_s)
            ta_buy, ta_sell, ta_signal, ta_summary_str = _ta_summary(macd_sig, rsi_sig, obv_sig, bb_sig)
    except Exception:
        pass

    # ── 박스권(장기 횡보) 탐지 (v41g 신규) ──
    try:
        box_info = _detect_box_range(close_s, high_s, low_s, vol_s)
    except Exception:
        box_info = {
            "박스권여부": "N", "박스기간(일)": None,
            "박스상단": None, "박스하단": None,
            "박스폭(%)": None, "박스내위치(%)": None,
            "박스이탈신호": "계산오류",
        }

    return {
        # 기존 가격 지표
        "현재가_YF":        int(round(cur)),
        "52주고가":         int(round(hi52)),
        "52주저가":         int(round(lo52)),
        "52주위치":         round(pos, 1),
        "52주수익률":       round(ret52, 1),
        "6개월수익률":      round(ret6m, 1),
        # 거래량 지표
        "거래대금5일(억)":  amt5,
        "거래대금20일(억)": amt20,
        "거래량비율VR(%)":  vr,
        # MACD
        "MACD선":          macd_line,
        "MACD시그널":       macd_sig_line,
        "MACD히스토":       macd_hist,
        "MACD신호":         macd_sig,
        # RSI
        "RSI14":           rsi_val,
        "RSI신호":          rsi_sig,
        # OBV
        "OBV추세":          obv_trend,
        "OBV신호":          obv_sig,
        # Bollinger Bands
        "BB상단":           bb_upper,
        "BB하단":           bb_lower,
        "BB%B":            bb_pct_b,
        "BB밴드폭(%)":      bb_width,
        "BB신호":           bb_sig,
        # 종합 기술신호
        "기술매수신호수":    ta_buy,
        "기술매도신호수":    ta_sell,
        "기술신호":          ta_signal,
        "기술지표요약":      ta_summary_str,
        # 박스권 분석 (v41g 신규)
        **box_info,
    }

def fetch_price_yfinance_batch(tickers, market_map=None) -> pd.DataFrame:
    if not HAS_YFINANCE:
        print("  ⚠ yfinance 미설치 → 모멘텀 중립값 사용")
        return pd.DataFrame()

    print(f"  [가격] yfinance 배치 다운로드 중 ({len(tickers)}개)...")
    if market_map is None:
        market_map = {}

    def _suffix(code):
        return ".KQ" if market_map.get(code,"KOSPI") == "KOSDAQ" else ".KS"

    yf_map  = {f"{c}{_suffix(c)}": c for c in tickers}
    yf_list = list(yf_map.keys())
    rows    = {}
    CHUNK   = 500

    chunks = [yf_list[i:i+CHUNK] for i in range(0, len(yf_list), CHUNK)]
    print(f"  배치: {len(chunks)}청크 × {CHUNK}개")

    all_chunks = []
    for ci, chunk in enumerate(chunks, 1):
        try:
            h = yf.download(chunk, period="1y", progress=False,
                            group_by="ticker", auto_adjust=True, threads=True)
            if h is not None and len(h) >= 5:
                all_chunks.append((chunk, h))
                print(f"  청크 {ci}/{len(chunks)} 완료")
            time.sleep(2)
        except Exception as e:
            print(f"  청크 {ci} 실패: {e}")
            time.sleep(5)

    processed = set()
    for chunk_list, hist_chunk in all_chunks:
        chunk_map = {t: yf_map[t] for t in chunk_list if t in yf_map}
        for yf_t, code in chunk_map.items():
            if code in processed:
                continue
            try:
                c_s, h_s, l_s, v_s = _safe_get_ohlc(hist_chunk, yf_t)
                if len(c_s) < 5:
                    continue
                rows[code] = _compute_price_row(c_s, h_s, l_s, v_s)
                processed.add(code)
            except Exception:
                pass

    if not rows:
        print("  ⚠ 가격 데이터 수집 실패 → 중립값 적용")
        return pd.DataFrame()

    df = pd.DataFrame.from_dict(rows, orient="index")
    df.index.name = "코드"
    df.index = df.index.astype(str)
    print(f"  ✓ 가격 수집 완료: {len(df)}개")
    return df

# ══════════════════════════════════════════════════════════
# 5. 한국 유니버스 수집
# ══════════════════════════════════════════════════════════

def fetch_kr_universe():
    tickers, markets = {}, {}
    EXCL = re.compile(r"스팩|SPAC|ETF|ETN|리츠|인프라|우$|우B$|우C$|우선주")

    # ── 방법1: FinanceDataReader (KRX 공식) ──
    if HAS_FDR:
        try:
            kospi  = fdr.StockListing("KOSPI") [["Code","Name","Market"]]
            kosdaq = fdr.StockListing("KOSDAQ")[["Code","Name","Market"]]
            for _, row in pd.concat([kospi, kosdaq]).iterrows():
                if not EXCL.search(str(row["Name"])):
                    tickers[str(row["Code"])] = str(row["Name"])
                    markets[str(row["Code"])] = str(row["Market"])
            if len(tickers) >= 500:
                print(f"  [유니버스] FDR 수집 완료: {len(tickers)}개")
                codes = list(tickers.keys())
                return codes, [tickers[c] for c in codes], [markets.get(c,"KOSPI") for c in codes]
            else:
                print(f"  [유니버스] FDR 결과 부족({len(tickers)}개) → KRX API 시도")
                tickers.clear(); markets.clear()
        except Exception as e:
            print(f"  [유니버스] FDR 실패: {e} → KRX API 시도")

    # ── 방법2: KRX 정보데이터시스템 직접 호출 ──
    def _fetch_krx(market_id, market_name):
        result = {}
        # KRX는 세션 쿠키 + generate_key 헤더가 필요 → 먼저 메인 페이지를 방문해 쿠키 획득
        try:
            sess = requests.Session()
            sess.get(
                "http://data.krx.co.kr/contents/MDC/MDI/mdiLoader/index.cmd",
                headers={"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)"},
                timeout=10,
            )
            url  = "http://data.krx.co.kr/comm/bldAttendant/getJsonData.cmd"
            body = {
                "bld":         "dbms/MDC/STAT/standard/MDCSTAT01901",
                "mktId":       market_id,
                "share":       "1",
                "money":       "1",
                "csvxls_isNo": "false",
            }
            hdrs = {
                "User-Agent":    "Mozilla/5.0 (Windows NT 10.0; Win64; x64)",
                "Referer":       "http://data.krx.co.kr/contents/MDC/MDI/mdiLoader/index.cmd",
                "Content-Type":  "application/x-www-form-urlencoded; charset=UTF-8",
                "Accept":        "application/json, text/javascript, */*; q=0.01",
                "X-Requested-With": "XMLHttpRequest",
            }
            resp = sess.post(url, data=body, headers=hdrs, timeout=20)
            resp.raise_for_status()
            if not resp.text.strip():
                raise ValueError("빈 응답")
            data = resp.json()
            for item in data.get("OutBlock_1", []):
                code = item.get("ISU_SRT_CD","").strip()
                name = item.get("ISU_ABBRV","").strip()
                if code and name and len(code)==6 and not EXCL.search(name):
                    result[code] = name
            print(f"  [유니버스] KRX {market_name}: {len(result)}개")
        except Exception as e:
            print(f"  [유니버스] KRX {market_name} 실패: {e}")
        return result

    kospi_d  = _fetch_krx("STK", "KOSPI")
    kosdaq_d = _fetch_krx("KSQ", "KOSDAQ")
    for code, name in kospi_d.items():
        tickers[code] = name; markets[code] = "KOSPI"
    for code, name in kosdaq_d.items():
        if code not in tickers:
            tickers[code] = name; markets[code] = "KOSDAQ"

    if len(tickers) >= 100:
        print(f"  [유니버스] KRX 수집 완료: {len(tickers)}개 (KOSPI {len(kospi_d)} + KOSDAQ {len(kosdaq_d)})")
    else:
        # ── 방법3: 네이버 시가총액 페이지 (최후 폴백) ──
        print(f"  [유니버스] KRX 결과 부족 → 네이버 수집 시도...")
        def _fetch_naver(sosok):
            result = {}
            for page in range(1, 300):
                try:
                    resp = requests.get(
                        f"https://finance.naver.com/sise/sise_market_sum.naver"
                        f"?sosok={sosok}&page={page}",
                        headers=NAVER_HEADERS, timeout=8)
                    html = _decode_html(resp.content)
                    found = re.findall(r'/item/main\.naver\?code=(\d{6})[^>]*>\s*([^<]{1,30})', html)
                    if not found:
                        break
                    new = 0
                    for code, name in found:
                        name = name.strip()
                        if name and len(code)==6 and code not in result and not EXCL.search(name):
                            result[code] = name; new += 1
                    if new == 0:
                        break
                    time.sleep(0.1)
                except Exception:
                    break
            return result

        nv_kospi  = _fetch_naver(0)
        nv_kosdaq = _fetch_naver(1)
        for code, name in nv_kospi.items():
            if code not in tickers:
                tickers[code] = name; markets[code] = "KOSPI"
        for code, name in nv_kosdaq.items():
            if code not in tickers:
                tickers[code] = name; markets[code] = "KOSDAQ"
        print(f"  [유니버스] 네이버 수집 완료: {len(tickers)}개")

    codes = list(tickers.keys())
    return codes, [tickers[c] for c in codes], [markets.get(c,"KOSPI") for c in codes]


# ══════════════════════════════════════════════════════════
# 6. 한국 통합 데이터 수집
# ══════════════════════════════════════════════════════════

def fetch_kr_all_data(dart: DartClient, tickers, names, markets, sample_size=None, no_cache=False):
    if sample_size:
        tickers = tickers[:sample_size]
        names   = names[:sample_size]
        markets = markets[:sample_size]

    market_map = dict(zip(tickers, markets))
    name_map   = dict(zip(tickers, names))

    # ★ 핵심: 유니버스 전체를 먼저 기본 DataFrame으로 생성
    # 수집 성공/실패와 무관하게 모든 종목이 결과에 포함됨
    base_df = pd.DataFrame({
        "코드":   tickers,
        "종목명": names,
        "시장":   markets,
    }).set_index("코드")
    base_df.index = base_df.index.astype(str)
    print(f"  [유니버스] 전체 {len(base_df)}개 종목으로 스크리닝 시작")

    # ① 네이버 보조 지표 (병렬)
    df_nv = fetch_naver_bulk(tickers, names, no_cache=no_cache)
    df_nv.index = df_nv.index.astype(str)
    # 네이버 수집 실패한 종목도 base_df 기준으로 유지
    df_nv = base_df.join(df_nv.drop(columns=["종목명"], errors="ignore"), how="left")
    market_s = pd.Series(market_map, name="시장")
    df_nv["시장"] = df_nv.index.map(market_s).fillna("KOSPI")
    print(f"  [네이버] 수집 후 유니버스 유지: {len(df_nv)}개")

    # ② yfinance 가격 (배치)
    df_price = fetch_price_yfinance_batch(tickers, market_map)

    # ③ DART 재무 — 병렬처리 + 캐시
    total_dart = len(tickers)
    cache_key_bulk = f"dart_bulk_{','.join(sorted(tickers[:20]))}"
    cached_dart = None if no_cache else cache_get(cache_key_bulk, tier="A")

    if cached_dart is not None:
        print(f"  [캐시] DART 재무 캐시 적중 → API 호출 생략 (약 {total_dart * 0.08 / 60:.0f}분 절약)")
        dart_rows = cached_dart
    else:
        eta_min = round(total_dart * 0.04 / 60, 1)  # 병렬처리로 절반
        print(f"  [DART] 병렬 재무 수집 중 ({total_dart}개) — 예상 약 {eta_min}분 소요...")
        dart.load_corp_codes()
        dart_rows = {}
        _dart_start = time.time()

        # ── 병렬 처리 (스레드 수: 8 — DART API 과부하 방지 최적값) ──
        DART_WORKERS = 8
        lock = __import__("threading").Lock()
        completed = [0]

        def _fetch_one(code):
            # 개별 종목 캐시 확인 — TIER-A (7일 유효: 분기 재무는 자주 안 바뀜)
            ck = f"dart_{code}"
            cached = cache_get(ck, tier="A")
            if cached is not None:
                return code, cached
            result = dart.fetch_financials_dart(code)
            cache_set(ck, result, tier="A")  # TIER-A: 7일 캐시
            time.sleep(0.05)       # DART API 과부하 방지
            return code, result

        if HAS_TQDM:
            pbar = tqdm(total=total_dart, desc="  DART 병렬", unit="종목",
                        bar_format="{l_bar}{bar}| {n_fmt}/{total_fmt} [{elapsed}<{remaining}]")

        with ThreadPoolExecutor(max_workers=DART_WORKERS) as executor:
            futures = {executor.submit(_fetch_one, code): code for code in tickers}
            for future in as_completed(futures):
                try:
                    code, result = future.result(timeout=30)
                    dart_rows[code] = result
                except Exception as e:
                    code = futures[future]
                    dart_rows[code] = {}
                with lock:
                    completed[0] += 1
                    n = completed[0]
                if HAS_TQDM:
                    pbar.update(1)
                elif n % 100 == 0:
                    elapsed = time.time() - _dart_start
                    remaining = elapsed / n * (total_dart - n)
                    print(f"  [DART] {n}/{total_dart} ({n/total_dart*100:.0f}%) "
                          f"— 경과 {elapsed/60:.1f}분 / 잔여 약 {remaining/60:.1f}분")

        if HAS_TQDM:
            pbar.close()

        elapsed_total = time.time() - _dart_start
        print(f"  [DART] 수집 완료 (소요 {elapsed_total/60:.1f}분, 평균 {elapsed_total/max(total_dart,1):.2f}초/종목)")

        # 전체 결과 캐시 저장 — TIER-A (7일 유효)
        cache_set(cache_key_bulk, dart_rows, tier="A")
        print(f"  [캐시] DART 결과 저장 완료 (TIER-A, 7일 유효)")

    df_dart = pd.DataFrame.from_dict(dart_rows, orient="index")
    df_dart.index.name = "코드"
    df_dart.index = df_dart.index.astype(str)
    cfo_count = df_dart["cfo"].notna().sum() if "cfo" in df_dart.columns else 0
    print(f"  ✓ DART 재무 수집 완료: {cfo_count}개 CFO 수집")

    # ④ 발행주식수 배치 수집 (v28 신규) ──────────────────────────────
    # DART /stockTotqySttus.json → istc_totqy (7일 캐시, 병렬처리)
    # 없으면 시총÷현재가 역산으로 fallback
    shares_map = {}
    try:
        shares_map = dart.get_shares_bulk(tickers, max_workers=30)
        print(f"  [주식수] DART 발행주식수 로드: {len(shares_map)}개")
    except Exception as e:
        print(f"  [주식수] DART 수집 실패, 역산으로 대체: {e}")

    # ⑤ 병합 (left join으로 유니버스 전체 유지)
    df = df_nv.join(df_dart, how="left")
    if not df_price.empty:
        df = df.join(df_price, how="left")
    else:
        for col in ["현재가_YF","52주고가","52주저가","52주위치","52주수익률","6개월수익률"]:
            df[col] = np.nan

    # 현재가 통합 (네이버 우선, YF 폴백)
    if "현재가_NV" in df.columns and "현재가_YF" in df.columns:
        df["현재가"] = df["현재가_NV"].combine_first(df["현재가_YF"])
    elif "현재가_NV" in df.columns:
        df["현재가"] = df["현재가_NV"]
    elif "현재가_YF" in df.columns:
        df["현재가"] = df["현재가_YF"]
    else:
        df["현재가"] = np.nan

    # ── 발행주식수 확정 (v28) ──
    # 우선순위: ① DART istc_totCnt → ② 시총÷현재가 역산
    # index.map() 결과는 ndarray → pd.Series로 명시적 변환
    _smap = shares_map  # {stock_code: int}
    df["shares_dart"] = pd.Series(
        [float(_smap[c]) if c in _smap and _smap[c] else np.nan for c in df.index],
        index=df.index,
        dtype="float64"
    )
    # 역산: 시총(억) × 1억 ÷ 현재가 (float64로 통일)
    cur_s  = pd.to_numeric(df["현재가"], errors="coerce")
    mc_s   = pd.to_numeric(df.get("시가총액(억)", pd.Series(np.nan, index=df.index)), errors="coerce")
    implied = (mc_s * 1e8 / cur_s).where(
        cur_s.notna() & mc_s.notna() & (cur_s > 0) & (mc_s > 0)
    ).round()  # float64 유지 (Int64 NA 혼용 방지)
    # DART 값 우선, 없으면 역산
    df["shares"] = df["shares_dart"].combine_first(implied)  # float64
    dart_ok  = df["shares_dart"].notna().sum()
    total_ok = df["shares"].notna().sum()
    print(f"  [주식수] 확정: DART {dart_ok}개 + 역산 {total_ok - dart_ok}개 = 총 {total_ok}개")

    # 52주위치 NaN → 중립
    df["52주위치"]   = df["52주위치"].fillna(50.0)
    df["52주수익률"] = df["52주수익률"].fillna(0.0)
    df["6개월수익률"] = df["6개월수익률"].fillna(0.0)

    # ROE: DART 우선, 없으면 네이버
    if "roe" in df.columns and "ROE_NV" in df.columns:
        df["ROE"] = df["roe"].combine_first(df["ROE_NV"])
    elif "roe" in df.columns:
        df["ROE"] = df["roe"]
    else:
        df["ROE"] = df.get("ROE_NV", np.nan)

    # ── ROA: DART 계산값 우선, 없으면 네이버 ──
    # DART: roa 컬럼 = net_income / total_assets × 100
    if "roa" in df.columns:
        dart_roa = pd.to_numeric(df["roa"], errors="coerce")
        # DART ROA 유효성 검사: -50~+50% 범위
        dart_roa = dart_roa.where((dart_roa >= -50) & (dart_roa <= 50))
    else:
        dart_roa = pd.Series(np.nan, index=df.index)

    nv_roa = pd.to_numeric(df.get("ROA", pd.Series(np.nan, index=df.index)), errors="coerce")
    nv_roa = nv_roa.where((nv_roa >= -50) & (nv_roa <= 50))

    df["ROA"] = dart_roa.combine_first(nv_roa)

    # DART에도 ROA 없으면 net_income / total_assets로 직접 계산
    if "net_income" in df.columns and "total_assets" in df.columns:
        mask_no_roa = df["ROA"].isna()
        ni  = pd.to_numeric(df.loc[mask_no_roa, "net_income"],   errors="coerce")
        ta  = pd.to_numeric(df.loc[mask_no_roa, "total_assets"], errors="coerce")
        calc_roa = (ni / ta * 100).where((ta > 0) & ta.notna() & ni.notna())
        calc_roa = calc_roa.where((calc_roa >= -50) & (calc_roa <= 50))
        df.loc[mask_no_roa, "ROA"] = calc_roa

    # 부채비율: DART 우선, 없으면 네이버
    if "debt_ratio" in df.columns and "부채비율_NV" in df.columns:
        df["부채비율"] = df["debt_ratio"].combine_first(df["부채비율_NV"])
    elif "debt_ratio" in df.columns:
        df["부채비율"] = df["debt_ratio"]
    else:
        df["부채비율"] = df.get("부채비율_NV", np.nan)

    # PCR = 시가총액 / CFO
    # 시가총액: 네이버 수집값 우선, 없으면 yfinance market_cap (원→억 변환)
    if "시가총액(억)" not in df.columns:
        df["시가총액(억)"] = np.nan

    # yfinance 배치에서 시가총액 보완 (네이버 누락 종목)
    nan_mc_codes = df[df["시가총액(억)"].isna()].index.tolist()
    if nan_mc_codes and HAS_YFINANCE:
        print(f"  [시가총액] yfinance 보완 수집 중 ({len(nan_mc_codes)}개 누락)...")
        def _fetch_marcap_yf(code, market):
            try:
                suffix = ".KS" if market == "KOSPI" else ".KQ"
                info   = yf.Ticker(code + suffix).fast_info
                mc     = getattr(info, "market_cap", np.nan)
                if mc and not np.isnan(mc) and mc > 0:
                    return code, round(mc / 1e8, 0)
            except Exception:
                pass
            return code, np.nan

        with ThreadPoolExecutor(max_workers=10) as ex:
            futs = {ex.submit(_fetch_marcap_yf, c,
                              market_map.get(c,"KOSPI")): c for c in nan_mc_codes}
            for fut in as_completed(futs):
                code, mc = fut.result()
                if mc and not np.isnan(mc):
                    df.loc[code, "시가총액(억)"] = mc
        filled = df.loc[nan_mc_codes, "시가총액(억)"].notna().sum()
        print(f"  ✓ 시가총액 보완: {filled}/{len(nan_mc_codes)}개")
    elif nan_mc_codes and not HAS_YFINANCE:
        print(f"  [시가총액] yfinance 미설치 → 시가총액 보완 스킵 ({len(nan_mc_codes)}개 NaN 유지)")

    # _market_cap_원 컬럼 추가 (엑셀 _marcap 함수용)
    df["_market_cap_원"] = pd.to_numeric(df["시가총액(억)"], errors="coerce") * 1e8

    df["PCR"] = np.nan
    if "시가총액(억)" in df.columns and "cfo" in df.columns:
        mask = df["시가총액(억)"].notna() & df["cfo"].notna() & (df["cfo"] > 0)
        if mask.any():
            pcr_vals = (df.loc[mask, "시가총액(억)"].astype(float) * 1e8) / df.loc[mask, "cfo"].astype(float)
            df["PCR"] = df["PCR"].astype(float)
            df.loc[mask, "PCR"] = pcr_vals

    return df

# ══════════════════════════════════════════════════════════
# 7. 미국 데이터 수집 (yfinance)
# ══════════════════════════════════════════════════════════

def compute_100pt_score(row: dict) -> dict:
    """
    100점 만점 기업 종합 평가 점수 계산.
    가능한 항목만 채점, 데이터 없는 항목은 업종 평균 추정 또는 부분점수 부여.
    반환: {항목별 점수, 합계, 등급, 투자의견}
    """
    def _v(key, default=None):
        """row에서 값 안전 추출"""
        v = row.get(key)
        if v is None or (isinstance(v, float) and np.isnan(v)):
            return default
        try:
            return float(v)
        except Exception:
            return default

    scores = {}

    # ────────────────────────────────────────────
    # 1. 수익성 (20점)
    # ────────────────────────────────────────────
    # 영업이익률 ≥15% → 5점, ≥10%→4, ≥5%→3, ≥0%→1, 적자→0
    opm = _v("영업이익률")
    if opm is None:
        # 영업이익/매출 역산
        oi = _v("op_income") or _v("영업이익_NV")
        rev = _v("매출액")
        if oi and rev and rev > 0:
            opm = oi / rev * 100
    if opm is not None:
        if   opm >= 15: s_opm = 5
        elif opm >= 10: s_opm = 4
        elif opm >=  5: s_opm = 3
        elif opm >=  0: s_opm = 1
        else:           s_opm = 0
    else:
        s_opm = 2  # 데이터 없음 → 중립

    # ROE ≥15%→5, ≥10%→4, ≥8%→3, ≥5%→1, <5%→0
    roe = _v("ROE") or _v("roe") or _v("ROE_NV")
    if roe is not None:
        if   roe >= 15: s_roe = 5
        elif roe >= 10: s_roe = 4
        elif roe >=  8: s_roe = 3
        elif roe >=  5: s_roe = 1
        else:           s_roe = 0
    else:
        s_roe = 2

    # ROA ≥7%→5, ≥5%→4, ≥3%→3, ≥1%→1, <1%→0
    roa = _v("ROA")
    if roa is not None:
        if   roa >= 7: s_roa = 5
        elif roa >= 5: s_roa = 4
        elif roa >= 3: s_roa = 3
        elif roa >= 1: s_roa = 1
        else:          s_roa = 0
    else:
        # ROA = ROE × 자기자본/총자산 근사 (ROE × (1-부채비율/100)/2)
        debt_r = _v("부채비율") or _v("debt_ratio")
        if roe and debt_r:
            eq_ratio = 1 / (1 + max(debt_r, 0) / 100)
            roa_est = roe * eq_ratio
            if   roa_est >= 7: s_roa = 5
            elif roa_est >= 5: s_roa = 4
            elif roa_est >= 3: s_roa = 3
            else:              s_roa = 2
        else:
            s_roa = 2

    # ROIC ≥10%→5 (ROE 대용: 데이터 없으면 ROE로 근사)
    # 실제 ROIC = NOPAT / 투하자본, 여기선 ROE*0.8 근사
    if roe is not None:
        roic_est = roe * 0.8
        if   roic_est >= 10: s_roic = 5
        elif roic_est >=  7: s_roic = 4
        elif roic_est >=  5: s_roic = 3
        elif roic_est >=  3: s_roic = 1
        else:                s_roic = 0
    else:
        s_roic = 2

    scores["수익성"] = s_opm + s_roe + s_roa + s_roic  # max 20

    # ────────────────────────────────────────────
    # 2. 성장성 (20점)
    # ────────────────────────────────────────────
    rev_gr = _v("매출성장률")
    if rev_gr is not None:
        if   rev_gr >= 20: s_rev_g = 5
        elif rev_gr >= 10: s_rev_g = 4
        elif rev_gr >=  5: s_rev_g = 3
        elif rev_gr >=  0: s_rev_g = 1
        else:              s_rev_g = 0
    else:
        s_rev_g = 2  # 중립

    # 영업이익 성장 (네이버 영업이익 증가율 근사: op_income vs rev_gr 대용)
    # 직접 데이터 없으므로 매출성장 × 1.2 (영업레버리지 근사) 사용
    if rev_gr is not None:
        oi_gr_est = rev_gr * 1.2
        if   oi_gr_est >= 20: s_oi_g = 5
        elif oi_gr_est >= 10: s_oi_g = 4
        elif oi_gr_est >=  5: s_oi_g = 3
        elif oi_gr_est >=  0: s_oi_g = 1
        else:                 s_oi_g = 0
    else:
        s_oi_g = 2

    # EPS 성장 (ROE 변화 없으면 매출성장과 동행 가정)
    eps = _v("EPS")
    per = _v("PER") or _v("per")
    cur_price = _v("현재가") or _v("현재가_NV") or _v("현재가_YF")
    # EPS 직접 성장률 데이터 없음 → ROE로 내부유보 성장율 추정
    # g = ROE × (1 - 배당성향)
    payout = _v("dividend_payout")
    if roe and roe > 0:
        retention = 1.0 - min(max((payout or 40) / 100, 0), 1)
        eps_growth_est = roe * retention
        if   eps_growth_est >= 15: s_eps_g = 5
        elif eps_growth_est >= 10: s_eps_g = 4
        elif eps_growth_est >=  7: s_eps_g = 3
        elif eps_growth_est >=  3: s_eps_g = 1
        else:                      s_eps_g = 0
    else:
        s_eps_g = 2

    # 신규 성장동력: PEG < 1 이거나 매출성장+모멘텀이 좋으면 인정
    # PEG = PER / EPS성장률 (EPS성장률 근사)
    if per and per > 0 and rev_gr and rev_gr > 0:
        peg_est = per / rev_gr
        s_growth_new = 5 if peg_est < 1 else (4 if peg_est < 1.5 else (3 if peg_est < 2 else (1 if peg_est < 3 else 0)))
    elif rev_gr and rev_gr >= 10:
        s_growth_new = 4  # 고성장 자체가 성장동력 있음을 시사
    else:
        s_growth_new = 2

    scores["성장성"] = s_rev_g + s_oi_g + s_eps_g + s_growth_new  # max 20

    # ────────────────────────────────────────────
    # 3. 재무안정성 (20점)
    # ────────────────────────────────────────────
    debt = _v("부채비율") or _v("debt_ratio")
    if debt is not None:
        if   debt <= 50:  s_debt = 5
        elif debt <= 100: s_debt = 4
        elif debt <= 150: s_debt = 3
        elif debt <= 200: s_debt = 2
        elif debt <= 300: s_debt = 1
        else:             s_debt = 0
    else:
        s_debt = 2

    cur_ratio = _v("current_ratio")
    if cur_ratio is not None:
        if   cur_ratio >= 200: s_cur = 5
        elif cur_ratio >= 150: s_cur = 4
        elif cur_ratio >= 120: s_cur = 3
        elif cur_ratio >= 100: s_cur = 1
        else:                  s_cur = 0
    else:
        s_cur = 2

    ic = _v("interest_coverage")
    if ic is not None:
        if   ic >= 10: s_ic = 5
        elif ic >=  5: s_ic = 4
        elif ic >=  3: s_ic = 3
        elif ic >=  1: s_ic = 1
        else:          s_ic = 0
    else:
        s_ic = 3  # 이자비용 없는 순현금 기업 가능성

    # 자본잠식 여부: equity > 0 이면 잠식 없음
    equity = _v("equity")
    if equity is not None:
        s_no_impair = 5 if equity > 0 else 0
    else:
        s_no_impair = 4  # 모름 → 대부분 잠식 없음

    scores["재무안정성"] = s_debt + s_cur + s_ic + s_no_impair  # max 20

    # ────────────────────────────────────────────
    # 4. 현금흐름 (20점)
    # ────────────────────────────────────────────
    cfo = _v("cfo")
    ni  = _v("net_income")
    fcf = _v("fcf")

    # CFO 연속 플러스 (단일 연도 데이터 → 현재 연도만 판단)
    if cfo is not None:
        s_cfo_pos = 5 if cfo > 0 else 0
    else:
        s_cfo_pos = 2

    # FCF 연속 플러스
    if fcf is not None:
        s_fcf_pos = 5 if fcf > 0 else (2 if fcf >= -abs(cfo or 1) * 0.3 else 0)
    else:
        s_fcf_pos = 2

    # CFO > 순이익 (이익 품질)
    if cfo is not None and ni is not None:
        s_cfo_ni = 5 if cfo >= ni else (3 if cfo >= ni * 0.7 else (1 if cfo > 0 else 0))
    else:
        s_cfo_ni = 2

    # 현금성자산 증가 (단일 시점 → 현금/총자산 비율로 대체)
    cash = _v("cash")
    ta   = _v("total_assets")
    if cash is not None and ta and ta > 0:
        cash_ratio = cash / ta * 100
        if   cash_ratio >= 20: s_cash_inc = 5
        elif cash_ratio >= 10: s_cash_inc = 4
        elif cash_ratio >=  5: s_cash_inc = 3
        else:                  s_cash_inc = 1
    else:
        s_cash_inc = 2

    scores["현금흐름"] = s_cfo_pos + s_fcf_pos + s_cfo_ni + s_cash_inc  # max 20

    # ────────────────────────────────────────────
    # 5. 밸류에이션 (10점)
    # ────────────────────────────────────────────
    # PER (업종 평균 대비) → 한국 상장사 평균 ~12x, 저평가 기준 10x 이하
    per_ref = _v("PER_결산") or per
    if per_ref is not None and per_ref > 0:
        if   per_ref <= 8:  s_per = 4
        elif per_ref <= 12: s_per = 3
        elif per_ref <= 15: s_per = 2
        elif per_ref <= 20: s_per = 1
        else:               s_per = 0
    else:
        s_per = 1  # 적자/음수 PER

    # PBR (업종 평균 대비)
    pbr = _v("PBR") or _v("pbr")
    if pbr is not None and pbr > 0:
        if   pbr <= 0.5: s_pbr = 3
        elif pbr <= 1.0: s_pbr = 3
        elif pbr <= 1.5: s_pbr = 2
        elif pbr <= 2.0: s_pbr = 1
        else:            s_pbr = 0
    else:
        s_pbr = 1

    # PEG (PER / EPS성장률)
    if per_ref and per_ref > 0 and rev_gr and rev_gr > 0:
        peg = per_ref / max(rev_gr, 1)
        if   peg < 0.5: s_peg = 3
        elif peg < 1.0: s_peg = 3
        elif peg < 1.5: s_peg = 2
        elif peg < 2.0: s_peg = 1
        else:           s_peg = 0
    else:
        s_peg = 1

    scores["밸류에이션"] = s_per + s_pbr + s_peg  # max 10

    # ────────────────────────────────────────────
    # 6. 경영 및 주주친화 (10점)
    # ────────────────────────────────────────────
    div_yield = _v("DIV") or 0

    # 배당: 꾸준함 → DIV수익률로 대용
    if   div_yield >= 3: s_div_mgmt = 3
    elif div_yield >= 2: s_div_mgmt = 3
    elif div_yield >= 1: s_div_mgmt = 2
    elif div_yield >  0: s_div_mgmt = 1
    else:                s_div_mgmt = 0

    # 자사주 매입/소각: 배당성향+배당수익률 조합으로 추정
    # (직접 데이터 없음) → 배당수익률 높으면 주주환원 적극적
    if div_yield >= 2 or (payout and 20 <= (payout or 0) <= 60):
        s_buyback = 3
    elif div_yield >= 1:
        s_buyback = 2
    else:
        s_buyback = 1

    # 횡령배임 없음: Altman + 이자보상 + FCF 복합 판단
    # Altman 단독이 아닌 복합 지표로 재무 건전성 평가
    altman  = _v("altman_z")
    ic_val  = _v("interest_coverage")
    nd      = _v("net_debt_ratio")   # 순부채비율
    ebitda  = _v("ebitda")
    rev     = _v("매출액")

    # 재무건전성 점수: Altman은 보조, 이자보상+FCF+부채를 주로 봄
    ic_ok  = ic_val is not None and ic_val >= 3
    fcf_ok = fcf is not None and fcf > 0
    debt_ok= debt is None or debt < 150

    if ic_ok and fcf_ok and debt_ok:
        s_no_fraud = 2   # 핵심 3개 모두 OK
    elif ic_ok and fcf_ok:
        s_no_fraud = 2   # 이자보상+FCF OK (부채 좀 있어도 실위험 낮음)
    elif ic_ok or fcf_ok:
        s_no_fraud = 1   # 둘 중 하나만 OK
    else:
        s_no_fraud = 0   # 둘 다 문제

    # 내부자 매수: FCF Yield 기반
    mkcap_100 = _v("시가총액(억)")
    if fcf and mkcap_100 and mkcap_100 > 0:
        fy = fcf / (mkcap_100 * 1e8) * 100
        s_insider = 2 if fy >= 5 else (1 if fy >= 2 else 0)
    else:
        s_insider = 1

    scores["경영주주친화"] = s_div_mgmt + s_buyback + s_no_fraud + s_insider  # max 10

    # ────────────────────────────────────────────
    # 7. 정성 보정 (±10점) — 정교화 버전
    # ────────────────────────────────────────────
    bonus = 0
    malus = 0
    bonus_notes = []
    malus_notes = []

    # ── 가산점 요인 ──

    # +5점: 독점적 해자 (영업이익률 25%↑ + ROE 20%↑)
    if opm and opm >= 25 and roe and roe >= 20:
        bonus += 5
        bonus_notes.append(f"독점적 해자(OPM {opm:.0f}%·ROE {roe:.0f}%)")
    elif opm and opm >= 20:
        bonus += 3
        bonus_notes.append(f"높은 수익성(OPM {opm:.0f}%)")
    elif opm and opm >= 15 and roe and roe >= 15:
        bonus += 2
        bonus_notes.append(f"양호한 수익성(OPM {opm:.0f}%·ROE {roe:.0f}%)")

    # +5점: 성장산업
    sector = str(row.get("업종", "") or "")
    growth_keywords = ["반도체","전기전자","IT","바이오","제약","에너지","전력","소프트웨어",
                       "인터넷","AI","통신","자동화","2차전지","배터리","방산","우주","로봇"]
    if any(kw in sector for kw in growth_keywords):
        bonus += 5
        bonus_notes.append(f"성장산업({sector})")
    elif rev_gr and rev_gr >= 20:
        bonus += 3
        bonus_notes.append(f"고성장(매출+{rev_gr:.0f}%)")
    elif rev_gr and rev_gr >= 10:
        bonus += 1
        bonus_notes.append(f"안정성장(매출+{rev_gr:.0f}%)")

    # ── 감점 요인 (복합 조건으로 오탐 방지) ──

    # 감점 A: 실질 부도위험
    # Altman 단독 X → Altman + 이자보상 + FCF 복합으로만 감점
    real_danger = False
    if altman is not None and altman < 1.8:
        if ic_val is not None and ic_val < 2:
            # Altman 낮고 이자보상도 낮음 → 실제 위험
            malus += 5
            malus_notes.append(f"부도위험(Z={altman:.1f}·이자보상{ic_val:.1f}배)")
            real_danger = True
        elif ic_val is not None and ic_val < 3 and fcf is not None and fcf < 0:
            # Altman 낮고 FCF도 음수
            malus += 4
            malus_notes.append(f"재무위험(Z={altman:.1f}·FCF음수)")
            real_danger = True
        elif ic_val is None or (ic_val >= 3 and (fcf is None or fcf >= 0)):
            # Altman 낮지만 이자보상OK·FCF OK → 업종 구조적 문제일 가능성
            malus += 1
            malus_notes.append(f"Altman낮음(Z={altman:.1f}, 단 이자보상·FCF 양호)")
    elif altman is not None and altman < 2.5:
        if ic_val is not None and ic_val < 2:
            malus += 2
            malus_notes.append(f"Altman회색+이자보상낮음(Z={altman:.1f}·IC{ic_val:.1f}배)")
        # 이자보상 OK면 회색지대여도 감점 없음

    # 감점 B: FCF 음수 (단독보다 부채와 결합 시 위험)
    if not real_danger:  # 위에서 이미 감점 안 했을 때만
        if fcf is not None and fcf < 0:
            if debt is not None and debt > 200:
                malus += 4
                malus_notes.append(f"FCF음수+고부채({debt:.0f}%)")
            elif debt is not None and debt > 150:
                malus += 2
                malus_notes.append(f"FCF음수+부채주의({debt:.0f}%)")
            elif ic_val is not None and ic_val < 2:
                malus += 3
                malus_notes.append(f"FCF음수+이자보상낮음({ic_val:.1f}배)")
            else:
                malus += 1
                malus_notes.append("FCF음수(단기, 이자보상 OK)")

    # 감점 C: 매출 감소 (성장성 역행)
    if rev_gr is not None and rev_gr < -10:
        malus += 3
        malus_notes.append(f"매출급감({rev_gr:.0f}%)")
    elif rev_gr is not None and rev_gr < 0:
        malus += 1
        malus_notes.append(f"매출감소({rev_gr:.0f}%)")

    # 감점 D: 순부채/EBITDA 과도 (실질 부채 부담)
    if ebitda and ebitda > 0 and nd is not None:
        # 순부채비율을 EBITDA 대용으로 활용
        if nd > 300:
            malus += 2
            malus_notes.append(f"순부채과다({nd:.0f}%)")

    bonus = min(bonus, 10)
    malus = min(malus, 10)
    scores["정성보정"] = bonus - malus

    # ────────────────────────────────────────────
    # 최종 합산
    # ────────────────────────────────────────────
    base = sum(v for k, v in scores.items() if k != "정성보정")
    total = min(max(base + scores["정성보정"], 0), 110)  # 보너스로 최대 110

    # 100점 기준 등급
    if   total >= 90: grade_100 = "★★★★★ 최상급"
    elif total >= 80: grade_100 = "★★★★  우수"
    elif total >= 70: grade_100 = "★★★   양호"
    elif total >= 60: grade_100 = "★★    보통"
    elif total >= 50: grade_100 = "★     주의"
    else:             grade_100 = "      투자재검토"

    # 5대 필수 체크 - 이모지 대신 실제 수치 표시
    # CFO: 실제값(억) + OK/NG
    chk1 = f"{round(cfo/1e8,1):+.1f}억 ✅" if cfo and cfo > 0 else (f"{round(cfo/1e8,1):+.1f}억 ❌" if cfo else "N/A")
    # 자본잠식: 자기자본 표시
    chk2 = f"✅ 정상" if equity and equity > 0 else ("❌ 잠식" if equity is not None else "N/A")
    # 부채비율: 실제 수치 표시
    chk3 = f"{debt:.0f}% ✅" if debt and debt < 150 else (f"{debt:.0f}% ⚠" if debt and debt < 300 else (f"{debt:.0f}% ❌" if debt else "N/A"))
    # ROE: 실제 수치 표시
    chk4 = f"{roe:.1f}% ✅" if roe and roe >= 10 else (f"{roe:.1f}% ⚠" if roe and roe >= 5 else (f"{roe:.1f}% ❌" if roe else "N/A"))
    # PER: 실제 수치 표시
    chk5 = f"{per_ref:.1f}x ✅" if per_ref and per_ref > 0 and per_ref < 15 else (f"{per_ref:.1f}x ⚠" if per_ref and per_ref > 0 and per_ref < 25 else (f"{per_ref:.1f}x ❌" if per_ref and per_ref > 0 else "적자/N/A"))

    bonus_str = ", ".join(bonus_notes) if bonus_notes else "-"
    malus_str = ", ".join(malus_notes) if malus_notes else "-"

    return {
        "100점_수익성":     scores["수익성"],
        "100점_성장성":     scores["성장성"],
        "100점_재무안정":   scores["재무안정성"],
        "100점_현금흐름":   scores["현금흐름"],
        "100점_밸류":       scores["밸류에이션"],
        "100점_주주친화":   scores["경영주주친화"],
        "100점_정성보정":   scores["정성보정"],
        "100점_합계":       total,
        "100점_등급":       grade_100,
        "체크_CFO":         chk1,
        "체크_자본잠식":    chk2,
        "체크_부채비율":    chk3,
        "체크_ROE":         chk4,
        "체크_PER":         chk5,
        "보너스근거":       bonus_str,
        "감점근거":         malus_str,
    }


# ══════════════════════════════════════════════════════════
# 8-B. 매매 시그널 엔진 (현재가 기반)
# ══════════════════════════════════════════════════════════
#
#  [시그널 체계]
#   ■■■ 강력매수    복합점수≥85 & 100점≥80 & 괴리율≥+20%
#   ■■  매수        복합점수≥75 & 100점≥70 & 괴리율≥+10%
#   ■   관심/비중확대 복합점수≥65 & 100점≥60 & 괴리율≥0%
#   ─   보유        조건 애매, 특이사항 없음
#   ▽   비중축소    복합점수<60 또는 괴리율<-10%
#   ▼▼  매도        복합점수<50 또는 괴리율<-20%
#   ▼▼▼ 즉시매도    하드필터 탈락 조건 근접
#
#  [가격 밴드] ── 현재가 기준 매수/손절, 적정가 기준 목표가
#   적극매수선   = 현재가 × 0.95  (지금보다 5% 더 빠지면 추가 매수)
#   매수선       = 현재가          (현재 매수 가능 기준가)
#   1차목표가    = 적정가 평균     (적정가 도달 시 30% 익절)
#   2차목표가    = 적정가 평균 × 1.15 (추가 상승 여력)
#   추격매수경보 = 적정가 평균 × 1.20 (이 가격 이상은 고평가)
#   손절선       = 현재가 × 0.85  (15% 하락 시 손절)
#
# ══════════════════════════════════════════════════════════

def compute_trading_signal(row: dict, score_100: dict) -> dict:
    """매매 시그널과 가격 밴드 계산"""

    def _v(key, default=None):
        v = row.get(key)
        if v is None or (isinstance(v, float) and np.isnan(v)):
            return default
        try:
            return float(v)
        except Exception:
            return default

    cur      = _v("현재가") or _v("현재가_NV") or _v("현재가_YF")
    avg_tgt  = _v("적정가_평균")
    per_tgt  = _v("적정가_PER")
    pbr_tgt  = _v("적정가_PBR")
    grm_tgt  = _v("적정가_Graham")
    # ── 업종 특수 구조 감지 (v30) ──
    # 금융·보험·지주는 일반기업 기준의 부채비율·FCF·Altman이 의미없음
    sector_str = str(row.get("업종", "") or "")
    is_finance_sector = any(kw in sector_str for kw in [
        "보험","은행","증권","카드","금융","지주","저축","캐피탈","리츠","자산운용"
    ])

    upside   = _v("괴리율(%)")
    comp     = _v("복합점수", 0)
    pt100    = score_100.get("100점_합계", 0)
    cfo      = _v("cfo")
    debt     = _v("부채비율") or _v("debt_ratio")
    altman   = _v("altman_z")
    pos52    = _v("52주위치", 50)
    ret6m    = _v("6개월수익률", 0)
    fcf      = _v("fcf")
    div      = _v("DIV", 0)
    ic       = _v("interest_coverage")

    # ── 위험 신호 감지 (3단계 복합 판단) ──
    danger_flags  = []
    caution_flags = []

    # ① CFO 음수 — 보험/금융은 투자CF가 크게 음수이므로 CFO만 본다
    if cfo is not None and cfo < 0:
        if is_finance_sector:
            caution_flags.append("CFO음수(금융업참고용)")  # 위험으로 처리 안 함
        elif debt is not None and debt > 150:
            danger_flags.append("CFO음수+부채주의")
        else:
            caution_flags.append("CFO음수(부채는 양호)")

    # ② 부채비율 — 보험/금융/지주는 구조적 고부채, 예외 처리
    if debt is not None and debt > 300:
        if is_finance_sector:
            caution_flags.append(f"부채비율{debt:.0f}%(금융업정상)")
        else:
            danger_flags.append(f"부채비율과다({debt:.0f}%)")
    elif debt is not None and debt > 200:
        if not is_finance_sector:
            caution_flags.append(f"부채비율높음({debt:.0f}%)")

    # ③ Altman Z — 금융/보험은 Altman 모델 적용 불가
    if altman is not None and not is_finance_sector:
        if altman < 1.8:
            if ic is not None and ic < 2:
                danger_flags.append(f"부도위험(Z={altman:.1f}·IC{ic:.1f}배)")
            elif fcf is not None and fcf < 0:
                danger_flags.append(f"재무위험(Z={altman:.1f}·FCF음수)")
            else:
                caution_flags.append(f"Altman낮음(Z={altman:.1f}, 이자보상·FCF 양호)")
        elif altman < 2.5:
            if ic is not None and ic < 2:
                caution_flags.append(f"Altman회색+IC낮음(Z={altman:.1f})")

    # ④ 이자보상배율 — 보험/금융은 제외
    if ic is not None and not is_finance_sector:
        if ic < 1.0:
            danger_flags.append(f"이자보상위험({ic:.1f}배)")
        elif ic < 2.0:
            caution_flags.append(f"이자보상낮음({ic:.1f}배)")

    # ⑤ FCF 음수+고부채 — 보험은 FCF 개념 다름
    if fcf is not None and fcf < 0 and debt is not None and debt > 200:
        if not is_finance_sector and "CFO음수+부채주의" not in danger_flags:
            danger_flags.append(f"FCF음수+고부채({debt:.0f}%)")

    # ⑥ 매출 급감
    rev_gr_s = _v("매출성장률")
    if rev_gr_s is not None and rev_gr_s < -15:
        danger_flags.append(f"매출급감({rev_gr_s:.0f}%)")
    elif rev_gr_s is not None and rev_gr_s < -5:
        caution_flags.append(f"매출감소({rev_gr_s:.0f}%)")

    # 위험 등급 결정
    if danger_flags:
        danger_level = "🔴 실위험"
        danger_str = "🔴 " + " / ".join(danger_flags)
        if caution_flags:
            danger_str += "  🟡 " + " / ".join(caution_flags)
    elif caution_flags:
        danger_level = "🟡 주의"
        danger_str = "🟡 " + " / ".join(caution_flags)
    else:
        danger_level = "없음"
        danger_str = "없음"

    # ── 가격 밴드 계산 ──
    # 매수선/적극매수선 → 현재가 기준
    # 목표가/추격매수경보 → 적정가 기준
    if cur and cur > 0:
        price_buy_strong = round(cur * 0.95)
        price_buy        = round(cur)
        stop_loss        = round(cur * 0.85)
    else:
        price_buy_strong = price_buy = stop_loss = None

    if avg_tgt and avg_tgt > 0:
        if avg_tgt >= cur:
            # ── 정상: 적정가 > 현재가 → 적정가 기반 목표가 ──
            price_target_1   = round(avg_tgt)
            price_target_2   = round(avg_tgt * 1.15)
            price_overbought = round(avg_tgt * 1.20)
        else:
            # ── v26 신규: 적정가 < 현재가 → 기술적 목표가 사용 ──
            # 현재가가 이미 퀀트 적정가를 초과한 상태
            # 1차목표가 = 현재가 × 1.10  (단기 +10%)
            # 2차목표가 = 현재가 × 1.20  (중기 +20%)
            # 추격매수경보 = 현재가 × 1.30  (과열 기준)
            # (적정가는 별도 컬럼에 유지, 가격밴드는 기술적 기준으로)
            price_target_1   = round(cur * 1.10)
            price_target_2   = round(cur * 1.20)
            price_overbought = round(cur * 1.30)
    else:
        price_target_1 = price_target_2 = price_overbought = None

    # ── 현재가 위치 진단 (적정가 대비) ──
    position_diag = "N/A"
    price_ok = False   # 가격 관점에서 매수 가능한지
    if cur and avg_tgt and avg_tgt > 0:
        ratio = cur / avg_tgt * 100
        if   ratio < 60:
            position_diag = "🟦 적극매수구간 (적정가 대비 40%↑ 저평가)"
            price_ok = True
        elif ratio < 80:
            position_diag = "🟩 매수구간 (적정가 대비 20~40% 저평가)"
            price_ok = True
        elif ratio < 95:
            position_diag = "🟩 매수구간 (적정가 대비 5~20% 저평가)"
            price_ok = True
        elif ratio <= 105:
            position_diag = "⬜ 적정구간 (적정가 근접, 보유)"
            price_ok = False
        elif ratio <= 120:
            position_diag = "🟨 소폭고평가 (비중축소 고려)"
            price_ok = False
        else:
            position_diag = "🟥 고평가구간 (매도 검토)"
            price_ok = False
    elif avg_tgt is None:
        # 적정가 없는 경우 가격 판단 보류
        price_ok = None

    # ── 시그널 결정 ──
    # 핵심 원칙: 1차목표가(적정가) < 현재가 이면 매수 시그널 불가
    signal = "─ 보유"
    signal_color = "FFCC00"

    # 1단계: 위험/하락 시그널 (가격 무관하게 우선 적용)
    if len(danger_flags) >= 2:
        signal = "▼▼▼ 즉시매도/회피"
        signal_color = "C00000"
    elif len(danger_flags) == 1:
        signal = "▼▼ 매도 검토"
        signal_color = "FF4444"
    elif comp < 50 or pt100 < 50:
        signal = "▼▼ 매도 검토"
        signal_color = "FF4444"
    elif upside is not None and upside < -20:
        # 고평가 20% 이상 → 매도
        signal = "▼▼ 매도 검토"
        signal_color = "FF4444"
    elif upside is not None and upside < -10:
        # 고평가 10~20% → 비중축소
        signal = "▽ 비중축소"
        signal_color = "FF9900"

    # 2단계: 매수 시그널 — 반드시 적정가 > 현재가 조건 충족해야 함
    elif price_ok:
        # 적정가가 현재가보다 높을 때만 매수 시그널 가능
        if comp >= 85 and pt100 >= 80 and (upside is None or upside >= 15):
            signal = "■■■ 강력매수"
            signal_color = "0070C0"
        elif comp >= 75 and pt100 >= 70 and (upside is None or upside >= 5):
            signal = "■■ 매수"
            signal_color = "2E75B6"
        elif comp >= 65 and pt100 >= 60:
            signal = "■ 관심/비중확대"
            signal_color = "92D050"
        else:
            signal = "─ 보유"
            signal_color = "FFCC00"
    elif price_ok is None:
        # 적정가 없음 → 점수로만 판단
        if comp >= 85 and pt100 >= 80:
            signal = "■■ 매수"       # 적정가 없으면 강력매수 대신 매수
            signal_color = "2E75B6"
        elif comp >= 75 and pt100 >= 70:
            signal = "■ 관심/비중확대"
            signal_color = "92D050"
        # 나머지는 보유 유지
    else:
        # price_ok=False: 적정가 근접/고평가 → 보유 또는 비중축소
        if upside is not None and upside < -10:
            signal = "▽ 비중축소"
            signal_color = "FF9900"
        # 나머지는 보유 유지

    # 주의 신호만 있을 때: 매수 시그널 한 단계 낮춤
    if caution_flags and not danger_flags:
        if signal == "■■■ 강력매수":
            signal = "■■ 매수"
            signal_color = "2E75B6"
        elif signal == "■■ 매수":
            signal = "■ 관심/비중확대"
            signal_color = "92D050"

    # ── 보유 전략 (현재가/목표가 관계 반영) ──
    hold_strategy = ""
    has_upside = price_target_1 and cur and price_target_1 > cur  # 목표가 > 현재가
    upside_pct = round((price_target_1 - cur) / cur * 100) if (price_target_1 and cur) else None
    tgt_str = f"{price_target_1:,}원" if price_target_1 else "N/A"
    stp_str = f"{stop_loss:,}원" if stop_loss else "-"

    if "강력매수" in signal:
        hold_strategy = f"분할매수 추천 (2~3회 진입) | 1차목표가 {tgt_str}(+{upside_pct}%) | 손절 {stp_str}"
    elif "■■ 매수" in signal:
        hold_strategy = f"매수 진입 | 1차목표가 {tgt_str} 도달 시 30% 익절 | 손절 {stp_str}"
    elif "관심" in signal:
        hold_strategy = f"소량 선취매 후 추가 확인 | 목표 {tgt_str}"
    elif "보유" in signal:
        if has_upside:
            hold_strategy = f"1차목표가({tgt_str}) 도달 시 30% 익절 고려"
        else:
            hold_strategy = f"적정가({tgt_str}) 이미 하회 중 — 추가매수 자제, 손절 {stp_str} 확인"
    elif "비중축소" in signal:
        hold_strategy = f"추가 매수 자제, 기존 보유분 30~50% 정리 | 손절 {stp_str}"
    elif "매도" in signal or "즉시매도" in signal:
        hold_strategy = f"손절선({stp_str}) 이탈 시 즉시 매도"

    # ── 기술적 지표 신호 읽기 (v32) ──
    ta_signal_str = str(row.get("기술신호", "") or "─ 중립")
    ta_summary_str = str(row.get("기술지표요약", "") or "")
    _ta_buy_raw  = row.get("기술매수신호수", 0)
    _ta_sell_raw = row.get("기술매도신호수", 0)
    ta_buy    = int(_ta_buy_raw)  if (_ta_buy_raw  is not None and _ta_buy_raw  == _ta_buy_raw)  else 0
    ta_sell   = int(_ta_sell_raw) if (_ta_sell_raw is not None and _ta_sell_raw == _ta_sell_raw) else 0
    macd_sig  = str(row.get("MACD신호", "중립") or "중립")
    rsi_val   = _v("RSI14")
    rsi_sig   = str(row.get("RSI신호", "중립") or "중립")
    bb_pct    = _v("BB%B")
    bb_sig    = str(row.get("BB신호", "중립") or "중립")
    obv_sig   = str(row.get("OBV신호", "중립") or "중립")

    # ── 핵심 투자 포인트 (기술지표 포함) ──
    inv_points = []
    if div and div >= 3:
        inv_points.append(f"배당수익률 {div:.1f}% (안정적 인컴)")
    if upside and upside >= 20:
        inv_points.append(f"적정가 대비 {upside:.0f}% 저평가")
    if pos52 <= 40:
        inv_points.append(f"52주 저점 근처 ({pos52:.0f}%)")
    if ret6m >= 15:
        inv_points.append(f"6개월 모멘텀 강함 (+{ret6m:.0f}%)")
    if fcf and fcf > 0 and cur:
        mkcap_100 = _v("시가총액(억)")
        if mkcap_100 and mkcap_100 > 0:
            fy = fcf / (mkcap_100 * 1e8) * 100
            if fy >= 5:
                inv_points.append(f"FCF Yield {fy:.1f}% (현금창출력 우수)")
    # 기술적 포인트 추가
    if rsi_val and rsi_val < 30:
        inv_points.append(f"RSI {rsi_val:.0f} 과매도 (반등 기대)")
    elif rsi_val and rsi_val > 70:
        inv_points.append(f"RSI {rsi_val:.0f} 과매수 (조정 주의)")
    if "골든크로스" in macd_sig or macd_sig == "매수":
        inv_points.append("MACD 골든크로스 (상승전환)")
    if bb_pct is not None and bb_pct < 0:
        inv_points.append(f"BB 하단밴드 이탈 (과매도, %B={bb_pct:.0f}%)")
    if ta_buy >= 3:
        inv_points.append(f"기술지표 {ta_buy}/4 매수신호 집중")
    # 박스권 돌파 포인트 (v41g 신규)
    box_sig_txt = str(row.get("박스이탈신호", "") or "")
    box_period  = row.get("박스기간(일)")
    if "모멘텀후보" in box_sig_txt:
        inv_points.append(f"📦 박스권({box_period}일) 상단 돌파+거래량급증 (뉴스·재료성 급등 후보)")
    elif "돌파대기" in box_sig_txt:
        inv_points.append(f"📦 박스권({box_period}일) 상단 근접 — 돌파 시 모멘텀 주목")
    if not inv_points:
        inv_points.append("기본 지표 충족")
    inv_point_str = " / ".join(inv_points)

    # ── 보유 전략 (기술신호 포함) ──
    has_upside = price_target_1 and cur and price_target_1 > cur
    upside_pct = round((price_target_1 - cur) / cur * 100) if (price_target_1 and cur) else None
    tgt_str = f"{price_target_1:,}원" if price_target_1 else "N/A"
    stp_str = f"{stop_loss:,}원" if stop_loss else "-"
    ta_addon = f" | 기술:{ta_signal_str}" if ta_signal_str and "중립" not in ta_signal_str else ""

    if "강력매수" in signal:
        hold_strategy = f"분할매수 추천 (2~3회 진입) | 1차목표가 {tgt_str}(+{upside_pct}%) | 손절 {stp_str}{ta_addon}"
    elif "■■ 매수" in signal:
        hold_strategy = f"매수 진입 | 1차목표가 {tgt_str} 도달 시 30% 익절 | 손절 {stp_str}{ta_addon}"
    elif "관심" in signal:
        hold_strategy = f"소량 선취매 후 추가 확인 | 목표 {tgt_str}{ta_addon}"
    elif "보유" in signal:
        if has_upside:
            hold_strategy = f"1차목표가({tgt_str}) 도달 시 30% 익절 고려{ta_addon}"
        else:
            hold_strategy = f"적정가({tgt_str}) 이미 하회 중 — 추가매수 자제, 손절 {stp_str} 확인{ta_addon}"
    elif "비중축소" in signal:
        hold_strategy = f"추가 매수 자제, 기존 보유분 30~50% 정리 | 손절 {stp_str}{ta_addon}"
    elif "매도" in signal or "즉시매도" in signal:
        hold_strategy = f"손절선({stp_str}) 이탈 시 즉시 매도{ta_addon}"
    else:
        hold_strategy = ""

    return {
        "매매시그널":      signal,
        "현재가위치":      position_diag,
        "적극매수선":      price_buy_strong,
        "매수선":          price_buy,
        "1차목표가":       price_target_1,
        "2차목표가":       price_target_2,
        "추격매수경보":    price_overbought,
        "손절선":          stop_loss,
        "보유전략":        hold_strategy,
        "핵심투자포인트":  inv_point_str,
        "위험신호":        danger_str,
        "_signal_color":   signal_color,
        # 기술신호 (v32)
        "기술신호":        ta_signal_str,
        "기술지표요약":    ta_summary_str,
    }


# ══════════════════════════════════════════════════════════
# 8. 하드 필터 (DART 재무 기준)
# ══════════════════════════════════════════════════════════

def apply_hard_filter(df: pd.DataFrame, args) -> pd.DataFrame:
    """
    하드 필터: 데이터가 있는 종목에만 조건 적용
    NaN(데이터 미수집) 종목은 무조건 통과 → 전체종목 시트에 포함
    금융·보험·지주 업종은 부채비율·이자보상 필터 면제 (v30)
    """
    n = len(df)
    cfo_col  = "cfo"  if "cfo"  in df.columns else None
    ni_col   = "net_income" if "net_income" in df.columns else None
    debt_col = "부채비율" if "부채비율" in df.columns else "debt_ratio"
    ic_col   = "interest_coverage" if "interest_coverage" in df.columns else None

    # 금융/보험 업종 마스크
    FINANCE_KW = ["보험","은행","증권","카드","금융","지주","저축","캐피탈","리츠","자산운용"]
    if "업종" in df.columns:
        is_finance = df["업종"].fillna("").apply(
            lambda s: any(kw in str(s) for kw in FINANCE_KW)
        )
    else:
        is_finance = pd.Series(False, index=df.index)

    mask = pd.Series(True, index=df.index)

    # CFO > 0 (데이터 있을 때만)
    if cfo_col:
        mask &= df[cfo_col].isna() | (df[cfo_col] > 0)
        if ni_col:
            both = df[cfo_col].notna() & df[ni_col].notna()
            mask &= ~both | (df[cfo_col] >= df[ni_col])

    # 부채비율 — 금융/보험 업종 면제
    if debt_col in df.columns:
        debt_filter = df[debt_col].isna() | (df[debt_col] < args.max_debt)
        mask &= debt_filter | is_finance  # 금융업은 통과

    # 이자보상배율 — 금융/보험 업종 면제
    if ic_col and ic_col in df.columns:
        ic_filter = df[ic_col].isna() | (df[ic_col] > args.min_ic)
        mask &= ic_filter | is_finance  # 금융업은 통과

    out = df[mask].copy()
    n_data   = (df[cfo_col].notna()).sum() if cfo_col else 0
    n_failed = n - len(out)
    print(f"  [필터] 하드 필터: {n}개 → {len(out)}개 통과 "
          f"(탈락 {n_failed}개 / DART 수집 {n_data}개)")
    return out

# ══════════════════════════════════════════════════════════
# 9. 7팩터 점수 계산
# ══════════════════════════════════════════════════════════

def compute_scores(df: pd.DataFrame, args) -> pd.DataFrame:
    result = df.copy()

    pbr_col = "PBR" if "PBR" in result.columns else "pbr"
    roe_col = "ROE" if "ROE" in result.columns else "roe"

    # PBR/ROE 조건은 점수에 반영 (필터로 탈락시키지 않음)
    # → 모든 종목이 전체종목 시트에 포함되고, 점수로 순위만 결정
    filtered = result.copy()
    print(f"  [스코어] 점수 계산 대상: {len(filtered)}개 전체 종목")

    if len(filtered) == 0:
        return pd.DataFrame()

    # ── ① 가치 점수 (PBR 30% + ROE 30% + FCF Yield 20% + EV/EBITDA 20%) ──
    def _pbr_s(v):
        if v is None or np.isnan(v): return 0
        if v <= 0.3: return 100
        if v >= 3.0: return 0
        return max(0, (1 - (v-0.3)/2.7) * 100)

    def _roe_s(v):
        if v is None or np.isnan(v): return 0
        if v >= 50: return 100
        if v <= 8:  return 0
        return max(0, (v-8)/42 * 100)

    def _fcf_yield_s(row):
        """FCF Yield = FCF / 시가총액 (%) - 높을수록 저평가"""
        fcf   = _safe(row.get("fcf"))
        mkcap = _safe(row.get("시가총액(억)"))
        if fcf is None or mkcap is None or mkcap <= 0:
            return 50  # 데이터 없으면 중립
        fy = fcf / (mkcap * 1e8) * 100  # 억→원 환산
        if   fy >= 10: return 100
        elif fy >=  8: return 90
        elif fy >=  5: return 75
        elif fy >=  3: return 55
        elif fy >=  0: return 35
        else:          return 10  # FCF 음수

    def _ev_ebitda_s(row):
        """EV/EBITDA - 낮을수록 저평가"""
        mkcap  = _safe(row.get("시가총액(억)"))
        tl     = _safe(row.get("total_liabilities"))
        cash   = _safe(row.get("cash"))
        ebitda = _safe(row.get("ebitda"))
        if mkcap is None or ebitda is None or ebitda <= 0:
            return 50  # 데이터 없으면 중립
        # EV = 시가총액 + 순부채
        debt_net = (tl or 0) - (cash or 0)
        ev = mkcap * 1e8 + debt_net
        ratio = ev / ebitda
        if   ratio <=  6: return 100
        elif ratio <=  8: return 85
        elif ratio <= 12: return 65
        elif ratio <= 16: return 45
        elif ratio <= 20: return 25
        else:             return 0

    filtered["가치점수"] = (
        filtered[pbr_col].apply(_pbr_s) * 0.30 +
        filtered[roe_col].apply(_roe_s) * 0.30 +
        filtered.apply(_fcf_yield_s, axis=1) * 0.20 +
        filtered.apply(_ev_ebitda_s, axis=1) * 0.20
    ).round(1)

    # ── ② 모멘텀 점수 (52주 위치 기반 비선형) ──
    def _mom_s(pos):
        if pos is None or np.isnan(pos): return 0
        pos = float(pos)
        bps = [(0,0),(40,85),(57.5,100),(75,85),(100,0)]
        for i in range(len(bps)-1):
            x0,y0 = bps[i]; x1,y1 = bps[i+1]
            if x0 <= pos <= x1:
                t = (pos-x0)/(x1-x0)
                return y0 + t*(y1-y0)
        return 0

    def _vol_mom_s(row):
        """거래량 모멘텀 점수 (yfinance에서 수집한 경우)"""
        vm = _safe(row.get("vol_ratio"))  # 20일 평균 대비 거래량 비율
        if vm is None:
            return 50  # 중립
        if   vm >= 3.0: return 100
        elif vm >= 2.0: return 85
        elif vm >= 1.5: return 70
        elif vm >= 1.0: return 50
        elif vm >= 0.7: return 30
        else:           return 10

    filtered["모멘텀점수"] = (
        filtered["52주위치"].apply(_mom_s) * 0.70 +
        filtered.apply(_vol_mom_s, axis=1)  * 0.30
    ).round(1)

    # ── ③ 배당 점수 (배당수익률 60% + 배당성향 지속가능성 40%) ──
    def _div_s(row):
        div    = _safe(row.get("DIV"), 0)
        payout = _safe(row.get("dividend_payout"))

        # 배당수익률 점수
        if   div >= 6: yield_s = 100
        elif div >= 4: yield_s = 85
        elif div >= 3: yield_s = 70
        elif div >= 2: yield_s = 55
        elif div >= 1: yield_s = 35
        elif div >  0: yield_s = 20
        else:          yield_s = 0

        # 배당성향 점수 (지속가능성)
        if payout is not None:
            if   20 <= payout <= 50: pay_s = 100  # 최적
            elif payout <= 70:       pay_s = 70
            elif payout <= 80:       pay_s = 40
            elif payout >  80:       pay_s = 10   # 삭감 위험
            else:                    pay_s = 50   # 저배당
        else:
            pay_s = 50  # 데이터 없으면 중립

        return round(yield_s * 0.60 + pay_s * 0.40, 1)

    filtered["배당점수"] = filtered.apply(_div_s, axis=1)

    # ── ④ 품질 점수 (DART 부채비율 우선) ──
    def _quality_s(row):
        debt = _safe(row.get("부채비율") or row.get("debt_ratio"))
        roa  = _safe(row.get("ROA"))
        opm  = _safe(row.get("영업이익률"))
        roe  = _safe(row.get(roe_col), 0)
        pbr  = _safe(row.get(pbr_col), 1)

        # 부채비율 점수 (DART 원본 우선)
        if debt is not None:
            if   debt <= 50:  debt_s = 100
            elif debt <= 100: debt_s = 85
            elif debt <= 150: debt_s = 65
            elif debt <= 200: debt_s = 45
            elif debt <= 300: debt_s = 25
            else:             debt_s = 0
        else:
            debt_s = 85 if pbr <= 1.0 else (65 if pbr <= 1.5 else 45)

        # ROA
        if roa is not None:
            if   roa >= 15: roa_s = 100
            elif roa >= 10: roa_s = 85
            elif roa >= 7:  roa_s = 70
            elif roa >= 5:  roa_s = 55
            elif roa >= 3:  roa_s = 40
            elif roa >= 0:  roa_s = 25
            else:           roa_s = 0
        else:
            roa_s = min(100, max(0, roe * 0.6))

        # 영업이익률
        if opm is not None:
            if   opm >= 20: opm_s = 100
            elif opm >= 15: opm_s = 85
            elif opm >= 10: opm_s = 70
            elif opm >= 5:  opm_s = 50
            elif opm >= 0:  opm_s = 30
            else:           opm_s = 0
        else:
            opm_s = min(100, max(0, roe * 1.5))

        return round(debt_s*0.45 + roa_s*0.35 + opm_s*0.20, 1)

    filtered["품질점수"] = filtered.apply(_quality_s, axis=1)

    # ── ⑤ 성장 점수 (DART 매출 데이터 + 네이버 성장률) ──
    def _growth_s(row):
        rev_gr = _safe(row.get("매출성장률"))
        roe    = _safe(row.get(roe_col), 0)
        per    = _safe(row.get("PER") or row.get("per"), 20)
        div    = _safe(row.get("DIV"), 0)

        if rev_gr is not None:
            if   rev_gr >= 30: rev_s = 100
            elif rev_gr >= 20: rev_s = 85
            elif rev_gr >= 10: rev_s = 70
            elif rev_gr >= 5:  rev_s = 55
            elif rev_gr >= 0:  rev_s = 40
            elif rev_gr >= -5: rev_s = 20
            else:              rev_s = 0
        else:
            proxy = roe * 0.6 + div * 2
            rev_s = 100 if proxy>=15 else (80 if proxy>=10 else (60 if proxy>=6 else (40 if proxy>=3 else 20)))

        eps_proxy = roe / max(per if per else 20, 1) * 100
        if   eps_proxy >= 5: eps_s = 100
        elif eps_proxy >= 3: eps_s = 80
        elif eps_proxy >= 1: eps_s = 60
        elif eps_proxy >= 0: eps_s = 40
        else:                eps_s = 20

        return round(rev_s*0.50 + eps_s*0.50, 1)

    filtered["성장점수"] = filtered.apply(_growth_s, axis=1)

    # ── ⑥ 수익성 점수 ──
    def _prof_s(row):
        opm  = _safe(row.get("영업이익률"))
        netm = _safe(row.get("순이익률"))
        roe  = _safe(row.get(roe_col), 0)
        pbr  = _safe(row.get(pbr_col), 1)
        per  = _safe(row.get("PER") or row.get("per"), 20)

        if opm is not None:
            if   opm>=20: op_s=100
            elif opm>=15: op_s=85
            elif opm>=10: op_s=70
            elif opm>=5:  op_s=50
            elif opm>=0:  op_s=30
            else:         op_s=0
        else:
            proxy = roe / max(pbr, 0.1) * 5
            op_s = 100 if proxy>=20 else (85 if proxy>=15 else (70 if proxy>=10 else (50 if proxy>=5 else 30)))

        if netm is not None:
            if   netm>=15: net_s=100
            elif netm>=10: net_s=85
            elif netm>=7:  net_s=70
            elif netm>=5:  net_s=55
            elif netm>=0:  net_s=35
            else:          net_s=0
        else:
            proxy2 = roe / max(pbr, 0.1) * 5 * 1.2
            net_s = 100 if proxy2>=25 else (85 if proxy2>=20 else (70 if proxy2>=15 else (50 if proxy2>=10 else 30)))

        # FCF 마진 실제값 우선, 없으면 PER 역수 추정
        fcf_margin = _safe(row.get("fcf_margin"))
        if fcf_margin is not None:
            if   fcf_margin >= 20: fcf_s = 100
            elif fcf_margin >= 15: fcf_s = 85
            elif fcf_margin >= 10: fcf_s = 70
            elif fcf_margin >=  5: fcf_s = 55
            elif fcf_margin >=  0: fcf_s = 35
            else:                  fcf_s = 0   # FCF 음수
        else:
            fcf_proxy = 100 / max(per if per else 20, 1)
            fcf_s = 100 if fcf_proxy>=15 else (85 if fcf_proxy>=10 else (70 if fcf_proxy>=5 else (50 if fcf_proxy>=3 else 30)))

        return round(op_s*0.45 + net_s*0.35 + fcf_s*0.20, 1)

    filtered["수익성점수"] = filtered.apply(_prof_s, axis=1)

    # ── ⑦ 안정 점수 ──
    def _stab_s(row):
        debt     = _safe(row.get("부채비율") or row.get("debt_ratio"))
        net_debt = _safe(row.get("net_debt_ratio"))
        cur_r    = _safe(row.get("current_ratio"))
        altman   = _safe(row.get("altman_z"))
        roe      = _safe(row.get(roe_col), 0)
        div      = _safe(row.get("DIV"), 0)
        ic       = _safe(row.get("interest_coverage"))

        # ① 부채 안전성 (순부채비율 우선, 없으면 부채비율)
        ref_debt = net_debt if net_debt is not None else debt
        if ref_debt is not None:
            if   ref_debt <  0:   debt_s = 100  # 순현금 기업
            elif ref_debt <= 30:  debt_s = 90
            elif ref_debt <= 70:  debt_s = 75
            elif ref_debt <= 100: debt_s = 55
            elif ref_debt <= 150: debt_s = 35
            elif ref_debt <= 200: debt_s = 20
            else:                 debt_s = 0
        else:
            debt_s = 50

        # 이자보상배율 보너스
        if ic is not None and ic > 0:
            debt_s = min(100, debt_s + (20 if ic>=10 else 10 if ic>=5 else 0))

        # ② 유동비율 점수
        if cur_r is not None:
            if   cur_r >= 200: cur_s = 100
            elif cur_r >= 150: cur_s = 80
            elif cur_r >= 120: cur_s = 60
            elif cur_r >= 100: cur_s = 40
            else:              cur_s = 10
        else:
            cur_s = 50  # 중립

        # ③ Altman Z-Score
        if altman is not None:
            if   altman >= 3.0: alt_s = 100
            elif altman >= 2.5: alt_s = 80
            elif altman >= 1.8: alt_s = 50  # 회색지대
            else:               alt_s = 10  # 부도 위험
        else:
            # 데이터 없으면 ROE로 대체 추정
            alt_s = 80 if roe>=15 else (60 if roe>=10 else (40 if roe>=5 else 20))

        # ④ 손실 위험 (ROE)
        if   roe>=20: loss_s=100
        elif roe>=15: loss_s=85
        elif roe>=10: loss_s=70
        elif roe>=8:  loss_s=55
        elif roe>=5:  loss_s=35
        else:         loss_s=10

        # ⑤ 배당 안정성
        div_payout = _safe(row.get("dividend_payout"))
        if div_payout is not None:
            # 20~60% 지속 가능, 80% 초과는 삭감 위험
            if   20 <= div_payout <= 60: div_s = 100
            elif div_payout < 20:        div_s = 60   # 주주환원 의지↓
            elif div_payout <= 80:       div_s = 40   # 다소 높음
            else:                        div_s = 10   # 삭감 위험
        else:
            # 배당성향 없으면 배당수익률로 대체
            if   div>=4: div_s=100
            elif div>=2: div_s=80
            elif div>=1: div_s=60
            elif div>0:  div_s=40
            else:        div_s=20

        # 가중합: 부채35% + 유동성15% + Altman20% + ROE15% + 배당15%
        return round(debt_s*0.35 + cur_s*0.15 + alt_s*0.20 + loss_s*0.15 + div_s*0.15, 1)

    filtered["안정점수"] = filtered.apply(_stab_s, axis=1)

    # ── ⑧ 복합 점수 ──
    ws = [args.w_value, args.w_mom, args.w_div, args.w_quality,
          args.w_growth, args.w_prof, args.w_stability]
    total_w = sum(ws)
    if abs(total_w - 100) > 0.1:
        ws = [w / total_w * 100 for w in ws]
    w_v,w_m,w_d,w_q,w_g,w_p,w_s = [w/100 for w in ws]

    filtered["복합점수"] = (
        filtered["가치점수"]   * w_v +
        filtered["모멘텀점수"] * w_m +
        filtered["배당점수"]   * w_d +
        filtered["품질점수"]   * w_q +
        filtered["성장점수"]   * w_g +
        filtered["수익성점수"] * w_p +
        filtered["안정점수"]   * w_s
    ).round(1)

    filtered["등급"] = filtered["복합점수"].apply(lambda s: grade_label(s)[0])

    # ── ⑨ 밸류에이션 적정가 3종 ──
    def _valuation(row):
        roe  = _safe(row.get("ROE") or row.get("roe"))
        cur  = _safe(row.get("현재가") or row.get("현재가_NV") or row.get("현재가_YF"))
        sector = str(row.get("업종", "") or "")
        name   = str(row.get("종목명", "") or "")

        # ── 현재가 정합성 검증 ──
        # 시가총액 ÷ 현재가 = 추정 주식수. 이 값이 비현실적이면 현재가 오류
        # (네이버 현재가 파싱 시 자릿수 밀림 오류 탐지)
        marcap_eok = _safe(row.get("시가총액(억)") or row.get("시총"))  # 억원
        cur_invalid = False
        if marcap_eok and marcap_eok > 0 and cur and cur > 0:
            implied_shares = marcap_eok * 1e8 / cur  # 추정 주식수
            # 정상 상장사 주식수: 약 10만주 ~ 500억주
            if implied_shares < 100_000 or implied_shares > 5e10:
                cur_invalid = True

        if cur is None or cur <= 0 or cur_invalid:
            reason = f"현재가오류(시총{marcap_eok:.0f}억÷현재가 비정상)" if cur_invalid else "현재가데이터없음"
            return pd.Series({
                "적정가_PER": None, "적정가_PBR": None,
                "적정가_Graham": None, "적정가_평균": None,
                "괴리율(%)": None, "적정가_사유": reason,
            })

        # ── PER/PBR 계산: BPS/EPS 직접 사용 (v30 핵심) ──
        # 우선순위:
        #   ① BPS_dart → 현재가÷BPS = PBR(현재가)  [가장 정확]
        #   ② EPS_dart → 현재가÷EPS = PER(현재가)  [가장 정확]
        #   ③ shares×현재가÷equity/ni               [2순위]
        #   ④ 결산 수집값 fallback
        bps_dart  = _safe(row.get("bps_dart"))
        eps_dart  = _safe(row.get("eps_dart"))
        equity_val = _safe(row.get("equity"))
        ni_val     = _safe(row.get("net_income"))
        marcap_eok = _safe(row.get("시가총액(억)") or row.get("시총"))

        # shares 안전 접근 (Int64/NA 방지)
        _shr_raw = row.get("shares")
        shares_val = None
        if _shr_raw is not None and getattr(_shr_raw, '__class__', None) and _shr_raw.__class__.__name__ != 'NAType':
            v = _safe(_shr_raw)
            if v and not (isinstance(v, float) and np.isnan(v)):
                shares_val = float(v)

        per_close = _safe(row.get("PER_결산"))
        pbr_close = _safe(row.get("PBR_결산"))
        per_nv    = _safe(row.get("PER") or row.get("per"))
        pbr_nv    = _safe(row.get("PBR") or row.get("pbr"))

        # ① BPS/EPS 기반 (가장 정확)
        pbr_cur = round(cur / bps_dart, 3) if (bps_dart and bps_dart > 0 and cur > 0) else None
        per_cur = round(cur / eps_dart, 2) if (eps_dart and eps_dart > 0 and cur > 0) else None

        # ② shares × 현재가 ÷ 재무수치 (BPS/EPS 없을 때)
        if pbr_cur is None or per_cur is None:
            mktcap_won = None
            if shares_val and shares_val > 0:
                mktcap_won = shares_val * cur
            elif marcap_eok and marcap_eok > 0:
                mktcap_won = marcap_eok * 1e8
            if mktcap_won and mktcap_won > 0:
                if pbr_cur is None and equity_val and equity_val > 0:
                    pbr_cur = round(mktcap_won / equity_val, 3)
                if per_cur is None and ni_val and ni_val > 0:
                    per_cur = round(mktcap_won / ni_val, 2)

        # ③ fallback: 결산 수집값
        pbr_raw = pbr_cur if (pbr_cur and 0 < pbr_cur < 100) else \
                  (pbr_close if (pbr_close and 0 < pbr_close < 30) else pbr_nv)
        per_raw = per_cur if (per_cur and 0 < per_cur) else \
                  (per_close if (per_close and 0 < per_close < 200) else per_nv)

        # ── 빈 셀 이유 추적 ──
        notes = []
        per = per_raw
        pbr = pbr_raw

        # PER 유효성
        if per is None:
            notes.append("PER없음")
        elif per <= 0:
            notes.append(f"PER음수(적자,{per:.0f})")
            per = None
        elif per > 80:
            # v26: PER>80은 EPS 신뢰불가 (적자직전/일시이익극소) → PBR 기반만 사용
            notes.append(f"PER과대({per:.0f}배→제외,PBR기반만)")
            per = None

        # PBR 유효성
        if pbr is None:
            notes.append("PBR없음")
        elif pbr <= 0:
            notes.append(f"PBR음수")
            pbr = None
        elif pbr > 30:
            notes.append(f"PBR과대({pbr:.1f}배→제외)")
            pbr = None

        # ── 업종별 적정 PER/PBR ──
        SECTOR_PER = {
            "반도체": 18, "전기전자": 16, "IT서비스": 20, "소프트웨어": 22,
            "바이오": 25, "제약": 20, "화장품": 18, "인터넷": 22,
            "2차전지": 20, "배터리": 20, "방산": 18, "우주": 20, "로봇": 22,
            "자동차부품": 10, "자동차": 10, "철강": 8, "화학": 10,
            "건설": 8, "유통": 12, "음식료": 14, "섬유": 10,
            "조선": 12, "해운": 7, "항공": 12,
            "금융": 8, "은행": 7, "보험": 8, "증권": 9, "지주": 8,
            "부동산": 12, "호텔": 14, "레저": 14,
            "에너지": 10, "전력": 10, "가스": 9, "기계": 12,
        }
        base_per = 12
        for kw, v in SECTOR_PER.items():
            if kw in sector:
                base_per = v
                break

        SECTOR_PBR = {
            "반도체": 2.0, "전기전자": 1.5, "IT서비스": 2.5, "소프트웨어": 3.0,
            "바이오": 3.0, "제약": 2.5, "화장품": 2.5, "인터넷": 3.0,
            "2차전지": 2.5, "배터리": 2.5, "방산": 2.0, "로봇": 2.5,
            "자동차부품": 1.0, "자동차": 1.0, "철강": 0.8, "화학": 1.2,
            "건설": 0.9, "유통": 1.2, "음식료": 1.5,
            "조선": 1.0, "해운": 0.9,
            "금융": 0.8, "은행": 0.7, "보험": 0.9, "증권": 1.0, "지주": 0.8,
            "에너지": 1.0, "전력": 1.0, "기계": 1.2,
        }
        base_pbr = 1.2
        for kw, v in SECTOR_PBR.items():
            if kw in sector:
                base_pbr = v
                break

        # ── EPS/BPS 역산 ──
        eps = cur / per if (per and per > 0) else None
        bps = cur / pbr if (pbr and pbr > 0) else None

        # ── PER 이상치 보정 ──
        # PER이 비정상적으로 높은데(>30배) ROE는 건전하면(≥8%)
        # → 일시적 이익 감소(사이클 저점)로 후행 PER 왜곡 → ROE 기반 EPS 재추정
        per_abnormal = False
        if per and per > 30 and roe and roe >= 8 and bps and bps > 0:
            # 정상 EPS ≈ BPS × ROE/100 (지속가능 이익 수준)
            eps_normalized = bps * roe / 100.0
            if eps_normalized > 0:
                eps = eps_normalized   # EPS를 정상화 값으로 교체
                per_abnormal = True
                notes.append(f"PER{per:.0f}배→ROE정상화")

        # ── 적정 PER 계산 (v26 개선) ──
        # [로직]
        # ① PER <= 업종평균×1.1: blended와 업종평균 중 작은값 → 단, per×0.75 하한 보장
        # ② PER > 업종평균×1.1 (프리미엄): per×0.85 상한 → 단, per×0.75 하한 보장
        # ③ per_abnormal(ROE 정상화): 업종평균 적용 (고성장 보정은 아래에서 추가)
        #
        # ★ 핵심: fair_per 하한 = per × 0.75 보장
        #   → 적정가(PER기반) = EPS × fair_per = (cur/per) × fair_per >= cur × 0.75
        #   → 1차목표가(=적정가)가 최소 현재가의 75% 이상 보장
        if per and per > 0 and not per_abnormal:
            blended = per * 0.6 + base_per * 0.4
            if per <= base_per * 1.1:
                fair_per = max(min(blended, base_per), per * 0.75)
            else:
                fair_per = max(min(blended, per * 0.85), per * 0.75)
        else:
            fair_per = base_per

        # ── 고성장 PER 보정 (v24) ──
        # ROE 정상화 EPS를 쓰는 경우(per_abnormal=True), ROE >15%이면
        # "고성장 프리미엄" → fair_per을 ROE 비례 상향 (업종평균 최대 2.5배 상한)
        if per_abnormal and roe and roe > 15:
            _roe_for_growth = min(roe, 25.0)
            roe_growth_bonus = min((_roe_for_growth - 15) * 0.5, base_per * 1.0)
            fair_per = min(fair_per + roe_growth_bonus, base_per * 2.5)
            notes.append(f"ROE{roe:.0f}%성장보정→fair_per{fair_per:.1f}")

        # ── 시장 프리미엄 과다 종목 감지 ──
        is_premium_stock = bool(pbr and pbr > 5 and per and per > 40)
        if is_premium_stock:
            notes.append("⚠고성장프리미엄(퀀트적정가참고용)")

        # ── 적정 PBR 계산 (v26 개선) ──
        # [케이스별 처리]
        # ① 고PBR(>3): BPS 기반 적정가 건너뜀 → PER/Graham 기반만 사용
        # ② 중PBR(1~3): fair_pbr = min(pbr×0.8, ROE혼합) → 하한 pbr×0.75 보장
        # ③ 저PBR(<1):
        #    - ROE>=5%: 기존 ROE 혼합 로직 → 하한 pbr×0.75 보장
        #    - ROE<5% (저수익/적자): fair_pbr = min(pbr, base_pbr) 직접 사용
        #      (현재 PBR을 적정하다 보고 업종평균 이하면 유지)
        #
        # ★ 핵심: fair_pbr 하한 = pbr × 0.75 보장 (ROE>=5% 구간)
        #   → 적정가(PBR기반) = BPS × fair_pbr = (cur/pbr) × fair_pbr >= cur × 0.75
        roe_capped = min(roe, 25.0) if (roe and roe > 0) else None
        roe_pbr = (roe_capped / 10.0) if roe_capped else 0.0

        skip_pbr_target = False
        if pbr and pbr > 0:
            if pbr > 3.0:
                skip_pbr_target = True
                notes.append(f"PBR{pbr:.1f}배(고PBR→PBR적정가제외)")
            elif pbr > 1.0:
                # 중PBR: ROE 혼합, 하한 pbr×0.75 보장
                raw_fair = min(pbr * 0.8, max(pbr * 0.5 + roe_pbr * 0.5, 0.5))
                fair_pbr = max(min(raw_fair, pbr), pbr * 0.75)
            else:
                # 저PBR
                if roe_capped is not None and roe_capped >= 5:
                    # 정상 수익성: ROE 혼합, 하한 pbr×0.75 보장
                    raw_fair = min(max(pbr * 0.5 + roe_pbr * 0.5, 0.3), base_pbr)
                    fair_pbr = max(raw_fair, pbr * 0.75)
                else:
                    # 저수익/적자(ROE<5%): 현재 PBR을 적정으로 인정, 업종평균 이하면 그대로
                    fair_pbr = min(pbr, base_pbr)
        else:
            fair_pbr = max(min(roe_pbr if roe_pbr > 0 else 0.5, base_pbr), 0.3)

        # ── 적정가 3종 계산 ──
        per_target    = round(eps * fair_per)  if (eps and eps > 0) else None
        pbr_target    = None if skip_pbr_target else (round(bps * fair_pbr) if (bps and bps > 0) else None)
        graham_target = None
        roe_for_graham = roe_capped or roe
        if eps and bps and eps > 0 and bps > 0 and roe_for_graham and roe_for_graham >= 8:
            graham_target = round((22.5 * eps * bps) ** 0.5)

        # ── 개별 필터: 현재가 대비 범위 체크 ──
        # 상한: PBR 낮을수록 적정가가 높게 나오는 게 정상 (저PBR 가치주)
        # 하한: 고PBR 성장주는 PER/Graham 기반이므로 하한 완화
        filtered_out = []
        upper_limit = 5.0 if (pbr and pbr < 0.5) else (4.0 if (pbr and pbr < 1.0) else 3.0)
        # 고PBR(>3)은 PER/Graham 기반만 쓰므로 하한을 넉넉하게 설정
        lower_limit = 0.10 if (pbr and pbr > 5.0) else (0.15 if (pbr and pbr > 3.0) else (0.20 if (pbr and pbr > 2.0) else 0.25))
        def _check(t, label):
            if t is None: return None
            if cur <= 0: return None
            ratio = t / cur
            if ratio < lower_limit:
                filtered_out.append(f"{label}필터(<{lower_limit}배,{ratio:.2f}배)")
                return None
            if ratio > upper_limit:
                filtered_out.append(f"{label}필터(>{upper_limit}배,{ratio:.1f}배)")
                return None
            return t

        per_t_f    = _check(per_target,    "PER")
        pbr_t_f    = _check(pbr_target,    "PBR")
        graham_t_f = _check(graham_target, "Graham")

        # ── 중앙값 기반 이상치 제거 후 가중 평균 ──
        # 3종 적정가가 서로 크게 어긋나면(데이터 오류 가능성)
        # 중앙값에서 ±50% 벗어난 값을 이상치로 제거
        raw_targets = [(per_t_f, 0.5, "PER"), (pbr_t_f, 0.35, "PBR"), (graham_t_f, 0.15, "Graham")]
        valid_vals = [(t, w, lbl) for t, w, lbl in raw_targets if t is not None]

        outlier_removed = []
        if len(valid_vals) >= 2:
            vals_only = sorted([t for t, _, _ in valid_vals])
            n = len(vals_only)
            median = vals_only[n//2] if n % 2 else (vals_only[n//2-1] + vals_only[n//2]) / 2
            # 중앙값 ±50% 범위만 채택
            kept = []
            for t, w, lbl in valid_vals:
                if 0.5 * median <= t <= 1.5 * median:
                    kept.append((t, w))
                else:
                    outlier_removed.append(f"{lbl}이상치({t:,}≠중앙값{round(median):,})")
            weighted = kept if kept else [(t, w) for t, w, _ in valid_vals]
        else:
            weighted = [(t, w) for t, w, _ in valid_vals]

        if not weighted:
            avg_target = None
            if not notes and not filtered_out:
                notes.append("계산불가(데이터부족)")
            if filtered_out:
                notes.extend(filtered_out)
            reason = " / ".join(notes) if notes else "알수없음"
        else:
            tw = sum(w for _, w in weighted)
            avg_target = round(sum(t * w for t, w in weighted) / tw)
            note_parts = []
            if notes: note_parts.extend(notes)
            if outlier_removed: note_parts.extend(outlier_removed)
            reason = "정상" if not note_parts else ("(" + ",".join(note_parts) + ")")

        # 사용된 PER/PBR 원본값을 사유에 항상 부기 (검증용)
        per_str = f"PER{per:.1f}" if per else "PER-"
        pbr_str = f"PBR{pbr:.2f}" if pbr else "PBR-"
        reason = f"[{per_str}/{pbr_str}] {reason}"

        # ── 괴리율 ──
        upside = None
        if avg_target and cur > 0:
            upside = round((avg_target - cur) / cur * 100, 1)
            # 저PBR 가치주는 괴리율 300%까지 정상, 일반 300% → 200%
            upside_max = 300 if (pbr and pbr < 0.5) else 200
            if upside < -80 or upside > upside_max:
                reason = f"[{per_str}/{pbr_str}] 괴리율극단({upside:.0f}%)→제외"
                upside = None
                avg_target = None

        return pd.Series({
            "적정가_PER":    per_t_f,
            "적정가_PBR":    pbr_t_f,
            "적정가_Graham": graham_t_f,
            "적정가_평균":   avg_target,
            "괴리율(%)":     upside,
            "적정가_사유":   reason,
        })

    val_df = filtered.apply(_valuation, axis=1)
    for col in val_df.columns:
        filtered[col] = val_df[col]

    # ── ⑩ 100점 만점 기업 종합 평가 ──
    print("  [100점] 기업 종합 평가 계산 중...")
    pt100_rows = filtered.apply(
        lambda row: pd.Series(compute_100pt_score(row.to_dict())), axis=1
    )
    for col in pt100_rows.columns:
        filtered[col] = pt100_rows[col]

    # ── ⑪ 매매 시그널 ──
    print("  [시그널] 매매 시그널 및 가격 밴드 계산 중...")
    def _sig_row(row):
        sc100 = {c: row.get(c) for c in pt100_rows.columns}
        return pd.Series(compute_trading_signal(row.to_dict(), sc100))
    sig_rows = filtered.apply(_sig_row, axis=1)
    for col in sig_rows.columns:
        filtered[col] = sig_rows[col]

    return filtered.sort_values("복합점수", ascending=False)

# ══════════════════════════════════════════════════════════
# 10. 엑셀 출력
# ══════════════════════════════════════════════════════════

def fetch_stock_news_claude(ticker: str, name: str, sector: str, top_n: int = 3) -> dict:
    """
    Claude API를 이용해 종목별 주간 핵심 이슈/뉴스 요약 + 감성 점수 생성
    - 실제 웹 검색(web_search tool)을 사용해 최신 뉴스를 찾아 요약
    - 같은 호출에서 감성 점수(-100~+100)까지 함께 받아 매수/매도 판단에 반영
      (-100=치명적 악재, 0=중립/뉴스없음, +100=강력한 호재)

    반환: {"text": "엑셀용 요약문", "sentiment": int}
    """
    try:
        import requests as _req
        today_str = datetime.today().strftime("%Y년 %m월 %d일")
        prompt = (
            f"오늘은 {today_str}입니다.\n"
            f"한국 주식 종목 '{name}'({ticker}, 업종: {sector})의 "
            f"최근 1주일 이내 핵심 투자 이슈를 {top_n}줄 이내로 간략히 요약해주세요.\n"
            f"형식: 첫 줄에 '감성점수: N' (N은 -100~+100 정수, 분식회계·횡령·상장폐지·"
            f"실적쇼크 등 치명적 악재는 -80 이하, 단순 우려는 -30~-10, 특이사항 없으면 0, "
            f"단순 호재는 +10~+30, 신규수주·실적서프라이즈 등 강한 호재는 +60 이상)\n"
            f"그 다음 줄부터 '① [날짜] 이슈내용 ② [날짜] 이슈내용 ...' 형태로 작성.\n"
            f"이슈가 없으면 '감성점수: 0' 다음 줄에 '특이사항 없음'으로 답하세요. "
            f"불필요한 설명 없이 이 형식만 정확히 지켜서 작성."
        )
        resp = _req.post(
            "https://api.anthropic.com/v1/messages",
            headers={"Content-Type": "application/json"},
            json={
                "model": "claude-sonnet-4-6",
                "max_tokens": 300,
                "tools": [{"type": "web_search_20250305", "name": "web_search"}],
                "messages": [{"role": "user", "content": prompt}]
            },
            timeout=30
        )
        data = resp.json()
        texts = [b.get("text","") for b in data.get("content",[]) if b.get("type")=="text"]
        result = " ".join(texts).strip()

        if not result:
            return {"text": "뉴스 수집 실패", "sentiment": 0}

        # "감성점수: N" 패턴 추출 (첫 줄 또는 본문 어디든)
        sentiment = 0
        m = re.search(r"감성점수\s*:?\s*([+-]?\d+)", result)
        if m:
            try:
                sentiment = max(-100, min(100, int(m.group(1))))
            except ValueError:
                sentiment = 0
            # 엑셀 요약문에서는 감성점수 줄을 제거해 가독성 유지
            result = re.sub(r"감성점수\s*:?\s*[+-]?\d+\s*", "", result, count=1).strip()

        return {"text": result if result else "특이사항 없음", "sentiment": sentiment}
    except Exception as e:
        return {"text": f"뉴스 수집 오류: {str(e)[:30]}", "sentiment": 0}


def fetch_news_batch(df_top: pd.DataFrame, max_stocks: int = 30) -> dict:
    """
    상위 종목의 뉴스를 병렬로 배치 수집
    max_stocks: API 비용 제한을 위해 상위 N개만 수집
    """
    from concurrent.futures import ThreadPoolExecutor, as_completed
    stocks = list(df_top.iterrows())[:max_stocks]
    news_map = {}

    print(f"  [뉴스] 상위 {len(stocks)}개 종목 뉴스 수집 중 (Claude API)...")

    def _fetch(item):
        code, row = item
        name   = str(row.get("종목명","") or "")
        sector = str(row.get("업종","") or "")
        news = fetch_stock_news_claude(str(code), name, sector)
        return str(code), news

    with ThreadPoolExecutor(max_workers=5) as ex:
        futs = {ex.submit(_fetch, item): item for item in stocks}
        done = 0
        for fut in as_completed(futs):
            code, news = fut.result()
            news_map[code] = news
            done += 1
            if done % 10 == 0:
                print(f"    {done}/{len(stocks)}개 완료...")

    print(f"  [뉴스] 수집 완료: {len(news_map)}개")
    return news_map


def build_excel(df_all, df_top, label, args, output_dir=BASE_DIR):
    if not HAS_OPENPYXL:
        print("  ⚠ openpyxl 미설치 → 엑셀 저장 불가")
        return None

    # v33: 날짜 고정 파일명 → 매일 같은 파일을 덮어씀 (작업스케줄러 자동실행 최적화)
    # 파일명 예: quant_KR_20260617.xlsx  (시각 없음 → 항상 같은 이름)
    fname = datetime.today().strftime(f"quant_{label}_%Y%m%d.xlsx")
    fpath = os.path.join(output_dir, fname)

    # ── 스타일 상수 ──
    KR_FONT    = "맑은 고딕"
    body_font  = Font(name=KR_FONT, size=10)
    _HDR_FILL  = PatternFill("solid", fgColor="1F3864")
    _ALT_FILL  = PatternFill("solid", fgColor="F5F5F5")
    _GRN_FILL  = PatternFill("solid", fgColor="E2EFDA")

    # ── 숫자 포맷 헬퍼 ──
    def _r1(v):
        if v is None: return None
        try: return round(float(v), 1)
        except: return None
    def _r2(v):
        if v is None: return None
        try: return round(float(v), 2)
        except: return None
    def _ri(v):
        if v is None: return None
        try: return int(round(float(v)))
        except: return None
    def _eok(v):
        """원 단위 → 억 단위 변환 (DART 재무 수치용)"""
        if v is None: return None
        try:
            fv = float(v)
            if abs(fv) > 1e8:          # 1억 이상이면 원 단위로 판단 → 억으로 변환
                return int(round(fv / 1e8))
            else:                       # 이미 억 단위이거나 작은 값
                return int(round(fv))
        except: return None

    # ── 다중 키 값 조회 ──
    def _gv(row, *keys):
        for k in keys:
            v = row.get(k)
            if v is not None and not (isinstance(v, float) and np.isnan(v)):
                try:
                    f = float(v)
                    if f != 0: return f
                except:
                    if str(v).strip(): return v
        return None

    # ── 시가총액 조회 ──
    def _marcap(row):
        mc = row.get("시가총액(억)") or row.get("시총")
        if mc: return _safe(mc)
        cur = _safe(row.get("현재가") or row.get("현재가_NV") or row.get("현재가_YF"))
        shr = _safe(row.get("shares"))
        if cur and shr: return round(cur * shr / 1e8, 1)
        return None

    wb = Workbook()
    ws_detail = wb.active          # 시트1: 스크리닝결과상세
    ws_guide   = wb.create_sheet()  # 시트2: 지표해석가이드
    ws_all     = wb.create_sheet()  # 시트3: 전체종목데이터
    ws_signal  = wb.create_sheet()  # 시트4: 매매시그널보드
    ws_news    = wb.create_sheet()  # 시트5: 종목별 주간뉴스 (v30 신규)

    # ── 뉴스 수집 (Claude API, 상위 종목 한정) ──
    # df_top에 이미 _apply_news_sentiment()에서 수집된 "뉴스감성점수"/"뉴스요약"이
    # 있으면 그걸 그대로 재사용 (API 중복호출·중복비용 방지). 없으면 새로 수집.
    news_map = {}
    if "뉴스요약" in df_top.columns and "뉴스감성점수" in df_top.columns:
        for code in df_top.index:
            news_map[str(code)] = {
                "text": df_top.at[code, "뉴스요약"],
                "sentiment": df_top.at[code, "뉴스감성점수"],
            }
    else:
        fetch_news = getattr(args, "fetch_news", True)
        news_top_n = min(getattr(args, "news_top", 30), len(df_top))
        if fetch_news and news_top_n > 0:
            try:
                news_map = fetch_news_batch(df_top, max_stocks=news_top_n)
            except Exception as e:
                print(f"  [뉴스] 수집 실패: {e}")

    # ── is_kr 먼저 정의 ──
    is_kr = label in ("KR","ALL")

    # ── 시트2: 스크리닝 결과 상세 (핵심 컬럼 통합) ──
    ws_detail.title = "스크리닝결과상세"

    # 제목행
    ws_detail.merge_cells("A1:Z1")
    t2 = ws_detail.cell(row=1, column=1,
        value=f"퀀트 스크리닝 상위 {args.top}종목 [{label}]  기준일: {datetime.today().strftime('%Y-%m-%d')}")
    t2.font  = Font(name=KR_FONT, bold=True, size=13, color="FFFFFF")
    t2.fill  = PatternFill("solid", fgColor="1F3864")
    t2.alignment = Alignment(horizontal="center", vertical="center")
    ws_detail.row_dimensions[1].height = 28

    ws_detail.merge_cells("A2:Z2")
    ws_detail.cell(row=2, column=1,
        value=f"가중치: 가치{args.w_value:.0f}% 모멘텀{args.w_mom:.0f}% "
              f"배당{args.w_div:.0f}% 품질{args.w_quality:.0f}% "
              f"성장{args.w_growth:.0f}% 수익성{args.w_prof:.0f}% "
              f"안정{args.w_stability:.0f}%").font = Font(name=KR_FONT, size=10, italic=True)

    # ── 헤더 그룹 정의 ──
    sub_headers = [
        # 기본 식별
        "순위","코드","종목명","시장","업종","현재가","시가총액(억)",
        # 7팩터 점수
        "복합점수","등급","가치","모멘텀","배당","품질","성장","수익성","안정",
        # 핵심 재무
        "PBR","ROE(%)","ROA(%)","DIV(%)","부채비율(%)","이자보상배율",
        "영업이익률(%)","FCF(억)","CFO(억)",
        # 적정가/밸류
        "적정가_PER","적정가_PBR","적정가_Graham","적정가_평균","괴리율(%)","적정가_사유",
        # 100점 평가
        "100점_합계","100점_등급",
        "100점_수익성","100점_성장성","100점_재무안정","100점_현금흐름","100점_밸류","100점_주주친화","100점_정성보정",
        # 5대 체크 (숫자+판정)
        "CFO(판정)","자본잠식(판정)","부채비율(판정)","ROE(판정)","PER(판정)",
        # 매매 시그널
        "매매시그널","현재가위치",
        "적극매수선","매수선","1차목표가","2차목표가","손절선",
        "보유전략","위험신호","보너스근거",
    ]

    # 헤더 색상 그룹
    HDR_GROUPS = {
        range(1,8):   "1F3864",  # 기본
        range(8,17):  "2E75B6",  # 7팩터
        range(17,26): "375623",  # 재무
        range(26,31): "7030A0",  # 적정가
        range(31,41): "833C00",  # 100점
        range(41,46): "404040",  # 체크
        range(46,57): "C00000",  # 시그널
    }
    def _hdr_color(j):
        for rng, color in HDR_GROUPS.items():
            if j in rng: return color
        return "1F3864"

    for j, h in enumerate(sub_headers, 1):
        c = ws_detail.cell(row=3, column=j, value=h)
        c.font      = Font(name=KR_FONT, bold=True, color="FFFFFF", size=9)
        c.fill      = PatternFill("solid", fgColor=_hdr_color(j))
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = _THIN
    ws_detail.row_dimensions[3].height = 30

    for ri, (code, row) in enumerate(df_top.iterrows(), 1):
        score     = _safe(row.get("복합점수"), 0)
        grade_txt, grade_color = grade_label(score)
        cur       = _gv(row, "현재가","현재가_NV","현재가_YF")
        pt100     = _safe(row.get("100점_합계"))
        sig_txt   = str(row.get("매매시그널","") or "")
        sig_color = str(row.get("_signal_color","FFCC00") or "FFCC00")

        row_vals = [
            # 기본
            ri, str(code),
            str(_gv(row,"종목명") or ""),
            str(row.get("시장","") or ""),
            str(row.get("업종","") or ""),
            _ri(cur),
            _ri(_marcap(row)),
            # 7팩터
            _r1(score), str(grade_txt),
            _r1(row.get("가치점수")),   _r1(row.get("모멘텀점수")),
            _r1(row.get("배당점수")),   _r1(row.get("품질점수")),
            _r1(row.get("성장점수")),   _r1(row.get("수익성점수")),
            _r1(row.get("안정점수")),
            # 재무
            _r2(row.get("PBR") or row.get("pbr")),
            _r1(row.get("ROE") or row.get("roe")),
            _r1(row.get("ROA")),
            _r1(row.get("DIV") or 0),
            _r1(row.get("부채비율") or row.get("debt_ratio")),
            _r1(row.get("interest_coverage")),
            _r1(row.get("영업이익률")),
            _eok(row.get("fcf")),
            _eok(row.get("cfo")),
            # 적정가
            _ri(row.get("적정가_PER")),
            _ri(row.get("적정가_PBR")),
            _ri(row.get("적정가_Graham")),
            _ri(row.get("적정가_평균")),
            _r1(row.get("괴리율(%)")),
            str(row.get("적정가_사유","") or ""),
            # 100점
            _r1(pt100),
            str(row.get("100점_등급","") or ""),
            _safe(row.get("100점_수익성")),
            _safe(row.get("100점_성장성")),
            _safe(row.get("100점_재무안정")),
            _safe(row.get("100점_현금흐름")),
            _safe(row.get("100점_밸류")),
            _safe(row.get("100점_주주친화")),
            _safe(row.get("100점_정성보정")),
            # 5대 체크 (숫자+판정)
            str(row.get("체크_CFO","") or ""),
            str(row.get("체크_자본잠식","") or ""),
            str(row.get("체크_부채비율","") or ""),
            str(row.get("체크_ROE","") or ""),
            str(row.get("체크_PER","") or ""),
            # 시그널
            sig_txt,
            str(row.get("현재가위치","") or ""),
            _ri(row.get("적극매수선")),
            _ri(row.get("매수선")),
            _ri(row.get("1차목표가")),
            _ri(row.get("2차목표가")),
            _ri(row.get("손절선")),
            str(row.get("보유전략","") or ""),
            str(row.get("위험신호","없음") or "없음"),
            str(row.get("보너스근거","") or ""),
        ]

        rf = _rank_fill(ri)
        for j, v in enumerate(row_vals, 1):
            c = ws_detail.cell(row=ri+3, column=j, value=v)
            c.fill   = rf
            c.border = _THIN
            c.font   = body_font
            c.alignment = Alignment(horizontal="center", vertical="center",
                                    wrap_text=(j >= 53))
            # 숫자 포맷
            fmt_map = {
                6:"#,##0", 7:"#,##0",
                8:"#,##0.0", 10:"#,##0.0", 11:"#,##0.0", 12:"#,##0.0",
                13:"#,##0.0", 14:"#,##0.0", 15:"#,##0.0", 16:"#,##0.0",
                17:"#,##0.00", 18:"#,##0.0", 19:"#,##0.0", 20:"#,##0.0",
                21:"#,##0.0", 22:"#,##0.0", 23:"#,##0.0", 24:"#,##0", 25:"#,##0",
                26:"#,##0", 27:"#,##0", 28:"#,##0", 29:"#,##0", 30:"#,##0.0",
                31:"#,##0.0",
                48:"#,##0", 49:"#,##0", 50:"#,##0", 51:"#,##0", 52:"#,##0",
            }
            if j in fmt_map and v is not None:
                c.number_format = fmt_map[j]
            # 100점 합계 색상
            if j == 31 and v is not None:
                try:
                    fv = float(v)
                    if fv >= 90:   c.fill = PatternFill("solid", fgColor="0070C0"); c.font = Font(name=KR_FONT, size=9, color="FFFFFF", bold=True)
                    elif fv >= 80: c.fill = PatternFill("solid", fgColor="2E75B6"); c.font = Font(name=KR_FONT, size=9, color="FFFFFF", bold=True)
                    elif fv >= 70: c.fill = PatternFill("solid", fgColor="92D050"); c.font = Font(name=KR_FONT, size=9, bold=True)
                    elif fv >= 60: c.fill = PatternFill("solid", fgColor="FFCC00"); c.font = Font(name=KR_FONT, size=9, bold=True)
                    else:          c.fill = PatternFill("solid", fgColor="FF9900"); c.font = Font(name=KR_FONT, size=9, bold=True)
                except Exception: pass
            # 매매시그널 색상
            if j == 46 and v:
                try:
                    c.fill = PatternFill("solid", fgColor=sig_color)
                    is_dark = sig_color in ("C00000","0070C0","2E75B6","FF4444")
                    c.font = Font(name=KR_FONT, size=9, bold=True,
                                  color="FFFFFF" if is_dark else "000000")
                except Exception: pass
            # 괴리율 색상
            if j == 30 and v is not None:
                try:
                    fv = float(v)
                    c.font = Font(name=KR_FONT, size=9,
                                  color="0070C0" if fv >= 0 else "C00000",
                                  bold=(abs(fv) >= 20))
                except Exception: pass

    ws_detail.freeze_panes = "D4"
    auto_col_width(ws_detail)

    # ── 시트3: 지표 해석 가이드 ──
    ws_guide.title = "지표해석가이드"

    # ── 헬퍼: 가이드 셀 쓰기 ──
    def _gw(row, col, value, bold=False, color="000000", bg=None, wrap=True, size=10, align="left"):
        c = ws_guide.cell(row=row, column=col, value=value)
        c.font      = Font(name=KR_FONT, bold=bold, color=color, size=size)
        c.fill      = PatternFill("solid", fgColor=bg) if bg else PatternFill("none")
        c.border    = _THIN
        c.alignment = Alignment(wrap_text=wrap, vertical="top",
                                horizontal="center" if align=="center" else "left")
        return c

    # 컬럼 너비 수동 설정
    col_widths = [6, 18, 38, 22, 22, 22, 30]
    for i, w in enumerate(col_widths, 1):
        ws_guide.column_dimensions[chr(64+i)].width = w

    # ── 색상 정의 ──
    C_TITLE   = "1F3864"   # 진남색 (대제목)
    C_SEC     = "2E75B6"   # 중간파랑 (섹션)
    C_GREEN   = "E2EFDA"   # 연초록 배경 (좋음)
    C_YELLOW  = "FFF2CC"   # 연노랑 배경 (보통/주의)
    C_RED     = "FCE4D6"   # 연빨강 배경 (위험)
    C_BLUE    = "DEEAF1"   # 연파랑 배경 (정보)
    C_GRAY    = "F5F5F5"   # 연회색 배경 (중립)
    C_ORANGE  = "FCE4D6"   # 주황 배경 (경고)

    # ── 대제목 ──
    ws_guide.merge_cells("A1:G1")
    t = ws_guide.cell(row=1, column=1, value="퀀트 스크리너 지표 해석 완전 가이드")
    t.font      = Font(name=KR_FONT, bold=True, size=16, color="FFFFFF")
    t.fill      = PatternFill("solid", fgColor=C_TITLE)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws_guide.row_dimensions[1].height = 36

    ws_guide.merge_cells("A2:G2")
    sub = ws_guide.cell(row=2, column=1,
        value=f"기준일: {datetime.today().strftime('%Y-%m-%d')}  |  각 지표의 의미, 계산식, 기준값, 투자 판단 활용법")
    sub.font      = Font(name=KR_FONT, size=10, color="FFFFFF", italic=True)
    sub.fill      = PatternFill("solid", fgColor="2E75B6")
    sub.alignment = Alignment(horizontal="center", vertical="center")
    ws_guide.row_dimensions[2].height = 22

    # ── 투자 판단 우선순위 박스 ──
    ws_guide.merge_cells("A3:G3")
    _gw(3, 1, "【 투자 판단 5단계 우선순위 】  1단계: 생존필터(CFO·자본잠식·부채·이자보상) → "
        "2단계: 수익성(ROE·영업이익률) → 3단계: 가격(괴리율·PBR·PER) → "
        "4단계: 현금흐름품질(FCF·CFO>순이익) → 5단계: 복합점수·100점 종합 후 매매시그널",
        bold=True, bg="FFF2CC", size=10, align="left")
    ws_guide.row_dimensions[3].height = 30

    cur_row = 4

    # ── 섹션 정의 ──
    SECTIONS = [

        # ─────────────────────────────────────────────
        ("■ 수익성 지표", None, None),
        # (번호, 지표명, 의미, 계산식, 좋은기준, 보통기준, 위험기준/주의사항)
        ("1", "ROE\n(자기자본이익률)",
         "주주가 맡긴 돈으로 얼마나 벌었는지.\n높을수록 자본 효율이 뛰어난 기업.\n버핏이 가장 중시하는 지표.",
         "순이익 ÷ 자기자본 × 100",
         "✅ 20% 이상\n→ 독점적 해자 가능성\n→ 가산점 +5점",
         "⚠ 10~20%\n→ 보유 가능 수준\n→ 업종 평균과 비교 필요",
         "❌ 10% 미만\n→ 자본 비효율\n→ 주의: 부채로 ROE 부풀리기 가능\n   → ROA와 함께 확인"),

        ("2", "ROA\n(총자산이익률)",
         "총자산 대비 수익률.\nROE가 높아도 ROA가 낮으면\n부채로 수익을 부풀린 것일 수 있음.",
         "순이익 ÷ 총자산 × 100",
         "✅ 7% 이상\n→ 자산 활용 우수",
         "⚠ 3~7%\n→ 보통 수준",
         "❌ 3% 미만\n→ 자산 생산성 낮음\n→ ROE와 괴리 크면 고부채 의심"),

        ("3", "영업이익률",
         "본업으로 매출의 몇 %를 버는지.\n업종별 편차가 매우 크므로\n반드시 동종업계와 비교.",
         "영업이익 ÷ 매출액 × 100",
         "✅ 25% 이상\n→ 독점적 해자 추정\n→ 가산점 +5점\n\n✅ 15~25%\n→ 경쟁우위",
         "⚠ 5~15%\n→ 업종 평균 수준",
         "❌ 5% 미만\n→ 수익성 낮음\n❌ 적자\n→ 하드필터 탈락 가능"),

        ("4", "순이익률",
         "최종적으로 남는 이익 비율.\n영업이익률과 큰 차이가 나면\n이자·세금 부담이 크다는 신호.",
         "당기순이익 ÷ 매출액 × 100",
         "✅ 8% 이상\n→ 우수",
         "⚠ 3~8%\n→ 보통",
         "❌ 3% 미만\n→ 위험\n→ 영업이익률 대비 낮으면\n   이자비용 과다 확인"),

        # ─────────────────────────────────────────────
        ("■ 밸류에이션 지표", None, None),

        ("5", "PBR\n(주가순자산비율)",
         "주가가 순자산의 몇 배인지.\n1배 미만 = 이론상 청산가치보다 싸게 거래.\n현재가·결산기준 두 가지로 산출.",
         "시가총액 ÷ 자기자본(순자산)\n\nPBR갭(%) = (현재가기준 - 결산기준)\n           ÷ 결산기준 × 100",
         "✅ 1배 미만\n→ 저평가 관심\n\nPBR갭 음수\n→ 주가가 자산 대비 내려옴",
         "⚠ 1~2배\n→ 적정 수준",
         "❌ 3배 이상\n→ 고평가 주의\n→ 단, 고ROE 기업은\n   높은 PBR이 정당화될 수 있음"),

        ("6", "PER\n(주가수익비율)",
         "이익 대비 주가 수준.\n낮을수록 싸게 거래되는 것.\n업종 평균과의 비교가 핵심.\n적자 기업은 의미 없음.",
         "주가 ÷ EPS(주당순이익)\n\nPER갭(%) = (현재가기준 - 결산기준)\n           ÷ 결산기준 × 100",
         "✅ 10배 미만\n→ 저평가 매수 선호",
         "⚠ 10~20배\n→ 적정 수준",
         "❌ 25배 이상\n→ 고평가 주의\n→ 성장주는 예외 적용\n   (PEG = PER ÷ 성장률 활용)"),

        ("7", "DIV\n(배당수익률)",
         "주가 대비 연간 배당금 비율.\n3% 이상이면 안정적 인컴 자산.\n배당성향 60% 초과 시 지속가능성 확인.",
         "주당배당금 ÷ 주가 × 100\n\n배당성향 = 배당금 ÷ 순이익 × 100",
         "✅ DIV 3% 이상\n→ 안정적 인컴\n✅ 배당성향 20~50%\n→ 지속 가능",
         "⚠ DIV 1~3%\n⚠ 배당성향 50~70%",
         "❌ 배당성향 80% 초과\n→ 삭감 위험\n❌ DIV 높은데 PER 낮으면\n→ 실적 악화 가능성 확인"),

        ("8", "EV/EBITDA",
         "부채까지 포함한 기업 전체 가치가\n영업이익+감가상각의 몇 배인지.\n부채 많은 기업 비교에 유용.",
         "EV = 시가총액 + 총부채 - 현금\nEBITDA = 영업이익 + 감가상각\n\nEV/EBITDA = EV ÷ EBITDA",
         "✅ 6배 미만\n→ 저평가",
         "⚠ 6~12배\n→ 적정",
         "❌ 15배 이상\n→ 고평가\n→ 장치산업(제조·통신)에서\n   PER보다 더 유용"),

        ("9", "적정가 3종 & 괴리율",
         "세 가지 방법으로 내재가치 추정 후\n평균을 현재가와 비교.\n괴리율이 클수록 저평가.",
         "적정가_PER  = EPS × 업종평균PER\n적정가_PBR  = BPS × 업종평균PBR\n적정가_Graham = √(22.5×EPS×BPS)\n\n괴리율(%) = (적정가평균 - 현재가)\n            ÷ 현재가 × 100",
         "✅ 괴리율 20% 이상\n→ 저평가 매수 기회\n→ 파란색 표시",
         "⚠ 괴리율 0~20%\n→ 소폭 저평가",
         "❌ 괴리율 음수\n→ 고평가\n→ 빨간색 표시\n→ -20% 이하면 매도 검토"),

        # ─────────────────────────────────────────────
        ("■ 재무안정성 지표", None, None),

        ("10", "부채비율",
         "자기자본 대비 총부채 비율.\n업종별 차이 매우 큼.\n금융·지주·건설은 구조적으로 높음.",
         "총부채 ÷ 자기자본 × 100",
         "✅ 100% 미만\n→ 안전\n→ 체크 ✅",
         "⚠ 100~200%\n→ 모니터링\n→ 체크 ⚠",
         "❌ 300% 이상\n→ 위험\n→ 체크 ❌\n→ FCF음수와 겹치면 즉시 경고"),

        ("11", "순부채비율",
         "현금을 제외한 실질 부채 수준.\n음수면 순현금 기업으로 매우 우량.\n부채비율보다 실질 위험을 잘 반영.",
         "(총부채 - 현금및현금성자산)\n÷ 자기자본 × 100",
         "✅ 음수 (순현금 기업)\n→ 최우량 등급",
         "⚠ 0~50%\n→ 안전권",
         "❌ 100% 이상\n→ 실질 부채 과다\n→ 300% 이상이면 감점 -2점"),

        ("12", "유동비율",
         "1년 내 갚아야 할 부채를\n1년 내 현금화 가능한 자산으로\n감당할 수 있는지.",
         "유동자산 ÷ 유동부채 × 100",
         "✅ 200% 이상\n→ 단기 안전",
         "⚠ 150~200%\n→ 양호",
         "❌ 100% 미만\n→ 단기 유동성 위기 가능"),

        ("13", "이자보상배율",
         "영업이익으로 이자를 몇 번 갚을 수 있는지.\n1배 미만 = 이자도 못 버는 한계기업.\nAltman Z보다 더 직관적인 위험 지표.",
         "영업이익 ÷ 이자비용",
         "✅ 5배 이상\n→ 안전",
         "⚠ 2~5배\n→ 양호",
         "❌ 1배 미만\n→ 한계기업\n→ 하드필터 탈락\n→ 실위험 판단 핵심 지표"),

        ("14", "Altman Z-Score",
         "5개 재무비율로 부도 가능성 예측.\n★ 단독 판단 금지 ★\n이자보상배율·FCF와 함께 봐야 함.\n금융·지주·부동산은 구조적으로 낮게 나옴.",
         "Z = 1.2×(운전자본/자산)\n  + 1.4×(이익잉여금/자산)\n  + 3.3×(EBIT/자산)\n  + 0.6×(자기자본/총부채)\n  + 1.0×(매출/자산)",
         "✅ 2.99 이상\n→ 안전지대",
         "⚠ 1.81~2.99\n→ 회색지대\n→ 이자보상 확인 필수",
         "❌ 1.81 미만 + 이자보상<2배\n→ 실제 위험 → 감점\n\n✅ 1.81 미만 + 이자보상≥3배\n→ 업종 구조적 문제\n→ 감점 최소화"),

        # ─────────────────────────────────────────────
        ("■ 현금흐름 지표", None, None),

        ("15", "CFO\n(영업현금흐름)",
         "본업에서 실제로 들어온 현금.\n순이익보다 조작하기 어려워\n이익 품질 판단에 핵심.\n반드시 양수여야 함 (하드필터).",
         "현금흐름표 → 영업활동현금흐름\n(손익계산서 순이익과 다름)",
         "✅ CFO > 0\n→ 생존 필터 통과\n✅ CFO > 순이익\n→ 이익 품질 최상",
         "⚠ CFO ≈ 순이익\n→ 보통 수준",
         "❌ CFO < 0\n→ 하드필터 탈락\n→ 순이익이 플러스여도 의미 없음\n→ '가짜 이익' 가능성"),

        ("16", "FCF\n(잉여현금흐름)",
         "영업현금에서 설비투자(CAPEX)를 뺀\n진짜 자유현금.\n배당·자사주·부채상환 가능한 돈.",
         "FCF = CFO - CAPEX(자본적지출)",
         "✅ FCF > 0\n→ 현금창출 우수\n→ 배당·성장 여력",
         "⚠ FCF ≈ 0\n→ 투자 확대기\n→ 성장주는 일시적 허용",
         "❌ FCF < 0 + 부채비율 > 200%\n→ 감점 -4점\n→ 자금조달(CB·유증) 위험\n\n❌ FCF < 0 + 이자보상 < 2배\n→ 감점 -3점"),

        ("17", "FCF Yield\n(잉여현금수익률)",
         "시가총액 대비 FCF 비율.\n5% 이상이면 현금창출력 우수.\n배당수익률보다 실질 주주환원을 잘 반영.",
         "FCF ÷ 시가총액 × 100",
         "✅ 5% 이상\n→ 매력적\n→ 핵심투자포인트 표시",
         "⚠ 2~5%\n→ 보통",
         "❌ 2% 미만 또는 음수\n→ 현금창출 부족"),

        ("18", "FCF마진",
         "매출액 대비 FCF 비율.\n사업모델의 현금 효율성을 측정.",
         "FCF ÷ 매출액 × 100",
         "✅ 10% 이상\n→ 우수한 사업모델",
         "⚠ 5~10%\n→ 보통",
         "❌ 5% 미만 또는 음수\n→ 자본집약도 높음"),

        # ─────────────────────────────────────────────
        ("■ 성장성 · 모멘텀 지표", None, None),

        ("19", "매출성장률",
         "전년 대비 매출 증가율.\n지속적인 성장이 핵심.\n3년 연속 감소는 사업 경쟁력 약화 신호.",
         "( 당기매출 - 전기매출 ) ÷ 전기매출 × 100",
         "✅ 20% 이상\n→ 고성장 (가산점 +3점)\n✅ 10~20%\n→ 성장 (가산점 +1점)",
         "⚠ 0~10%\n→ 안정성장",
         "❌ -5% 이하\n→ 매출감소 (감점 -1점)\n❌ -10% 이하\n→ 매출급감 (감점 -3점)"),

        ("20", "52주위치(%)",
         "52주 최저~최고가 구간에서\n현재가가 어디에 위치하는지.\n0%=52주 최저가, 100%=52주 최고가.",
         "( 현재가 - 52주최저 )\n÷ ( 52주최고 - 52주최저 ) × 100",
         "✅ 0~30%\n→ 52주 저점 근처\n→ 역발상 매수 기회\n→ 핵심투자포인트 표시",
         "⚠ 30~70%\n→ 중간 구간",
         "❌ 80% 이상\n→ 52주 고점 근처\n→ 추격매수 주의"),

        ("21", "6개월수익률",
         "최근 6개월 주가 상승률.\n모멘텀 팩터의 핵심 지표.\n강한 모멘텀은 추세 지속 경향.",
         "( 현재가 - 6개월전가 ) ÷ 6개월전가 × 100",
         "✅ 15% 이상\n→ 강한 모멘텀\n→ 핵심투자포인트 표시",
         "⚠ 0~15%\n→ 중립",
         "❌ -10% 이하\n→ 하락 추세"),

        # ─────────────────────────────────────────────
        ("■ 100점 만점 종합 평가", None, None),

        ("22", "100점_합계\n& 등급",
         "6개 카테고리 합산 + 정성 보정.\n복합점수(7팩터)와 함께 투자 판단.",
         "수익성20 + 성장성20 + 재무안정20\n+ 현금흐름20 + 밸류10 + 주주친화10\n± 정성보정(최대 ±10점)",
         "✅ 90점 이상 → ★★★★★ 최상급\n✅ 80~89점 → ★★★★  우수",
         "⚠ 70~79점 → ★★★ 양호\n⚠ 60~69점 → ★★  보통",
         "❌ 50~59점 → ★ 주의\n❌ 50점 미만 → 투자재검토"),

        ("23", "정성보정\n(±10점)",
         "데이터로 잡기 어려운 질적 요소를\n가감점으로 반영.",
         "가산: 독점해자+5, 성장산업+5, 고성장+3\n감점: 부도위험-5, FCF음수+고부채-4,\n      매출급감-3, 순부채과다-2",
         "✅ +5점 이상\n→ 강한 경쟁우위 보유",
         "⚠ 0점\n→ 중립",
         "❌ -5점 이하\n→ 실질 위험 존재\n→ 감점 근거 컬럼 반드시 확인"),

        # ─────────────────────────────────────────────
        ("■ 매매시그널 & 가격 밴드", None, None),

        ("24", "매매시그널\n(7단계)",
         "복합점수·100점·괴리율·위험신호를\n종합해 자동 산출하는 투자 의견.\n주의: 참고용이며 최종 판단은 직접.",
         "■■■ 강력매수: 복합≥85 & 100점≥80\n■■  매수:    복합≥75 & 100점≥70\n■   관심:    복합≥65 & 100점≥60\n─   보유:    특이사항 없음\n▽   비중축소: 괴리율 < -10%\n▼▼  매도검토: 실위험 1개 이상\n▼▼▼ 즉시매도: 실위험 2개 이상",
         "✅ ■■■ 강력매수\n→ 분할매수 2~3회 진입",
         "⚠ ─ 보유\n→ 1차목표가까지 보유",
         "❌ ▼▼ 매도검토\n→ 손절선 이탈 시 매도\n❌ ▼▼▼ 즉시매도\n→ 신규 진입 금지"),

        ("25", "가격 밴드\n(6단계)",
         "현재가·적정가 기준으로\n매수·매도 가격선 자동 계산.\n손절선은 반드시 설정 후 진입.",
         "적극매수선 = 현재가 × 0.95\n매수선     = 현재가\n1차목표가  = 적정가 평균\n2차목표가  = 적정가 × 1.15\n추격매수경보 = 적정가 × 1.20\n손절선     = 현재가 × 0.85",
         "✅ 현재가 < 적극매수선\n→ 적극 진입 구간",
         "⚠ 현재가 = 매수선 근처\n→ 현재 매수 가능",
         "❌ 현재가 > 추격매수경보\n→ 고평가, 신규 진입 자제\n❌ 현재가 < 손절선\n→ 즉시 손절"),

        ("26", "위험신호\n(3단계)",
         "개별 지표가 아닌 복합 조건으로\n실질 위험을 판단.\n오탐(False Positive) 최소화.",
         "🔴 실위험: 복합 조건 충족\n   (Altman<1.8 AND IC<2배 등)\n🟡 주의: 단일 조건 (실위험 낮음)\n⬜ 없음: 모두 정상",
         "⬜ 없음\n→ 위험 요소 없음",
         "🟡 주의\n→ 해당 지표 추가 확인\n→ 매수시그널 1단계 하향",
         "🔴 실위험 1개 → 매도검토\n🔴 실위험 2개↑ → 즉시매도\n→ 감점근거 컬럼 반드시 확인"),

        # ─────────────────────────────────────────────
        ("■ 7팩터 복합점수 구성", None, None),

        ("27", "가치점수\n(Value)",
         "PBR·ROE·FCF수익률·EV/EBITDA\n조합으로 계산한 가치 점수.",
         "PBR 30% + ROE 30%\n+ FCF Yield 20% + EV/EBITDA 20%",
         "✅ PBR≤0.5, FCFYield≥8%\n→ 80점 이상 목표",
         "⚠ PBR 1~2배 수준",
         "❌ PBR≥3.0\n❌ EV/EBITDA≥20"),

        ("28", "모멘텀점수\n(Momentum)",
         "52주 위치와 6개월 수익률 기반.\n추세 추종 전략의 핵심 요소.",
         "52주위치 70% + 6개월수익률 30%",
         "✅ 52주 40~75% 구간\n+ 6개월 수익률 양호",
         "⚠ 52주 위치 중립 구간",
         "❌ 52주 0% 또는 100% 극단\n→ 과매도/과매수 판단 필요"),

        ("29", "품질점수\n(Quality)",
         "재무 건전성 종합.\n부채·ROA·영업이익률 기반.",
         "부채비율 + ROA + 영업이익률\n3개 지표 가중 평균",
         "✅ 부채비율≤50%\n✅ ROA≥7%",
         "⚠ 부채비율 100~200%",
         "❌ 부채비율≥300%\n❌ ROA 음수"),

        ("30", "안정점수\n(Stability)",
         "재무 안정성 종합.\nAltman Z·유동비율·ROE·배당 포함.",
         "순부채비율35% + 유동비율15%\n+ AltmanZ20% + ROE15% + 배당15%",
         "✅ Altman≥3.0\n✅ 유동비율≥150%",
         "⚠ Altman 회색지대",
         "❌ Altman≤1.8 (복합 확인)\n❌ 유동비율≤100%"),

        # ─────────────────────────────────────────────
        ("■ 박스권(장기 횡보) 분석 — v41g 신규", None, None),

        ("31", "박스권여부",
         "최근 가격이 일정 구간(60~120거래일,\n약 3~6개월) 안에서 횡보했는지 여부.\n1개월 미만 단기 등락은 박스권으로 보지 않음.",
         "120일→90일→60일 순으로 검사:\n가격폭≤35% + 추세(드리프트)≤12%\n+ 종가의 90%이상이 밴드 안에 머묾",
         "✅ Y\n→ 장기 횡보 구간 형성\n→ 돌파 시 추세 신뢰도 높음",
         "─ N\n→ 추세장이거나 등락이 큰 종목",
         "박스권 여부만으로는 매수신호 아님\n→ 박스이탈신호와 함께 판단"),

        ("32", "박스기간(일)\n박스상단·박스하단",
         "박스권으로 인정된 구간의 길이(거래일)와\n그 구간의 저항선(상단)·지지선(하단).\n길수록(120일 쪽일수록) 돌파 시 의미가 큼.",
         "구간 내 최고가 = 박스상단\n구간 내 최저가 = 박스하단",
         "✅ 120일 박스\n→ 6개월짜리 강한 매물대\n→ 돌파 시 강한 추세 기대",
         "⚠ 60일 박스\n→ 3개월짜리, 상대적으로 약함",
         "박스하단 이탈 시 지지선 붕괴\n→ 추가 하락 가능성 점검"),

        ("33", "박스폭(%)\n박스내위치(%)",
         "박스 상단·하단의 가격 차이(%)와\n현재가가 박스 안 어디에 있는지(0~100%).\n0%=박스하단, 100%=박스상단.",
         "박스폭 = (박스상단-박스하단)÷박스하단×100\n박스내위치 = (현재가-박스하단)\n            ÷(박스상단-박스하단)×100",
         "✅ 박스내위치 80%이상\n→ 상단 돌파 임박, 주목 구간",
         "⚠ 박스내위치 20~80%\n→ 박스 중간, 관망",
         "❌ 박스내위치 20%이하\n→ 하단 지지 테스트 중\n→ 이탈 시 손절 고려"),

        ("34", "박스이탈신호",
         "박스권을 실제로 벗어났는지,\n거래량까지 함께 터졌는지를 종합 판정.\n뉴스·재료 발생 시 거래량과 함께\n박스 상단을 뚫는 패턴을 포착하기 위한 지표.",
         "상단돌파 + (5일/20일 거래대금≥1.5배)\n  → 뉴스성모멘텀후보\n상단돌파(거래량 미동반) → 확인필요\n하단이탈 / 상단근접 / 하단근접 / 박스권유지",
         "✅ 상단돌파+거래량급증\n(뉴스성모멘텀후보)\n→ 핵심투자포인트에도 자동 표시\n→ 재료성 급등 후보, 관심 1순위",
         "⚠ 상단돌파(확인필요)\n→ 거래량 미동반, 가짜돌파 주의\n⚠ 상단근접(돌파대기)\n→ 분할 매수 후보 관찰",
         "❌ 하단이탈(주의)\n→ 지지선 붕괴, 추가 하락 경계\n❌ 데이터부족/박스아님\n→ 추세장이거나 상장초기"),
    ]

    # ── 헤더 행 출력 ──
    HDR_ROW = cur_row
    hdr_vals = ["No.", "지표명", "의미 및 설명", "계산식", "✅ 좋음 (기준값)", "⚠ 보통 (주의)", "❌ 위험 / 주의사항"]
    for ci, v in enumerate(hdr_vals, 1):
        _gw(HDR_ROW, ci, v, bold=True, color="FFFFFF", bg=C_TITLE, size=10, align="center")
    ws_guide.row_dimensions[HDR_ROW].height = 22
    cur_row += 1

    # ── 섹션 & 데이터 행 출력 ──
    for item in SECTIONS:
        if item[1] is None:
            # 섹션 구분 행
            ws_guide.merge_cells(f"A{cur_row}:G{cur_row}")
            c = ws_guide.cell(row=cur_row, column=1, value=item[0])
            c.font      = Font(name=KR_FONT, bold=True, size=11, color="FFFFFF")
            c.fill      = PatternFill("solid", fgColor=C_SEC)
            c.border    = _THIN
            c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            ws_guide.row_dimensions[cur_row].height = 22
            cur_row += 1
        else:
            no, name, desc, formula, good, mid, bad = item
            row_data = [no, name, desc, formula, good, mid, bad]
            # 행 높이 = 내용에 따라 조정
            ws_guide.row_dimensions[cur_row].height = 75

            for ci, v in enumerate(row_data, 1):
                if ci == 1:   # No.
                    _gw(cur_row, ci, v, bold=True, bg=C_GRAY, size=10, align="center")
                elif ci == 2: # 지표명
                    _gw(cur_row, ci, v, bold=True, bg=C_BLUE, size=10)
                elif ci == 3: # 의미
                    _gw(cur_row, ci, v, size=10)
                elif ci == 4: # 계산식
                    c = ws_guide.cell(row=cur_row, column=ci, value=v)
                    c.font      = Font(name="Courier New", size=9, color="2E4057")
                    c.fill      = PatternFill("solid", fgColor="F0F4F8")
                    c.border    = _THIN
                    c.alignment = Alignment(wrap_text=True, vertical="top")
                elif ci == 5: # 좋음
                    _gw(cur_row, ci, v, size=10, bg=C_GREEN)
                elif ci == 6: # 보통
                    _gw(cur_row, ci, v, size=10, bg=C_YELLOW)
                elif ci == 7: # 위험
                    _gw(cur_row, ci, v, size=10, bg=C_RED)
            cur_row += 1

    # ── 하단 요약 박스 ──
    ws_guide.merge_cells(f"A{cur_row}:G{cur_row}")
    _gw(cur_row, 1,
        "【 실전 투자 체크리스트 】  "
        "① CFO > 0  "
        "② 자본잠식 없음  "
        "③ 부채비율 < 150%  "
        "④ ROE ≥ 10%  "
        "⑤ PER 업종 평균 이하  "
        "⑥ FCF > 0  "
        "⑦ 이자보상배율 ≥ 3배  "
        "→ 7개 모두 통과하면 부실기업 대부분 걸러짐",
        bold=True, bg="D9EAD3", size=10)
    ws_guide.row_dimensions[cur_row].height = 28

    ws_guide.freeze_panes = "A5"

    # ══════════════════════════════════════════════════════════
    # 전체종목데이터 시트 — 전체 컬럼(104개) 상세 가이드 (v41g 신규)
    # ══════════════════════════════════════════════════════════
    cur_row += 2
    ws_guide.merge_cells(f"A{cur_row}:G{cur_row}")
    t2 = ws_guide.cell(row=cur_row, column=1,
        value="【 전체종목데이터 시트 — 전체 컬럼 상세 가이드 (총 104개 컬럼, 시트 내 등장 순서대로) 】")
    t2.font      = Font(name=KR_FONT, bold=True, size=13, color="FFFFFF")
    t2.fill      = PatternFill("solid", fgColor=C_TITLE)
    t2.alignment = Alignment(horizontal="center", vertical="center")
    ws_guide.row_dimensions[cur_row].height = 26
    cur_row += 1

    ws_guide.merge_cells(f"C{cur_row}:E{cur_row}")
    ws_guide.merge_cells(f"F{cur_row}:G{cur_row}")
    _gw(cur_row, 1, "No.",   bold=True, color="FFFFFF", bg=C_SEC, size=10, align="center")
    _gw(cur_row, 2, "컬럼명", bold=True, color="FFFFFF", bg=C_SEC, size=10, align="center")
    _gw(cur_row, 3, "설명(의미)", bold=True, color="FFFFFF", bg=C_SEC, size=10, align="center")
    _gw(cur_row, 6, "해석가이드 (기준값 · 활용법)", bold=True, color="FFFFFF", bg=C_SEC, size=10, align="center")
    ws_guide.row_dimensions[cur_row].height = 20
    cur_row += 1

    FULL_COL_GUIDE = [
        ("SEC", "■ 기본정보 (1~8번)"),
        (1,  "순위",        "복합점수 내림차순 정렬 순서.",
             "낮을수록(1위에 가까울수록) 종합 매력도가 높은 종목."),
        (2,  "코드",        "종목코드(6자리).",
             "매매 시 종목 식별자. 변하지 않는 고유값."),
        (3,  "종목명",      "종목 이름.",
             "그대로 참고."),
        (4,  "시장",        "KOSPI / KOSDAQ 구분.",
             "KOSDAQ은 일반적으로 변동성·리스크가 더 큰 편."),
        (5,  "업종",        "업종(섹터) 분류명.",
             "PER·PBR 등을 볼 때 반드시 업종 평균과 비교."),
        (6,  "현재가",      "최근 거래일 종가(원).",
             "매수선·손절선 등 가격 밴드 비교의 기준값."),
        (7,  "시가총액(억)", "발행주식수 × 현재가 (억원).",
             "1,000억 미만 소형주는 변동성·유동성 리스크 체크 필요."),
        (8,  "발행주식수",  "총 발행 주식 수.",
             "PBR·PER 등 1주당 지표를 직접 재계산할 때 사용."),

        ("SEC", "■ 거래량 · 기술지표 (9~28번)"),
        (9,  "거래대금5일(억)",  "최근 5거래일 평균 거래대금.", "단기 관심도·수급 강도를 보여줌."),
        (10, "거래대금20일(억)", "최근 20거래일 평균 거래대금.", "중기 평균 수급 규모, 5일값과 비교 기준."),
        (11, "거래량비율VR(%)", "5일 거래대금 ÷ 20일 거래대금 × 100.", "100% 초과면 최근 거래 활발(관심 집중), 50% 미만이면 관심 저하."),
        (12, "MACD선",      "12일EMA - 26일EMA.", "추세 방향성을 보는 기초선."),
        (13, "MACD시그널",  "MACD선의 9일 지수이동평균.", "MACD선과의 교차로 매매 시점을 포착."),
        (14, "MACD히스토",  "MACD선 - MACD시그널.", "0선을 위로 교차 = 골든크로스(매수), 아래로 = 데드크로스(매도)."),
        (15, "MACD신호",    "히스토그램 부호 교차 기반 매수·매도 판정.", "'매수'(골든크로스 발생 시점) 직후가 진입 타이밍."),
        (16, "RSI14",       "14일 상대강도지수 (0~100).", "30 이하 과매도(반등 기대), 70 이상 과매수(조정 주의)."),
        (17, "RSI신호",     "RSI14 기반 매수·매도 판정.", "'매수' = 과매도 진입권."),
        (18, "OBV추세",     "누적거래량(OBV)이 5일 전보다 상승/하락했는지.", "가격과 거래량의 동행 여부를 확인."),
        (19, "OBV신호",     "OBV 5일선과 20일선 비교 판정.", "가격 상승 + OBV매수 동반 시 신뢰도 ↑ (가짜 상승 배제)."),
        (20, "BB상단",      "볼린저밴드 상단 (20일, ±2표준편차).", "단기 과열 기준선."),
        (21, "BB하단",      "볼린저밴드 하단.", "단기 과매도 기준선."),
        (22, "BB%B",        "밴드 내 현재가 상대위치 (0~100%).", "0% 이하 하단이탈(과매도), 100% 이상 상단이탈(과매수)."),
        (23, "BB밴드폭(%)", "(상단-하단) ÷ 중간(20일 평균) × 100.", "폭이 좁아지는 '스퀴즈'는 곧 변동성 확대(돌파)가 임박했다는 신호."),
        (24, "BB신호",      "%B 기반 매수·매도 판정.", "하단이탈 = '매수' 신호."),
        (25, "기술매수신호수", "MACD/RSI/OBV/BB 중 매수신호 개수 (0~4).", "3개 이상이면 기술적 매수신호 집중 구간."),
        (26, "기술매도신호수", "위 4개 지표 중 매도신호 개수.", "3개 이상이면 기술적 매도신호 집중 구간."),
        (27, "기술신호",    "매수/매도신호수를 종합한 한 줄 평가.", "'⚡ 강력매수'가 가장 강한 기술적 매수 신호."),
        (28, "기술지표요약", "4개 지표 개별 신호를 한 줄로 표기.", "지표들이 서로 엇갈릴 때 한눈에 확인하는 용도."),

        ("SEC", "■ 박스권(장기 횡보) 분석 — v41g 신규 (29~35번)"),
        (29, "박스권여부",   "최근 60~120거래일(약 3~6개월) 가격이 횡보했는지 여부.", "위 [박스권 분석] 섹션 31번 항목 참고. 1개월 미만 단기 등락은 제외."),
        (30, "박스기간(일)", "박스권으로 인정된 구간 길이(60/90/120일).", "길수록(120일 쪽) 박스가 더 견고하다고 봄 → 돌파 시 추세 신뢰도 ↑."),
        (31, "박스상단",     "박스 구간의 저항선(구간 내 최고가).", "돌파 시 1차 목표가·추격매수 판단의 기준선."),
        (32, "박스하단",     "박스 구간의 지지선(구간 내 최저가).", "이탈 시 손절·위험신호로 즉시 활용."),
        (33, "박스폭(%)",    "(박스상단-박스하단) ÷ 박스하단 × 100.", "폭이 좁을수록(20% 이하) 돌파 시 폭발력이 큰 경향."),
        (34, "박스내위치(%)", "현재가가 박스 안에서 차지하는 상대 위치 (0~100%).", "80% 이상이면 상단 돌파 임박 구간으로 주목."),
        (35, "박스이탈신호", "박스권 + 거래량 종합 판정 결과.", "'상단돌파+거래량급증(뉴스성모멘텀후보)'가 핵심 관찰 대상 — 재료·뉴스 발생 시 자주 나타나는 패턴."),

        ("SEC", "■ 밸류에이션 (36~45번)"),
        (36, "PBR(현재가)", "현재가 기준 주가순자산비율.", "1배 미만이면 저평가권 (항목5 참고)."),
        (37, "PBR(결산기준)", "최근 결산 시점 기준 PBR.", "현재가기준과 비교해 결산 이후 갭을 확인."),
        (38, "PBR갭(%)",   "(현재가PBR-결산PBR)÷결산PBR×100.", "음수면 결산 발표 이후 주가가 더 빠진 상태."),
        (39, "ROE(%)",     "자기자본이익률.", "항목1 참고 — 20% 이상이면 가산점 대상."),
        (40, "ROA(%)",     "총자산이익률.", "항목2 참고 — 7% 이상 우수, ROE와 괴리 크면 고부채 의심."),
        (41, "PER(현재가)", "현재가 기준 주가수익비율.", "10배 미만이면 저평가권 (항목6 참고)."),
        (42, "PER(결산기준)", "결산 시점 기준 PER.", "현재가기준과 비교해 실적 발표 전후 갭 확인."),
        (43, "PER갭(%)",   "PBR갭과 동일한 계산 로직.", "음수면 실적 발표 이후 주가 하락."),
        (44, "DIV(%)",     "배당수익률.", "항목7 참고 — 3% 이상이면 안정적 인컴 매력."),
        (45, "배당성향(%)", "배당금 ÷ 순이익 × 100.", "80% 초과 시 배당 삭감 위험 (항목7 참고)."),

        ("SEC", "■ 수익성 · 재무안정성 (46~51번)"),
        (46, "영업이익률(%)", "본업 수익성.", "항목3 참고 — 반드시 업종 평균과 비교."),
        (47, "순이익률(%)",   "최종 이익 비율.", "항목4 참고 — 영업이익률과 격차 크면 이자·세금 부담 확인."),
        (48, "부채비율(%)",   "총부채÷자기자본×100.", "항목10 참고 — 100% 미만 안전."),
        (49, "순부채비율(%)", "현금 제외 실질 부채 수준.", "항목11 참고 — 음수면 순현금기업(최우량)."),
        (50, "유동비율(%)",   "단기 채무상환능력.", "항목12 참고 — 200% 이상 단기 안전."),
        (51, "Altman-Z",      "부도위험 예측 점수.", "항목14 참고 — 단독판단 금지, 이자보상배율과 함께 확인."),

        ("SEC", "■ 현금흐름 · 이자보상 (52~59번)"),
        (52, "FCF(억)",       "영업현금흐름-CAPEX.", "항목16 참고 — 양수면 배당·성장 여력 있음."),
        (53, "FCF마진(%)",    "FCF÷매출액×100.", "항목18 참고 — 10% 이상 우수한 사업모델."),
        (54, "FCF Yield(%)",  "FCF÷시가총액×100.", "항목17 참고 — 5% 이상 매력적."),
        (55, "EBITDA(억)",    "영업이익+감가상각.", "EV/EBITDA 산출에 사용되는 분모값."),
        (56, "EV/EBITDA",     "기업가치÷EBITDA.", "항목8 참고 — 6배 미만 저평가."),
        (57, "매출성장률(%)", "전년 대비 매출 증가율.", "항목19 참고 — 20% 이상 고성장(가산점)."),
        (58, "CFO(억)",       "영업활동현금흐름.", "항목15 참고 — 반드시 양수(하드필터, 음수면 탈락)."),
        (59, "이자보상배율",  "영업이익÷이자비용.", "항목13 참고 — 1배 미만이면 한계기업."),

        ("SEC", "■ 모멘텀 (60~62번)"),
        (60, "52주위치(%)",   "52주 최저~최고 구간 내 현재가 위치.", "항목20 참고 — 0~30%면 역발상 매수 관심권."),
        (61, "52주수익률(%)", "52주 전 대비 수익률.", "장기 추세를 가늠하는 보조 지표."),
        (62, "6개월수익률(%)", "최근 6개월 수익률.", "항목21 참고 — 15% 이상이면 강한 모멘텀."),

        ("SEC", "■ 7팩터 복합점수 구성 (63~70번)"),
        (63, "가치점수",   "PBR·ROE·FCF Yield·EV/EBITDA 조합.", "항목27 참고."),
        (64, "모멘텀점수", "52주위치·6개월수익률 조합.", "항목28 참고."),
        (65, "배당점수",   "DIV·배당성향 기반 점수.", "배당 매력도를 0~100점으로 환산."),
        (66, "품질점수",   "부채비율·ROA·영업이익률 조합.", "항목29 참고."),
        (67, "성장점수",   "매출성장률 등 성장성 지표 기반.", "고성장 종목일수록 점수가 높음."),
        (68, "수익성점수", "ROE·ROA·영업이익률·순이익률 종합.", "수익성 카테고리 전반을 점수화."),
        (69, "안정점수",   "Altman Z·유동비율·ROE·배당 조합.", "항목30 참고."),
        (70, "복합점수",   "7개 팩터 점수의 가중합.", "이 시트 '순위' 정렬 기준이 되는 핵심 점수."),

        ("SEC", "■ 적정가 3종 & 등급 (71~77번)"),
        (71, "적정가_PER",    "EPS×업종평균PER로 산출한 적정가.", "항목9 참고."),
        (72, "적정가_PBR",    "BPS×업종평균PBR로 산출한 적정가.", "항목9 참고."),
        (73, "적정가_Graham", "√(22.5×EPS×BPS) 그레이엄 공식 적정가.", "항목9 참고 — 보수적인 추정값."),
        (74, "적정가_평균",   "위 3종 적정가의 평균.", "현재가와 비교해 괴리율을 계산하는 기준값."),
        (75, "괴리율(%)",     "(적정가평균-현재가)÷현재가×100.", "항목9 참고 — 20% 이상이면 저평가 매수 기회로 파란색 표시."),
        (76, "적정가_사유",   "적정가 산출에 어떤 방식이 쓰였는지(데이터 결측 시 대체방식 등) 설명.", "EPS·BPS 등 원천 데이터가 없을 때 어떤 방법으로 대체했는지 확인용."),
        (77, "등급",          "복합점수 기준 별 등급(grade_label).", "◆◆◆ A+가 최상위 등급."),

        ("SEC", "■ 100점 종합평가 (78~86번)"),
        (78, "100점_합계",        "6개 카테고리 합산+정성보정.", "항목22 참고."),
        (79, "100점_등급",        "100점 합계 기준 별점 등급(★).", "90점 이상 ★★★★★ 최상급."),
        (80, "100점_수익성(20)",  "수익성 카테고리 배점(만점 20).", "ROE·ROA·영업이익률·순이익률 기반."),
        (81, "100점_성장성(20)",  "성장성 카테고리 배점(만점 20).", "매출성장률 등 기반."),
        (82, "100점_재무안정(20)", "재무안정 카테고리 배점(만점 20).", "부채비율·유동비율·Altman-Z 기반."),
        (83, "100점_현금흐름(20)", "현금흐름 카테고리 배점(만점 20).", "CFO·FCF 기반."),
        (84, "100점_밸류(10)",    "밸류에이션 카테고리 배점(만점 10).", "PER·PBR 기반."),
        (85, "100점_주주친화(10)", "주주환원 카테고리 배점(만점 10).", "DIV·배당성향 기반."),
        (86, "100점_정성보정",   "데이터로 잡기 힘든 질적 요소 가감점.", "항목23 참고 — ±10점 한도."),

        ("SEC", "■ 판정 플래그 & 근거 (87~93번)"),
        (87, "CFO(판정)",     "CFO 하드필터 통과 여부(✅/❌).", "❌면 100점 평가에서도 큰 감점 대상."),
        (88, "자본잠식(판정)", "자본잠식 여부 체크.", "❌면 즉시 투자 재검토 대상."),
        (89, "부채비율(판정)", "부채비율 등급(✅/⚠/❌).", "항목10 기준 동일."),
        (90, "ROE(판정)",     "ROE 등급(✅/⚠/❌).", "항목1 기준 동일."),
        (91, "PER(판정)",     "PER 등급(✅/⚠/❌).", "항목6 기준 동일."),
        (92, "보너스근거",    "100점_정성보정에서 가산된 이유 텍스트.", "독점해자·고성장 등 가산 사유를 그대로 기록."),
        (93, "감점근거",      "100점_정성보정에서 감점된 이유 텍스트.", "부도위험·FCF음수 등 감점 사유를 그대로 기록."),

        ("SEC", "■ 매매시그널 · 가격밴드 · 전략 (94~104번)"),
        (94, "매매시그널",     "7단계 자동 매매 의견.", "항목24 참고."),
        (95, "현재가위치",     "가격밴드 대비 현재가가 어디에 있는지 설명.", "적극매수선·매수선 등과 비교한 위치 텍스트."),
        (96, "적극매수선",     "현재가×0.95.", "항목25 참고."),
        (97, "매수선",         "현재가 그대로.", "항목25 참고."),
        (98, "1차목표가",      "적정가 평균.", "항목25 참고."),
        (99, "2차목표가",      "적정가×1.15.", "항목25 참고."),
        (100,"추격매수경보",   "적정가×1.20.", "항목25 참고 — 이 가격을 넘으면 추격매수 자제."),
        (101,"손절선",         "현재가×0.85.", "항목25 참고 — 반드시 진입 전에 확인."),
        (102,"보유전략",       "매매시그널에 따른 구체적 행동 가이드 문장.", "분할매수·익절·손절 시점을 구체적으로 안내."),
        (103,"핵심투자포인트", "해당 종목의 매수 근거를 모아놓은 한 줄 요약.", "박스권 돌파(뉴스성모멘텀후보) 등 v41g 신규 포인트도 여기 자동 표시."),
        (104,"위험신호",       "복합 조건 기반 실위험 판정(🔴/🟡/⬜).", "항목26 참고."),
    ]

    for item in FULL_COL_GUIDE:
        if item[0] == "SEC":
            ws_guide.merge_cells(f"A{cur_row}:G{cur_row}")
            c = ws_guide.cell(row=cur_row, column=1, value=item[1])
            c.font      = Font(name=KR_FONT, bold=True, size=11, color="FFFFFF")
            c.fill      = PatternFill("solid", fgColor=C_SEC)
            c.border    = _THIN
            c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            ws_guide.row_dimensions[cur_row].height = 20
            cur_row += 1
        else:
            no, name, desc, interp = item
            ws_guide.merge_cells(f"C{cur_row}:E{cur_row}")
            ws_guide.merge_cells(f"F{cur_row}:G{cur_row}")
            _gw(cur_row, 1, no,     bold=True, bg=C_GRAY, size=9, align="center")
            _gw(cur_row, 2, name,   bold=True, bg=C_BLUE, size=9)
            _gw(cur_row, 3, desc,   size=9)
            _gw(cur_row, 6, interp, size=9, bg=C_YELLOW)
            ws_guide.row_dimensions[cur_row].height = 30
            cur_row += 1

    ws_guide.freeze_panes = "A5"


    # ── 시트4: 전체 종목 데이터 ──
    ws_all.title = "전체종목데이터"
    ws_all_hdr_font = Font(name=KR_FONT, bold=True, color="FFFFFF", size=9)
    ws_all_body_font = Font(name=KR_FONT, size=9)

    # 제목행
    ws_all.merge_cells("A1:BW1")
    t_all = ws_all.cell(row=1, column=1,
        value=f"전체 종목 데이터 [{label}]  기준일: {datetime.today().strftime('%Y-%m-%d')}  (복합점수 내림차순)")
    t_all.font  = Font(name=KR_FONT, bold=True, size=12, color="FFFFFF")
    t_all.fill  = PatternFill("solid", fgColor="1F3864")
    t_all.alignment = Alignment(horizontal="center", vertical="center")
    ws_all.row_dimensions[1].height = 26

    # 전체종목 헤더 (기술지표 추가)
    all_headers = ["순위","코드","종목명","시장","업종","현재가",
                   "시가총액(억)" if is_kr else "시총",
                   "발행주식수",
                   # 거래량
                   "거래대금5일(억)","거래대금20일(억)","거래량비율VR(%)",
                   # 기술지표 (v32)
                   "MACD선","MACD시그널","MACD히스토","MACD신호",
                   "RSI14","RSI신호",
                   "OBV추세","OBV신호",
                   "BB상단","BB하단","BB%B","BB밴드폭(%)","BB신호",
                   "기술매수신호수","기술매도신호수","기술신호","기술지표요약",
                   # 박스권 분석 (v41g 신규) — 장기 횡보 구간 + 뉴스성 돌파 후보 탐지
                   "박스권여부","박스기간(일)","박스상단","박스하단",
                   "박스폭(%)","박스내위치(%)","박스이탈신호",
                   # 밸류에이션
                   "PBR(현재가)","PBR(결산기준)","PBR갭(%)",
                   "ROE(%)","ROA(%)",
                   "PER(현재가)","PER(결산기준)","PER갭(%)",
                   "DIV(%)","배당성향(%)",
                   "영업이익률(%)","순이익률(%)","부채비율(%)","순부채비율(%)",
                   "유동비율(%)","Altman-Z",
                   "FCF(억)" if is_kr else "FCF","FCF마진(%)","FCF Yield(%)",
                   "EBITDA(억)" if is_kr else "EBITDA","EV/EBITDA",
                   "매출성장률(%)",
                   "CFO(억)" if is_kr else "CFO",
                   "이자보상배율",
                   "52주위치(%)","52주수익률(%)","6개월수익률(%)",
                   "가치점수","모멘텀점수","배당점수",
                   "품질점수","성장점수","수익성점수","안정점수",
                   "복합점수",
                   "적정가_PER","적정가_PBR","적정가_Graham","적정가_평균","괴리율(%)","적정가_사유",
                   "등급",
                   "100점_합계","100점_등급",
                   "100점_수익성(20)","100점_성장성(20)","100점_재무안정(20)",
                   "100점_현금흐름(20)","100점_밸류(10)","100점_주주친화(10)","100점_정성보정",
                   "CFO(판정)","자본잠식(판정)","부채비율(판정)","ROE(판정)","PER(판정)",
                   "보너스근거","감점근거",
                   "매매시그널","현재가위치",
                   "적극매수선","매수선","1차목표가","2차목표가","추격매수경보","손절선",
                   "보유전략","핵심투자포인트","위험신호"]

    # 헤더 색상 그룹 (v41g: 박스권 컬럼 추가로 전체 재구성 + 누락 구간 보완)
    _ALL_HDR_GROUPS = {
        range(1,9):    "1F3864",   # 기본정보 (순위~발행주식수)
        range(9,29):   "2E75B6",   # 거래량 + 기술지표(MACD/RSI/OBV/BB)
        range(29,36):  "BF8F00",   # 박스권 분석 (v41g 신규)
        range(36,46):  "375623",   # 밸류에이션 (PBR~배당성향)
        range(46,52):  "7030A0",   # 수익성·재무안정성 (영업이익률~Altman-Z)
        range(52,60):  "833C00",   # 현금흐름·이자보상 (FCF~이자보상배율)
        range(60,63):  "C00000",   # 모멘텀 (52주위치~6개월수익률)
        range(63,71):  "4A4A4A",   # 7팩터 복합점수
        range(71,78):  "2E5C8A",   # 적정가 3종 & 등급
        range(78,87):  "538DD5",   # 100점 종합평가
        range(87,94):  "A6A6A6",   # 판정 플래그 & 근거
        range(94,105): "9C0006",   # 매매시그널·가격밴드·전략
    }
    def _all_hdr_color(j):
        for rng, color in _ALL_HDR_GROUPS.items():
            if j in rng: return color
        return "1F3864"

    # 컬럼명 → 표시 위치를 동적으로 찾기 (컬럼이 추가/이동되어도 깨지지 않도록)
    def _hidx(name, default=None):
        try:
            return all_headers.index(name) + 1
        except ValueError:
            return default

    IDX_GAP      = _hidx("괴리율(%)")        # 괴리율 색상 강조
    IDX_PT100    = _hidx("100점_합계")        # 100점 합계 색상 강조
    IDX_SIGNAL   = _hidx("매매시그널")        # 매매시그널 색상 강조
    IDX_BOX_SIG  = _hidx("박스이탈신호")      # 박스권 돌파 신호 색상 강조
    IDX_BOX_YN   = _hidx("박스권여부")        # 박스권 여부 색상 강조
    IDX_WRAP_FROM = _hidx("적정가_사유", 999)  # 이 컬럼부터 줄바꿈 표시

    for j, h in enumerate(all_headers, 1):
        c = ws_all.cell(row=2, column=j, value=h)
        c.font      = ws_all_hdr_font
        c.fill      = PatternFill("solid", fgColor=_all_hdr_color(j))
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = _THIN
    ws_all.row_dimensions[2].height = 28

    _fmt_by_name = {
        "현재가":"#,##0", "시가총액(억)":"#,##0", "시총":"#,##0", "발행주식수":"#,##0",
        "거래대금5일(억)":"#,##0.0", "거래대금20일(억)":"#,##0.0", "거래량비율VR(%)":"#,##0.0",
        "MACD선":"#,##0.0", "MACD시그널":"#,##0.0", "MACD히스토":"#,##0.0",
        "RSI14":"#,##0.0",
        "BB상단":"#,##0", "BB하단":"#,##0", "BB%B":"#,##0.0", "BB밴드폭(%)":"#,##0.0",
        "기술매수신호수":"#,##0", "기술매도신호수":"#,##0",
        # 박스권 (v41g 신규)
        "박스기간(일)":"#,##0", "박스상단":"#,##0", "박스하단":"#,##0",
        "박스폭(%)":"#,##0.0", "박스내위치(%)":"#,##0.0",
        # 밸류에이션
        "PBR(현재가)":"#,##0.00", "PBR(결산기준)":"#,##0.00", "PBR갭(%)":"#,##0.0",
        "ROE(%)":"#,##0.0", "ROA(%)":"#,##0.0",
        "PER(현재가)":"#,##0.0", "PER(결산기준)":"#,##0.0", "PER갭(%)":"#,##0.0",
        "DIV(%)":"#,##0.0", "배당성향(%)":"#,##0.0",
        "영업이익률(%)":"#,##0.0", "순이익률(%)":"#,##0.0",
        "부채비율(%)":"#,##0.0", "순부채비율(%)":"#,##0.0", "유동비율(%)":"#,##0.0",
        "Altman-Z":"#,##0.00",
        "FCF(억)":"#,##0", "FCF":"#,##0", "FCF마진(%)":"#,##0.0", "FCF Yield(%)":"#,##0.0",
        "EBITDA(억)":"#,##0", "EBITDA":"#,##0", "EV/EBITDA":"#,##0.0",
        "매출성장률(%)":"#,##0.0",
        "CFO(억)":"#,##0", "CFO":"#,##0", "이자보상배율":"#,##0.0",
        "52주위치(%)":"#,##0.0", "52주수익률(%)":"#,##0.0", "6개월수익률(%)":"#,##0.0",
        "가치점수":"#,##0.0", "모멘텀점수":"#,##0.0", "배당점수":"#,##0.0",
        "품질점수":"#,##0.0", "성장점수":"#,##0.0", "수익성점수":"#,##0.0", "안정점수":"#,##0.0",
        "복합점수":"#,##0.0",
        "적정가_PER":"#,##0", "적정가_PBR":"#,##0", "적정가_Graham":"#,##0", "적정가_평균":"#,##0",
        "괴리율(%)":"#,##0.0",
        "100점_합계":"#,##0.0",
        "100점_수익성(20)":"#,##0", "100점_성장성(20)":"#,##0", "100점_재무안정(20)":"#,##0",
        "100점_현금흐름(20)":"#,##0", "100점_밸류(10)":"#,##0", "100점_주주친화(10)":"#,##0",
        "100점_정성보정":"#,##0",
        "적극매수선":"#,##0", "매수선":"#,##0", "1차목표가":"#,##0", "2차목표가":"#,##0",
        "추격매수경보":"#,##0", "손절선":"#,##0",
    }
    _all_fmt = {idx: fmt for idx, fmt in
                ((_hidx(name), fmt) for name, fmt in _fmt_by_name.items())
                if idx is not None}

    for ri_all, (code_a, row_a) in enumerate(df_all.iterrows(), 1):
        sc_a  = _safe(row_a.get("복합점수"), 0)
        cur_a = _gv(row_a, "현재가","현재가_NV","현재가_YF")
        mc_a  = _marcap(row_a)
        fcf_a = _safe(row_a.get("fcf"))
        eb_a  = _safe(row_a.get("ebitda"))
        # 원 단위 → 억 단위 변환
        fcf_a_eok = _eok(fcf_a)
        eb_a_eok  = _eok(eb_a)
        cfo_a_eok = _eok(row_a.get("cfo"))
        # FCF Yield: FCF(원) / 시총(원) × 100
        fy_a = _r1(fcf_a / (mc_a * 1e8) * 100) if (fcf_a and mc_a and mc_a > 0) else None
        ev_a = None
        if eb_a and eb_a > 0 and mc_a:
            tl_a = _safe(row_a.get("total_liabilities")) or 0
            ca_a = _safe(row_a.get("cash")) or 0
            ev_a = _r1((mc_a * 1e8 + tl_a - ca_a) / eb_a)
        grd_a, grd_color_a = grade_label(sc_a)

        # ── PBR/PER 현재가 기준 재계산 (v28: 주식수 기반) ──
        # 우선순위: ① shares×현재가÷재무수치  ② 시총÷재무수치  ③ 수집값
        _shr_raw = row_a.get("shares")
        shr_a = _safe(_shr_raw) if (_shr_raw is not None and not (hasattr(_shr_raw, '__class__') and _shr_raw.__class__.__name__ == 'NAType')) else _safe(row_a.get("shares_dart"))
        if shr_a is not None and not (isinstance(shr_a, float) and np.isnan(shr_a)):
            shr_a = float(shr_a)
        else:
            shr_a = None
        equity_a = _safe(row_a.get("equity"))
        ni_a     = _safe(row_a.get("net_income"))

        mktcap_won_a = None
        if shr_a and shr_a > 0 and cur_a and cur_a > 0:
            mktcap_won_a = shr_a * cur_a                      # 주식수 × 현재가(원)
        elif mc_a and mc_a > 0:
            mktcap_won_a = mc_a * 1e8                         # 시총(억) → 원

        if mktcap_won_a and equity_a and equity_a > 0:
            pbr_na = round(mktcap_won_a / equity_a, 2)
        else:
            pbr_na = _r2(row_a.get("PBR") or row_a.get("pbr"))  # fallback

        if mktcap_won_a and ni_a and ni_a > 0:
            per_na = round(mktcap_won_a / ni_a, 1)
        else:
            per_na = _r1(row_a.get("PER") or row_a.get("per"))  # fallback

        pbr_ra = _r2(row_a.get("PBR_결산"))
        per_ra = _r1(row_a.get("PER_결산"))
        pbr_ga = _r1((pbr_na-pbr_ra)/abs(pbr_ra)*100) if (pbr_na and pbr_ra and pbr_ra!=0) else None
        per_ga = _r1((per_na-per_ra)/abs(per_ra)*100) if (per_na and per_ra and per_ra>0) else None
        pt100_a   = _safe(row_a.get("100점_합계"))
        sig_txt_a = str(row_a.get("매매시그널","") or "")
        sig_clr_a = str(row_a.get("_signal_color","FFCC00") or "FFCC00")

        all_vals = [
            ri_all,
            str(code_a),
            str(row_a.get("종목명","") or ""),
            str(row_a.get("시장","") or ""),
            str(row_a.get("업종","") or ""),
            _ri(cur_a), _ri(mc_a), _ri(shr_a),
            # 거래량
            _r1(row_a.get("거래대금5일(억)")),
            _r1(row_a.get("거래대금20일(억)")),
            _r1(row_a.get("거래량비율VR(%)")),
            # 기술지표 (v32)
            _r2(row_a.get("MACD선")),
            _r2(row_a.get("MACD시그널")),
            _r2(row_a.get("MACD히스토")),
            str(row_a.get("MACD신호","") or ""),
            _r1(row_a.get("RSI14")),
            str(row_a.get("RSI신호","") or ""),
            str(row_a.get("OBV추세","") or ""),
            str(row_a.get("OBV신호","") or ""),
            _ri(row_a.get("BB상단")),
            _ri(row_a.get("BB하단")),
            _r1(row_a.get("BB%B")),
            _r1(row_a.get("BB밴드폭(%)")),
            str(row_a.get("BB신호","") or ""),
            row_a.get("기술매수신호수"),
            row_a.get("기술매도신호수"),
            str(row_a.get("기술신호","") or ""),
            str(row_a.get("기술지표요약","") or ""),
            # 박스권 분석 (v41g 신규)
            str(row_a.get("박스권여부","N") or "N"),
            row_a.get("박스기간(일)"),
            _ri(row_a.get("박스상단")),
            _ri(row_a.get("박스하단")),
            _r1(row_a.get("박스폭(%)")),
            _r1(row_a.get("박스내위치(%)")),
            str(row_a.get("박스이탈신호","") or ""),
            # 밸류에이션
            pbr_na, pbr_ra, pbr_ga,
            _r1(row_a.get("ROE") or row_a.get("roe")),
            _r1(row_a.get("ROA")),
            per_na, per_ra, per_ga,
            _r1(row_a.get("DIV") or 0),
            _r1(row_a.get("dividend_payout")),
            _r1(row_a.get("영업이익률")),
            _r1(row_a.get("순이익률")),
            _r1(row_a.get("부채비율") or row_a.get("debt_ratio")),
            _r1(row_a.get("net_debt_ratio")),
            _r1(row_a.get("current_ratio")),
            _r2(row_a.get("altman_z")),
            fcf_a_eok, _r1(row_a.get("fcf_margin")), fy_a,
            eb_a_eok, ev_a,
            _r1(row_a.get("매출성장률")),
            cfo_a_eok,
            _r1(row_a.get("interest_coverage")),
            _r1(row_a.get("52주위치")),
            _r1(row_a.get("52주수익률")),
            _r1(row_a.get("6개월수익률")),
            _r1(row_a.get("가치점수")), _r1(row_a.get("모멘텀점수")),
            _r1(row_a.get("배당점수")), _r1(row_a.get("품질점수")),
            _r1(row_a.get("성장점수")), _r1(row_a.get("수익성점수")),
            _r1(row_a.get("안정점수")),
            _r1(sc_a),
            _ri(row_a.get("적정가_PER")), _ri(row_a.get("적정가_PBR")),
            _ri(row_a.get("적정가_Graham")), _ri(row_a.get("적정가_평균")),
            _r1(row_a.get("괴리율(%)")),
            str(row_a.get("적정가_사유","") or ""),
            str(grd_a),
            # 100점 평가
            _r1(pt100_a),
            str(row_a.get("100점_등급","") or ""),
            _safe(row_a.get("100점_수익성")), _safe(row_a.get("100점_성장성")),
            _safe(row_a.get("100점_재무안정")), _safe(row_a.get("100점_현금흐름")),
            _safe(row_a.get("100점_밸류")), _safe(row_a.get("100점_주주친화")),
            _safe(row_a.get("100점_정성보정")),
            str(row_a.get("체크_CFO","") or ""),
            str(row_a.get("체크_자본잠식","") or ""),
            str(row_a.get("체크_부채비율","") or ""),
            str(row_a.get("체크_ROE","") or ""),
            str(row_a.get("체크_PER","") or ""),
            str(row_a.get("보너스근거","") or ""),
            str(row_a.get("감점근거","") or ""),
            # 매매 시그널
            sig_txt_a,
            str(row_a.get("현재가위치","") or ""),
            _ri(row_a.get("적극매수선")),
            _ri(row_a.get("매수선")),
            _ri(row_a.get("1차목표가")),
            _ri(row_a.get("2차목표가")),
            _ri(row_a.get("추격매수경보")),
            _ri(row_a.get("손절선")),
            str(row_a.get("보유전략","") or ""),
            str(row_a.get("핵심투자포인트","") or ""),
            str(row_a.get("위험신호","없음") or "없음"),
        ]

        # 상위 N종목은 강조색, 나머지는 교번색
        if ri_all <= args.top:
            row_fill = _rank_fill(ri_all)
        elif ri_all % 2 == 0:
            row_fill = _ALT_FILL
        else:
            row_fill = PatternFill("solid", fgColor="FFFFFF")

        for j, v in enumerate(all_vals, 1):
            c = ws_all.cell(row=ri_all+2, column=j, value=v)
            c.font   = ws_all_body_font
            c.fill   = row_fill
            c.border = _THIN
            c.alignment = Alignment(horizontal="center", vertical="center",
                                    wrap_text=(j >= IDX_WRAP_FROM))
            fmt_a = _all_fmt.get(j)
            if fmt_a and v is not None:
                c.number_format = fmt_a
            # 괴리율 색상
            if j == IDX_GAP and v is not None:
                try:
                    fv = float(v)
                    c.font = Font(name=KR_FONT, size=9,
                                  color="0070C0" if fv >= 0 else "C00000",
                                  bold=(abs(fv) >= 20))
                except Exception: pass
            # 100점 합계 색상
            if j == IDX_PT100 and v is not None:
                try:
                    fv = float(v)
                    if fv >= 90:   c.fill = PatternFill("solid", fgColor="0070C0"); c.font = Font(name=KR_FONT, size=9, color="FFFFFF", bold=True)
                    elif fv >= 80: c.fill = PatternFill("solid", fgColor="2E75B6"); c.font = Font(name=KR_FONT, size=9, color="FFFFFF", bold=True)
                    elif fv >= 70: c.fill = PatternFill("solid", fgColor="92D050"); c.font = Font(name=KR_FONT, size=9, bold=True)
                    elif fv >= 60: c.fill = PatternFill("solid", fgColor="FFCC00"); c.font = Font(name=KR_FONT, size=9, bold=True)
                    else:          c.fill = PatternFill("solid", fgColor="FF9900"); c.font = Font(name=KR_FONT, size=9, bold=True)
                except Exception: pass
            # 매매시그널 색상
            if j == IDX_SIGNAL and v:
                try:
                    c.fill = PatternFill("solid", fgColor=sig_clr_a)
                    is_dark = sig_clr_a in ("C00000","0070C0","2E75B6","FF4444")
                    c.font = Font(name=KR_FONT, size=9, bold=True,
                                  color="FFFFFF" if is_dark else "000000")
                except Exception: pass
            # 박스권여부 색상 (v41g 신규) — 박스권 종목을 한눈에 구분
            if j == IDX_BOX_YN and v == "Y":
                try:
                    c.fill = PatternFill("solid", fgColor="FFF2CC")
                    c.font = Font(name=KR_FONT, size=9, bold=True, color="806000")
                except Exception: pass
            # 박스이탈신호 색상 (v41g 신규) — 뉴스성 모멘텀 후보는 강조 표시
            if j == IDX_BOX_SIG and v:
                try:
                    sv = str(v)
                    if "모멘텀후보" in sv:
                        c.fill = PatternFill("solid", fgColor="FF9900")
                        c.font = Font(name=KR_FONT, size=9, bold=True, color="FFFFFF")
                    elif "돌파대기" in sv:
                        c.fill = PatternFill("solid", fgColor="FFD966")
                        c.font = Font(name=KR_FONT, size=9, bold=True)
                    elif "이탈" in sv:
                        c.fill = PatternFill("solid", fgColor="FCE4D6")
                        c.font = Font(name=KR_FONT, size=9, color="C00000")
                except Exception: pass

    ws_all.freeze_panes = "C3"
    auto_col_width(ws_all)
    print(f"  📋 전체 종목 시트: {len(df_all)}개 종목 × {len(all_headers)}컬럼")

    # ── 시트5: 매매시그널 보드 ──
    ws_signal.title = "매매시그널보드"
    ws_signal_hdr_font  = Font(name=KR_FONT, bold=True, color="FFFFFF", size=11)
    ws_signal_body_font = Font(name=KR_FONT, size=10)

    # 제목
    ws_signal.merge_cells("A1:O1")
    title_cell = ws_signal.cell(row=1, column=1,
        value=f"📊 매매시그널 보드  [{label}]  기준일: {datetime.today().strftime('%Y-%m-%d')}")
    title_cell.font = Font(name=KR_FONT, bold=True, size=14, color="FFFFFF")
    title_cell.fill = PatternFill("solid", fgColor="1F3864")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws_signal.row_dimensions[1].height = 32

    sig_headers = [
        "순위","코드","종목명","현재가",
        "매매시그널","현재가위치",
        "100점","복합점수",
        "적극매수선","매수선","1차목표가","2차목표가","손절선",
        "보유전략","위험신호"
    ]
    SIG_FILLS = {
        "강력매수": PatternFill("solid", fgColor="0070C0"),
        "■■ 매수":  PatternFill("solid", fgColor="2E75B6"),
        "관심":     PatternFill("solid", fgColor="92D050"),
        "보유":     PatternFill("solid", fgColor="FFCC00"),
        "비중축소": PatternFill("solid", fgColor="FF9900"),
        "매도":     PatternFill("solid", fgColor="FF4444"),
        "즉시매도": PatternFill("solid", fgColor="C00000"),
    }

    for j, h in enumerate(sig_headers, 1):
        c = ws_signal.cell(row=2, column=j, value=h)
        c.font = ws_signal_hdr_font
        c.fill = PatternFill("solid", fgColor="2E75B6")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _THIN
    ws_signal.row_dimensions[2].height = 24

    for ri_s, (code_s, row_s) in enumerate(df_top.iterrows(), 1):
        sig_txt  = str(row_s.get("매매시그널","") or "")
        pos_txt  = str(row_s.get("현재가위치","") or "")
        cur_s    = _gv(row_s, "현재가","현재가_NV","현재가_YF")
        pt100_s  = _r1(row_s.get("100점_합계"))
        comp_s   = _r1(row_s.get("복합점수", 0))

        # 시그널별 배경색
        sig_fill = None
        for kw, fill in SIG_FILLS.items():
            if kw in sig_txt:
                sig_fill = fill
                break
        if sig_fill is None:
            sig_fill = PatternFill("solid", fgColor="F5F5F5")

        sig_row_vals = [
            ri_s,
            str(code_s),
            str(row_s.get("종목명","") or ""),
            _ri(cur_s),
            sig_txt,
            pos_txt,
            pt100_s,
            comp_s,
            _ri(row_s.get("적극매수선")),
            _ri(row_s.get("매수선")),
            _ri(row_s.get("1차목표가")),
            _ri(row_s.get("2차목표가")),
            _ri(row_s.get("손절선")),
            str(row_s.get("보유전략","") or ""),
            str(row_s.get("위험신호","없음") or "없음"),
        ]

        is_danger = "위험" in str(row_s.get("위험신호","")) or "즉시" in sig_txt

        for j, v in enumerate(sig_row_vals, 1):
            c = ws_signal.cell(row=ri_s+2, column=j, value=v)
            c.border = _THIN
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=(j >= 14))
            if j == 5:  # 시그널 컬럼 색상
                c.fill = sig_fill
                is_dark = any(kw in sig_txt for kw in ("강력매수","매도","즉시"))
                c.font = Font(name=KR_FONT, size=10, bold=True,
                              color="FFFFFF" if is_dark else "000000")
            elif is_danger and j == 15:
                c.fill = PatternFill("solid", fgColor="FFE7E7")
                c.font = Font(name=KR_FONT, size=10, color="C00000", bold=True)
            else:
                c.fill = _GRN_FILL if ri_s % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
                c.font = ws_signal_body_font
            if j in (4, 9, 10, 11, 12, 13) and v is not None:
                c.number_format = "#,##0"

    ws_signal.freeze_panes = "A3"
    auto_col_width(ws_signal)
    print(f"  📋 매매시그널 보드: {len(df_top)}개 종목")

    # ── 시트5: 종목별 주간 뉴스/이슈 (v30 신규) ──
    ws_news.title = "종목별주간이슈"
    ws_news.merge_cells("A1:F1")
    _nc = ws_news.cell(row=1, column=1,
        value=f"📰 종목별 주간 핵심 이슈  [{label}]  기준일: {datetime.today().strftime('%Y-%m-%d')}"
              f"  (Claude AI 웹검색 기반 자동 요약)")
    _nc.font  = Font(name=KR_FONT, bold=True, size=12, color="FFFFFF")
    _nc.fill  = PatternFill("solid", fgColor="1F3864")
    _nc.alignment = Alignment(horizontal="center", vertical="center")
    ws_news.row_dimensions[1].height = 26

    _news_hdrs = ["순위","코드","종목명","업종","복합점수","감성점수","주간 핵심 이슈 (Claude AI 자동 요약)"]
    _news_hdr_colors = ["1F3864","1F3864","1F3864","1F3864","2E75B6","2E75B6","7030A0"]
    for j, (h, col) in enumerate(zip(_news_hdrs, _news_hdr_colors), 1):
        c = ws_news.cell(row=2, column=j, value=h)
        c.font      = Font(name=KR_FONT, bold=True, color="FFFFFF", size=10)
        c.fill      = PatternFill("solid", fgColor=col)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = _THIN
    ws_news.row_dimensions[2].height = 24
    ws_news.column_dimensions["F"].width = 80  # 뉴스 컬럼 넓게

    for ri_n, (code_n, row_n) in enumerate(df_top.iterrows(), 1):
        score_n = _safe(row_n.get("복합점수"), 0)
        news_item = news_map.get(str(code_n), {"text": "뉴스 미수집", "sentiment": 0})
        # 구버전 호환: 혹시 str로 저장된 경우도 안전하게 처리
        if isinstance(news_item, dict):
            news_txt = news_item.get("text", "뉴스 미수집")
            sentiment = news_item.get("sentiment", 0)
        else:
            news_txt = str(news_item)
            sentiment = 0
        _alt = _GRN_FILL if ri_n % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
        row_news = [
            ri_n,
            str(code_n),
            str(row_n.get("종목명","") or ""),
            str(row_n.get("업종","") or ""),
            _r1(score_n),
            sentiment,
            news_txt,
        ]
        for j, v in enumerate(row_news, 1):
            c = ws_news.cell(row=ri_n+2, column=j, value=v)
            c.border    = _THIN
            c.fill      = _alt
            c.font      = Font(name=KR_FONT, size=9)
            # 감성점수 컬럼: 양수=파랑, 음수=빨강 강조
            if j == 6 and isinstance(v, (int, float)):
                c.font = Font(name=KR_FONT, size=9, bold=True,
                              color="0070C0" if v > 0 else ("C00000" if v < 0 else "808080"))
            c.alignment = Alignment(
                horizontal="center" if j < 6 else "left",
                vertical="top",
                wrap_text=(j == 6)
            )
        ws_news.row_dimensions[ri_n+2].height = max(
            30, min(15 * (len(news_txt)//60 + 1), 90)
        )

    ws_news.freeze_panes = "A3"
    for j in range(1, 6):
        auto_col_width(ws_news)
    print(f"  📰 종목별 주간이슈 시트: {len(news_map)}개 종목 수록")

    wb.save(fpath)
    print(f"\n  ✅ 엑셀 저장: {fpath}")
    return fpath

def save_signal_json(df_top, label, args, output_dir=BASE_DIR):
    try:
        stocks = []
        for rank, (code, row) in enumerate(df_top.iterrows(), 1):
            score = float(_safe(row.get("복합점수"), 0))
            cur   = _safe(row.get("현재가") or row.get("현재가_NV") or row.get("현재가_YF"), 0)
            pt100 = float(_safe(row.get("100점_합계"), 0))
            if score < 50 or not cur:
                continue
            stocks.append({
                "rank":         rank,
                "code":         str(code),
                "name":         str(row.get("종목명","")),
                "market":       str(row.get("시장","")),
                "sector":       str(row.get("업종","")),
                "price":        int(cur) if cur else 0,
                # ── 복합점수 (7팩터) ──
                "composite_score": round(score, 1),
                "grade":        str(row.get("등급","")),
                # ── 100점 종합 평가 ──
                "score_100": {
                    "total":        round(pt100, 1),
                    "grade":        str(row.get("100점_등급","")),
                    "수익성":       _safe(row.get("100점_수익성")),
                    "성장성":       _safe(row.get("100점_성장성")),
                    "재무안정":     _safe(row.get("100점_재무안정")),
                    "현금흐름":     _safe(row.get("100점_현금흐름")),
                    "밸류에이션":   _safe(row.get("100점_밸류")),
                    "주주친화":     _safe(row.get("100점_주주친화")),
                    "정성보정":     _safe(row.get("100점_정성보정")),
                },
                # ── 5대 필수 체크 ──
                "health_check": {
                    "cfo_positive":    str(row.get("체크_CFO","")),
                    "no_impairment":   str(row.get("체크_자본잠식","")),
                    "debt_safe":       str(row.get("체크_부채비율","")),
                    "roe_ok":          str(row.get("체크_ROE","")),
                    "per_ok":          str(row.get("체크_PER","")),
                },
                # ── 매매 시그널 ──
                "trading": {
                    "signal":          str(row.get("매매시그널","")),
                    "position_diag":   str(row.get("현재가위치","")),
                    "price_band": {
                        "strong_buy":  _safe(row.get("적극매수선")),
                        "buy":         _safe(row.get("매수선")),
                        "target_1":    _safe(row.get("1차목표가")),
                        "target_2":    _safe(row.get("2차목표가")),
                        "overbought":  _safe(row.get("추격매수경보")),
                        "stop_loss":   _safe(row.get("손절선")),
                    },
                    "strategy":        str(row.get("보유전략","")),
                    "key_points":      str(row.get("핵심투자포인트","")),
                    "risk_flags":      str(row.get("위험신호","없음")),
                    "bonus_reason":    str(row.get("보너스근거","")),
                    "malus_reason":    str(row.get("감점근거","")),
                },
                # ── 기본 재무 ──
                "fundamentals": {
                    "pbr":          _safe(row.get("PBR") or row.get("pbr")),
                    "roe":          _safe(row.get("ROE") or row.get("roe")),
                    "div":          _safe(row.get("DIV"), 0),
                    "debt_ratio":   _safe(row.get("부채비율") or row.get("debt_ratio")),
                    "interest_coverage": _safe(row.get("interest_coverage")),
                    "cfo":          _safe(row.get("cfo")),
                    "fcf":          _safe(row.get("fcf")),
                },
                # ── 밸류에이션 ──
                "valuation": {
                    "per_target":    _safe(row.get("적정가_PER")),
                    "pbr_target":    _safe(row.get("적정가_PBR")),
                    "graham_target": _safe(row.get("적정가_Graham")),
                    "avg_target":    _safe(row.get("적정가_평균")),
                    "upside_pct":    _safe(row.get("괴리율(%)")),
                },
            })

        data = {
            "scan_date":  datetime.today().strftime("%Y-%m-%d"),
            "scan_time":  datetime.now().strftime("%H:%M"),
            "screener":   f"QuantScreener_{VERSION}_{label}",
            "top_stocks": stocks,
            "filter_conditions": {
                "min_roe": args.min_roe, "max_pbr": args.max_pbr,
                "max_debt": args.max_debt, "min_ic": args.min_ic, "top_n": args.top,
            }
        }

        fname  = datetime.today().strftime(f"매매신호_{label}_%Y%m%d.json")
        fpath  = os.path.join(output_dir, fname)
        with open(fpath, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        print(f"  ✅ 매매신호 JSON: {fpath} ({len(stocks)}개)")
    except Exception as e:
        print(f"  ⚠ JSON 저장 실패: {e}")

# ══════════════════════════════════════════════════════════
# 11. 터미널 요약 출력
# ══════════════════════════════════════════════════════════

def print_summary(df_top, label, args):
    print()
    print("=" * 110)
    print(f"  ★ 퀀트 스크리닝 상위 {args.top}종목 [{label}]  {VERSION}")
    print(f"  기준: PBR≤{args.max_pbr} / ROE≥{args.min_roe}% / 부채비율<{args.max_debt}% / 이자보상>={args.min_ic}배")
    print("=" * 110)
    print(f"  {'순위':<4} {'코드':<8} {'종목명':<14} {'현재가':>8} {'PBR':>5} "
          f"{'ROE%':>7} {'DIV%':>6} {'복합':>6} {'100점':>5} {'시그널':<15} "
          f"{'1차목표가':>10} {'괴리율%':>8} {'손절선':>8}")
    print("  " + "-" * 106)

    for rank, (code, row) in enumerate(df_top.iterrows(), 1):
        name   = str(row.get("종목명",""))[:12]
        cur    = _safe(row.get("현재가") or row.get("현재가_NV") or row.get("현재가_YF"), 0)
        pbr    = _safe(row.get("PBR") or row.get("pbr"), 0)
        roe    = _safe(row.get("ROE") or row.get("roe"), 0)
        div    = _safe(row.get("DIV"), 0.0)
        score  = _safe(row.get("복합점수"), 0)
        pt100  = _safe(row.get("100점_합계"), 0)
        signal = str(row.get("매매시그널","─ 보유") or "─ 보유")[:14]
        tgt1   = _safe(row.get("1차목표가"))
        upside = _safe(row.get("괴리율(%)"))
        stop   = _safe(row.get("손절선"))
        try:
            cur_str  = f"{int(cur):,}"
        except Exception:
            cur_str  = "N/A"
        try:
            tgt_str  = f"{int(tgt1):,}" if tgt1 else "N/A"
        except Exception:
            tgt_str  = "N/A"
        try:
            stop_str = f"{int(stop):,}" if stop else "N/A"
        except Exception:
            stop_str = "N/A"
        upside_str = f"{upside:+.1f}%" if upside is not None else "N/A"
        print(f"  {rank:<4} {code:<8} {name:<14} {cur_str:>8} {pbr:>5.2f} "
              f"{roe:>6.1f}% {div:>5.2f}% {score:>6.1f} {pt100:>5.0f} {signal:<15} "
              f"{tgt_str:>10} {upside_str:>8} {stop_str:>8}")
    print()

# ══════════════════════════════════════════════════════════
# 12. Google Drive 업로드 + 메인 실행
#     (v33: GitHub Actions + Google Drive 자동화)
# ══════════════════════════════════════════════════════════


# ══════════════════════════════════════════════════════════
# 파일명 래퍼 — 항상 날짜+시분초로 저장 (중복 방지)
#
#  v35 build_excel / save_signal_json 는 내부적으로
#  %Y%m%d (날짜만) 파일명을 사용해 같은 날 실행하면 덮어씀.
#  v36 에서는 %Y%m%d_%H%M%S 형식을 강제해 항상 새 파일 생성.
#
#  파일명 예시:
#    quant_KR_20260618_091530.xlsx   ← 9시15분30초 실행분
#    quant_KR_20260618_143022.xlsx   ← 14시30분22초 실행분
#    매매신호_KR_20260618_091530.json
#    backtest_KR_20260618_091530.xlsx
# ══════════════════════════════════════════════════════════

def _ts() -> str:
    """현재 실행 타임스탬프 (날짜+시분초), 한 번 생성 후 모듈 변수로 고정"""
    return datetime.now().strftime("%Y%m%d_%H%M%S")

# 실행 시작 시각을 모듈 수준에서 한 번만 기록
# → 같은 실행에서 엑셀과 JSON 파일명이 동일한 타임스탬프를 가짐
_RUN_TS = datetime.now().strftime("%Y%m%d_%H%M%S")


def build_excel_v36(df_all, df_top, label, args, output_dir=None):
    """
    v35 build_excel을 호출한 뒤 파일명을 날짜+시분초로 변경.
    - v35 내부에서는 quant_{label}_%Y%m%d.xlsx 로 저장됨
    - 저장 직후 quant_{label}_%Y%m%d_%H%M%S.xlsx 로 rename
    """
    if not HAS_V35:
        return None
    out_dir = output_dir or BASE_DIR

    # v35 원본 저장 (날짜만 파일명)
    result_path = build_excel(df_all, df_top, label, args, out_dir)
    if not result_path or not os.path.exists(result_path):
        return result_path

    # 날짜만 파일명 → 날짜+시분초 파일명으로 rename
    date_only = datetime.now().strftime(f"quant_{label}_%Y%m%d.xlsx")
    new_name  = f"quant_{label}_{_RUN_TS}.xlsx"
    new_path  = os.path.join(out_dir, new_name)

    # 혹시 target 파일이 이미 있으면 삭제 후 rename
    if os.path.exists(new_path):
        os.remove(new_path)
    try:
        os.rename(result_path, new_path)
        print(f"  📁 파일명 확정: {new_name}")
    except Exception as e:
        print(f"  ⚠ 파일명 변경 실패({e}) → 원본 유지: {result_path}")
        return result_path
    return new_path


def save_signal_json_v36(df_top, label, args, output_dir=None):
    """
    v35 save_signal_json을 호출한 뒤 파일명을 날짜+시분초로 변경.
    - v35 내부: 매매신호_{label}_%Y%m%d.json
    - rename 후: 매매신호_{label}_%Y%m%d_%H%M%S.json
    """
    if not HAS_V35:
        return
    out_dir = output_dir or BASE_DIR

    save_signal_json(df_top, label, args, out_dir)

    old_name = datetime.now().strftime(f"매매신호_{label}_%Y%m%d.json")
    old_path = os.path.join(out_dir, old_name)
    new_name = f"매매신호_{label}_{_RUN_TS}.json"
    new_path = os.path.join(out_dir, new_name)

    if not os.path.exists(old_path):
        return
    if os.path.exists(new_path):
        os.remove(new_path)
    try:
        os.rename(old_path, new_path)
        print(f"  📁 파일명 확정: {new_name}")
    except Exception as e:
        print(f"  ⚠ 파일명 변경 실패({e}) → 원본 유지: {old_path}")


# ══════════════════════════════════════════════════════════
# 인수 파서 (한국 전용으로 단순화)
# ══════════════════════════════════════════════════════════
def parse_args():
    p = argparse.ArgumentParser(
        description=f"퀀트 주식 스크리너 {VERSION} — 한국 KOSPI/KOSDAQ 전용"
    )
    # 스크리닝 파라미터
    p.add_argument("--top",          type=int,   default=20,    help="상위 종목 수 (실전 확정값: 20)")
    p.add_argument("--min-roe",      type=float, default=8.0)
    p.add_argument("--max-pbr",      type=float, default=4.0)
    p.add_argument("--max-debt",     type=float, default=150.0)
    p.add_argument("--min-ic",       type=float, default=2.0)
    p.add_argument("--min-per",      type=float, default=0.0)
    p.add_argument("--max-per",      type=float, default=999.0)
    # 팩터 가중치
    p.add_argument("--w-value",      type=float, default=20.0)
    p.add_argument("--w-mom",        type=float, default=20.0)
    p.add_argument("--w-div",        type=float, default=5.0)
    p.add_argument("--w-quality",    type=float, default=15.0)
    p.add_argument("--w-growth",     type=float, default=15.0)
    p.add_argument("--w-prof",       type=float, default=15.0)
    p.add_argument("--w-stability",  type=float, default=10.0)
    # 실행 옵션
    p.add_argument("--output-dir",   type=str,   default=BASE_DIR)
    p.add_argument("--no-cache",     action="store_true", help="캐시 무시")
    p.add_argument("--auto",         action="store_true", help="비대화형 자동 실행")
    p.add_argument("--scope",        type=str,   default="2",
                   choices=["1","2","3"], help="1=100개 2=300개 3=전체")
    # v36 신규
    p.add_argument("--backtest",     action="store_true", help="백테스트 전용")
    p.add_argument("--trade",        action="store_true", help="KIS 자동매매 실행 (한국투자증권 REST API)")
    p.add_argument("--trade-only",   action="store_true",
                   help="스크리닝 없이 저장된 매매신호_KR_*.json만 읽어 KIS 주문 실행 (장중 별도 스케줄용)")
    p.add_argument("--monitor",      action="store_true", help="모니터링 전용")
    p.add_argument("--factor-model", action="store_true", help="Z-score 팩터모델 적용")
    p.add_argument("--no-cap-weight", dest="cap_weight", action="store_false", default=True,
                   help="스크리닝 시총가중 모멘텀 비활성 (기본은 활성 — 실전 확정값)")
    p.add_argument("--cap-strength", type=float, default=0.7,
                   help="스크리닝 시총가중 강도 0.0~1.0 (실전 확정값: 0.7)")
    p.add_argument("--no-pullback", dest="pullback", action="store_false", default=True,
                   help="스크리닝 눌림목 보너스 비활성 (기본은 활성 — 실전 확정값)")
    p.add_argument("--no-news-sentiment", dest="news_sentiment", action="store_false",
                   default=True,
                   help="뉴스 감성분석 반영 비활성 (기본은 활성 — Claude API 비용 발생)")
    p.add_argument("--no-telegram",  action="store_true", help="텔레그램 알림 비활성")
    p.add_argument("--bt-start",     type=str,   default="2020-01-01")
    p.add_argument("--bt-end",       type=str,   default="")
    p.add_argument("--bt-is",        type=int,   default=12, help="IS 개월수")
    p.add_argument("--bt-oos",       type=int,   default=3,  help="OOS 개월수")
    p.add_argument("--bt-fee-pct", type=float, default=0.015,
                   help="편도 거래수수료(%%, 기본 0.015%%)")
    p.add_argument("--bt-tax-pct", type=float, default=0.18,
                   help="매도 시 거래세(%%, 기본 0.18%% — 2025년 코스피/코스닥 공통)")
    p.add_argument("--bt-slippage-pct", type=float, default=0.20,
                   help="편도 슬리피지(%%, 기본 0.20%% — 체결가가 불리하게 밀리는 효과 근사)")
    p.add_argument("--bt-no-cost", action="store_true",
                   help="거래비용을 0으로 두고 실행 (예전 버전과 비교용, 실전 추정에는 쓰지 말 것)")
    p.add_argument("--bt-cap-weight", action="store_true",
                   help="백테스트 모멘텀에 시가총액 가중 적용 (대형주 쏠림장 대응)")
    p.add_argument("--bt-cap-strength", type=float, default=0.5,
                   help="시총가중 강도 0.0~1.0 (0.5=0.5~1.5배 완만, 0.7=0.3~1.7배 강함, 1.0=0~2배 최강)")
    p.add_argument("--bt-topn-compare", type=str, default="",
                   help="여러 top_n 비교 (예: '5,10,20') — 지정 시 --top 옵션은 무시됨")
    p.add_argument("--bt-basket", action="store_true",
                   help="포트폴리오(바스켓) 단위 시뮬레이션 — 실제 KIS 자동매매와 동일한 "
                        "합산수익률 손절/익절 판단 (run_walkforward 대신 사용)")
    p.add_argument("--bt-stop-loss", type=float, default=7.0,
                   help="바스켓 모드 손절 기준(%%) — KISAutoTrader 기본값과 동일")
    p.add_argument("--bt-tp-min", type=float, default=5.0,
                   help="바스켓 모드 익절 시작 기준(%%)")
    p.add_argument("--bt-tp-max", type=float, default=30.0,
                   help="바스켓 모드 익절 종료 기준(%%)")
    p.add_argument("--bt-fundamental-filter", action="store_true",
                   help="백테스트 진입에 펀더멘털 필터 적용 (PER≤30/ROE≥5%%/부채비율≤200%%) "
                        "— DART 재무수집이 추가되어 실행시간이 늘어남")
    p.add_argument("--bt-pullback-ab", action="store_true",
                   help="눌림목(Pull-back) 보너스 A/B 비교 — 동일 데이터로 ON/OFF만 다르게 "
                        "비교해 표본 차이 없이 순수 효과 측정")
    p.add_argument("--bt-gridsearch", action="store_true",
                   help="여러 (top_n, 손절%%, 익절범위) 조합을 한 번에 비교 — 바스켓 모드 전용")
    return p.parse_args()


# ══════════════════════════════════════════════════════════
# 실행 함수
# ══════════════════════════════════════════════════════════
def _run_backtest(args):
    print("\n" + "=" * 65)
    print(f"  📊 백테스트 모드 [{args.bt_start} ~ {args.bt_end or '오늘'}]")
    if getattr(args, "bt_cap_weight", False):
        s = getattr(args, "bt_cap_strength", 0.5)
        lo, hi = round(1 - s, 2), round(1 + s, 2)
        print(f"  ★ 시가총액 가중 모멘텀 적용 (강도 {s} → {lo}~{hi}배, 대형주 쏠림장 대응)")
    print("=" * 65)
    if not HAS_V35:
        print("  ⚠ quant_screener_v35.py 필요")
        return

    tickers, _, markets_list = fetch_kr_universe()

    # ── 코스피/코스닥 비율대로 섞어서 200개 샘플링 ──
    # (단순히 tickers[:200] 하면 KOSPI가 먼저 정렬돼 있어 KOSDAQ이 전혀 안 뽑힘)
    # fetch_kr_universe 내부에서 FDR이 시총순으로 정렬한 순서를 그대로 유지하므로
    # kospi_codes[:n] / kosdaq_codes[:n] 은 각 시장의 시총 상위 n개와 동일함
    kospi_codes  = [c for c, m in zip(tickers, markets_list) if "KOSDAQ" not in str(m).upper()]
    kosdaq_codes = [c for c, m in zip(tickers, markets_list) if "KOSDAQ" in str(m).upper()]

    sample_size = 200
    kospi_n  = min(len(kospi_codes),  round(sample_size * len(kospi_codes)
                                              / max(len(tickers), 1)))
    kosdaq_n = min(len(kosdaq_codes), sample_size - kospi_n)
    codes_200 = kospi_codes[:kospi_n] + kosdaq_codes[:kosdaq_n]

    print(f"  [백테스트] 표본 구성: 전체 {len(tickers)}개 중 KOSPI {kospi_n}개 + "
          f"KOSDAQ {kosdaq_n}개 = {len(codes_200)}개 (시총 상위 순)")

    markets_dict = dict(zip(tickers, markets_list))   # {code: "KOSPI"|"KOSDAQ"}

    # 시총가중 옵션 쓸 때만 DART 클라이언트 생성 (불필요한 키 로드 방지)
    dart_for_bt = None
    if getattr(args, "bt_cap_weight", False):
        dart_for_bt = DartClient(load_dart_key())

    bt = BacktestEngine(
        universe_codes=codes_200,
        start=args.bt_start,
        end=args.bt_end or None,
        is_months=args.bt_is,
        oos_months=args.bt_oos,
        top_n=args.top,
        markets=markets_dict,
        cap_weighted_momentum=getattr(args, "bt_cap_weight", False),
        cap_weight_strength=getattr(args, "bt_cap_strength", 0.5),
        dart_client=dart_for_bt,
        fee_pct=0.0 if getattr(args, "bt_no_cost", False) else getattr(args, "bt_fee_pct", 0.015),
        tax_pct=0.0 if getattr(args, "bt_no_cost", False) else getattr(args, "bt_tax_pct", 0.18),
        slippage_pct=0.0 if getattr(args, "bt_no_cost", False) else getattr(args, "bt_slippage_pct", 0.20),
    )
    price_df = bt.load_price_data()
    if price_df.empty:
        print("  ⚠ 가격 데이터 없음")
        return

    # ── 펀더멘털 필터용 재무데이터 수집 (선택적 — 시간이 오래 걸림) ──
    fundamentals_df = None
    if getattr(args, "bt_fundamental_filter", False):
        print("  [백테스트] 펀더멘털 필터용 재무데이터 수집 중 (PER·ROE·부채비율)...")
        try:
            dart_for_fund = dart_for_bt or DartClient(load_dart_key())
            names_200 = [c for c in codes_200]  # placeholder names, fetch_kr_all_data만 재무용으로 사용
            df_fund_raw = fetch_kr_all_data(
                dart_for_fund, codes_200, names_200,
                [markets_dict.get(c, "KOSPI") for c in codes_200],
                sample_size=None, no_cache=False
            )
            fundamentals_df = df_fund_raw[["PER", "ROE", "부채비율"]].copy() \
                if all(c in df_fund_raw.columns for c in ["PER", "ROE", "부채비율"]) else None
            if fundamentals_df is not None:
                print(f"  [백테스트] 펀더멘털 필터 적용: PER≤30, ROE≥5%, 부채비율≤200% "
                      f"({len(fundamentals_df)}개 종목 데이터 확보)")
            else:
                print("  ⚠ 펀더멘털 컬럼 없음 → 필터 미적용")
        except Exception as e:
            print(f"  ⚠ 펀더멘털 데이터 수집 실패({e}) → 필터 미적용")

    # BacktestEngine에 펀더멘털 데이터를 인스턴스 속성으로 저장해
    # compute_factor_scores에서 fundamentals 인자로 전달받을 수 있게 함
    bt._fundamentals_cache = fundamentals_df

    # ── 눌림목(Pull-back) A/B 비교 모드 ──
    if getattr(args, "bt_pullback_ab", False):
        ab_result = bt.run_pullback_ab_test(price_df)
        bt.save_pullback_ab_to_excel(ab_result)
        if not args.no_telegram:
            m = MonitorEngine()
            diff = ab_result.get("diff", {})
            m.send(
                f"📊 <b>눌림목 A/B 비교 [KR]</b>\n"
                f"CAGR 차이: {diff.get('CAGR(%)',0):+.2f}%p\n"
                f"Sharpe 차이: {diff.get('Sharpe_OOS',0):+.2f}\n"
                f"MDD 차이: {diff.get('MDD(%)',0):+.2f}%p"
            )
        return

    # ── 그리드서치 모드 (바스켓 전용) ──
    if getattr(args, "bt_gridsearch", False):
        grid_result = bt.run_grid_search(price_df)
        bt.save_grid_search_to_excel(grid_result)
        if not args.no_telegram and grid_result.get("results"):
            m = MonitorEngine()
            best = grid_result["results"][0]
            m.send(
                f"📊 <b>그리드서치 결과 [KR]</b>\n"
                f"🏆 최고: top{best['top_n']} / 손절-{best['손절(%)']}% / "
                f"익절{best['익절범위']}%\n"
                f"CAGR {best['CAGR(%)']:+.2f}% | MDD {best['MDD(%)']:.2f}%"
            )
        return

    # ── 바스켓(포트폴리오) 단위 시뮬레이션 모드 — 실제 KIS 자동매매와 동일 규칙 ──
    if getattr(args, "bt_basket", False):
        result = bt.run_basket_simulation(
            price_df,
            stop_loss_pct=getattr(args, "bt_stop_loss", 7.0),
            take_profit_min_pct=getattr(args, "bt_tp_min", 2.0),
            take_profit_max_pct=getattr(args, "bt_tp_max", 25.0),
        )
        if result:
            fpath = bt.save_basket_to_excel(result)
            if not args.no_telegram:
                m = MonitorEngine()
                s = result.get("summary", {})
                m.send(
                    f"📊 <b>바스켓 백테스트 결과 [KR]</b>\n"
                    f"CAGR: {s.get('CAGR(%)',0):+.1f}% | MDD: {s.get('MDD(%)',0):.1f}%\n"
                    f"매도 {s.get('매도건수',0)}건 (손절{s.get('손절건수',0)}/익절{s.get('익절건수',0)}) | "
                    f"승률: {s.get('승률(%)',0):.1f}%"
                )
        return

    # ── top_n 비교 모드 ──
    topn_compare = getattr(args, "bt_topn_compare", "")
    if topn_compare:
        try:
            top_n_list = [int(x.strip()) for x in topn_compare.split(",") if x.strip()]
        except ValueError:
            print(f"  ⚠ --bt-topn-compare 형식 오류: '{topn_compare}' (예: '5,10,20')")
            return

        comp_result = bt.run_topn_comparison(price_df, top_n_list)
        fpath = bt.save_comparison_to_excel(comp_result)

        if not args.no_telegram and comp_result.get("comparison"):
            m = MonitorEngine()
            lines = ["📊 <b>포트폴리오 집중도 비교 [KR]</b>"]
            for c in comp_result["comparison"]:
                lines.append(
                    f"top{c['top_n']}: CAGR {c['CAGR(%)']:+.1f}% | "
                    f"Sharpe {c['Sharpe_OOS']:.2f} | MDD {c['MDD(%)']:.1f}%"
                )
            m.send("\n".join(lines))
        return

    # ── 단일 top_n 실행 (기존 동작) ──
    result = bt.run_walkforward(price_df)
    if result:
        fpath = bt.save_to_excel(result)
        if not args.no_telegram:
            m = MonitorEngine()
            s = result.get("summary", {})
            m.send(
                f"📊 <b>백테스트 결과 [KR]</b>\n"
                f"CAGR: {s.get('CAGR(%)',0):+.1f}% | "
                f"Sharpe: {s.get('Sharpe_OOS',0):.2f}\n"
                f"MDD: {s.get('MDD(%)',0):.1f}% | "
                f"승률: {s.get('승률(%)',0):.1f}%\n"
                f"{s.get('과적합판단','')}"
            )


def _find_latest_signal_json(output_dir: str) -> Optional[str]:
    """output_dir 안에서 가장 최근(파일명 timestamp 기준) 매매신호_KR_*.json 찾기"""
    candidates = [
        f for f in os.listdir(output_dir)
        if f.startswith("매매신호_KR_") and f.endswith(".json")
    ]
    if not candidates:
        return None
    # 파일명에 박힌 타임스탬프(YYYYMMDD_HHMMSS) 기준 내림차순 정렬
    candidates.sort(reverse=True)
    return os.path.join(output_dir, candidates[0])


def _load_signal_json_as_df(json_path: str) -> pd.DataFrame:
    """
    save_signal_json_v36 으로 저장된 매매신호 JSON을 읽어
    KISAutoTrader.execute_signals 가 기대하는 컬럼명의 DataFrame으로 복원.
    인덱스 = 종목코드.
    """
    with open(json_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    rows = {}
    for s in data.get("top_stocks", []):
        code = str(s.get("code", ""))
        if not code:
            continue
        trading = s.get("trading", {})
        rows[code] = {
            "종목명":      s.get("name", ""),
            "매매시그널":  trading.get("signal", ""),
            "복합점수":    s.get("composite_score", 0),
            "100점_합계":  s.get("score_100", {}).get("total", 0),
            "괴리율(%)":   s.get("valuation", {}).get("upside_pct"),
        }
    return pd.DataFrame.from_dict(rows, orient="index")


def _run_trade_only(args):
    """
    스크리닝 없이, 가장 최근에 저장된 매매신호_KR_*.json만 읽어서
    KIS 자동매매(매수/손절/익절)만 실행.
    장중 시간에 별도 스케줄로 돌리기 위한 진입점.
    """
    print("\n  [매매전용] 저장된 매매신호로 KIS 자동매매 실행")

    json_path = _find_latest_signal_json(args.output_dir)
    if not json_path:
        print(f"  ⚠ {args.output_dir} 안에 매매신호_KR_*.json 파일이 없습니다.")
        print("     먼저 스크리닝(--auto 또는 일반 실행)을 돌려 신호 파일을 생성하세요.")
        sys.exit(1)

    print(f"  [매매전용] 신호 파일: {os.path.basename(json_path)}")
    df_top = _load_signal_json_as_df(json_path)
    if df_top.empty:
        print("  ⚠ 신호 파일에 유효한 종목이 없습니다.")
        return

    monitor = None if args.no_telegram else MonitorEngine()
    trader  = KISAutoTrader()

    bal       = trader.get_balance()
    total_cap = bal.get("순자산", 10_000_000)
    executed  = trader.execute_signals(df_top, total_cap)

    if monitor and executed:
        monitor.notify_trade(executed, reinvest_pool=trader._reinvest_pool)

    print(f"\n  ✅ 매매전용 실행 완료: {len(executed)}건")


def _run_monitor(args):
    print("\n  [모니터] 모니터링 전용 실행")
    monitor = MonitorEngine()
    trader  = KISAutoTrader()
    trader.sync_positions_from_broker()   # 실제 계좌 잔고 기준으로 보유종목 표시
    monitor.notify_positions(trader)
    monitor.send(f"ℹ {trader.status_summary()}")


def _apply_news_sentiment(df_top: pd.DataFrame, args) -> pd.DataFrame:
    """
    상위 종목에 대해 뉴스 감성점수를 수집하고 매매시그널에 반영.

    ── 반영 규칙 ──
    1) 뉴스감성점수 <= -60 (치명적 악재: 분식회계·횡령·상장폐지 등)
       → 매매시그널을 강제로 "🚫 악재감지 매수금지"로 변경 (기존 매수 시그널 무시)
    2) -60 < 뉴스감성점수 < 0 (경미한 악재)
       → 복합점수에서 소폭 차감 (최대 -5점)
    3) 뉴스감성점수 > 0 (호재)
       → 복합점수에 소폭 가산 (최대 +5점), 단 매수 여부 자체를 새로 만들지는 않음
       (이미 매수 시그널이 아니었던 종목을 호재만으로 매수로 바꾸지 않음 — 안전상 비대칭 설계)

    --news-sentiment 옵션이 꺼져있으면 수집을 건너뛰고 원본 그대로 반환.
    """
    if not getattr(args, "news_sentiment", True):
        return df_top
    if df_top.empty:
        return df_top

    try:
        news_map = fetch_news_batch(df_top, max_stocks=min(len(df_top), 20))
    except Exception as e:
        print(f"  [뉴스감성] 수집 실패 → 감성 미반영: {e}")
        return df_top

    df_top = df_top.copy()
    df_top["뉴스감성점수"] = 0
    df_top["뉴스요약"] = ""

    blocked = []
    for code in df_top.index:
        item = news_map.get(str(code), {"text": "", "sentiment": 0})
        sentiment = item.get("sentiment", 0) if isinstance(item, dict) else 0
        text = item.get("text", "") if isinstance(item, dict) else str(item)
        df_top.at[code, "뉴스감성점수"] = sentiment
        df_top.at[code, "뉴스요약"] = text

        score_col = "강화복합점수" if "강화복합점수" in df_top.columns else "복합점수"
        if sentiment <= -60:
            # 치명적 악재 → 매수 시그널 강제 차단
            if "매매시그널" in df_top.columns:
                df_top.at[code, "매매시그널"] = "🚫 악재감지(거래중단)"
            blocked.append((code, sentiment))
        elif sentiment < 0:
            df_top.at[code, score_col] = max(0, df_top.at[code, score_col] + sentiment * 0.083)  # -60→-5점
        elif sentiment > 0:
            df_top.at[code, score_col] = df_top.at[code, score_col] + min(sentiment, 60) * 0.083  # +60→+5점

    if blocked:
        print(f"  [뉴스감성] ⚠ 치명적 악재 감지로 매수금지 처리: "
              f"{', '.join(f'{c}({s})' for c, s in blocked)}")
    else:
        print(f"  [뉴스감성] 반영 완료 ({len(df_top)}종목, 치명적 악재 없음)")

    return df_top


def _run_scan(args, scope: str):
    """한국 KOSPI/KOSDAQ 스크리닝 + 신규 엔진 통합"""
    if not HAS_V35:
        print("  ⚠ quant_screener_v35.py 없음")
        return

    kr_size = {"1": 100, "2": 300, "3": None}[scope]

    monitor = None if args.no_telegram else MonitorEngine()
    trader  = KISAutoTrader() if args.trade else None
    # 실전 확정: top20 + 시총가중 강도 0.7 + 눌림목 보너스 (--no-cap-weight/--no-pullback 로 끌 수 있음)
    factor  = FactorModel(sector_neutral=True,
                          cap_weighted_momentum=getattr(args, "cap_weight", True),
                          cap_weight_strength=getattr(args, "cap_strength", 0.7),
                          enable_pullback=getattr(args, "pullback", True)) if args.factor_model else None

    print("\n" + "─" * 55)
    print("  [한국 KOSPI/KOSDAQ 스크리닝]")
    print("─" * 55)

    # v35 데이터 수집
    dart_key = load_dart_key()
    dart     = DartClient(dart_key)
    tickers, names, markets = fetch_kr_universe()
    df_raw = fetch_kr_all_data(
        dart, tickers, names, markets, kr_size,
        no_cache=getattr(args, "no_cache", False)
    )
    df_raw = compute_scores(df_raw, args)

    # Z-score 팩터 모델 적용
    if factor:
        print("  [팩터모델] Z-score 섹터중립화 적용 중...")
        df_raw = factor.enhance_scores(df_raw)

    df_filtered = apply_hard_filter(df_raw, args)
    if df_filtered.empty:
        print("  ⚠ 필터 통과 종목 없음")
        return

    # 강화복합점수 있으면 재정렬
    sort_col = "강화복합점수" if "강화복합점수" in df_filtered.columns else "복합점수"
    df_filtered = df_filtered.sort_values(sort_col, ascending=False)
    df_top      = df_filtered.head(args.top)

    print(f"  📊 통과: {len(df_filtered)}개 / 전체 {len(df_raw)}개 → 상위 {args.top}개")
    print_summary(df_top, "한국 KOSPI/KOSDAQ", args)

    # ── 뉴스·이슈 감성 반영 (매수/매도 판단 전에 미리 수집) ──
    # 강한 악재(감성점수 <= -60)가 감지된 종목은 매매시그널에서 즉시 차단
    df_top = _apply_news_sentiment(df_top, args)

    # 엑셀·JSON 저장 (날짜+시분초 파일명 — 실행마다 새 파일 생성)
    build_excel_v36(df_raw, df_top, "KR", args, args.output_dir)
    save_signal_json_v36(df_top, "KR", args, args.output_dir)

    # 텔레그램
    if monitor:
        df_prev = monitor.load_yesterday()
        monitor.notify_signal_change(df_top, df_prev)
        monitor.notify_screening(df_top)
        monitor.track(df_top)

    # KIS 자동매매
    if trader:
        print("\n  [KIS] 자동매매 실행...")
        bal       = trader.get_balance()
        total_cap = bal.get("순자산", 10_000_000)
        executed  = trader.execute_signals(df_top, total_cap)
        if monitor and executed:
            monitor.notify_trade(executed, reinvest_pool=trader._reinvest_pool)

    print("\n  ✅ 완료")
    print("  📌 신규 기능:")
    print(f"     팩터모델:    {'✅' if factor else '─ (--factor-model)'}")
    print(f"     KIS 자동매매: {'✅' if trader else '─ (--trade / kis_config.json 설정)'}")
    print(f"     텔레그램:    {'✅' if (monitor and monitor.enabled) else '─ (telegram_config.json 설정)'}")


# ══════════════════════════════════════════════════════════
# 대화형 메뉴
# ══════════════════════════════════════════════════════════
def interactive_menu():
    print()
    print("  ┌──────────────────────────────────────────────────────────┐")
    print(f"  │  퀀트 주식 스크리너 {VERSION}  (한국 KOSPI/KOSDAQ 전용)  │")
    print("  │  단일파일 통합 + 백테스트 + Z팩터 + KIS자동매매 + 텔레그램  │")
    print("  └──────────────────────────────────────────────────────────┘")
    print()
    cache_show_status()
    cache_clear(days_old=8)
    print()

    print("  [실행 모드]")
    print("  1. 스크리닝  (v35 전체 + 팩터모델 강화)")
    print("  2. 스크리닝 + KIS 자동매매  (한국투자증권 REST API, 64bit)")
    print("  3. 백테스트  (Walk-forward IS/OOS)")
    print("  4. 모니터링  (텔레그램 포지션 현황)")
    print("  5. 설정 안내  (KIS API / 텔레그램)")
    while True:
        mode = input("\n  선택 (1~5): ").strip()
        if mode in ("1","2","3","4","5"):
            break
        print("  1~5 중 하나를 입력하세요.")

    if mode == "5":
        print()
        print("  ── KIS API 설정 (한국투자증권, 64bit Python 지원) ──")
        print(f"  → 설정파일: {os.path.join(BASE_DIR, 'kis_config.json')}")
        print("  → https://apiportal.koreainvestment.com 에서 앱키 무료 발급")
        print("  → app_key / app_secret / account_no 입력 후 저장")
        print("  → is_real: false (모의) → true (실전) 로 변경하면 실전 주문")
        print()
        print("  ── 손절·익절·재투자 설정 (포트폴리오 바스켓 단위) ──")
        print("  → 판단 단위: 종목 개별이 아니라 보유종목 전체 합산 수익률")
        print("  → stop_loss_pct: 손절 기준(%) — 기본 7.0 (합산수익률 이하면 전종목 동시매도)")
        print("  → take_profit_min_pct ~ take_profit_max_pct: 익절 범위(%) — 기본 2.0~25.0 (실전 확정값)")
        print("     (합산수익률이 이 범위에 들면 전종목 동시매도, 범위 밖이면 전부 보유유지)")
        print("  → base_invest_amount: 최초 매수 예산(원) — 기본 10,000,000")
        print("  → reinvest_mode: true면 바스켓 매도 회수금을 다음 매수 예산으로 누적 재투자")
        print("     예) 5종목×200만원 → 합산+10%매도 → 회수1,100만원 → 다음 4종목×275만원")
        print("     (trades/reinvest_pool.json 에 저장되어 다음 실행에서 이어서 사용)")
        print()
        print("  ── 텔레그램 설정 ──")
        print(f"  → 설정파일: {os.path.join(BASE_DIR, 'telegram_config.json')}")
        print("  → @BotFather → /newbot → 토큰 발급")
        print("  → https://api.telegram.org/bot<TOKEN>/getUpdates → chat_id 확인")
        sys.exit(0)

    if mode == "3":
        print()
        start_input = input("  백테스트 시작일 (YYYY-MM-DD, 엔터=기본 2020-01-01): ").strip()
        bt_start = start_input if start_input else "2020-01-01"
        print(f"    ※ 참고: IS(학습기간) 12개월이 지나야 검증 결과가 나옵니다.")
        print(f"      → 시작일 {bt_start} 기준, 실제 매매 결과는 "
              f"{(pd.Timestamp(bt_start) + pd.DateOffset(months=12)).date()}부터 표시됩니다.")

        print()
        print("    ※ 합성 진입설계 (실익 보강, 항상 자동 적용됨):")
        print("      ① 이동평균 정배열(20>60>120일선) 아니면 진입점수 50% 감점")
        print("      ② 시장레짐 필터: KOSPI가 200일선 아래(위험장)면 매수종목수 절반 축소")
        fund_yn = input("    ③ 펀더멘털 필터도 적용? (PER≤30/ROE≥5%/부채비율≤200%, "
                       "재무수집 추가로 시간 더 걸림, y/n, 기본 n): ").strip().lower()

        print()
        print("    백테스트 방식 선택:")
        print("      1. 분기 리밸런싱 (기존) — 3개월마다 강제로 top_n 종목 교체")
        print("      2. 포트폴리오 바스켓 (신규) — 실제 KIS 자동매매와 동일하게,")
        print("         보유종목 전체 합산 수익률이 손절/익절 범위에 들 때만 매도")
        print("      3. 눌림목 A/B 비교 — 동일 데이터로 눌림목 ON/OFF만 비교 (정확한 효과측정)")
        print("      4. 그리드서치 — top_n×손절×익절 여러 조합을 한 번에 비교")
        method_choice = input("    선택 (1~4, 엔터=1): ").strip()

        if method_choice == "3":
            return "backtest", {"bt_start": bt_start, "pullback_ab": True,
                                "fundamental_filter": fund_yn == "y"}

        if method_choice == "4":
            return "backtest", {"bt_start": bt_start, "gridsearch": True,
                                "fundamental_filter": fund_yn == "y"}

        if method_choice == "2":
            stop_in = input("    손절 기준(%, 양수로 입력 — 예: 7) (엔터=7.0): ").strip()
            tpmin_in = input("    익절 시작 기준(%) (엔터=5.0): ").strip()
            tpmax_in = input("    익절 종료 기준(%) (엔터=30.0): ").strip()
            # 음수로 입력해도(예: -3) 절댓값으로 안전하게 변환 — 표시 시 중복 마이너스 방지
            stop_loss_val = abs(float(stop_in)) if stop_in else 7.0
            return "backtest", {
                "bt_start": bt_start,
                "basket_mode": True,
                "stop_loss": stop_loss_val,
                "tp_min": float(tpmin_in) if tpmin_in else 5.0,
                "tp_max": float(tpmax_in) if tpmax_in else 30.0,
                "fundamental_filter": fund_yn == "y",
            }

        cap_yn = input("  시가총액 가중 모멘텀 적용? (대형주 쏠림장 대응, y/n, 기본 n): ").strip().lower()
        cap_strength = 0.5
        if cap_yn == "y":
            print("    가중 강도 선택:")
            print("      1. 완만 (0.5~1.5배) — 기본값")
            print("      2. 강함 (0.3~1.7배)")
            print("      3. 최강 (0.0~2.0배) — 소형주 영향력 거의 제거")
            strength_choice = input("    선택 (1~3, 엔터=1): ").strip()
            cap_strength = {"1": 0.5, "2": 0.7, "3": 1.0}.get(strength_choice, 0.5)
        compare_yn = input("  포트폴리오 집중도 비교 실행? (top5/10/20 동시비교, y/n, 기본 n): ").strip().lower()
        topn_str = ""
        if compare_yn == "y":
            topn_str = input("  비교할 top_n 목록 (예: 5,10,20 / 엔터=기본값): ").strip() or "5,10,20"
        return "backtest", {"cap_weight": cap_yn == "y", "cap_strength": cap_strength,
                            "topn_compare": topn_str, "bt_start": bt_start,
                            "fundamental_filter": fund_yn == "y"}
    if mode == "4":
        return "monitor", None

    print()
    print("  [스크리닝 범위]")
    print("  1. 빠른 테스트  (KOSPI+KOSDAQ 100개,  약 10~15분)")
    print("  2. 중간 범위    (KOSPI+KOSDAQ 300개,  약 30~40분)")
    print("  3. 전체 스크리닝 (전 종목,              수 시간)")
    while True:
        scope = input("\n  선택 (1/2/3): ").strip()
        if scope in ("1","2","3"):
            break

    factor_yn = input("\n  Z-score 팩터모델 적용? (y/n, 기본 n): ").strip().lower()
    news_yn   = input("  뉴스·이슈 감성분석 반영? (Claude API 비용발생, y/n, 기본 y): ").strip().lower()
    trade_yn  = ("y" if mode == "2"
                 else input("  KIS 자동매매 실행? (y/n, 기본 n): ").strip().lower())

    return scope, {"factor_model": factor_yn == "y", "trade": trade_yn == "y",
                   "news_sentiment": news_yn != "n"}


# ══════════════════════════════════════════════════════════
# 메인
# ══════════════════════════════════════════════════════════
def main():
    print()
    print("=" * 65)
    print(f"  퀀트 스크리너 {VERSION}  —  한국 KOSPI/KOSDAQ 전용")
    print("  단일파일 통합 + 백테스트 + Z팩터 + KIS자동매매(64bit) + 텔레그램")
    print("=" * 65)

    args = parse_args()

    # 단독 모드
    if args.backtest:
        _run_backtest(args)
        return
    if args.monitor:
        _run_monitor(args)
        return
    if args.trade_only:
        _run_trade_only(args)
        return

    if args.auto:
        # 비대화형 자동 실행
        scope = args.scope
        print(f"\n  🤖 자동 실행  [{datetime.now().strftime('%Y-%m-%d %H:%M')}]")
        cache_show_status()
        cache_clear(days_old=8)
    else:
        # 대화형 메뉴
        result = interactive_menu()
        if result[0] == "backtest":
            bt_opts = result[1] or {}
            args.bt_cap_weight   = bt_opts.get("cap_weight", False)
            args.bt_cap_strength = bt_opts.get("cap_strength", 0.5)
            args.bt_topn_compare = bt_opts.get("topn_compare", "")
            args.bt_start        = bt_opts.get("bt_start", args.bt_start)
            args.bt_basket       = bt_opts.get("basket_mode", False)
            args.bt_stop_loss    = bt_opts.get("stop_loss", 7.0)
            args.bt_tp_min       = bt_opts.get("tp_min", 5.0)
            args.bt_tp_max       = bt_opts.get("tp_max", 30.0)
            args.bt_fundamental_filter = bt_opts.get("fundamental_filter", False)
            args.bt_pullback_ab = bt_opts.get("pullback_ab", False)
            args.bt_gridsearch  = bt_opts.get("gridsearch", False)
            _run_backtest(args)
            return
        if result[0] == "monitor":
            _run_monitor(args)
            return

        scope, opts = result
        opts = opts or {}
        if opts.get("factor_model"):
            args.factor_model = True
        if opts.get("trade"):
            args.trade = True
        if "news_sentiment" in opts:
            args.news_sentiment = opts["news_sentiment"]

    _run_scan(args, scope)

    # ── 결과 파일이 실제로 생성됐는지 최종 확인 ──
    # (필터 통과 0개 등으로 build_excel이 호출되지 않은 채 끝나는 경우 감지)
    today_pattern = datetime.now().strftime("%Y%m%d")
    produced = [f for f in os.listdir(args.output_dir)
                if f.startswith("quant_KR_") and today_pattern in f and f.endswith(".xlsx")]
    if not produced:
        print()
        print("  ⚠ 경고: 오늘자 quant_KR_*.xlsx 파일이 생성되지 않았습니다.")
        print("     스크리닝 필터를 통과한 종목이 없었거나, 도중에 중단되었을 수 있습니다.")
        print("     위 로그를 확인하세요.")
        sys.exit(1)


if __name__ == "__main__":
    main()
