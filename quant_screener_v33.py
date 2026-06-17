"""
==============================================================
  퀀트 주식 스크리너 v4.0 (통합 최종판)
  
  ★ 한국: DART API(재무 원본) + 네이버 금융(배당/업종/PER/PBR)
           + yfinance 배치 다운로드(52주 가격/모멘텀)
  ★ 미국: yfinance (재무 + 가격)
  
  ▶ 7팩터 통합 점수 (한국) / 4팩터 (미국)
    가치20% + 모멘텀20% + 배당5% + 품질15% + 성장15% + 수익성15% + 안정10%
  
  ▶ 병렬처리 (ThreadPoolExecutor) - 속도 3~4배 향상
  ▶ 하드 필터: CFO>0, 부채비율<150%, 이자보상배율>2
  ▶ 스케줄러 모드 (매주 금요일 자동 실행)
  ▶ JSON 매매신호 파일 자동 생성
  ▶ 엑셀 3시트 출력 (스크리닝결과 / 상위종목상세 / 지표가이드)
  
  필요 패키지:
    pip install finance-datareader yfinance requests beautifulsoup4
                pandas numpy openpyxl tqdm dart-fss
==============================================================
"""

import os, sys, re, time, json, zipfile, io, warnings, argparse, traceback, hashlib
from datetime import datetime, timedelta
from concurrent.futures import ThreadPoolExecutor, as_completed

# ── 저장 폴더: 스크립트 파일이 있는 폴더로 자동 설정 ──
BASE_DIR  = os.path.dirname(os.path.abspath(__file__))
CACHE_DIR = os.path.join(BASE_DIR, ".cache")
os.makedirs(CACHE_DIR, exist_ok=True)

import requests
import pandas as pd
import numpy as np

try:
    from bs4 import BeautifulSoup
    HAS_BS4 = True
except ImportError:
    HAS_BS4 = False

try:
    import yfinance as yf
    HAS_YFINANCE = True
except ImportError:
    HAS_YFINANCE = False

try:
    import FinanceDataReader as fdr
    HAS_FDR = True
except ImportError:
    HAS_FDR = False

try:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side, GradientFill
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.merge import MergedCell
    from openpyxl.formatting.rule import ColorScaleRule
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    from tqdm import tqdm
    HAS_TQDM = True
except ImportError:
    HAS_TQDM = False

warnings.filterwarnings("ignore")

# ══════════════════════════════════════════════════════════
# 패키지 상태 사전 안내
# ══════════════════════════════════════════════════════════
def _print_pkg_status():
    missing = []
    if not HAS_BS4:      missing.append("beautifulsoup4  (네이버 스크래핑)")
    if not HAS_YFINANCE: missing.append("yfinance        (가격/모멘텀)")
    if not HAS_FDR:      missing.append("finance-datareader (유니버스 수집 1순위)")
    if not HAS_OPENPYXL: missing.append("openpyxl        (엑셀 저장)")
    if not HAS_TQDM:     missing.append("tqdm            (진행률 바)")
    if missing:
        print("\n  ⚠ 미설치 패키지 감지 → 해당 기능은 중립값/폴백으로 대체됩니다.")
        print("    pip install " + " ".join([m.split()[0] for m in missing]))
        for m in missing:
            print(f"    - {m}")
        print()

# ══════════════════════════════════════════════════════════
# 0. 상수 / 스타일 / 설정
# ══════════════════════════════════════════════════════════
VERSION = "v5.2.0"  # v33: 3계층 캐시(DART 7일/네이버 1일/현재가 매일갱신) + Windows 작업스케줄러 자동실행 + 고정파일명 엑셀 덮어쓰기

_HDR_FILL  = PatternFill("solid", fgColor="1F3864") if HAS_OPENPYXL else None
_SUB_FILL  = PatternFill("solid", fgColor="2E75B6") if HAS_OPENPYXL else None
_GLD_FILL  = PatternFill("solid", fgColor="FFD700") if HAS_OPENPYXL else None
_SLV_FILL  = PatternFill("solid", fgColor="C0C0C0") if HAS_OPENPYXL else None
_BRZ_FILL  = PatternFill("solid", fgColor="CD7F32") if HAS_OPENPYXL else None
_GRN_FILL  = PatternFill("solid", fgColor="E2EFDA") if HAS_OPENPYXL else None
_ALT_FILL  = PatternFill("solid", fgColor="F5F5F5") if HAS_OPENPYXL else None
_RED_FILL  = PatternFill("solid", fgColor="FFE7E7") if HAS_OPENPYXL else None
_YEL_FILL  = PatternFill("solid", fgColor="FFF2CC") if HAS_OPENPYXL else None
_THIN = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
) if HAS_OPENPYXL else None

NAVER_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/120.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "ko-KR,ko;q=0.9",
    "Referer": "https://finance.naver.com",
}

# DART 계정과목 코드 매핑
# account_id (IFRS/DART 표준) + account_nm (한국어 계정명) 모두 지원
DART_ACCOUNTS = {
    "영업활동현금흐름": [
        "ifrs-full_CashFlowsFromUsedInOperatingActivities",
        "dart_CashFlowsFromOperatingActivities",
        "영업활동으로인한현금흐름", "영업활동현금흐름", "영업활동으로인현금흐름",
    ],
    "투자활동현금흐름": [
        "ifrs-full_CashFlowsFromUsedInInvestingActivities",
        "dart_CashFlowsFromInvestingActivities",
        "투자활동으로인한현금흐름", "투자활동현금흐름", "투자활동으로인현금흐름",
    ],
    "당기순이익": [
        "ifrs-full_ProfitLoss", "dart_ProfitLoss",
        "당기순이익", "당기순손익", "연결당기순이익", "당기순이익(손실)",
    ],
    "매출총이익": [
        "ifrs-full_GrossProfit", "dart_GrossProfit",
        "매출총이익", "매출총손익",
    ],
    "영업이익": [
        "ifrs-full_ProfitLossFromOperatingActivities",
        "dart_OperatingIncomeLoss",
        "영업이익", "영업손익", "영업이익(손실)",
    ],
    "매출액": [
        "ifrs-full_Revenue", "dart_Revenue",
        "매출액", "수익(매출액)", "영업수익", "매출",
    ],
    "이자비용": [
        "ifrs-full_FinanceCosts", "dart_FinanceCosts",
        "이자비용", "금융비용", "이자비용합계",
    ],
    "감가상각비": [
        "ifrs-full_DepreciationAndAmortisationExpense",
        "dart_DepreciationAndAmortisation",
        "감가상각비", "감가상각비및상각비", "유형자산감가상각비",
    ],
    "총자산": [
        "ifrs-full_Assets", "dart_Assets",
        "자산총계", "총자산",
    ],
    "총부채": [
        "ifrs-full_Liabilities", "dart_Liabilities",
        "부채총계", "총부채",
    ],
    "자기자본": [
        "ifrs-full_Equity", "dart_Equity",
        "자본총계", "자기자본", "지배기업소유주지분",
    ],
    "현금및현금성자산": [
        "ifrs-full_CashAndCashEquivalents",
        "dart_CashAndCashEquivalents",
        "현금및현금성자산", "현금및현금등가물",
    ],
    "유동자산": [
        "ifrs-full_CurrentAssets", "dart_CurrentAssets",
        "유동자산", "유동자산합계",
    ],
    "유동부채": [
        "ifrs-full_CurrentLiabilities", "dart_CurrentLiabilities",
        "유동부채", "유동부채합계",
    ],
    "배당금지급": [
        "ifrs-full_DividendsPaid", "dart_DividendsPaid",
        "배당금지급", "배당금의지급", "현금배당금지급",
        "배당금지급(현금배당)", "주주에대한배당금지급",
    ],
    # ── v30: 주당 지표 (PBR/PER 현재가 기준 계산에 직접 사용) ──
    "주당순이익": [
        "ifrs-full_BasicEarningsLossPerShare",
        "ifrs-full_EarningsPerShare",
        "dart_EarningsPerShare",
        "기본주당이익(손실)", "기본주당순이익", "주당순이익(기본)",
        "주당순이익", "기본EPS",
    ],
    "주당순자산": [
        "dart_BookValuePerShare",
        "ifrs-full_NetAssetsPerShare",
        "주당순자산", "주당장부금액", "주당순자산가치",
        "주당자산가치", "기본BPS",
    ],
}

# 계정명 → 표준키 역방향 매핑 (빠른 검색용)
_ACNT_NM_MAP = {}
for _k, _ids in DART_ACCOUNTS.items():
    for _id in _ids:
        _ACNT_NM_MAP[_id] = _k
        _ACNT_NM_MAP[_id.replace(" ", "")] = _k

# ══════════════════════════════════════════════════════════
# 1. 공통 유틸
# ══════════════════════════════════════════════════════════

def _safe(v, default=None):
    try:
        if v is None or (isinstance(v, float) and np.isnan(v)):
            return default
        return v
    except Exception:
        return default

def _fmt(v, digits=2, suffix=""):
    try:
        if v is None or (isinstance(v, float) and np.isnan(v)):
            return "N/A"
        return f"{v:.{digits}f}{suffix}"
    except Exception:
        return "N/A"

def safe_div(a, b, default=np.nan):
    try:
        if b == 0 or pd.isna(b) or pd.isna(a):
            return default
        return a / b
    except Exception:
        return default

def _clamp(v, lo, hi):
    """네이버 파싱값 이상치 필터: lo~hi 범위 밖이면 None 반환"""
    if v is None:
        return None
    try:
        fv = float(v)
        if np.isnan(fv):
            return None
        return fv if lo <= fv <= hi else None
    except Exception:
        return None

def percentile_rank(series: pd.Series) -> pd.Series:
    return series.rank(pct=True) * 100

def grade_label(score):
    if score >= 85:   return "◆◆◆ A+ 탁월", "FF4444"
    elif score >= 75: return "◆◆  A  우수",  "FF9900"
    elif score >= 65: return "◆   B  양호",  "FFCC00"
    elif score >= 50: return "    C  보통",   "92D050"
    else:             return "    D  검토",   "BFBFBF"

def _rank_fill(rank):
    if not HAS_OPENPYXL: return None
    if rank == 1:   return _GLD_FILL
    elif rank == 2: return _SLV_FILL
    elif rank == 3: return _BRZ_FILL
    elif rank <= 5: return _GRN_FILL
    else:           return _ALT_FILL

def auto_col_width(ws, min_w=8, max_w=45):
    if not HAS_OPENPYXL: return
    col_widths = {}
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell, MergedCell) or cell.value is None:
                continue
            text = str(cell.value)
            w = sum(2 if ord(c) > 127 else 1 for c in text) + 2
            col_letter = get_column_letter(cell.column)
            col_widths[col_letter] = max(col_widths.get(col_letter, 0), w)
    for col_letter, w in col_widths.items():
        ws.column_dimensions[col_letter].width = min(max(w, min_w), max_w)

# ══════════════════════════════════════════════════════════
# 캐시 시스템 v33 — 3계층 TTL (데이터 성격별 유효기간 분리)
#
#  ★ TIER_A (재무제표)  : TTL 7일  → DART 분기 데이터, 거의 안 변함
#  ★ TIER_B (보조지표)  : TTL 1일  → 네이버 업종/배당/결산PER/PBR
#  ★ TIER_C (현재가격)  : TTL 없음 → 매 실행마다 항상 새로 수집
#
#  cache_get(key, tier)  / cache_set(key, data, tier)
#  tier = "A" | "B"  (TIER_C는 캐시 자체를 사용하지 않음)
# ══════════════════════════════════════════════════════════

_TIER_TTL = {"A": 7, "B": 1}   # 단위: 일

def _cache_path(key: str, tier: str = "B") -> str:
    """캐시 파일 경로 반환 — tier별 폴더로 분리"""
    safe_key = hashlib.md5(key.encode()).hexdigest()[:12]
    tier_dir = os.path.join(CACHE_DIR, f"tier_{tier}")
    os.makedirs(tier_dir, exist_ok=True)
    return os.path.join(tier_dir, f"{safe_key}.json")

def cache_get(key: str, tier: str = "B"):
    """캐시에서 데이터 로드 — tier별 TTL 적용"""
    ttl_days = _TIER_TTL.get(tier, 1)
    path = _cache_path(key, tier)
    if not os.path.exists(path):
        return None
    try:
        mtime = os.path.getmtime(path)
        age_days = (time.time() - mtime) / 86400
        if age_days > ttl_days:
            return None   # 만료
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return None

def cache_set(key: str, data, tier: str = "B") -> None:
    """캐시에 데이터 저장"""
    path = _cache_path(key, tier)
    try:
        with open(path, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False)
    except Exception:
        pass

def cache_clear(days_old: int = 8) -> None:
    """오래된 캐시 파일 자동 삭제 (tier 폴더 포함)"""
    deleted = 0
    cutoff = time.time()
    try:
        for tier_name in ["tier_A", "tier_B", ""]:
            target_dir = os.path.join(CACHE_DIR, tier_name) if tier_name else CACHE_DIR
            if not os.path.isdir(target_dir):
                continue
            for fname in os.listdir(target_dir):
                if not fname.endswith(".json"):
                    continue
                fpath = os.path.join(target_dir, fname)
                try:
                    if (cutoff - os.path.getmtime(fpath)) / 86400 >= days_old:
                        os.remove(fpath)
                        deleted += 1
                except Exception:
                    pass
        if deleted:
            print(f"  [캐시] 오래된 캐시 {deleted}개 삭제")
    except Exception:
        pass

def cache_show_status() -> None:
    """캐시 상태 요약 출력 (v33: tier별)"""
    for tier in ["A", "B"]:
        tier_dir = os.path.join(CACHE_DIR, f"tier_{tier}")
        if not os.path.isdir(tier_dir):
            continue
        files = [f for f in os.listdir(tier_dir) if f.endswith(".json")]
        ttl = _TIER_TTL[tier]
        valid = 0
        for f in files:
            try:
                age = (time.time() - os.path.getmtime(os.path.join(tier_dir, f))) / 86400
                if age <= ttl:
                    valid += 1
            except Exception:
                pass
        label = "DART 재무(7일)" if tier == "A" else "네이버 보조(1일)"
        print(f"  💾 캐시 TIER-{tier} [{label}]: {valid}/{len(files)}개 유효")

def load_dart_key() -> str:
    # 항상 스크립트 폴더에 저장 (실행 위치 무관)
    key_file = os.path.join(BASE_DIR, "dart_api.txt")
    if os.path.exists(key_file):
        with open(key_file) as f:
            key = f.read().strip()
        if key:
            print(f"  ✅ DART API 키 자동 로드 완료")
            return key
    print("\n  DART API 키를 입력하세요 (opendart.fss.or.kr 에서 발급, 무료):")
    print(f"  (한 번 입력하면 {key_file} 에 저장되어 다음부터 자동 로드됩니다)")
    key = input("  키 입력: ").strip()
    if key:
        with open(key_file, "w") as f:
            f.write(key)
        print(f"  💾 저장 완료 → 앞으로 자동 로드됩니다")
    return key

def parse_args():
    p = argparse.ArgumentParser(description=f"퀀트 주식 스크리너 {VERSION}")
    p.add_argument("--top",         type=int,   default=20)
    p.add_argument("--min-roe",     type=float, default=8.0)
    p.add_argument("--max-pbr",     type=float, default=4.0)
    p.add_argument("--min-per",     type=float, default=0.0)
    p.add_argument("--max-per",     type=float, default=999.0)
    p.add_argument("--max-debt",    type=float, default=150.0)
    p.add_argument("--min-ic",      type=float, default=2.0)
    p.add_argument("--w-value",     type=float, default=20.0)
    p.add_argument("--w-mom",       type=float, default=20.0)
    p.add_argument("--w-div",       type=float, default=5.0)
    p.add_argument("--w-quality",   type=float, default=15.0)
    p.add_argument("--w-growth",    type=float, default=15.0)
    p.add_argument("--w-prof",      type=float, default=15.0)
    p.add_argument("--w-stability", type=float, default=10.0)
    p.add_argument("--output-dir",  type=str,   default=BASE_DIR)
    p.add_argument("--no-cache",    action="store_true", help="캐시 무시하고 새로 수집")
    return p.parse_args()

# ══════════════════════════════════════════════════════════
# 2. DART 클라이언트 (한국 재무 원본)
# ══════════════════════════════════════════════════════════

class DartClient:
    BASE = "https://opendart.fss.or.kr/api"

    def __init__(self, api_key: str):
        self.api_key  = api_key
        self._corp_map = {}

    def load_corp_codes(self) -> dict:
        if self._corp_map:
            return self._corp_map

        # 캐시 파일 확인 (당일 다운로드한 파일 재사용)
        cache_path = os.path.join(BASE_DIR, ".cache", "dart_corp_codes.json")
        if os.path.exists(cache_path):
            try:
                mtime = os.path.getmtime(cache_path)
                # 7일 이내 캐시 재사용
                if time.time() - mtime < 7 * 86400:
                    with open(cache_path, "r", encoding="utf-8") as f:
                        self._corp_map = json.load(f)
                    print(f"  [DART] 기업 코드 캐시 로드: {len(self._corp_map)}개")
                    return self._corp_map
            except Exception:
                pass

        print("  [DART] 기업 고유번호 목록 다운로드 중...")
        last_err = None
        for attempt in range(3):
            try:
                resp = requests.get(
                    f"{self.BASE}/corpCode.xml",
                    params={"crtfc_key": self.api_key},
                    timeout=30
                )
                # 응답이 ZIP인지 확인
                if resp.content[:2] != b'PK':
                    raise ValueError(f"ZIP 응답 아님 (HTTP {resp.status_code}): {resp.text[:200]}")

                with zipfile.ZipFile(io.BytesIO(resp.content)) as z:
                    xml_data = z.read("CORPCODE.xml").decode("utf-8")
                import xml.etree.ElementTree as ET
                root = ET.fromstring(xml_data)
                for item in root.findall("list"):
                    stock_code = item.findtext("stock_code","").strip()
                    corp_code  = item.findtext("corp_code","").strip()
                    if stock_code and len(stock_code) == 6:
                        self._corp_map[stock_code] = corp_code
                print(f"  [DART] 기업 코드 로드: {len(self._corp_map)}개")

                # 캐시 저장
                try:
                    os.makedirs(os.path.dirname(cache_path), exist_ok=True)
                    with open(cache_path, "w", encoding="utf-8") as f:
                        json.dump(self._corp_map, f, ensure_ascii=False)
                except Exception:
                    pass
                return self._corp_map

            except Exception as e:
                last_err = e
                print(f"  [DART] 기업코드 다운로드 실패 (시도 {attempt+1}/3): {e}")
                time.sleep(3)

        print(f"  ⚠ DART 기업코드 로드 실패 — 재무 데이터 수집 불가: {last_err}")
        return self._corp_map

    def get_financial(self, corp_code: str, year: str, reprt_code="11011") -> dict:
        url = f"{self.BASE}/fnlttSinglAcntAll.json"
        for fs_div in ("CFS", "OFS"):
            params = {"crtfc_key": self.api_key, "corp_code": corp_code,
                      "bsns_year": year, "reprt_code": reprt_code, "fs_div": fs_div}
            try:
                resp = requests.get(url, params=params, timeout=15)
                data = resp.json()
                if data.get("status") == "000" and data.get("list"):
                    result = {}
                    for row in data["list"]:
                        acnt_id = row.get("account_id", "").strip()
                        acnt_nm = row.get("account_nm", "").strip()
                        thstrm  = row.get("thstrm_amount", "").replace(",", "").strip()
                        if not thstrm:
                            continue
                        try:
                            val = float(thstrm)
                        except ValueError:
                            continue

                        # ① account_id 역방향 매핑 (가장 정확)
                        std_key = _ACNT_NM_MAP.get(acnt_id) or _ACNT_NM_MAP.get(acnt_id.replace(" ",""))
                        if std_key and std_key not in result:
                            result[std_key] = val
                            continue

                        # ② account_nm 역방향 매핑
                        std_key = _ACNT_NM_MAP.get(acnt_nm) or _ACNT_NM_MAP.get(acnt_nm.replace(" ",""))
                        if std_key and std_key not in result:
                            result[std_key] = val
                            continue

                        # ③ account_nm 부분 포함 매칭 (회사별 자유 명칭 대응)
                        for kor_name, ids in DART_ACCOUNTS.items():
                            if kor_name not in result:
                                for alias in ids:
                                    if len(alias) > 4 and alias in acnt_nm:
                                        result[kor_name] = val
                                        break

                    if result:
                        return result
            except Exception:
                pass
        return {}

    def get_latest_year(self) -> str:
        now = datetime.now()
        return str(now.year - 1) if now.month >= 4 else str(now.year - 2)

    def get_shares_outstanding(self, corp_code: str) -> int | None:
        """DART /company.json에서 발행주식총수(istc_totCnt) 조회"""
        url = f"{self.BASE}/company.json"
        try:
            resp = requests.get(
                url,
                params={"crtfc_key": self.api_key, "corp_code": corp_code},
                timeout=10
            )
            data = resp.json()
            if data.get("status") == "000":
                cnt_raw = data.get("istc_totCnt") or ""
                # 쉼표·공백 등 비숫자 문자 제거 후 정수 변환
                cnt = "".join(c for c in str(cnt_raw) if c.isdigit())
                if cnt:
                    val = int(cnt)
                    # 비현실적 주식수 필터 (1천주 미만 or 1조주 초과)
                    if 1_000 <= val <= 1_000_000_000_000:
                        return val
        except Exception:
            pass
        return None

    def get_shares_bulk(self, stock_codes: list, max_workers: int = 30) -> dict:
        """발행주식수 배치 수집 — 7일 캐시 적용, 병렬처리
        반환: {stock_code: shares(int)} 딕셔너리
        """
        cache_path = os.path.join(BASE_DIR, ".cache", "dart_shares.json")

        # ── 캐시 로드 (7일 이내) ──
        cached = {}
        if os.path.exists(cache_path):
            try:
                mtime = os.path.getmtime(cache_path)
                if time.time() - mtime < 7 * 86400:
                    with open(cache_path, "r", encoding="utf-8") as f:
                        cached = json.load(f)
                    # 정수 타입 보정
                    cached = {k: int(v) for k, v in cached.items() if v}
            except Exception:
                cached = {}

        corp_map = self.load_corp_codes()
        need = [c for c in stock_codes if c not in cached]

        if need:
            print(f"  [주식수] DART 배치 수집: {len(need)}개 종목 (미캐시)")
            from concurrent.futures import ThreadPoolExecutor, as_completed

            def _fetch_one(code):
                corp_code = corp_map.get(code)
                if not corp_code:
                    return code, None
                shares = self.get_shares_outstanding(corp_code)
                return code, shares

            results = {}
            with ThreadPoolExecutor(max_workers=max_workers) as ex:
                futures = {ex.submit(_fetch_one, c): c for c in need}
                done = 0
                for fut in as_completed(futures):
                    code, shares = fut.result()
                    if shares:
                        results[code] = shares
                    done += 1
                    if done % 200 == 0:
                        print(f"    {done}/{len(need)} 완료...")

            cached.update(results)
            ok = sum(1 for v in results.values() if v)
            print(f"  [주식수] 수집 완료: {ok}/{len(need)}개 성공")

            # 캐시 저장
            try:
                os.makedirs(os.path.dirname(cache_path), exist_ok=True)
                with open(cache_path, "w", encoding="utf-8") as f:
                    json.dump({k: str(v) for k, v in cached.items()}, f, ensure_ascii=False)
            except Exception:
                pass
        else:
            print(f"  [주식수] 캐시에서 로드: {len(cached)}개 종목")

        return cached

    def fetch_financials_dart(self, stock_code: str) -> dict:
        keys = [
            "cfo","cfi","net_income","gross_profit","total_assets",
            "total_liabilities","equity","op_income","interest_expense",
            "depreciation","cash","current_assets","current_liabilities",
            "dividends_paid",
            "eps_dart","bps_dart",           # v30: 주당순이익/주당순자산
            # 파생 지표
            "debt_ratio","net_debt_ratio","current_ratio","interest_coverage",
            "roe","gpa","fcf","fcf_margin","ebitda","altman_z",
            "dividend_payout","vol_ratio",
        ]
        empty = {k: np.nan for k in keys}
        corp_map  = self.load_corp_codes()
        corp_code = corp_map.get(stock_code)
        if not corp_code:
            return empty
        year = self.get_latest_year()
        fs = self.get_financial(corp_code, year)
        if not fs:
            fs = self.get_financial(corp_code, str(int(year)-1))
        if not fs:
            return empty
        r = empty.copy()

        # ── 원본 항목 ──
        r["cfo"]                = fs.get("영업활동현금흐름", np.nan)
        r["cfi"]                = fs.get("투자활동현금흐름", np.nan)
        r["net_income"]         = fs.get("당기순이익",       np.nan)
        r["gross_profit"]       = fs.get("매출총이익",       np.nan)
        r["total_assets"]       = fs.get("총자산",           np.nan)
        r["total_liabilities"]  = fs.get("총부채",           np.nan)
        r["equity"]             = fs.get("자기자본",         np.nan)
        r["op_income"]          = fs.get("영업이익",         np.nan)
        r["interest_expense"]   = fs.get("이자비용",         np.nan)
        r["depreciation"]       = fs.get("감가상각비",       np.nan)
        r["cash"]               = fs.get("현금및현금성자산", np.nan)
        r["current_assets"]     = fs.get("유동자산",         np.nan)
        r["current_liabilities"]= fs.get("유동부채",         np.nan)
        r["dividends_paid"]     = fs.get("배당금지급",       np.nan)
        # v30: 주당 지표 (원/주) → PBR/PER 현재가 기준 계산의 핵심
        r["eps_dart"]           = fs.get("주당순이익",       np.nan)
        r["bps_dart"]           = fs.get("주당순자산",       np.nan)

        # ── 파생 지표 계산 ──
        eq  = r["equity"]
        tl  = r["total_liabilities"]
        ta  = r["total_assets"]
        ni  = r["net_income"]
        cfo = r["cfo"]
        cfi = r["cfi"]
        opm = r["op_income"]
        ie  = r["interest_expense"]
        ca  = r["current_assets"]
        cl  = r["current_liabilities"]
        dep = r["depreciation"]
        cash= r["cash"]

        def _ok(*vals):
            return all(v is not None and not np.isnan(v) for v in vals)

        # 부채비율
        if _ok(tl, eq) and eq > 0:
            r["debt_ratio"] = (tl / eq) * 100

        # 순부채비율 = (총부채 - 현금) / 자기자본
        if _ok(tl, cash, eq) and eq > 0:
            r["net_debt_ratio"] = ((tl - cash) / eq) * 100

        # 유동비율
        if _ok(ca, cl) and cl > 0:
            r["current_ratio"] = (ca / cl) * 100

        # 이자보상배율
        if _ok(opm, ie) and ie > 0:
            r["interest_coverage"] = opm / abs(ie)

        # ROE
        if _ok(ni, eq) and eq > 0:
            r["roe"] = (ni / eq) * 100

        # GP/A
        if _ok(r["gross_profit"], ta) and ta > 0:
            r["gpa"] = r["gross_profit"] / ta

        # FCF = CFO + CFI (투자CF는 음수가 정상 → CAPEX 근사)
        # 더 정확: FCF = CFO - |CAPEX|, CAPEX ≈ |CFI| 근사 사용
        if _ok(cfo, cfi):
            capex_est = abs(cfi)  # 투자CF 절댓값 = 대략적 CAPEX
            r["fcf"] = cfo - capex_est

        # FCF 마진 = FCF / 매출총이익(근사)
        gp = r["gross_profit"]
        if _ok(r["fcf"], gp) and gp > 0:
            r["fcf_margin"] = (r["fcf"] / gp) * 100

        # EBITDA = 영업이익 + 감가상각비
        if _ok(opm, dep):
            r["ebitda"] = opm + abs(dep)
        elif _ok(opm):
            r["ebitda"] = opm  # 감가상각 없으면 영업이익으로 대체

        # Altman Z-Score (비금융업 기준)
        # Z = 1.2*X1 + 1.4*X2 + 3.3*X3 + 0.6*X4 + 1.0*X5
        # X1=운전자본/총자산, X2=이익잉여금/총자산(ROE 근사),
        # X3=EBIT/총자산, X4=자기자본/총부채, X5=매출/총자산(GP 근사)
        if _ok(ca, cl, ta, ni, opm, eq, tl, gp) and ta > 0 and tl > 0:
            x1 = (ca - cl) / ta
            x2 = (ni / ta) if ta > 0 else 0          # 순이익/총자산 (이익잉여금 근사)
            x3 = opm / ta
            x4 = eq / tl
            x5 = gp / ta                              # 매출총이익/총자산 (매출 근사)
            r["altman_z"] = round(1.2*x1 + 1.4*x2 + 3.3*x3 + 0.6*x4 + 1.0*x5, 2)

        # 배당성향 = 배당금지급 / 순이익
        div_paid = r["dividends_paid"]
        if _ok(div_paid, ni) and ni > 0:
            r["dividend_payout"] = abs(div_paid) / ni * 100

        return r

# ══════════════════════════════════════════════════════════
# 3. 네이버 금융 스크래핑 (배당/업종/PER/PBR/영업이익률 등 보조지표)
# ══════════════════════════════════════════════════════════


def _decode_html(content: bytes) -> str:
    """네이버 금융 HTML 바이트 → 문자열 (cp949 완전 지원)"""
    for enc in ("cp949", "euc-kr", "utf-8"):
        try:
            return content.decode(enc)
        except (UnicodeDecodeError, LookupError):
            continue
    return content.decode("utf-8", errors="ignore")


def _clean_str(s: str) -> str:
    """인코딩 깨짐 문자(□, ?, 특수문자 덩어리) 감지 후 빈문자열 반환"""
    if not s:
        return s
    # 유니코드 대체문자(U+FFFD), 알 수 없는 문자 비율이 높으면 빈값
    bad = sum(1 for c in s if ord(c) in (0xFFFD, 0x25A1, 0x3013))
    if bad / max(len(s), 1) > 0.3:
        return ""
    return s.strip()

def _safe_num(raw_text):
    if not raw_text:
        return None
    txt = raw_text.replace(",","").replace("%","").replace("▲","").replace("▼","-").strip()
    txt = re.sub(r"[^\d.\-]","", txt)
    try:
        return float(txt)
    except ValueError:
        return None

def _safe_price(raw_text):
    if not raw_text:
        return None
    nums = re.findall(r"[\d,]+", raw_text.replace(" ",""))
    if not nums:
        return None
    try:
        candidates = [int(n.replace(",","")) for n in nums]
        valid = [v for v in candidates if 100 <= v <= 20_000_000]
        return valid[0] if valid else None
    except (ValueError, IndexError):
        return None

def _parse_naver_current_price(code: str):
    try:
        url  = f"https://finance.naver.com/item/sise.naver?code={code}"
        resp = requests.get(url, headers=NAVER_HEADERS, timeout=5)
        html = _decode_html(resp.content)
        m = re.search(r'<strong[^>]*id="_nowVal"[^>]*>([\d,]+)</strong>', html)
        if m:
            return int(m.group(1).replace(",",""))
        soup = BeautifulSoup(html, "html.parser") if HAS_BS4 else None
        if soup:
            tag = soup.find(id="_nowVal") or soup.find("strong", class_="tah p11")
            if tag:
                v = _safe_price(tag.get_text(strip=True))
                if v:
                    return v
    except Exception:
        pass
    return None

def _get_latest_annual_idx(soup):
    try:
        th_eps = soup.find("th", class_="th_cop_anal17")
        if not th_eps:
            return 0
        table = th_eps.find_parent("table")
        if not table:
            return 0
        thead = table.find("thead")
        if not thead:
            return 0
        all_trs = thead.find_all("tr")
        year_pat = re.compile(r'^\d{4}\.\d{2}')
        annual_cols = 4
        for tr in all_trs:
            for cell in tr.find_all(["th","td"]):
                txt = cell.get_text(strip=True)
                if "연간" in txt:
                    try:
                        annual_cols = int(cell.get("colspan", 4))
                    except (ValueError, TypeError):
                        annual_cols = 4
                    break
        for tr in all_trs:
            cells     = tr.find_all(["th","td"])
            year_texts = [c.get_text(strip=True) for c in cells
                          if year_pat.match(c.get_text(strip=True))]
            if len(year_texts) < 2:
                continue
            ann_years = year_texts[:annual_cols]
            confirmed_idx = -1
            for i, t in enumerate(ann_years):
                if "(E)" not in t:
                    confirmed_idx = i
            return confirmed_idx if confirmed_idx >= 0 else len(ann_years) - 1
    except Exception:
        pass
    return 0

def _parse_naver_stock(code: str) -> dict:
    """네이버 금융에서 보조 재무지표 수집 (배당/업종/PER/영업이익률/ROA 등)"""
    result = {"code": code}
    if not HAS_BS4:
        return result
    for attempt in range(3):
        try:
            # 현재가
            url  = f"https://finance.naver.com/item/main.naver?code={code}"
            resp = requests.get(url, headers=NAVER_HEADERS, timeout=6)
            html = _decode_html(resp.content)
            soup = BeautifulSoup(html, "html.parser")

            # 현재가
            cur_price = None
            for sel in [("strong","tah p11"), ("p","no_today")]:
                tag = soup.find(sel[0], class_=sel[1])
                if tag:
                    v = _safe_price(tag.get_text(strip=True))
                    if v:
                        cur_price = v
                        break
            if not cur_price:
                cur_price = _parse_naver_current_price(code)
            if cur_price:
                result["current_price"] = cur_price

            # 시가총액 (현재가 정합성 검증에 사용)
            marcap_eok_nv = None
            for th in soup.find_all("th"):
                if th.get_text(strip=True) == "시가총액":
                    td = th.find_next_sibling("td")
                    if td:
                        raw = td.get_text(strip=True)
                        if "조" in raw:
                            m = re.search(r"([\d.]+)조", raw)
                            if m:
                                marcap_eok_nv = int(float(m.group(1)) * 10000)
                        elif "억" in raw:
                            m = re.search(r"([\d,]+)억", raw)
                            if m:
                                marcap_eok_nv = int(m.group(1).replace(",",""))
                    break

            # ── 현재가 정합성 검증 (시가총액 기반) ──
            if cur_price and marcap_eok_nv and marcap_eok_nv > 0:
                implied_shares = marcap_eok_nv * 1e8 / cur_price
                if implied_shares < 100_000 or implied_shares > 5e10:
                    # 비현실적 주식수 → 현재가 오류
                    result["current_price"] = None
                    result["price_error"] = f"현재가오류({cur_price:,}원,추정주식수{implied_shares:.0f}주)"
            if marcap_eok_nv:
                result["market_cap_억"] = marcap_eok_nv

            # 업종
            for em in soup.find_all("em", class_="coinfo01"):
                result["sector"] = _clean_str(em.get_text(strip=True))
                break
            if "sector" not in result:
                a_sec = soup.find("a", href=re.compile(r"sise_group_detail"))
                if a_sec:
                    result["sector"] = _clean_str(a_sec.get_text(strip=True))

            # 재무 분석 지표 (CLASS_MAP 방식)
            # ★ EPS/BPS(th_cop_anal17/18)는 네이버에서 억원 단위로 표시되어
            #   자릿수 오류가 생기므로 수집하지 않음 (결산 PER/PBR에서 역산)
            annual_idx = _get_latest_annual_idx(soup)
            CLASS_MAP = {
                "th_cop_anal11": "debt_ratio_nv",   # 부채비율 (네이버)
                "th_cop_anal13": "op_margin",        # 영업이익률
                "th_cop_anal14": "net_margin",       # 순이익률
                "th_cop_anal15": "roe_nv",           # ROE (네이버)
                "th_cop_anal16": "roa",              # ROA
                # th_cop_anal17(EPS), th_cop_anal18(BPS) 제거 — 단위 오류
                "th_cop_anal19": "dps",              # DPS
                "th_cop_anal20": "per_결산",         # PER(결산)
                "th_cop_anal21": "pbr_결산",         # PBR(결산)
                "th_cop_anal22": "div",              # 배당수익률
            }
            for cls, field in CLASS_MAP.items():
                th = soup.find("th", class_=cls)
                if not th:
                    continue
                tds = []
                sib = th.find_next_sibling("td")
                while sib is not None and sib.name == "td":
                    tds.append(sib)
                    sib = sib.find_next_sibling("td")
                if not tds:
                    continue
                idx = annual_idx if annual_idx < len(tds) else (len(tds)-1)
                v   = _safe_num(tds[idx].get_text(strip=True))
                if v is None:
                    continue
                # ── 필드별 유효성 검사 ──
                if field == "roa":
                    # ROA는 -50% ~ +50% 범위 (단위가 % 아닌 소수면 *100)
                    if abs(v) > 100:
                        continue   # 비정상값 제거
                    result[field] = v
                elif field == "roe_nv":
                    if abs(v) > 200:
                        continue
                    result[field] = v
                elif field == "debt_ratio_nv":
                    if v < 0 or v > 10000:
                        continue
                    result[field] = v
                elif field in ("per_결산", "pbr_결산"):
                    if v <= 0 or v > 500:
                        continue
                    result[field] = v
                elif field == "dps":
                    result[field] = int(v) if v > 0 else 0
                else:
                    result[field] = v

            # 매출액·영업이익 (텍스트 매칭)
            TEXT_MAP = {"매출액": "revenue", "영업이익": "op_profit"}
            for label, field in TEXT_MAP.items():
                for th in soup.find_all("th"):
                    if th.get_text(strip=True) == label:
                        tds = []
                        sib = th.find_next_sibling("td")
                        while sib and sib.name == "td":
                            tds.append(sib)
                            sib = sib.find_next_sibling("td")
                        if tds:
                            idx = annual_idx if annual_idx < len(tds) else (len(tds)-1)
                            v = _safe_num(tds[idx].get_text(strip=True))
                            if v is not None:
                                result[field] = v
                            if field == "revenue" and idx > 0:
                                pv = _safe_num(tds[idx-1].get_text(strip=True))
                                if pv and pv > 0:
                                    result["revenue_prev"] = pv
                        break

            # 매출 성장률
            rev   = result.get("revenue")
            rev_p = result.get("revenue_prev")
            if rev and rev_p and rev_p > 0:
                result["revenue_growth"] = round((rev - rev_p) / rev_p * 100, 1)

            # PBR/PER: 결산 기준 사용 (EPS/BPS 파싱 제거로 현재가 기준 재계산 불가)
            # 결산 PER/PBR이 가장 신뢰할 수 있는 값
            result["per"] = result.get("per_결산")
            result["pbr"] = result.get("pbr_결산")

            # EPS/BPS 컬럼도 제거 (오염된 값 전파 방지)
            result.pop("eps", None)
            result.pop("bps", None)

            return result

        except requests.exceptions.Timeout:
            time.sleep(1)
        except Exception as e:
            if attempt == 2:
                return {"code": code, "error": str(e)}
            time.sleep(0.5)
    return {"code": code, "error": "max retries"}

def fetch_naver_bulk(tickers, names, max_workers=20, no_cache=False) -> pd.DataFrame:
    """네이버 금융 병렬 수집 (캐시 적용)
    v33 변경:
      - 업종/배당/결산PER·PBR 등 보조지표 → TIER-B 캐시 (1일 유효)
      - 현재가(current_price) → 캐시 우회, 매 실행마다 항상 새로 수집 (TIER-C)
    """
    total_nv = len(tickers)

    # ── TIER-B 캐시 확인 (현재가 제외 보조지표) ──
    cache_key_nv = f"naver_bulk_{','.join(sorted(tickers[:20]))}_n{total_nv}"
    cached_meta = None  # 현재가 제외 보조지표 캐시
    if not no_cache:
        cached_meta = cache_get(cache_key_nv, tier="B")

    # ── TIER-C: 현재가는 항상 새로 수집 ──
    print(f"  [네이버] 현재가 실시간 수집 중 ({total_nv}개, 병렬처리)...")

    def _fetch_price_only(code):
        """현재가만 빠르게 수집 (TIER-C — 항상 최신)"""
        price = _parse_naver_current_price(code)
        time.sleep(0.02)
        return code, price

    price_map = {}
    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        futs = {executor.submit(_fetch_price_only, c): c for c in tickers}
        done = 0
        for fut in as_completed(futs):
            code, price = fut.result()
            if price:
                price_map[code] = price
            done += 1
            if done % 500 == 0:
                print(f"  [현재가] {done}/{total_nv} 수집 중...")
    ok_cnt = sum(1 for v in price_map.values() if v)
    print(f"  ✓ 현재가 수집 완료: {ok_cnt}/{total_nv}개")

    # 보조지표: 캐시 있으면 재사용, 없으면 새로 수집
    if cached_meta is not None:
        print(f"  [캐시] 네이버 보조지표 TIER-B 캐시 적중 → 재수집 생략 (업종/배당/PER/PBR 등)")
        try:
            df = pd.DataFrame(cached_meta).set_index("코드")
            # 현재가만 오늘 수집값으로 교체
            for code, price in price_map.items():
                if code in df.index:
                    df.loc[code, "현재가_NV"] = price
            return df
        except Exception:
            pass  # 캐시 파싱 실패 → 전체 재수집

    # 전체 재수집 (보조지표 캐시 없을 때)
    print(f"  [네이버] 보조지표 전체 수집 중 ({total_nv}개)...")
    code_to_name = dict(zip(tickers, names))
    results = {}
    _nv_done = [0]

    if HAS_TQDM:
        pbar = tqdm(total=total_nv, desc="  네이버 수집", unit="종목",
                    bar_format="{l_bar}{bar}| {n_fmt}/{total_fmt} [{elapsed}<{remaining}]")

    def _fetch_one(code):
        r = _parse_naver_stock(code)
        time.sleep(0.03)
        return code, r

    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        futures = {executor.submit(_fetch_one, c): c for c in tickers}
        for future in as_completed(futures):
            code, data = future.result()
            results[code] = data
            if HAS_TQDM:
                pbar.update(1)
            else:
                _nv_done[0] += 1
                if _nv_done[0] % 200 == 0:
                    print(f"  [네이버] {_nv_done[0]}/{total_nv} ({_nv_done[0]/total_nv*100:.0f}%) 수집 중...")

    if HAS_TQDM:
        pbar.close()

    rows = []
    for code in tickers:
        d = results.get(code, {})
        name = code_to_name.get(code) or d.get("name", code)
        rows.append({
            "코드":          code,
            "종목명":        _clean_str(str(name)) if name else "",
            "시가총액(억)":  _safe(d.get("market_cap_억")),
            "업종":          d.get("sector",""),
            "현재가_NV":     _safe(price_map.get(code) or d.get("current_price")),
            "PBR_결산":      _safe(d.get("pbr_결산")),
            "PBR":           _safe(d.get("pbr")),
            "PER_결산":      _safe(d.get("per_결산")),
            "PER":           _safe(d.get("per")),
            # EPS/BPS 제거 — 네이버 억원 단위 파싱 오류로 신뢰 불가
            "ROE_NV":        _clamp(d.get("roe_nv"),  -200, 200),
            "ROA":           _clamp(d.get("roa"),     -50,  50),   # ROA: -50%~+50% 범위
            "DIV":           _safe(d.get("div"), 0.0),
            "영업이익률":    _clamp(d.get("op_margin"),  -100, 100),
            "순이익률":      _clamp(d.get("net_margin"), -200, 200),
            "부채비율_NV":   _clamp(d.get("debt_ratio_nv"), 0, 5000),
            "매출액":        _safe(d.get("revenue")),
            "매출성장률":    _safe(d.get("revenue_growth")),
            "영업이익_NV":   _safe(d.get("op_profit")),
        })

    df = pd.DataFrame(rows).set_index("코드")
    print(f"  ✓ 네이버 수집 완료: {df['PBR'].notna().sum()}개 PBR 수집")

    # TIER-B 캐시 저장 (현재가 제외 보조지표만 저장)
    try:
        df_cache = df.copy()
        df_cache["현재가_NV"] = None  # 현재가는 캐시에 저장 안 함
        cache_set(cache_key_nv, df_cache.reset_index().to_dict(orient="records"), tier="B")
        print(f"  [캐시] 네이버 보조지표 TIER-B 저장 (1일 유효, 현재가 제외)")
    except Exception:
        pass

    return df

# ══════════════════════════════════════════════════════════
# 4. yfinance 가격/모멘텀 (배치 다운로드)
# ══════════════════════════════════════════════════════════

def _safe_get_ohlc(hist, yf_ticker):
    """Close/High/Low/Volume 시리즈를 안전하게 추출"""
    t = yf_ticker

    def _extract(sub):
        c = sub["Close"].dropna()
        h = sub["High"].dropna()
        l = sub["Low"].dropna()
        v = sub["Volume"].dropna() if "Volume" in sub.columns else pd.Series(dtype=float)
        return c, h, l, v

    if not isinstance(hist.columns, pd.MultiIndex):
        return _extract(hist)
    lv0 = hist.columns.get_level_values(0).unique()
    lv1 = hist.columns.get_level_values(1).unique()
    if t in lv0:
        sub = hist[t]
        if isinstance(sub, pd.DataFrame):
            return _extract(sub)
    if t in lv1:
        c = hist["Close"][t].dropna()
        h = hist["High"][t].dropna()
        l = hist["Low"][t].dropna()
        v = hist["Volume"][t].dropna() if "Volume" in hist.columns.get_level_values(0) else pd.Series(dtype=float)
        return c, h, l, v
    for lv in (0, 1):
        try:
            sub = hist.xs(t, level=lv, axis=1)
            if isinstance(sub, pd.DataFrame) and "Close" in sub.columns:
                return _extract(sub)
        except Exception:
            pass
    raise KeyError(f"{t} not found in columns")


# ── 기술적 지표 계산 함수 (외부 라이브러리 불필요 — 순수 pandas/numpy) ──

def _ta_macd(close):
    """MACD(12,26,9) — 히스토그램 부호 교차로 매수/매도 판단"""
    if len(close) < 30:
        return None, None, None, "중립"
    ema12 = close.ewm(span=12, adjust=False).mean()
    ema26 = close.ewm(span=26, adjust=False).mean()
    macd  = ema12 - ema26
    sig   = macd.ewm(span=9, adjust=False).mean()
    hist  = macd - sig
    cur_h, prev_h = float(hist.iloc[-1]), float(hist.iloc[-2])
    # 히스토그램 부호 교차
    if prev_h <= 0 and cur_h > 0:   sig_txt = "매수"       # 골든크로스
    elif prev_h >= 0 and cur_h < 0: sig_txt = "매도"       # 데드크로스
    elif cur_h > 0:                  sig_txt = "매수유지"
    elif cur_h < 0:                  sig_txt = "매도유지"
    else:                            sig_txt = "중립"
    return round(float(macd.iloc[-1]),2), round(float(sig.iloc[-1]),2), round(cur_h,2), sig_txt


def _ta_rsi(close, period=14):
    """RSI(14) — 30 이하 과매도=매수, 70 이상 과매수=매도"""
    if len(close) < period + 1:
        return None, "중립"
    delta = close.diff()
    gain  = delta.clip(lower=0)
    loss  = (-delta).clip(lower=0)
    avg_g = gain.ewm(alpha=1/period, adjust=False).mean()
    avg_l = loss.ewm(alpha=1/period, adjust=False).mean()
    rs    = avg_g / avg_l.replace(0, np.nan)
    rsi   = 100 - (100 / (1 + rs))
    val   = float(rsi.iloc[-1])
    if   val < 30: sig = "매수"       # 과매도
    elif val > 70: sig = "매도"       # 과매수
    elif val < 45: sig = "매수유지"
    elif val > 55: sig = "매도유지"
    else:          sig = "중립"
    return round(val, 1), sig


def _ta_obv(close, volume):
    """OBV — 5일/20일 OBV 이동평균 비교로 추세 판단"""
    if len(close) < 20 or volume is None or len(volume) < 20:
        return "중립", "중립"
    try:
        # 인덱스 정렬
        common = close.index.intersection(volume.index)
        c = close.reindex(common)
        v = volume.reindex(common)
        direction = np.sign(c.diff().fillna(0))
        obv  = (direction * v).cumsum()
        obv5  = float(obv.rolling(5).mean().iloc[-1])
        obv20 = float(obv.rolling(20).mean().iloc[-1])
        trend = "상승" if float(obv.iloc[-1]) > float(obv.iloc[-5]) else "하락"
        if   obv5 > obv20 * 1.02: sig = "매수"
        elif obv5 < obv20 * 0.98: sig = "매도"
        else:                      sig = "중립"
        return trend, sig
    except Exception:
        return "중립", "중립"


def _ta_bb(close, period=20, std_mult=2.0):
    """Bollinger Bands(20,2) — %B 기준 매수/매도"""
    if len(close) < period:
        return None, None, None, None, "중립"
    ma    = close.rolling(period).mean()
    std   = close.rolling(period).std()
    upper = ma + std_mult * std
    lower = ma - std_mult * std
    cur   = float(close.iloc[-1])
    u     = float(upper.iloc[-1])
    l     = float(lower.iloc[-1])
    pct_b = (cur - l) / (u - l) if (u - l) > 0 else 0.5
    # 밴드폭 = (상단-하단)/중간×100 (%)  — 수치가 낮으면 스퀴즈
    bw    = (u - l) / float(ma.iloc[-1]) * 100 if float(ma.iloc[-1]) > 0 else 0
    if   pct_b < 0.0: sig = "매수"       # 하단 이탈
    elif pct_b > 1.0: sig = "매도"       # 상단 이탈
    elif pct_b < 0.2: sig = "매수유지"
    elif pct_b > 0.8: sig = "매도유지"
    else:             sig = "중립"
    return round(u), round(l), round(pct_b*100,1), round(bw,1), sig


def _ta_summary(macd_sig, rsi_sig, obv_sig, bb_sig):
    """4개 지표 종합 → 기술신호 + 매수/매도 카운트"""
    sigs = [macd_sig, rsi_sig, obv_sig, bb_sig]
    buy  = sum(1 for s in sigs if "매수" in s)
    sell = sum(1 for s in sigs if "매도" in s)
    if   buy >= 3:  overall = "⚡ 강력매수"
    elif buy == 2:  overall = "▲ 매수"
    elif sell >= 3: overall = "🔻 강력매도"
    elif sell == 2: overall = "▼ 매도"
    else:           overall = "─ 중립"
    summary = f"MACD:{macd_sig} RSI:{rsi_sig} OBV:{obv_sig} BB:{bb_sig}"
    return buy, sell, overall, summary


def _compute_price_row(close_s, high_s, low_s, vol_s=None):
    """가격/모멘텀/거래량/기술지표(MACD·RSI·OBV·BB) 통합 계산 (v32)"""
    cur   = float(close_s.iloc[-1])
    hi52  = float(high_s.max())
    lo52  = float(low_s.min())
    pos   = (cur - lo52) / (hi52 - lo52) * 100 if hi52 > lo52 else 50.0
    first = float(close_s.iloc[0])
    ret52 = (cur / first - 1) * 100 if first > 0 else 0.0
    idx_6m = max(0, len(close_s) - 126)
    ret6m = (cur / float(close_s.iloc[idx_6m]) - 1) * 100 if float(close_s.iloc[idx_6m]) > 0 else 0.0

    # ── 거래량 지표 ──
    amt5 = amt20 = vr = None
    if vol_s is not None and len(vol_s) >= 5:
        try:
            common_idx = close_s.index.intersection(vol_s.index)
            if len(common_idx) >= 5:
                c_aln = close_s.reindex(common_idx)
                v_aln = vol_s.reindex(common_idx)
                amt_s = c_aln * v_aln
                n = len(amt_s)
                amt5  = round(float(amt_s.iloc[-5:].mean()) / 1e8, 1)
                amt20 = round(float(amt_s.iloc[-min(20,n):].mean()) / 1e8, 1)
                if amt20 and amt20 > 0:
                    vr = round(amt5 / amt20 * 100, 1)
        except Exception:
            pass

    # ── 4대 기술지표 계산 ──
    macd_line = macd_sig_line = macd_hist = None
    macd_sig = rsi_val = "중립"
    obv_trend = obv_sig = "중립"
    bb_upper = bb_lower = bb_pct_b = bb_width = None
    bb_sig = "중립"
    ta_buy = ta_sell = 0
    ta_signal = "─ 중립"
    ta_summary_str = ""

    try:
        if len(close_s) >= 30:
            macd_line, macd_sig_line, macd_hist, macd_sig = _ta_macd(close_s)
            rsi_val, rsi_sig = _ta_rsi(close_s)
            obv_trend, obv_sig = _ta_obv(close_s, vol_s)
            bb_upper, bb_lower, bb_pct_b, bb_width, bb_sig = _ta_bb(close_s)
            ta_buy, ta_sell, ta_signal, ta_summary_str = _ta_summary(macd_sig, rsi_sig, obv_sig, bb_sig)
    except Exception:
        pass

    return {
        # 기존 가격 지표
        "현재가_YF":        int(round(cur)),
        "52주고가":         int(round(hi52)),
        "52주저가":         int(round(lo52)),
        "52주위치":         round(pos, 1),
        "52주수익률":       round(ret52, 1),
        "6개월수익률":      round(ret6m, 1),
        # 거래량 지표
        "거래대금5일(억)":  amt5,
        "거래대금20일(억)": amt20,
        "거래량비율VR(%)":  vr,
        # MACD
        "MACD선":          macd_line,
        "MACD시그널":       macd_sig_line,
        "MACD히스토":       macd_hist,
        "MACD신호":         macd_sig,
        # RSI
        "RSI14":           rsi_val,
        "RSI신호":          rsi_sig,
        # OBV
        "OBV추세":          obv_trend,
        "OBV신호":          obv_sig,
        # Bollinger Bands
        "BB상단":           bb_upper,
        "BB하단":           bb_lower,
        "BB%B":            bb_pct_b,
        "BB밴드폭(%)":      bb_width,
        "BB신호":           bb_sig,
        # 종합 기술신호
        "기술매수신호수":    ta_buy,
        "기술매도신호수":    ta_sell,
        "기술신호":          ta_signal,
        "기술지표요약":      ta_summary_str,
    }

def fetch_price_yfinance_batch(tickers, market_map=None) -> pd.DataFrame:
    if not HAS_YFINANCE:
        print("  ⚠ yfinance 미설치 → 모멘텀 중립값 사용")
        return pd.DataFrame()

    print(f"  [가격] yfinance 배치 다운로드 중 ({len(tickers)}개)...")
    if market_map is None:
        market_map = {}

    def _suffix(code):
        return ".KQ" if market_map.get(code,"KOSPI") == "KOSDAQ" else ".KS"

    yf_map  = {f"{c}{_suffix(c)}": c for c in tickers}
    yf_list = list(yf_map.keys())
    rows    = {}
    CHUNK   = 500

    chunks = [yf_list[i:i+CHUNK] for i in range(0, len(yf_list), CHUNK)]
    print(f"  배치: {len(chunks)}청크 × {CHUNK}개")

    all_chunks = []
    for ci, chunk in enumerate(chunks, 1):
        try:
            h = yf.download(chunk, period="1y", progress=False,
                            group_by="ticker", auto_adjust=True, threads=True)
            if h is not None and len(h) >= 5:
                all_chunks.append((chunk, h))
                print(f"  청크 {ci}/{len(chunks)} 완료")
            time.sleep(2)
        except Exception as e:
            print(f"  청크 {ci} 실패: {e}")
            time.sleep(5)

    processed = set()
    for chunk_list, hist_chunk in all_chunks:
        chunk_map = {t: yf_map[t] for t in chunk_list if t in yf_map}
        for yf_t, code in chunk_map.items():
            if code in processed:
                continue
            try:
                c_s, h_s, l_s, v_s = _safe_get_ohlc(hist_chunk, yf_t)
                if len(c_s) < 5:
                    continue
                rows[code] = _compute_price_row(c_s, h_s, l_s, v_s)
                processed.add(code)
            except Exception:
                pass

    if not rows:
        print("  ⚠ 가격 데이터 수집 실패 → 중립값 적용")
        return pd.DataFrame()

    df = pd.DataFrame.from_dict(rows, orient="index")
    df.index.name = "코드"
    df.index = df.index.astype(str)
    print(f"  ✓ 가격 수집 완료: {len(df)}개")
    return df

# ══════════════════════════════════════════════════════════
# 5. 한국 유니버스 수집
# ══════════════════════════════════════════════════════════

def fetch_kr_universe():
    tickers, markets = {}, {}
    EXCL = re.compile(r"스팩|SPAC|ETF|ETN|리츠|인프라|우$|우B$|우C$|우선주")

    # ── 방법1: FinanceDataReader (KRX 공식) ──
    if HAS_FDR:
        try:
            kospi  = fdr.StockListing("KOSPI") [["Code","Name","Market"]]
            kosdaq = fdr.StockListing("KOSDAQ")[["Code","Name","Market"]]
            for _, row in pd.concat([kospi, kosdaq]).iterrows():
                if not EXCL.search(str(row["Name"])):
                    tickers[str(row["Code"])] = str(row["Name"])
                    markets[str(row["Code"])] = str(row["Market"])
            if len(tickers) >= 500:
                print(f"  [유니버스] FDR 수집 완료: {len(tickers)}개")
                codes = list(tickers.keys())
                return codes, [tickers[c] for c in codes], [markets.get(c,"KOSPI") for c in codes]
            else:
                print(f"  [유니버스] FDR 결과 부족({len(tickers)}개) → KRX API 시도")
                tickers.clear(); markets.clear()
        except Exception as e:
            print(f"  [유니버스] FDR 실패: {e} → KRX API 시도")

    # ── 방법2: KRX 정보데이터시스템 직접 호출 ──
    def _fetch_krx(market_id, market_name):
        result = {}
        # KRX는 세션 쿠키 + generate_key 헤더가 필요 → 먼저 메인 페이지를 방문해 쿠키 획득
        try:
            sess = requests.Session()
            sess.get(
                "http://data.krx.co.kr/contents/MDC/MDI/mdiLoader/index.cmd",
                headers={"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)"},
                timeout=10,
            )
            url  = "http://data.krx.co.kr/comm/bldAttendant/getJsonData.cmd"
            body = {
                "bld":         "dbms/MDC/STAT/standard/MDCSTAT01901",
                "mktId":       market_id,
                "share":       "1",
                "money":       "1",
                "csvxls_isNo": "false",
            }
            hdrs = {
                "User-Agent":    "Mozilla/5.0 (Windows NT 10.0; Win64; x64)",
                "Referer":       "http://data.krx.co.kr/contents/MDC/MDI/mdiLoader/index.cmd",
                "Content-Type":  "application/x-www-form-urlencoded; charset=UTF-8",
                "Accept":        "application/json, text/javascript, */*; q=0.01",
                "X-Requested-With": "XMLHttpRequest",
            }
            resp = sess.post(url, data=body, headers=hdrs, timeout=20)
            resp.raise_for_status()
            if not resp.text.strip():
                raise ValueError("빈 응답")
            data = resp.json()
            for item in data.get("OutBlock_1", []):
                code = item.get("ISU_SRT_CD","").strip()
                name = item.get("ISU_ABBRV","").strip()
                if code and name and len(code)==6 and not EXCL.search(name):
                    result[code] = name
            print(f"  [유니버스] KRX {market_name}: {len(result)}개")
        except Exception as e:
            print(f"  [유니버스] KRX {market_name} 실패: {e}")
        return result

    kospi_d  = _fetch_krx("STK", "KOSPI")
    kosdaq_d = _fetch_krx("KSQ", "KOSDAQ")
    for code, name in kospi_d.items():
        tickers[code] = name; markets[code] = "KOSPI"
    for code, name in kosdaq_d.items():
        if code not in tickers:
            tickers[code] = name; markets[code] = "KOSDAQ"

    if len(tickers) >= 100:
        print(f"  [유니버스] KRX 수집 완료: {len(tickers)}개 (KOSPI {len(kospi_d)} + KOSDAQ {len(kosdaq_d)})")
    else:
        # ── 방법3: 네이버 시가총액 페이지 (최후 폴백) ──
        print(f"  [유니버스] KRX 결과 부족 → 네이버 수집 시도...")
        def _fetch_naver(sosok):
            result = {}
            for page in range(1, 300):
                try:
                    resp = requests.get(
                        f"https://finance.naver.com/sise/sise_market_sum.naver"
                        f"?sosok={sosok}&page={page}",
                        headers=NAVER_HEADERS, timeout=8)
                    html = _decode_html(resp.content)
                    found = re.findall(r'/item/main\.naver\?code=(\d{6})[^>]*>\s*([^<]{1,30})', html)
                    if not found:
                        break
                    new = 0
                    for code, name in found:
                        name = name.strip()
                        if name and len(code)==6 and code not in result and not EXCL.search(name):
                            result[code] = name; new += 1
                    if new == 0:
                        break
                    time.sleep(0.1)
                except Exception:
                    break
            return result

        nv_kospi  = _fetch_naver(0)
        nv_kosdaq = _fetch_naver(1)
        for code, name in nv_kospi.items():
            if code not in tickers:
                tickers[code] = name; markets[code] = "KOSPI"
        for code, name in nv_kosdaq.items():
            if code not in tickers:
                tickers[code] = name; markets[code] = "KOSDAQ"
        print(f"  [유니버스] 네이버 수집 완료: {len(tickers)}개")

    codes = list(tickers.keys())
    return codes, [tickers[c] for c in codes], [markets.get(c,"KOSPI") for c in codes]


# ══════════════════════════════════════════════════════════
# 6. 한국 통합 데이터 수집
# ══════════════════════════════════════════════════════════

def fetch_kr_all_data(dart: DartClient, tickers, names, markets, sample_size=None, no_cache=False):
    if sample_size:
        tickers = tickers[:sample_size]
        names   = names[:sample_size]
        markets = markets[:sample_size]

    market_map = dict(zip(tickers, markets))
    name_map   = dict(zip(tickers, names))

    # ★ 핵심: 유니버스 전체를 먼저 기본 DataFrame으로 생성
    # 수집 성공/실패와 무관하게 모든 종목이 결과에 포함됨
    base_df = pd.DataFrame({
        "코드":   tickers,
        "종목명": names,
        "시장":   markets,
    }).set_index("코드")
    base_df.index = base_df.index.astype(str)
    print(f"  [유니버스] 전체 {len(base_df)}개 종목으로 스크리닝 시작")

    # ① 네이버 보조 지표 (병렬)
    df_nv = fetch_naver_bulk(tickers, names, no_cache=no_cache)
    df_nv.index = df_nv.index.astype(str)
    # 네이버 수집 실패한 종목도 base_df 기준으로 유지
    df_nv = base_df.join(df_nv.drop(columns=["종목명"], errors="ignore"), how="left")
    market_s = pd.Series(market_map, name="시장")
    df_nv["시장"] = df_nv.index.map(market_s).fillna("KOSPI")
    print(f"  [네이버] 수집 후 유니버스 유지: {len(df_nv)}개")

    # ② yfinance 가격 (배치)
    df_price = fetch_price_yfinance_batch(tickers, market_map)

    # ③ DART 재무 — 병렬처리 + 캐시
    total_dart = len(tickers)
    cache_key_bulk = f"dart_bulk_{','.join(sorted(tickers[:20]))}"
    cached_dart = None if no_cache else cache_get(cache_key_bulk, tier="A")

    if cached_dart is not None:
        print(f"  [캐시] DART 재무 캐시 적중 → API 호출 생략 (약 {total_dart * 0.08 / 60:.0f}분 절약)")
        dart_rows = cached_dart
    else:
        eta_min = round(total_dart * 0.04 / 60, 1)  # 병렬처리로 절반
        print(f"  [DART] 병렬 재무 수집 중 ({total_dart}개) — 예상 약 {eta_min}분 소요...")
        dart.load_corp_codes()
        dart_rows = {}
        _dart_start = time.time()

        # ── 병렬 처리 (스레드 수: 8 — DART API 과부하 방지 최적값) ──
        DART_WORKERS = 8
        lock = __import__("threading").Lock()
        completed = [0]

        def _fetch_one(code):
            # 개별 종목 캐시 확인 — TIER-A (7일 유효: 분기 재무는 자주 안 바뀜)
            ck = f"dart_{code}"
            cached = cache_get(ck, tier="A")
            if cached is not None:
                return code, cached
            result = dart.fetch_financials_dart(code)
            cache_set(ck, result, tier="A")  # TIER-A: 7일 캐시
            time.sleep(0.05)       # DART API 과부하 방지
            return code, result

        if HAS_TQDM:
            pbar = tqdm(total=total_dart, desc="  DART 병렬", unit="종목",
                        bar_format="{l_bar}{bar}| {n_fmt}/{total_fmt} [{elapsed}<{remaining}]")

        with ThreadPoolExecutor(max_workers=DART_WORKERS) as executor:
            futures = {executor.submit(_fetch_one, code): code for code in tickers}
            for future in as_completed(futures):
                try:
                    code, result = future.result(timeout=30)
                    dart_rows[code] = result
                except Exception as e:
                    code = futures[future]
                    dart_rows[code] = {}
                with lock:
                    completed[0] += 1
                    n = completed[0]
                if HAS_TQDM:
                    pbar.update(1)
                elif n % 100 == 0:
                    elapsed = time.time() - _dart_start
                    remaining = elapsed / n * (total_dart - n)
                    print(f"  [DART] {n}/{total_dart} ({n/total_dart*100:.0f}%) "
                          f"— 경과 {elapsed/60:.1f}분 / 잔여 약 {remaining/60:.1f}분")

        if HAS_TQDM:
            pbar.close()

        elapsed_total = time.time() - _dart_start
        print(f"  [DART] 수집 완료 (소요 {elapsed_total/60:.1f}분, 평균 {elapsed_total/max(total_dart,1):.2f}초/종목)")

        # 전체 결과 캐시 저장 — TIER-A (7일 유효)
        cache_set(cache_key_bulk, dart_rows, tier="A")
        print(f"  [캐시] DART 결과 저장 완료 (TIER-A, 7일 유효)")

    df_dart = pd.DataFrame.from_dict(dart_rows, orient="index")
    df_dart.index.name = "코드"
    df_dart.index = df_dart.index.astype(str)
    cfo_count = df_dart["cfo"].notna().sum() if "cfo" in df_dart.columns else 0
    print(f"  ✓ DART 재무 수집 완료: {cfo_count}개 CFO 수집")

    # ④ 발행주식수 배치 수집 (v28 신규) ──────────────────────────────
    # DART /company.json → istc_totCnt (7일 캐시, 병렬처리)
    # 없으면 시총÷현재가 역산으로 fallback
    shares_map = {}
    try:
        shares_map = dart.get_shares_bulk(tickers, max_workers=30)
        print(f"  [주식수] DART 발행주식수 로드: {len(shares_map)}개")
    except Exception as e:
        print(f"  [주식수] DART 수집 실패, 역산으로 대체: {e}")

    # ⑤ 병합 (left join으로 유니버스 전체 유지)
    df = df_nv.join(df_dart, how="left")
    if not df_price.empty:
        df = df.join(df_price, how="left")
    else:
        for col in ["현재가_YF","52주고가","52주저가","52주위치","52주수익률","6개월수익률"]:
            df[col] = np.nan

    # 현재가 통합 (네이버 우선, YF 폴백)
    if "현재가_NV" in df.columns and "현재가_YF" in df.columns:
        df["현재가"] = df["현재가_NV"].combine_first(df["현재가_YF"])
    elif "현재가_NV" in df.columns:
        df["현재가"] = df["현재가_NV"]
    elif "현재가_YF" in df.columns:
        df["현재가"] = df["현재가_YF"]
    else:
        df["현재가"] = np.nan

    # ── 발행주식수 확정 (v28) ──
    # 우선순위: ① DART istc_totCnt → ② 시총÷현재가 역산
    # index.map() 결과는 ndarray → pd.Series로 명시적 변환
    _smap = shares_map  # {stock_code: int}
    df["shares_dart"] = pd.Series(
        [float(_smap[c]) if c in _smap and _smap[c] else np.nan for c in df.index],
        index=df.index,
        dtype="float64"
    )
    # 역산: 시총(억) × 1억 ÷ 현재가 (float64로 통일)
    cur_s  = pd.to_numeric(df["현재가"], errors="coerce")
    mc_s   = pd.to_numeric(df.get("시가총액(억)", pd.Series(np.nan, index=df.index)), errors="coerce")
    implied = (mc_s * 1e8 / cur_s).where(
        cur_s.notna() & mc_s.notna() & (cur_s > 0) & (mc_s > 0)
    ).round()  # float64 유지 (Int64 NA 혼용 방지)
    # DART 값 우선, 없으면 역산
    df["shares"] = df["shares_dart"].combine_first(implied)  # float64
    dart_ok  = df["shares_dart"].notna().sum()
    total_ok = df["shares"].notna().sum()
    print(f"  [주식수] 확정: DART {dart_ok}개 + 역산 {total_ok - dart_ok}개 = 총 {total_ok}개")

    # 52주위치 NaN → 중립
    df["52주위치"]   = df["52주위치"].fillna(50.0)
    df["52주수익률"] = df["52주수익률"].fillna(0.0)
    df["6개월수익률"] = df["6개월수익률"].fillna(0.0)

    # ROE: DART 우선, 없으면 네이버
    if "roe" in df.columns and "ROE_NV" in df.columns:
        df["ROE"] = df["roe"].combine_first(df["ROE_NV"])
    elif "roe" in df.columns:
        df["ROE"] = df["roe"]
    else:
        df["ROE"] = df.get("ROE_NV", np.nan)

    # ── ROA: DART 계산값 우선, 없으면 네이버 ──
    # DART: roa 컬럼 = net_income / total_assets × 100
    if "roa" in df.columns:
        dart_roa = pd.to_numeric(df["roa"], errors="coerce")
        # DART ROA 유효성 검사: -50~+50% 범위
        dart_roa = dart_roa.where((dart_roa >= -50) & (dart_roa <= 50))
    else:
        dart_roa = pd.Series(np.nan, index=df.index)

    nv_roa = pd.to_numeric(df.get("ROA", pd.Series(np.nan, index=df.index)), errors="coerce")
    nv_roa = nv_roa.where((nv_roa >= -50) & (nv_roa <= 50))

    df["ROA"] = dart_roa.combine_first(nv_roa)

    # DART에도 ROA 없으면 net_income / total_assets로 직접 계산
    if "net_income" in df.columns and "total_assets" in df.columns:
        mask_no_roa = df["ROA"].isna()
        ni  = pd.to_numeric(df.loc[mask_no_roa, "net_income"],   errors="coerce")
        ta  = pd.to_numeric(df.loc[mask_no_roa, "total_assets"], errors="coerce")
        calc_roa = (ni / ta * 100).where((ta > 0) & ta.notna() & ni.notna())
        calc_roa = calc_roa.where((calc_roa >= -50) & (calc_roa <= 50))
        df.loc[mask_no_roa, "ROA"] = calc_roa

    # 부채비율: DART 우선, 없으면 네이버
    if "debt_ratio" in df.columns and "부채비율_NV" in df.columns:
        df["부채비율"] = df["debt_ratio"].combine_first(df["부채비율_NV"])
    elif "debt_ratio" in df.columns:
        df["부채비율"] = df["debt_ratio"]
    else:
        df["부채비율"] = df.get("부채비율_NV", np.nan)

    # PCR = 시가총액 / CFO
    # 시가총액: 네이버 수집값 우선, 없으면 yfinance market_cap (원→억 변환)
    if "시가총액(억)" not in df.columns:
        df["시가총액(억)"] = np.nan

    # yfinance 배치에서 시가총액 보완 (네이버 누락 종목)
    nan_mc_codes = df[df["시가총액(억)"].isna()].index.tolist()
    if nan_mc_codes and HAS_YFINANCE:
        print(f"  [시가총액] yfinance 보완 수집 중 ({len(nan_mc_codes)}개 누락)...")
        def _fetch_marcap_yf(code, market):
            try:
                suffix = ".KS" if market == "KOSPI" else ".KQ"
                info   = yf.Ticker(code + suffix).fast_info
                mc     = getattr(info, "market_cap", np.nan)
                if mc and not np.isnan(mc) and mc > 0:
                    return code, round(mc / 1e8, 0)
            except Exception:
                pass
            return code, np.nan

        with ThreadPoolExecutor(max_workers=10) as ex:
            futs = {ex.submit(_fetch_marcap_yf, c,
                              market_map.get(c,"KOSPI")): c for c in nan_mc_codes}
            for fut in as_completed(futs):
                code, mc = fut.result()
                if mc and not np.isnan(mc):
                    df.loc[code, "시가총액(억)"] = mc
        filled = df.loc[nan_mc_codes, "시가총액(억)"].notna().sum()
        print(f"  ✓ 시가총액 보완: {filled}/{len(nan_mc_codes)}개")
    elif nan_mc_codes and not HAS_YFINANCE:
        print(f"  [시가총액] yfinance 미설치 → 시가총액 보완 스킵 ({len(nan_mc_codes)}개 NaN 유지)")

    # _market_cap_원 컬럼 추가 (엑셀 _marcap 함수용)
    df["_market_cap_원"] = pd.to_numeric(df["시가총액(억)"], errors="coerce") * 1e8

    df["PCR"] = np.nan
    if "시가총액(억)" in df.columns and "cfo" in df.columns:
        mask = df["시가총액(억)"].notna() & df["cfo"].notna() & (df["cfo"] > 0)
        if mask.any():
            pcr_vals = (df.loc[mask, "시가총액(억)"].astype(float) * 1e8) / df.loc[mask, "cfo"].astype(float)
            df["PCR"] = df["PCR"].astype(float)
            df.loc[mask, "PCR"] = pcr_vals

    return df

# ══════════════════════════════════════════════════════════
# 7. 미국 데이터 수집 (yfinance)
# ══════════════════════════════════════════════════════════

def get_us_universe():
    print("  [US] S&P500 종목 수집 중...")
    try:
        tables = pd.read_html("https://en.wikipedia.org/wiki/List_of_S%26P_500_companies")
        sp500  = tables[0][["Symbol","Security","GICS Sector"]].copy()
        sp500.columns = ["ticker","name","sector"]
        sp500["ticker"] = sp500["ticker"].str.replace(".","–", regex=False)
        print(f"  [US] S&P500: {len(sp500)}개")
        return sp500
    except Exception:
        fallback = [
            ("AAPL","Apple","Technology"),("MSFT","Microsoft","Technology"),
            ("GOOGL","Alphabet","Technology"),("AMZN","Amazon","Consumer"),
            ("NVDA","NVIDIA","Technology"),("JPM","JPMorgan","Financial"),
            ("JNJ","J&J","Healthcare"),("PG","P&G","Consumer"),
            ("KO","Coca-Cola","Consumer"),("XOM","Exxon","Energy"),
        ]
        return pd.DataFrame(fallback, columns=["ticker","name","sector"])

def fetch_us_stock(ticker: str) -> dict:
    r = {k: np.nan for k in [
        "cfo","net_income","debt_ratio","interest_coverage","roe",
        "gross_profit","total_assets","gpa","PBR","PCR","DIV",
        "6개월수익률","52주위치","52주수익률","현재가","ROE","부채비율"
    ]}
    try:
        tk  = yf.Ticker(ticker)
        cf  = tk.cashflow
        bs  = tk.balance_sheet
        inc = tk.income_stmt

        def get_v(df, keys):
            if df is None or df.empty: return np.nan
            for k in keys:
                matches = [i for i in df.index if k.lower() in i.lower()]
                if matches:
                    try: return float(df.loc[matches[0]].iloc[0])
                    except: pass
            return np.nan

        r["cfo"]          = get_v(cf,  ["Operating Cash Flow","Cash From Operations"])
        r["net_income"]   = get_v(inc, ["Net Income"])
        r["gross_profit"] = get_v(inc, ["Gross Profit"])
        op_income         = get_v(inc, ["EBIT","Operating Income"])
        interest_exp      = get_v(inc, ["Interest Expense"])
        total_liab        = get_v(bs,  ["Total Liabilities Net Minority Interest","Total Liabilities"])
        equity            = get_v(bs,  ["Stockholders Equity","Total Equity"])
        r["total_assets"] = get_v(bs,  ["Total Assets"])

        if not np.isnan(total_liab) and not np.isnan(equity) and equity > 0:
            r["debt_ratio"] = (total_liab / equity) * 100
            r["부채비율"]   = r["debt_ratio"]
        if not np.isnan(op_income) and not np.isnan(interest_exp) and interest_exp < 0:
            r["interest_coverage"] = op_income / abs(interest_exp)
        if not np.isnan(r["net_income"]) and not np.isnan(equity) and equity > 0:
            r["roe"] = (r["net_income"] / equity) * 100
            r["ROE"] = r["roe"]
        if not np.isnan(r["gross_profit"]) and not np.isnan(r["total_assets"]) and r["total_assets"]>0:
            r["gpa"] = r["gross_profit"] / r["total_assets"]

        info      = tk.fast_info
        cur_price = getattr(info, "last_price", np.nan)
        r["현재가"] = cur_price
        r["PBR"]   = getattr(info, "price_to_book", np.nan)
        mkt_cap    = getattr(info, "market_cap", np.nan)
        if not np.isnan(mkt_cap) and not np.isnan(r["cfo"]) and r["cfo"] > 0:
            r["PCR"] = mkt_cap / r["cfo"]

        end  = datetime.today()
        hist = tk.history(start=(end-timedelta(days=400)).strftime("%Y-%m-%d"),
                          end=end.strftime("%Y-%m-%d"))
        if not hist.empty and len(hist) >= 20:
            p_now = hist["Close"].iloc[-1]
            p_6m  = hist["Close"].iloc[max(0, len(hist)-126)]
            hi52  = hist["Close"].tail(252).max()
            lo52  = hist["Close"].tail(252).min()
            r["6개월수익률"] = safe_div(p_now - p_6m, p_6m) * 100
            r["52주위치"]    = safe_div(p_now - lo52, hi52 - lo52) * 100
            r["52주수익률"]  = safe_div(p_now - hist["Close"].iloc[0], hist["Close"].iloc[0]) * 100
            r["현재가"]      = p_now

        divs = tk.dividends
        if not divs.empty:
            ann_div = divs[divs.index >= (datetime.today()-timedelta(days=365))].sum()
            if not np.isnan(cur_price) and cur_price > 0:
                r["DIV"] = ann_div / cur_price * 100

    except Exception:
        pass
    return r

def fetch_us_all_data(sample_size=None) -> pd.DataFrame:
    universe = get_us_universe()
    if sample_size:
        universe = universe.head(sample_size)

    rows  = []
    total = len(universe)
    print(f"  [US] 데이터 수집 시작 ({total}개)...")

    if HAS_TQDM:
        pbar = tqdm(total=total, desc="  US 수집", unit="종목",
                    bar_format="{l_bar}{bar}| {n_fmt}/{total_fmt} [{elapsed}<{remaining}]")

    for i, (_, row) in enumerate(universe.iterrows()):
        data = fetch_us_stock(row["ticker"])
        data["코드"]    = row["ticker"]
        data["종목명"]  = row["name"]
        data["시장"]    = "US"
        data["업종"]    = row.get("sector","")
        rows.append(data)
        if HAS_TQDM:
            pbar.update(1)
        elif i > 0 and i % 50 == 0:
            print(f"  [US] {i}/{total} ({i/total*100:.0f}%) 수집 중...")
        time.sleep(0.05)

    if HAS_TQDM:
        pbar.close()

    df = pd.DataFrame(rows).set_index("코드")
    df["52주위치"]   = df["52주위치"].fillna(50.0)
    df["52주수익률"] = df["52주수익률"].fillna(0.0)
    df["6개월수익률"] = df["6개월수익률"].fillna(0.0)
    df["DIV"]        = df["DIV"].fillna(0.0)
    print(f"  ✓ US 수집 완료: {len(df)}개")
    return df

# ══════════════════════════════════════════════════════════
# 8-A. 100점 만점 기업 종합 분석 (Ssun 평가표 기반)
# ══════════════════════════════════════════════════════════
#
#  [배점 구조]
#   수익성       20점  (영업이익률5 ROE5 ROA5 ROIC5)
#   성장성       20점  (매출성장5 영업이익성장5 EPS성장5 신규성장동력5)
#   재무안정성   20점  (부채비율5 유동비율5 이자보상5 자본잠식5)
#   현금흐름     20점  (CFO연속5 FCF연속5 CFO>순이익5 현금증가5)
#   밸류에이션   10점  (PER4 PBR3 PEG3)
#   경영/주주친화 10점  (배당3 자사주3 횡령배임2 내부자매수2)
#   ────────────────
#   정성 보정    ±10점 (시장지배력+5 성장산업+5 소송위험-5 CB남발-5)
#
# ══════════════════════════════════════════════════════════

def compute_100pt_score(row: dict) -> dict:
    """
    100점 만점 기업 종합 평가 점수 계산.
    가능한 항목만 채점, 데이터 없는 항목은 업종 평균 추정 또는 부분점수 부여.
    반환: {항목별 점수, 합계, 등급, 투자의견}
    """
    def _v(key, default=None):
        """row에서 값 안전 추출"""
        v = row.get(key)
        if v is None or (isinstance(v, float) and np.isnan(v)):
            return default
        try:
            return float(v)
        except Exception:
            return default

    scores = {}

    # ────────────────────────────────────────────
    # 1. 수익성 (20점)
    # ────────────────────────────────────────────
    # 영업이익률 ≥15% → 5점, ≥10%→4, ≥5%→3, ≥0%→1, 적자→0
    opm = _v("영업이익률")
    if opm is None:
        # 영업이익/매출 역산
        oi = _v("op_income") or _v("영업이익_NV")
        rev = _v("매출액")
        if oi and rev and rev > 0:
            opm = oi / rev * 100
    if opm is not None:
        if   opm >= 15: s_opm = 5
        elif opm >= 10: s_opm = 4
        elif opm >=  5: s_opm = 3
        elif opm >=  0: s_opm = 1
        else:           s_opm = 0
    else:
        s_opm = 2  # 데이터 없음 → 중립

    # ROE ≥15%→5, ≥10%→4, ≥8%→3, ≥5%→1, <5%→0
    roe = _v("ROE") or _v("roe") or _v("ROE_NV")
    if roe is not None:
        if   roe >= 15: s_roe = 5
        elif roe >= 10: s_roe = 4
        elif roe >=  8: s_roe = 3
        elif roe >=  5: s_roe = 1
        else:           s_roe = 0
    else:
        s_roe = 2

    # ROA ≥7%→5, ≥5%→4, ≥3%→3, ≥1%→1, <1%→0
    roa = _v("ROA")
    if roa is not None:
        if   roa >= 7: s_roa = 5
        elif roa >= 5: s_roa = 4
        elif roa >= 3: s_roa = 3
        elif roa >= 1: s_roa = 1
        else:          s_roa = 0
    else:
        # ROA = ROE × 자기자본/총자산 근사 (ROE × (1-부채비율/100)/2)
        debt_r = _v("부채비율") or _v("debt_ratio")
        if roe and debt_r:
            eq_ratio = 1 / (1 + max(debt_r, 0) / 100)
            roa_est = roe * eq_ratio
            if   roa_est >= 7: s_roa = 5
            elif roa_est >= 5: s_roa = 4
            elif roa_est >= 3: s_roa = 3
            else:              s_roa = 2
        else:
            s_roa = 2

    # ROIC ≥10%→5 (ROE 대용: 데이터 없으면 ROE로 근사)
    # 실제 ROIC = NOPAT / 투하자본, 여기선 ROE*0.8 근사
    if roe is not None:
        roic_est = roe * 0.8
        if   roic_est >= 10: s_roic = 5
        elif roic_est >=  7: s_roic = 4
        elif roic_est >=  5: s_roic = 3
        elif roic_est >=  3: s_roic = 1
        else:                s_roic = 0
    else:
        s_roic = 2

    scores["수익성"] = s_opm + s_roe + s_roa + s_roic  # max 20

    # ────────────────────────────────────────────
    # 2. 성장성 (20점)
    # ────────────────────────────────────────────
    rev_gr = _v("매출성장률")
    if rev_gr is not None:
        if   rev_gr >= 20: s_rev_g = 5
        elif rev_gr >= 10: s_rev_g = 4
        elif rev_gr >=  5: s_rev_g = 3
        elif rev_gr >=  0: s_rev_g = 1
        else:              s_rev_g = 0
    else:
        s_rev_g = 2  # 중립

    # 영업이익 성장 (네이버 영업이익 증가율 근사: op_income vs rev_gr 대용)
    # 직접 데이터 없으므로 매출성장 × 1.2 (영업레버리지 근사) 사용
    if rev_gr is not None:
        oi_gr_est = rev_gr * 1.2
        if   oi_gr_est >= 20: s_oi_g = 5
        elif oi_gr_est >= 10: s_oi_g = 4
        elif oi_gr_est >=  5: s_oi_g = 3
        elif oi_gr_est >=  0: s_oi_g = 1
        else:                 s_oi_g = 0
    else:
        s_oi_g = 2

    # EPS 성장 (ROE 변화 없으면 매출성장과 동행 가정)
    eps = _v("EPS")
    per = _v("PER") or _v("per")
    cur_price = _v("현재가") or _v("현재가_NV") or _v("현재가_YF")
    # EPS 직접 성장률 데이터 없음 → ROE로 내부유보 성장율 추정
    # g = ROE × (1 - 배당성향)
    payout = _v("dividend_payout")
    if roe and roe > 0:
        retention = 1.0 - min(max((payout or 40) / 100, 0), 1)
        eps_growth_est = roe * retention
        if   eps_growth_est >= 15: s_eps_g = 5
        elif eps_growth_est >= 10: s_eps_g = 4
        elif eps_growth_est >=  7: s_eps_g = 3
        elif eps_growth_est >=  3: s_eps_g = 1
        else:                      s_eps_g = 0
    else:
        s_eps_g = 2

    # 신규 성장동력: PEG < 1 이거나 매출성장+모멘텀이 좋으면 인정
    # PEG = PER / EPS성장률 (EPS성장률 근사)
    if per and per > 0 and rev_gr and rev_gr > 0:
        peg_est = per / rev_gr
        s_growth_new = 5 if peg_est < 1 else (4 if peg_est < 1.5 else (3 if peg_est < 2 else (1 if peg_est < 3 else 0)))
    elif rev_gr and rev_gr >= 10:
        s_growth_new = 4  # 고성장 자체가 성장동력 있음을 시사
    else:
        s_growth_new = 2

    scores["성장성"] = s_rev_g + s_oi_g + s_eps_g + s_growth_new  # max 20

    # ────────────────────────────────────────────
    # 3. 재무안정성 (20점)
    # ────────────────────────────────────────────
    debt = _v("부채비율") or _v("debt_ratio")
    if debt is not None:
        if   debt <= 50:  s_debt = 5
        elif debt <= 100: s_debt = 4
        elif debt <= 150: s_debt = 3
        elif debt <= 200: s_debt = 2
        elif debt <= 300: s_debt = 1
        else:             s_debt = 0
    else:
        s_debt = 2

    cur_ratio = _v("current_ratio")
    if cur_ratio is not None:
        if   cur_ratio >= 200: s_cur = 5
        elif cur_ratio >= 150: s_cur = 4
        elif cur_ratio >= 120: s_cur = 3
        elif cur_ratio >= 100: s_cur = 1
        else:                  s_cur = 0
    else:
        s_cur = 2

    ic = _v("interest_coverage")
    if ic is not None:
        if   ic >= 10: s_ic = 5
        elif ic >=  5: s_ic = 4
        elif ic >=  3: s_ic = 3
        elif ic >=  1: s_ic = 1
        else:          s_ic = 0
    else:
        s_ic = 3  # 이자비용 없는 순현금 기업 가능성

    # 자본잠식 여부: equity > 0 이면 잠식 없음
    equity = _v("equity")
    if equity is not None:
        s_no_impair = 5 if equity > 0 else 0
    else:
        s_no_impair = 4  # 모름 → 대부분 잠식 없음

    scores["재무안정성"] = s_debt + s_cur + s_ic + s_no_impair  # max 20

    # ────────────────────────────────────────────
    # 4. 현금흐름 (20점)
    # ────────────────────────────────────────────
    cfo = _v("cfo")
    ni  = _v("net_income")
    fcf = _v("fcf")

    # CFO 연속 플러스 (단일 연도 데이터 → 현재 연도만 판단)
    if cfo is not None:
        s_cfo_pos = 5 if cfo > 0 else 0
    else:
        s_cfo_pos = 2

    # FCF 연속 플러스
    if fcf is not None:
        s_fcf_pos = 5 if fcf > 0 else (2 if fcf >= -abs(cfo or 1) * 0.3 else 0)
    else:
        s_fcf_pos = 2

    # CFO > 순이익 (이익 품질)
    if cfo is not None and ni is not None:
        s_cfo_ni = 5 if cfo >= ni else (3 if cfo >= ni * 0.7 else (1 if cfo > 0 else 0))
    else:
        s_cfo_ni = 2

    # 현금성자산 증가 (단일 시점 → 현금/총자산 비율로 대체)
    cash = _v("cash")
    ta   = _v("total_assets")
    if cash is not None and ta and ta > 0:
        cash_ratio = cash / ta * 100
        if   cash_ratio >= 20: s_cash_inc = 5
        elif cash_ratio >= 10: s_cash_inc = 4
        elif cash_ratio >=  5: s_cash_inc = 3
        else:                  s_cash_inc = 1
    else:
        s_cash_inc = 2

    scores["현금흐름"] = s_cfo_pos + s_fcf_pos + s_cfo_ni + s_cash_inc  # max 20

    # ────────────────────────────────────────────
    # 5. 밸류에이션 (10점)
    # ────────────────────────────────────────────
    # PER (업종 평균 대비) → 한국 상장사 평균 ~12x, 저평가 기준 10x 이하
    per_ref = _v("PER_결산") or per
    if per_ref is not None and per_ref > 0:
        if   per_ref <= 8:  s_per = 4
        elif per_ref <= 12: s_per = 3
        elif per_ref <= 15: s_per = 2
        elif per_ref <= 20: s_per = 1
        else:               s_per = 0
    else:
        s_per = 1  # 적자/음수 PER

    # PBR (업종 평균 대비)
    pbr = _v("PBR") or _v("pbr")
    if pbr is not None and pbr > 0:
        if   pbr <= 0.5: s_pbr = 3
        elif pbr <= 1.0: s_pbr = 3
        elif pbr <= 1.5: s_pbr = 2
        elif pbr <= 2.0: s_pbr = 1
        else:            s_pbr = 0
    else:
        s_pbr = 1

    # PEG (PER / EPS성장률)
    if per_ref and per_ref > 0 and rev_gr and rev_gr > 0:
        peg = per_ref / max(rev_gr, 1)
        if   peg < 0.5: s_peg = 3
        elif peg < 1.0: s_peg = 3
        elif peg < 1.5: s_peg = 2
        elif peg < 2.0: s_peg = 1
        else:           s_peg = 0
    else:
        s_peg = 1

    scores["밸류에이션"] = s_per + s_pbr + s_peg  # max 10

    # ────────────────────────────────────────────
    # 6. 경영 및 주주친화 (10점)
    # ────────────────────────────────────────────
    div_yield = _v("DIV") or 0

    # 배당: 꾸준함 → DIV수익률로 대용
    if   div_yield >= 3: s_div_mgmt = 3
    elif div_yield >= 2: s_div_mgmt = 3
    elif div_yield >= 1: s_div_mgmt = 2
    elif div_yield >  0: s_div_mgmt = 1
    else:                s_div_mgmt = 0

    # 자사주 매입/소각: 배당성향+배당수익률 조합으로 추정
    # (직접 데이터 없음) → 배당수익률 높으면 주주환원 적극적
    if div_yield >= 2 or (payout and 20 <= (payout or 0) <= 60):
        s_buyback = 3
    elif div_yield >= 1:
        s_buyback = 2
    else:
        s_buyback = 1

    # 횡령배임 없음: Altman + 이자보상 + FCF 복합 판단
    # Altman 단독이 아닌 복합 지표로 재무 건전성 평가
    altman  = _v("altman_z")
    ic_val  = _v("interest_coverage")
    nd      = _v("net_debt_ratio")   # 순부채비율
    ebitda  = _v("ebitda")
    rev     = _v("매출액")

    # 재무건전성 점수: Altman은 보조, 이자보상+FCF+부채를 주로 봄
    ic_ok  = ic_val is not None and ic_val >= 3
    fcf_ok = fcf is not None and fcf > 0
    debt_ok= debt is None or debt < 150

    if ic_ok and fcf_ok and debt_ok:
        s_no_fraud = 2   # 핵심 3개 모두 OK
    elif ic_ok and fcf_ok:
        s_no_fraud = 2   # 이자보상+FCF OK (부채 좀 있어도 실위험 낮음)
    elif ic_ok or fcf_ok:
        s_no_fraud = 1   # 둘 중 하나만 OK
    else:
        s_no_fraud = 0   # 둘 다 문제

    # 내부자 매수: FCF Yield 기반
    mkcap_100 = _v("시가총액(억)")
    if fcf and mkcap_100 and mkcap_100 > 0:
        fy = fcf / (mkcap_100 * 1e8) * 100
        s_insider = 2 if fy >= 5 else (1 if fy >= 2 else 0)
    else:
        s_insider = 1

    scores["경영주주친화"] = s_div_mgmt + s_buyback + s_no_fraud + s_insider  # max 10

    # ────────────────────────────────────────────
    # 7. 정성 보정 (±10점) — 정교화 버전
    # ────────────────────────────────────────────
    bonus = 0
    malus = 0
    bonus_notes = []
    malus_notes = []

    # ── 가산점 요인 ──

    # +5점: 독점적 해자 (영업이익률 25%↑ + ROE 20%↑)
    if opm and opm >= 25 and roe and roe >= 20:
        bonus += 5
        bonus_notes.append(f"독점적 해자(OPM {opm:.0f}%·ROE {roe:.0f}%)")
    elif opm and opm >= 20:
        bonus += 3
        bonus_notes.append(f"높은 수익성(OPM {opm:.0f}%)")
    elif opm and opm >= 15 and roe and roe >= 15:
        bonus += 2
        bonus_notes.append(f"양호한 수익성(OPM {opm:.0f}%·ROE {roe:.0f}%)")

    # +5점: 성장산업
    sector = str(row.get("업종", "") or "")
    growth_keywords = ["반도체","전기전자","IT","바이오","제약","에너지","전력","소프트웨어",
                       "인터넷","AI","통신","자동화","2차전지","배터리","방산","우주","로봇"]
    if any(kw in sector for kw in growth_keywords):
        bonus += 5
        bonus_notes.append(f"성장산업({sector})")
    elif rev_gr and rev_gr >= 20:
        bonus += 3
        bonus_notes.append(f"고성장(매출+{rev_gr:.0f}%)")
    elif rev_gr and rev_gr >= 10:
        bonus += 1
        bonus_notes.append(f"안정성장(매출+{rev_gr:.0f}%)")

    # ── 감점 요인 (복합 조건으로 오탐 방지) ──

    # 감점 A: 실질 부도위험
    # Altman 단독 X → Altman + 이자보상 + FCF 복합으로만 감점
    real_danger = False
    if altman is not None and altman < 1.8:
        if ic_val is not None and ic_val < 2:
            # Altman 낮고 이자보상도 낮음 → 실제 위험
            malus += 5
            malus_notes.append(f"부도위험(Z={altman:.1f}·이자보상{ic_val:.1f}배)")
            real_danger = True
        elif ic_val is not None and ic_val < 3 and fcf is not None and fcf < 0:
            # Altman 낮고 FCF도 음수
            malus += 4
            malus_notes.append(f"재무위험(Z={altman:.1f}·FCF음수)")
            real_danger = True
        elif ic_val is None or (ic_val >= 3 and (fcf is None or fcf >= 0)):
            # Altman 낮지만 이자보상OK·FCF OK → 업종 구조적 문제일 가능성
            malus += 1
            malus_notes.append(f"Altman낮음(Z={altman:.1f}, 단 이자보상·FCF 양호)")
    elif altman is not None and altman < 2.5:
        if ic_val is not None and ic_val < 2:
            malus += 2
            malus_notes.append(f"Altman회색+이자보상낮음(Z={altman:.1f}·IC{ic_val:.1f}배)")
        # 이자보상 OK면 회색지대여도 감점 없음

    # 감점 B: FCF 음수 (단독보다 부채와 결합 시 위험)
    if not real_danger:  # 위에서 이미 감점 안 했을 때만
        if fcf is not None and fcf < 0:
            if debt is not None and debt > 200:
                malus += 4
                malus_notes.append(f"FCF음수+고부채({debt:.0f}%)")
            elif debt is not None and debt > 150:
                malus += 2
                malus_notes.append(f"FCF음수+부채주의({debt:.0f}%)")
            elif ic_val is not None and ic_val < 2:
                malus += 3
                malus_notes.append(f"FCF음수+이자보상낮음({ic_val:.1f}배)")
            else:
                malus += 1
                malus_notes.append("FCF음수(단기, 이자보상 OK)")

    # 감점 C: 매출 감소 (성장성 역행)
    if rev_gr is not None and rev_gr < -10:
        malus += 3
        malus_notes.append(f"매출급감({rev_gr:.0f}%)")
    elif rev_gr is not None and rev_gr < 0:
        malus += 1
        malus_notes.append(f"매출감소({rev_gr:.0f}%)")

    # 감점 D: 순부채/EBITDA 과도 (실질 부채 부담)
    if ebitda and ebitda > 0 and nd is not None:
        # 순부채비율을 EBITDA 대용으로 활용
        if nd > 300:
            malus += 2
            malus_notes.append(f"순부채과다({nd:.0f}%)")

    bonus = min(bonus, 10)
    malus = min(malus, 10)
    scores["정성보정"] = bonus - malus

    # ────────────────────────────────────────────
    # 최종 합산
    # ────────────────────────────────────────────
    base = sum(v for k, v in scores.items() if k != "정성보정")
    total = min(max(base + scores["정성보정"], 0), 110)  # 보너스로 최대 110

    # 100점 기준 등급
    if   total >= 90: grade_100 = "★★★★★ 최상급"
    elif total >= 80: grade_100 = "★★★★  우수"
    elif total >= 70: grade_100 = "★★★   양호"
    elif total >= 60: grade_100 = "★★    보통"
    elif total >= 50: grade_100 = "★     주의"
    else:             grade_100 = "      투자재검토"

    # 5대 필수 체크 - 이모지 대신 실제 수치 표시
    # CFO: 실제값(억) + OK/NG
    chk1 = f"{round(cfo/1e8,1):+.1f}억 ✅" if cfo and cfo > 0 else (f"{round(cfo/1e8,1):+.1f}억 ❌" if cfo else "N/A")
    # 자본잠식: 자기자본 표시
    chk2 = f"✅ 정상" if equity and equity > 0 else ("❌ 잠식" if equity is not None else "N/A")
    # 부채비율: 실제 수치 표시
    chk3 = f"{debt:.0f}% ✅" if debt and debt < 150 else (f"{debt:.0f}% ⚠" if debt and debt < 300 else (f"{debt:.0f}% ❌" if debt else "N/A"))
    # ROE: 실제 수치 표시
    chk4 = f"{roe:.1f}% ✅" if roe and roe >= 10 else (f"{roe:.1f}% ⚠" if roe and roe >= 5 else (f"{roe:.1f}% ❌" if roe else "N/A"))
    # PER: 실제 수치 표시
    chk5 = f"{per_ref:.1f}x ✅" if per_ref and per_ref > 0 and per_ref < 15 else (f"{per_ref:.1f}x ⚠" if per_ref and per_ref > 0 and per_ref < 25 else (f"{per_ref:.1f}x ❌" if per_ref and per_ref > 0 else "적자/N/A"))

    bonus_str = ", ".join(bonus_notes) if bonus_notes else "-"
    malus_str = ", ".join(malus_notes) if malus_notes else "-"

    return {
        "100점_수익성":     scores["수익성"],
        "100점_성장성":     scores["성장성"],
        "100점_재무안정":   scores["재무안정성"],
        "100점_현금흐름":   scores["현금흐름"],
        "100점_밸류":       scores["밸류에이션"],
        "100점_주주친화":   scores["경영주주친화"],
        "100점_정성보정":   scores["정성보정"],
        "100점_합계":       total,
        "100점_등급":       grade_100,
        "체크_CFO":         chk1,
        "체크_자본잠식":    chk2,
        "체크_부채비율":    chk3,
        "체크_ROE":         chk4,
        "체크_PER":         chk5,
        "보너스근거":       bonus_str,
        "감점근거":         malus_str,
    }


# ══════════════════════════════════════════════════════════
# 8-B. 매매 시그널 엔진 (현재가 기반)
# ══════════════════════════════════════════════════════════
#
#  [시그널 체계]
#   ■■■ 강력매수    복합점수≥85 & 100점≥80 & 괴리율≥+20%
#   ■■  매수        복합점수≥75 & 100점≥70 & 괴리율≥+10%
#   ■   관심/비중확대 복합점수≥65 & 100점≥60 & 괴리율≥0%
#   ─   보유        조건 애매, 특이사항 없음
#   ▽   비중축소    복합점수<60 또는 괴리율<-10%
#   ▼▼  매도        복합점수<50 또는 괴리율<-20%
#   ▼▼▼ 즉시매도    하드필터 탈락 조건 근접
#
#  [가격 밴드] ── 현재가 기준 매수/손절, 적정가 기준 목표가
#   적극매수선   = 현재가 × 0.95  (지금보다 5% 더 빠지면 추가 매수)
#   매수선       = 현재가          (현재 매수 가능 기준가)
#   1차목표가    = 적정가 평균     (적정가 도달 시 30% 익절)
#   2차목표가    = 적정가 평균 × 1.15 (추가 상승 여력)
#   추격매수경보 = 적정가 평균 × 1.20 (이 가격 이상은 고평가)
#   손절선       = 현재가 × 0.85  (15% 하락 시 손절)
#
# ══════════════════════════════════════════════════════════

def compute_trading_signal(row: dict, score_100: dict) -> dict:
    """매매 시그널과 가격 밴드 계산"""

    def _v(key, default=None):
        v = row.get(key)
        if v is None or (isinstance(v, float) and np.isnan(v)):
            return default
        try:
            return float(v)
        except Exception:
            return default

    cur      = _v("현재가") or _v("현재가_NV") or _v("현재가_YF")
    avg_tgt  = _v("적정가_평균")
    per_tgt  = _v("적정가_PER")
    pbr_tgt  = _v("적정가_PBR")
    grm_tgt  = _v("적정가_Graham")
    # ── 업종 특수 구조 감지 (v30) ──
    # 금융·보험·지주는 일반기업 기준의 부채비율·FCF·Altman이 의미없음
    sector_str = str(row.get("업종", "") or "")
    is_finance_sector = any(kw in sector_str for kw in [
        "보험","은행","증권","카드","금융","지주","저축","캐피탈","리츠","자산운용"
    ])

    upside   = _v("괴리율(%)")
    comp     = _v("복합점수", 0)
    pt100    = score_100.get("100점_합계", 0)
    cfo      = _v("cfo")
    debt     = _v("부채비율") or _v("debt_ratio")
    altman   = _v("altman_z")
    pos52    = _v("52주위치", 50)
    ret6m    = _v("6개월수익률", 0)
    fcf      = _v("fcf")
    div      = _v("DIV", 0)
    ic       = _v("interest_coverage")

    # ── 위험 신호 감지 (3단계 복합 판단) ──
    danger_flags  = []
    caution_flags = []

    # ① CFO 음수 — 보험/금융은 투자CF가 크게 음수이므로 CFO만 본다
    if cfo is not None and cfo < 0:
        if is_finance_sector:
            caution_flags.append("CFO음수(금융업참고용)")  # 위험으로 처리 안 함
        elif debt is not None and debt > 150:
            danger_flags.append("CFO음수+부채주의")
        else:
            caution_flags.append("CFO음수(부채는 양호)")

    # ② 부채비율 — 보험/금융/지주는 구조적 고부채, 예외 처리
    if debt is not None and debt > 300:
        if is_finance_sector:
            caution_flags.append(f"부채비율{debt:.0f}%(금융업정상)")
        else:
            danger_flags.append(f"부채비율과다({debt:.0f}%)")
    elif debt is not None and debt > 200:
        if not is_finance_sector:
            caution_flags.append(f"부채비율높음({debt:.0f}%)")

    # ③ Altman Z — 금융/보험은 Altman 모델 적용 불가
    if altman is not None and not is_finance_sector:
        if altman < 1.8:
            if ic is not None and ic < 2:
                danger_flags.append(f"부도위험(Z={altman:.1f}·IC{ic:.1f}배)")
            elif fcf is not None and fcf < 0:
                danger_flags.append(f"재무위험(Z={altman:.1f}·FCF음수)")
            else:
                caution_flags.append(f"Altman낮음(Z={altman:.1f}, 이자보상·FCF 양호)")
        elif altman < 2.5:
            if ic is not None and ic < 2:
                caution_flags.append(f"Altman회색+IC낮음(Z={altman:.1f})")

    # ④ 이자보상배율 — 보험/금융은 제외
    if ic is not None and not is_finance_sector:
        if ic < 1.0:
            danger_flags.append(f"이자보상위험({ic:.1f}배)")
        elif ic < 2.0:
            caution_flags.append(f"이자보상낮음({ic:.1f}배)")

    # ⑤ FCF 음수+고부채 — 보험은 FCF 개념 다름
    if fcf is not None and fcf < 0 and debt is not None and debt > 200:
        if not is_finance_sector and "CFO음수+부채주의" not in danger_flags:
            danger_flags.append(f"FCF음수+고부채({debt:.0f}%)")

    # ⑥ 매출 급감
    rev_gr_s = _v("매출성장률")
    if rev_gr_s is not None and rev_gr_s < -15:
        danger_flags.append(f"매출급감({rev_gr_s:.0f}%)")
    elif rev_gr_s is not None and rev_gr_s < -5:
        caution_flags.append(f"매출감소({rev_gr_s:.0f}%)")

    # 위험 등급 결정
    if danger_flags:
        danger_level = "🔴 실위험"
        danger_str = "🔴 " + " / ".join(danger_flags)
        if caution_flags:
            danger_str += "  🟡 " + " / ".join(caution_flags)
    elif caution_flags:
        danger_level = "🟡 주의"
        danger_str = "🟡 " + " / ".join(caution_flags)
    else:
        danger_level = "없음"
        danger_str = "없음"

    # ── 가격 밴드 계산 ──
    # 매수선/적극매수선 → 현재가 기준
    # 목표가/추격매수경보 → 적정가 기준
    if cur and cur > 0:
        price_buy_strong = round(cur * 0.95)
        price_buy        = round(cur)
        stop_loss        = round(cur * 0.85)
    else:
        price_buy_strong = price_buy = stop_loss = None

    if avg_tgt and avg_tgt > 0:
        if avg_tgt >= cur:
            # ── 정상: 적정가 > 현재가 → 적정가 기반 목표가 ──
            price_target_1   = round(avg_tgt)
            price_target_2   = round(avg_tgt * 1.15)
            price_overbought = round(avg_tgt * 1.20)
        else:
            # ── v26 신규: 적정가 < 현재가 → 기술적 목표가 사용 ──
            # 현재가가 이미 퀀트 적정가를 초과한 상태
            # 1차목표가 = 현재가 × 1.10  (단기 +10%)
            # 2차목표가 = 현재가 × 1.20  (중기 +20%)
            # 추격매수경보 = 현재가 × 1.30  (과열 기준)
            # (적정가는 별도 컬럼에 유지, 가격밴드는 기술적 기준으로)
            price_target_1   = round(cur * 1.10)
            price_target_2   = round(cur * 1.20)
            price_overbought = round(cur * 1.30)
    else:
        price_target_1 = price_target_2 = price_overbought = None

    # ── 현재가 위치 진단 (적정가 대비) ──
    position_diag = "N/A"
    price_ok = False   # 가격 관점에서 매수 가능한지
    if cur and avg_tgt and avg_tgt > 0:
        ratio = cur / avg_tgt * 100
        if   ratio < 60:
            position_diag = "🟦 적극매수구간 (적정가 대비 40%↑ 저평가)"
            price_ok = True
        elif ratio < 80:
            position_diag = "🟩 매수구간 (적정가 대비 20~40% 저평가)"
            price_ok = True
        elif ratio < 95:
            position_diag = "🟩 매수구간 (적정가 대비 5~20% 저평가)"
            price_ok = True
        elif ratio <= 105:
            position_diag = "⬜ 적정구간 (적정가 근접, 보유)"
            price_ok = False
        elif ratio <= 120:
            position_diag = "🟨 소폭고평가 (비중축소 고려)"
            price_ok = False
        else:
            position_diag = "🟥 고평가구간 (매도 검토)"
            price_ok = False
    elif avg_tgt is None:
        # 적정가 없는 경우 가격 판단 보류
        price_ok = None

    # ── 시그널 결정 ──
    # 핵심 원칙: 1차목표가(적정가) < 현재가 이면 매수 시그널 불가
    signal = "─ 보유"
    signal_color = "FFCC00"

    # 1단계: 위험/하락 시그널 (가격 무관하게 우선 적용)
    if len(danger_flags) >= 2:
        signal = "▼▼▼ 즉시매도/회피"
        signal_color = "C00000"
    elif len(danger_flags) == 1:
        signal = "▼▼ 매도 검토"
        signal_color = "FF4444"
    elif comp < 50 or pt100 < 50:
        signal = "▼▼ 매도 검토"
        signal_color = "FF4444"
    elif upside is not None and upside < -20:
        # 고평가 20% 이상 → 매도
        signal = "▼▼ 매도 검토"
        signal_color = "FF4444"
    elif upside is not None and upside < -10:
        # 고평가 10~20% → 비중축소
        signal = "▽ 비중축소"
        signal_color = "FF9900"

    # 2단계: 매수 시그널 — 반드시 적정가 > 현재가 조건 충족해야 함
    elif price_ok:
        # 적정가가 현재가보다 높을 때만 매수 시그널 가능
        if comp >= 85 and pt100 >= 80 and (upside is None or upside >= 15):
            signal = "■■■ 강력매수"
            signal_color = "0070C0"
        elif comp >= 75 and pt100 >= 70 and (upside is None or upside >= 5):
            signal = "■■ 매수"
            signal_color = "2E75B6"
        elif comp >= 65 and pt100 >= 60:
            signal = "■ 관심/비중확대"
            signal_color = "92D050"
        else:
            signal = "─ 보유"
            signal_color = "FFCC00"
    elif price_ok is None:
        # 적정가 없음 → 점수로만 판단
        if comp >= 85 and pt100 >= 80:
            signal = "■■ 매수"       # 적정가 없으면 강력매수 대신 매수
            signal_color = "2E75B6"
        elif comp >= 75 and pt100 >= 70:
            signal = "■ 관심/비중확대"
            signal_color = "92D050"
        # 나머지는 보유 유지
    else:
        # price_ok=False: 적정가 근접/고평가 → 보유 또는 비중축소
        if upside is not None and upside < -10:
            signal = "▽ 비중축소"
            signal_color = "FF9900"
        # 나머지는 보유 유지

    # 주의 신호만 있을 때: 매수 시그널 한 단계 낮춤
    if caution_flags and not danger_flags:
        if signal == "■■■ 강력매수":
            signal = "■■ 매수"
            signal_color = "2E75B6"
        elif signal == "■■ 매수":
            signal = "■ 관심/비중확대"
            signal_color = "92D050"

    # ── 보유 전략 (현재가/목표가 관계 반영) ──
    hold_strategy = ""
    has_upside = price_target_1 and cur and price_target_1 > cur  # 목표가 > 현재가
    upside_pct = round((price_target_1 - cur) / cur * 100) if (price_target_1 and cur) else None
    tgt_str = f"{price_target_1:,}원" if price_target_1 else "N/A"
    stp_str = f"{stop_loss:,}원" if stop_loss else "-"

    if "강력매수" in signal:
        hold_strategy = f"분할매수 추천 (2~3회 진입) | 1차목표가 {tgt_str}(+{upside_pct}%) | 손절 {stp_str}"
    elif "■■ 매수" in signal:
        hold_strategy = f"매수 진입 | 1차목표가 {tgt_str} 도달 시 30% 익절 | 손절 {stp_str}"
    elif "관심" in signal:
        hold_strategy = f"소량 선취매 후 추가 확인 | 목표 {tgt_str}"
    elif "보유" in signal:
        if has_upside:
            hold_strategy = f"1차목표가({tgt_str}) 도달 시 30% 익절 고려"
        else:
            hold_strategy = f"적정가({tgt_str}) 이미 하회 중 — 추가매수 자제, 손절 {stp_str} 확인"
    elif "비중축소" in signal:
        hold_strategy = f"추가 매수 자제, 기존 보유분 30~50% 정리 | 손절 {stp_str}"
    elif "매도" in signal or "즉시매도" in signal:
        hold_strategy = f"손절선({stp_str}) 이탈 시 즉시 매도"

    # ── 기술적 지표 신호 읽기 (v32) ──
    ta_signal_str = str(row.get("기술신호", "") or "─ 중립")
    ta_summary_str = str(row.get("기술지표요약", "") or "")
    _ta_buy_raw  = row.get("기술매수신호수", 0)
    _ta_sell_raw = row.get("기술매도신호수", 0)
    ta_buy    = int(_ta_buy_raw)  if (_ta_buy_raw  is not None and _ta_buy_raw  == _ta_buy_raw)  else 0
    ta_sell   = int(_ta_sell_raw) if (_ta_sell_raw is not None and _ta_sell_raw == _ta_sell_raw) else 0
    macd_sig  = str(row.get("MACD신호", "중립") or "중립")
    rsi_val   = _v("RSI14")
    rsi_sig   = str(row.get("RSI신호", "중립") or "중립")
    bb_pct    = _v("BB%B")
    bb_sig    = str(row.get("BB신호", "중립") or "중립")
    obv_sig   = str(row.get("OBV신호", "중립") or "중립")

    # ── 핵심 투자 포인트 (기술지표 포함) ──
    inv_points = []
    if div and div >= 3:
        inv_points.append(f"배당수익률 {div:.1f}% (안정적 인컴)")
    if upside and upside >= 20:
        inv_points.append(f"적정가 대비 {upside:.0f}% 저평가")
    if pos52 <= 40:
        inv_points.append(f"52주 저점 근처 ({pos52:.0f}%)")
    if ret6m >= 15:
        inv_points.append(f"6개월 모멘텀 강함 (+{ret6m:.0f}%)")
    if fcf and fcf > 0 and cur:
        mkcap_100 = _v("시가총액(억)")
        if mkcap_100 and mkcap_100 > 0:
            fy = fcf / (mkcap_100 * 1e8) * 100
            if fy >= 5:
                inv_points.append(f"FCF Yield {fy:.1f}% (현금창출력 우수)")
    # 기술적 포인트 추가
    if rsi_val and rsi_val < 30:
        inv_points.append(f"RSI {rsi_val:.0f} 과매도 (반등 기대)")
    elif rsi_val and rsi_val > 70:
        inv_points.append(f"RSI {rsi_val:.0f} 과매수 (조정 주의)")
    if "골든크로스" in macd_sig or macd_sig == "매수":
        inv_points.append("MACD 골든크로스 (상승전환)")
    if bb_pct is not None and bb_pct < 0:
        inv_points.append(f"BB 하단밴드 이탈 (과매도, %B={bb_pct:.0f}%)")
    if ta_buy >= 3:
        inv_points.append(f"기술지표 {ta_buy}/4 매수신호 집중")
    if not inv_points:
        inv_points.append("기본 지표 충족")
    inv_point_str = " / ".join(inv_points)

    # ── 보유 전략 (기술신호 포함) ──
    has_upside = price_target_1 and cur and price_target_1 > cur
    upside_pct = round((price_target_1 - cur) / cur * 100) if (price_target_1 and cur) else None
    tgt_str = f"{price_target_1:,}원" if price_target_1 else "N/A"
    stp_str = f"{stop_loss:,}원" if stop_loss else "-"
    ta_addon = f" | 기술:{ta_signal_str}" if ta_signal_str and "중립" not in ta_signal_str else ""

    if "강력매수" in signal:
        hold_strategy = f"분할매수 추천 (2~3회 진입) | 1차목표가 {tgt_str}(+{upside_pct}%) | 손절 {stp_str}{ta_addon}"
    elif "■■ 매수" in signal:
        hold_strategy = f"매수 진입 | 1차목표가 {tgt_str} 도달 시 30% 익절 | 손절 {stp_str}{ta_addon}"
    elif "관심" in signal:
        hold_strategy = f"소량 선취매 후 추가 확인 | 목표 {tgt_str}{ta_addon}"
    elif "보유" in signal:
        if has_upside:
            hold_strategy = f"1차목표가({tgt_str}) 도달 시 30% 익절 고려{ta_addon}"
        else:
            hold_strategy = f"적정가({tgt_str}) 이미 하회 중 — 추가매수 자제, 손절 {stp_str} 확인{ta_addon}"
    elif "비중축소" in signal:
        hold_strategy = f"추가 매수 자제, 기존 보유분 30~50% 정리 | 손절 {stp_str}{ta_addon}"
    elif "매도" in signal or "즉시매도" in signal:
        hold_strategy = f"손절선({stp_str}) 이탈 시 즉시 매도{ta_addon}"
    else:
        hold_strategy = ""

    return {
        "매매시그널":      signal,
        "현재가위치":      position_diag,
        "적극매수선":      price_buy_strong,
        "매수선":          price_buy,
        "1차목표가":       price_target_1,
        "2차목표가":       price_target_2,
        "추격매수경보":    price_overbought,
        "손절선":          stop_loss,
        "보유전략":        hold_strategy,
        "핵심투자포인트":  inv_point_str,
        "위험신호":        danger_str,
        "_signal_color":   signal_color,
        # 기술신호 (v32)
        "기술신호":        ta_signal_str,
        "기술지표요약":    ta_summary_str,
    }


# ══════════════════════════════════════════════════════════
# 8. 하드 필터 (DART 재무 기준)
# ══════════════════════════════════════════════════════════

def apply_hard_filter(df: pd.DataFrame, args) -> pd.DataFrame:
    """
    하드 필터: 데이터가 있는 종목에만 조건 적용
    NaN(데이터 미수집) 종목은 무조건 통과 → 전체종목 시트에 포함
    금융·보험·지주 업종은 부채비율·이자보상 필터 면제 (v30)
    """
    n = len(df)
    cfo_col  = "cfo"  if "cfo"  in df.columns else None
    ni_col   = "net_income" if "net_income" in df.columns else None
    debt_col = "부채비율" if "부채비율" in df.columns else "debt_ratio"
    ic_col   = "interest_coverage" if "interest_coverage" in df.columns else None

    # 금융/보험 업종 마스크
    FINANCE_KW = ["보험","은행","증권","카드","금융","지주","저축","캐피탈","리츠","자산운용"]
    if "업종" in df.columns:
        is_finance = df["업종"].fillna("").apply(
            lambda s: any(kw in str(s) for kw in FINANCE_KW)
        )
    else:
        is_finance = pd.Series(False, index=df.index)

    mask = pd.Series(True, index=df.index)

    # CFO > 0 (데이터 있을 때만)
    if cfo_col:
        mask &= df[cfo_col].isna() | (df[cfo_col] > 0)
        if ni_col:
            both = df[cfo_col].notna() & df[ni_col].notna()
            mask &= ~both | (df[cfo_col] >= df[ni_col])

    # 부채비율 — 금융/보험 업종 면제
    if debt_col in df.columns:
        debt_filter = df[debt_col].isna() | (df[debt_col] < args.max_debt)
        mask &= debt_filter | is_finance  # 금융업은 통과

    # 이자보상배율 — 금융/보험 업종 면제
    if ic_col and ic_col in df.columns:
        ic_filter = df[ic_col].isna() | (df[ic_col] > args.min_ic)
        mask &= ic_filter | is_finance  # 금융업은 통과

    out = df[mask].copy()
    n_data   = (df[cfo_col].notna()).sum() if cfo_col else 0
    n_failed = n - len(out)
    print(f"  [필터] 하드 필터: {n}개 → {len(out)}개 통과 "
          f"(탈락 {n_failed}개 / DART 수집 {n_data}개)")
    return out

# ══════════════════════════════════════════════════════════
# 9. 7팩터 점수 계산
# ══════════════════════════════════════════════════════════

def compute_scores(df: pd.DataFrame, args) -> pd.DataFrame:
    result = df.copy()

    pbr_col = "PBR" if "PBR" in result.columns else "pbr"
    roe_col = "ROE" if "ROE" in result.columns else "roe"

    # PBR/ROE 조건은 점수에 반영 (필터로 탈락시키지 않음)
    # → 모든 종목이 전체종목 시트에 포함되고, 점수로 순위만 결정
    filtered = result.copy()
    print(f"  [스코어] 점수 계산 대상: {len(filtered)}개 전체 종목")

    if len(filtered) == 0:
        return pd.DataFrame()

    # ── ① 가치 점수 (PBR 30% + ROE 30% + FCF Yield 20% + EV/EBITDA 20%) ──
    def _pbr_s(v):
        if v is None or np.isnan(v): return 0
        if v <= 0.3: return 100
        if v >= 3.0: return 0
        return max(0, (1 - (v-0.3)/2.7) * 100)

    def _roe_s(v):
        if v is None or np.isnan(v): return 0
        if v >= 50: return 100
        if v <= 8:  return 0
        return max(0, (v-8)/42 * 100)

    def _fcf_yield_s(row):
        """FCF Yield = FCF / 시가총액 (%) - 높을수록 저평가"""
        fcf   = _safe(row.get("fcf"))
        mkcap = _safe(row.get("시가총액(억)"))
        if fcf is None or mkcap is None or mkcap <= 0:
            return 50  # 데이터 없으면 중립
        fy = fcf / (mkcap * 1e8) * 100  # 억→원 환산
        if   fy >= 10: return 100
        elif fy >=  8: return 90
        elif fy >=  5: return 75
        elif fy >=  3: return 55
        elif fy >=  0: return 35
        else:          return 10  # FCF 음수

    def _ev_ebitda_s(row):
        """EV/EBITDA - 낮을수록 저평가"""
        mkcap  = _safe(row.get("시가총액(억)"))
        tl     = _safe(row.get("total_liabilities"))
        cash   = _safe(row.get("cash"))
        ebitda = _safe(row.get("ebitda"))
        if mkcap is None or ebitda is None or ebitda <= 0:
            return 50  # 데이터 없으면 중립
        # EV = 시가총액 + 순부채
        debt_net = (tl or 0) - (cash or 0)
        ev = mkcap * 1e8 + debt_net
        ratio = ev / ebitda
        if   ratio <=  6: return 100
        elif ratio <=  8: return 85
        elif ratio <= 12: return 65
        elif ratio <= 16: return 45
        elif ratio <= 20: return 25
        else:             return 0

    filtered["가치점수"] = (
        filtered[pbr_col].apply(_pbr_s) * 0.30 +
        filtered[roe_col].apply(_roe_s) * 0.30 +
        filtered.apply(_fcf_yield_s, axis=1) * 0.20 +
        filtered.apply(_ev_ebitda_s, axis=1) * 0.20
    ).round(1)

    # ── ② 모멘텀 점수 (52주 위치 기반 비선형) ──
    def _mom_s(pos):
        if pos is None or np.isnan(pos): return 0
        pos = float(pos)
        bps = [(0,0),(40,85),(57.5,100),(75,85),(100,0)]
        for i in range(len(bps)-1):
            x0,y0 = bps[i]; x1,y1 = bps[i+1]
            if x0 <= pos <= x1:
                t = (pos-x0)/(x1-x0)
                return y0 + t*(y1-y0)
        return 0

    def _vol_mom_s(row):
        """거래량 모멘텀 점수 (yfinance에서 수집한 경우)"""
        vm = _safe(row.get("vol_ratio"))  # 20일 평균 대비 거래량 비율
        if vm is None:
            return 50  # 중립
        if   vm >= 3.0: return 100
        elif vm >= 2.0: return 85
        elif vm >= 1.5: return 70
        elif vm >= 1.0: return 50
        elif vm >= 0.7: return 30
        else:           return 10

    filtered["모멘텀점수"] = (
        filtered["52주위치"].apply(_mom_s) * 0.70 +
        filtered.apply(_vol_mom_s, axis=1)  * 0.30
    ).round(1)

    # ── ③ 배당 점수 (배당수익률 60% + 배당성향 지속가능성 40%) ──
    def _div_s(row):
        div    = _safe(row.get("DIV"), 0)
        payout = _safe(row.get("dividend_payout"))

        # 배당수익률 점수
        if   div >= 6: yield_s = 100
        elif div >= 4: yield_s = 85
        elif div >= 3: yield_s = 70
        elif div >= 2: yield_s = 55
        elif div >= 1: yield_s = 35
        elif div >  0: yield_s = 20
        else:          yield_s = 0

        # 배당성향 점수 (지속가능성)
        if payout is not None:
            if   20 <= payout <= 50: pay_s = 100  # 최적
            elif payout <= 70:       pay_s = 70
            elif payout <= 80:       pay_s = 40
            elif payout >  80:       pay_s = 10   # 삭감 위험
            else:                    pay_s = 50   # 저배당
        else:
            pay_s = 50  # 데이터 없으면 중립

        return round(yield_s * 0.60 + pay_s * 0.40, 1)

    filtered["배당점수"] = filtered.apply(_div_s, axis=1)

    # ── ④ 품질 점수 (DART 부채비율 우선) ──
    def _quality_s(row):
        debt = _safe(row.get("부채비율") or row.get("debt_ratio"))
        roa  = _safe(row.get("ROA"))
        opm  = _safe(row.get("영업이익률"))
        roe  = _safe(row.get(roe_col), 0)
        pbr  = _safe(row.get(pbr_col), 1)

        # 부채비율 점수 (DART 원본 우선)
        if debt is not None:
            if   debt <= 50:  debt_s = 100
            elif debt <= 100: debt_s = 85
            elif debt <= 150: debt_s = 65
            elif debt <= 200: debt_s = 45
            elif debt <= 300: debt_s = 25
            else:             debt_s = 0
        else:
            debt_s = 85 if pbr <= 1.0 else (65 if pbr <= 1.5 else 45)

        # ROA
        if roa is not None:
            if   roa >= 15: roa_s = 100
            elif roa >= 10: roa_s = 85
            elif roa >= 7:  roa_s = 70
            elif roa >= 5:  roa_s = 55
            elif roa >= 3:  roa_s = 40
            elif roa >= 0:  roa_s = 25
            else:           roa_s = 0
        else:
            roa_s = min(100, max(0, roe * 0.6))

        # 영업이익률
        if opm is not None:
            if   opm >= 20: opm_s = 100
            elif opm >= 15: opm_s = 85
            elif opm >= 10: opm_s = 70
            elif opm >= 5:  opm_s = 50
            elif opm >= 0:  opm_s = 30
            else:           opm_s = 0
        else:
            opm_s = min(100, max(0, roe * 1.5))

        return round(debt_s*0.45 + roa_s*0.35 + opm_s*0.20, 1)

    filtered["품질점수"] = filtered.apply(_quality_s, axis=1)

    # ── ⑤ 성장 점수 (DART 매출 데이터 + 네이버 성장률) ──
    def _growth_s(row):
        rev_gr = _safe(row.get("매출성장률"))
        roe    = _safe(row.get(roe_col), 0)
        per    = _safe(row.get("PER") or row.get("per"), 20)
        div    = _safe(row.get("DIV"), 0)

        if rev_gr is not None:
            if   rev_gr >= 30: rev_s = 100
            elif rev_gr >= 20: rev_s = 85
            elif rev_gr >= 10: rev_s = 70
            elif rev_gr >= 5:  rev_s = 55
            elif rev_gr >= 0:  rev_s = 40
            elif rev_gr >= -5: rev_s = 20
            else:              rev_s = 0
        else:
            proxy = roe * 0.6 + div * 2
            rev_s = 100 if proxy>=15 else (80 if proxy>=10 else (60 if proxy>=6 else (40 if proxy>=3 else 20)))

        eps_proxy = roe / max(per if per else 20, 1) * 100
        if   eps_proxy >= 5: eps_s = 100
        elif eps_proxy >= 3: eps_s = 80
        elif eps_proxy >= 1: eps_s = 60
        elif eps_proxy >= 0: eps_s = 40
        else:                eps_s = 20

        return round(rev_s*0.50 + eps_s*0.50, 1)

    filtered["성장점수"] = filtered.apply(_growth_s, axis=1)

    # ── ⑥ 수익성 점수 ──
    def _prof_s(row):
        opm  = _safe(row.get("영업이익률"))
        netm = _safe(row.get("순이익률"))
        roe  = _safe(row.get(roe_col), 0)
        pbr  = _safe(row.get(pbr_col), 1)
        per  = _safe(row.get("PER") or row.get("per"), 20)

        if opm is not None:
            if   opm>=20: op_s=100
            elif opm>=15: op_s=85
            elif opm>=10: op_s=70
            elif opm>=5:  op_s=50
            elif opm>=0:  op_s=30
            else:         op_s=0
        else:
            proxy = roe / max(pbr, 0.1) * 5
            op_s = 100 if proxy>=20 else (85 if proxy>=15 else (70 if proxy>=10 else (50 if proxy>=5 else 30)))

        if netm is not None:
            if   netm>=15: net_s=100
            elif netm>=10: net_s=85
            elif netm>=7:  net_s=70
            elif netm>=5:  net_s=55
            elif netm>=0:  net_s=35
            else:          net_s=0
        else:
            proxy2 = roe / max(pbr, 0.1) * 5 * 1.2
            net_s = 100 if proxy2>=25 else (85 if proxy2>=20 else (70 if proxy2>=15 else (50 if proxy2>=10 else 30)))

        # FCF 마진 실제값 우선, 없으면 PER 역수 추정
        fcf_margin = _safe(row.get("fcf_margin"))
        if fcf_margin is not None:
            if   fcf_margin >= 20: fcf_s = 100
            elif fcf_margin >= 15: fcf_s = 85
            elif fcf_margin >= 10: fcf_s = 70
            elif fcf_margin >=  5: fcf_s = 55
            elif fcf_margin >=  0: fcf_s = 35
            else:                  fcf_s = 0   # FCF 음수
        else:
            fcf_proxy = 100 / max(per if per else 20, 1)
            fcf_s = 100 if fcf_proxy>=15 else (85 if fcf_proxy>=10 else (70 if fcf_proxy>=5 else (50 if fcf_proxy>=3 else 30)))

        return round(op_s*0.45 + net_s*0.35 + fcf_s*0.20, 1)

    filtered["수익성점수"] = filtered.apply(_prof_s, axis=1)

    # ── ⑦ 안정 점수 ──
    def _stab_s(row):
        debt     = _safe(row.get("부채비율") or row.get("debt_ratio"))
        net_debt = _safe(row.get("net_debt_ratio"))
        cur_r    = _safe(row.get("current_ratio"))
        altman   = _safe(row.get("altman_z"))
        roe      = _safe(row.get(roe_col), 0)
        div      = _safe(row.get("DIV"), 0)
        ic       = _safe(row.get("interest_coverage"))

        # ① 부채 안전성 (순부채비율 우선, 없으면 부채비율)
        ref_debt = net_debt if net_debt is not None else debt
        if ref_debt is not None:
            if   ref_debt <  0:   debt_s = 100  # 순현금 기업
            elif ref_debt <= 30:  debt_s = 90
            elif ref_debt <= 70:  debt_s = 75
            elif ref_debt <= 100: debt_s = 55
            elif ref_debt <= 150: debt_s = 35
            elif ref_debt <= 200: debt_s = 20
            else:                 debt_s = 0
        else:
            debt_s = 50

        # 이자보상배율 보너스
        if ic is not None and ic > 0:
            debt_s = min(100, debt_s + (20 if ic>=10 else 10 if ic>=5 else 0))

        # ② 유동비율 점수
        if cur_r is not None:
            if   cur_r >= 200: cur_s = 100
            elif cur_r >= 150: cur_s = 80
            elif cur_r >= 120: cur_s = 60
            elif cur_r >= 100: cur_s = 40
            else:              cur_s = 10
        else:
            cur_s = 50  # 중립

        # ③ Altman Z-Score
        if altman is not None:
            if   altman >= 3.0: alt_s = 100
            elif altman >= 2.5: alt_s = 80
            elif altman >= 1.8: alt_s = 50  # 회색지대
            else:               alt_s = 10  # 부도 위험
        else:
            # 데이터 없으면 ROE로 대체 추정
            alt_s = 80 if roe>=15 else (60 if roe>=10 else (40 if roe>=5 else 20))

        # ④ 손실 위험 (ROE)
        if   roe>=20: loss_s=100
        elif roe>=15: loss_s=85
        elif roe>=10: loss_s=70
        elif roe>=8:  loss_s=55
        elif roe>=5:  loss_s=35
        else:         loss_s=10

        # ⑤ 배당 안정성
        div_payout = _safe(row.get("dividend_payout"))
        if div_payout is not None:
            # 20~60% 지속 가능, 80% 초과는 삭감 위험
            if   20 <= div_payout <= 60: div_s = 100
            elif div_payout < 20:        div_s = 60   # 주주환원 의지↓
            elif div_payout <= 80:       div_s = 40   # 다소 높음
            else:                        div_s = 10   # 삭감 위험
        else:
            # 배당성향 없으면 배당수익률로 대체
            if   div>=4: div_s=100
            elif div>=2: div_s=80
            elif div>=1: div_s=60
            elif div>0:  div_s=40
            else:        div_s=20

        # 가중합: 부채35% + 유동성15% + Altman20% + ROE15% + 배당15%
        return round(debt_s*0.35 + cur_s*0.15 + alt_s*0.20 + loss_s*0.15 + div_s*0.15, 1)

    filtered["안정점수"] = filtered.apply(_stab_s, axis=1)

    # ── ⑧ 복합 점수 ──
    ws = [args.w_value, args.w_mom, args.w_div, args.w_quality,
          args.w_growth, args.w_prof, args.w_stability]
    total_w = sum(ws)
    if abs(total_w - 100) > 0.1:
        ws = [w / total_w * 100 for w in ws]
    w_v,w_m,w_d,w_q,w_g,w_p,w_s = [w/100 for w in ws]

    filtered["복합점수"] = (
        filtered["가치점수"]   * w_v +
        filtered["모멘텀점수"] * w_m +
        filtered["배당점수"]   * w_d +
        filtered["품질점수"]   * w_q +
        filtered["성장점수"]   * w_g +
        filtered["수익성점수"] * w_p +
        filtered["안정점수"]   * w_s
    ).round(1)

    filtered["등급"] = filtered["복합점수"].apply(lambda s: grade_label(s)[0])

    # ── ⑨ 밸류에이션 적정가 3종 ──
    def _valuation(row):
        roe  = _safe(row.get("ROE") or row.get("roe"))
        cur  = _safe(row.get("현재가") or row.get("현재가_NV") or row.get("현재가_YF"))
        sector = str(row.get("업종", "") or "")
        name   = str(row.get("종목명", "") or "")

        # ── 현재가 정합성 검증 ──
        # 시가총액 ÷ 현재가 = 추정 주식수. 이 값이 비현실적이면 현재가 오류
        # (네이버 현재가 파싱 시 자릿수 밀림 오류 탐지)
        marcap_eok = _safe(row.get("시가총액(억)") or row.get("시총"))  # 억원
        cur_invalid = False
        if marcap_eok and marcap_eok > 0 and cur and cur > 0:
            implied_shares = marcap_eok * 1e8 / cur  # 추정 주식수
            # 정상 상장사 주식수: 약 10만주 ~ 500억주
            if implied_shares < 100_000 or implied_shares > 5e10:
                cur_invalid = True

        if cur is None or cur <= 0 or cur_invalid:
            reason = f"현재가오류(시총{marcap_eok:.0f}억÷현재가 비정상)" if cur_invalid else "현재가데이터없음"
            return pd.Series({
                "적정가_PER": None, "적정가_PBR": None,
                "적정가_Graham": None, "적정가_평균": None,
                "괴리율(%)": None, "적정가_사유": reason,
            })

        # ── PER/PBR 계산: BPS/EPS 직접 사용 (v30 핵심) ──
        # 우선순위:
        #   ① BPS_dart → 현재가÷BPS = PBR(현재가)  [가장 정확]
        #   ② EPS_dart → 현재가÷EPS = PER(현재가)  [가장 정확]
        #   ③ shares×현재가÷equity/ni               [2순위]
        #   ④ 결산 수집값 fallback
        bps_dart  = _safe(row.get("bps_dart"))
        eps_dart  = _safe(row.get("eps_dart"))
        equity_val = _safe(row.get("equity"))
        ni_val     = _safe(row.get("net_income"))
        marcap_eok = _safe(row.get("시가총액(억)") or row.get("시총"))

        # shares 안전 접근 (Int64/NA 방지)
        _shr_raw = row.get("shares")
        shares_val = None
        if _shr_raw is not None and getattr(_shr_raw, '__class__', None) and _shr_raw.__class__.__name__ != 'NAType':
            v = _safe(_shr_raw)
            if v and not (isinstance(v, float) and np.isnan(v)):
                shares_val = float(v)

        per_close = _safe(row.get("PER_결산"))
        pbr_close = _safe(row.get("PBR_결산"))
        per_nv    = _safe(row.get("PER") or row.get("per"))
        pbr_nv    = _safe(row.get("PBR") or row.get("pbr"))

        # ① BPS/EPS 기반 (가장 정확)
        pbr_cur = round(cur / bps_dart, 3) if (bps_dart and bps_dart > 0 and cur > 0) else None
        per_cur = round(cur / eps_dart, 2) if (eps_dart and eps_dart > 0 and cur > 0) else None

        # ② shares × 현재가 ÷ 재무수치 (BPS/EPS 없을 때)
        if pbr_cur is None or per_cur is None:
            mktcap_won = None
            if shares_val and shares_val > 0:
                mktcap_won = shares_val * cur
            elif marcap_eok and marcap_eok > 0:
                mktcap_won = marcap_eok * 1e8
            if mktcap_won and mktcap_won > 0:
                if pbr_cur is None and equity_val and equity_val > 0:
                    pbr_cur = round(mktcap_won / equity_val, 3)
                if per_cur is None and ni_val and ni_val > 0:
                    per_cur = round(mktcap_won / ni_val, 2)

        # ③ fallback: 결산 수집값
        pbr_raw = pbr_cur if (pbr_cur and 0 < pbr_cur < 100) else \
                  (pbr_close if (pbr_close and 0 < pbr_close < 30) else pbr_nv)
        per_raw = per_cur if (per_cur and 0 < per_cur) else \
                  (per_close if (per_close and 0 < per_close < 200) else per_nv)

        # ── 빈 셀 이유 추적 ──
        notes = []
        per = per_raw
        pbr = pbr_raw

        # PER 유효성
        if per is None:
            notes.append("PER없음")
        elif per <= 0:
            notes.append(f"PER음수(적자,{per:.0f})")
            per = None
        elif per > 80:
            # v26: PER>80은 EPS 신뢰불가 (적자직전/일시이익극소) → PBR 기반만 사용
            notes.append(f"PER과대({per:.0f}배→제외,PBR기반만)")
            per = None

        # PBR 유효성
        if pbr is None:
            notes.append("PBR없음")
        elif pbr <= 0:
            notes.append(f"PBR음수")
            pbr = None
        elif pbr > 30:
            notes.append(f"PBR과대({pbr:.1f}배→제외)")
            pbr = None

        # ── 업종별 적정 PER/PBR ──
        SECTOR_PER = {
            "반도체": 18, "전기전자": 16, "IT서비스": 20, "소프트웨어": 22,
            "바이오": 25, "제약": 20, "화장품": 18, "인터넷": 22,
            "2차전지": 20, "배터리": 20, "방산": 18, "우주": 20, "로봇": 22,
            "자동차부품": 10, "자동차": 10, "철강": 8, "화학": 10,
            "건설": 8, "유통": 12, "음식료": 14, "섬유": 10,
            "조선": 12, "해운": 7, "항공": 12,
            "금융": 8, "은행": 7, "보험": 8, "증권": 9, "지주": 8,
            "부동산": 12, "호텔": 14, "레저": 14,
            "에너지": 10, "전력": 10, "가스": 9, "기계": 12,
        }
        base_per = 12
        for kw, v in SECTOR_PER.items():
            if kw in sector:
                base_per = v
                break

        SECTOR_PBR = {
            "반도체": 2.0, "전기전자": 1.5, "IT서비스": 2.5, "소프트웨어": 3.0,
            "바이오": 3.0, "제약": 2.5, "화장품": 2.5, "인터넷": 3.0,
            "2차전지": 2.5, "배터리": 2.5, "방산": 2.0, "로봇": 2.5,
            "자동차부품": 1.0, "자동차": 1.0, "철강": 0.8, "화학": 1.2,
            "건설": 0.9, "유통": 1.2, "음식료": 1.5,
            "조선": 1.0, "해운": 0.9,
            "금융": 0.8, "은행": 0.7, "보험": 0.9, "증권": 1.0, "지주": 0.8,
            "에너지": 1.0, "전력": 1.0, "기계": 1.2,
        }
        base_pbr = 1.2
        for kw, v in SECTOR_PBR.items():
            if kw in sector:
                base_pbr = v
                break

        # ── EPS/BPS 역산 ──
        eps = cur / per if (per and per > 0) else None
        bps = cur / pbr if (pbr and pbr > 0) else None

        # ── PER 이상치 보정 ──
        # PER이 비정상적으로 높은데(>30배) ROE는 건전하면(≥8%)
        # → 일시적 이익 감소(사이클 저점)로 후행 PER 왜곡 → ROE 기반 EPS 재추정
        per_abnormal = False
        if per and per > 30 and roe and roe >= 8 and bps and bps > 0:
            # 정상 EPS ≈ BPS × ROE/100 (지속가능 이익 수준)
            eps_normalized = bps * roe / 100.0
            if eps_normalized > 0:
                eps = eps_normalized   # EPS를 정상화 값으로 교체
                per_abnormal = True
                notes.append(f"PER{per:.0f}배→ROE정상화")

        # ── 적정 PER 계산 (v26 개선) ──
        # [로직]
        # ① PER <= 업종평균×1.1: blended와 업종평균 중 작은값 → 단, per×0.75 하한 보장
        # ② PER > 업종평균×1.1 (프리미엄): per×0.85 상한 → 단, per×0.75 하한 보장
        # ③ per_abnormal(ROE 정상화): 업종평균 적용 (고성장 보정은 아래에서 추가)
        #
        # ★ 핵심: fair_per 하한 = per × 0.75 보장
        #   → 적정가(PER기반) = EPS × fair_per = (cur/per) × fair_per >= cur × 0.75
        #   → 1차목표가(=적정가)가 최소 현재가의 75% 이상 보장
        if per and per > 0 and not per_abnormal:
            blended = per * 0.6 + base_per * 0.4
            if per <= base_per * 1.1:
                fair_per = max(min(blended, base_per), per * 0.75)
            else:
                fair_per = max(min(blended, per * 0.85), per * 0.75)
        else:
            fair_per = base_per

        # ── 고성장 PER 보정 (v24) ──
        # ROE 정상화 EPS를 쓰는 경우(per_abnormal=True), ROE >15%이면
        # "고성장 프리미엄" → fair_per을 ROE 비례 상향 (업종평균 최대 2.5배 상한)
        if per_abnormal and roe and roe > 15:
            _roe_for_growth = min(roe, 25.0)
            roe_growth_bonus = min((_roe_for_growth - 15) * 0.5, base_per * 1.0)
            fair_per = min(fair_per + roe_growth_bonus, base_per * 2.5)
            notes.append(f"ROE{roe:.0f}%성장보정→fair_per{fair_per:.1f}")

        # ── 시장 프리미엄 과다 종목 감지 ──
        is_premium_stock = bool(pbr and pbr > 5 and per and per > 40)
        if is_premium_stock:
            notes.append("⚠고성장프리미엄(퀀트적정가참고용)")

        # ── 적정 PBR 계산 (v26 개선) ──
        # [케이스별 처리]
        # ① 고PBR(>3): BPS 기반 적정가 건너뜀 → PER/Graham 기반만 사용
        # ② 중PBR(1~3): fair_pbr = min(pbr×0.8, ROE혼합) → 하한 pbr×0.75 보장
        # ③ 저PBR(<1):
        #    - ROE>=5%: 기존 ROE 혼합 로직 → 하한 pbr×0.75 보장
        #    - ROE<5% (저수익/적자): fair_pbr = min(pbr, base_pbr) 직접 사용
        #      (현재 PBR을 적정하다 보고 업종평균 이하면 유지)
        #
        # ★ 핵심: fair_pbr 하한 = pbr × 0.75 보장 (ROE>=5% 구간)
        #   → 적정가(PBR기반) = BPS × fair_pbr = (cur/pbr) × fair_pbr >= cur × 0.75
        roe_capped = min(roe, 25.0) if (roe and roe > 0) else None
        roe_pbr = (roe_capped / 10.0) if roe_capped else 0.0

        skip_pbr_target = False
        if pbr and pbr > 0:
            if pbr > 3.0:
                skip_pbr_target = True
                notes.append(f"PBR{pbr:.1f}배(고PBR→PBR적정가제외)")
            elif pbr > 1.0:
                # 중PBR: ROE 혼합, 하한 pbr×0.75 보장
                raw_fair = min(pbr * 0.8, max(pbr * 0.5 + roe_pbr * 0.5, 0.5))
                fair_pbr = max(min(raw_fair, pbr), pbr * 0.75)
            else:
                # 저PBR
                if roe_capped is not None and roe_capped >= 5:
                    # 정상 수익성: ROE 혼합, 하한 pbr×0.75 보장
                    raw_fair = min(max(pbr * 0.5 + roe_pbr * 0.5, 0.3), base_pbr)
                    fair_pbr = max(raw_fair, pbr * 0.75)
                else:
                    # 저수익/적자(ROE<5%): 현재 PBR을 적정으로 인정, 업종평균 이하면 그대로
                    fair_pbr = min(pbr, base_pbr)
        else:
            fair_pbr = max(min(roe_pbr if roe_pbr > 0 else 0.5, base_pbr), 0.3)

        # ── 적정가 3종 계산 ──
        per_target    = round(eps * fair_per)  if (eps and eps > 0) else None
        pbr_target    = None if skip_pbr_target else (round(bps * fair_pbr) if (bps and bps > 0) else None)
        graham_target = None
        roe_for_graham = roe_capped or roe
        if eps and bps and eps > 0 and bps > 0 and roe_for_graham and roe_for_graham >= 8:
            graham_target = round((22.5 * eps * bps) ** 0.5)

        # ── 개별 필터: 현재가 대비 범위 체크 ──
        # 상한: PBR 낮을수록 적정가가 높게 나오는 게 정상 (저PBR 가치주)
        # 하한: 고PBR 성장주는 PER/Graham 기반이므로 하한 완화
        filtered_out = []
        upper_limit = 5.0 if (pbr and pbr < 0.5) else (4.0 if (pbr and pbr < 1.0) else 3.0)
        # 고PBR(>3)은 PER/Graham 기반만 쓰므로 하한을 넉넉하게 설정
        lower_limit = 0.10 if (pbr and pbr > 5.0) else (0.15 if (pbr and pbr > 3.0) else (0.20 if (pbr and pbr > 2.0) else 0.25))
        def _check(t, label):
            if t is None: return None
            if cur <= 0: return None
            ratio = t / cur
            if ratio < lower_limit:
                filtered_out.append(f"{label}필터(<{lower_limit}배,{ratio:.2f}배)")
                return None
            if ratio > upper_limit:
                filtered_out.append(f"{label}필터(>{upper_limit}배,{ratio:.1f}배)")
                return None
            return t

        per_t_f    = _check(per_target,    "PER")
        pbr_t_f    = _check(pbr_target,    "PBR")
        graham_t_f = _check(graham_target, "Graham")

        # ── 중앙값 기반 이상치 제거 후 가중 평균 ──
        # 3종 적정가가 서로 크게 어긋나면(데이터 오류 가능성)
        # 중앙값에서 ±50% 벗어난 값을 이상치로 제거
        raw_targets = [(per_t_f, 0.5, "PER"), (pbr_t_f, 0.35, "PBR"), (graham_t_f, 0.15, "Graham")]
        valid_vals = [(t, w, lbl) for t, w, lbl in raw_targets if t is not None]

        outlier_removed = []
        if len(valid_vals) >= 2:
            vals_only = sorted([t for t, _, _ in valid_vals])
            n = len(vals_only)
            median = vals_only[n//2] if n % 2 else (vals_only[n//2-1] + vals_only[n//2]) / 2
            # 중앙값 ±50% 범위만 채택
            kept = []
            for t, w, lbl in valid_vals:
                if 0.5 * median <= t <= 1.5 * median:
                    kept.append((t, w))
                else:
                    outlier_removed.append(f"{lbl}이상치({t:,}≠중앙값{round(median):,})")
            weighted = kept if kept else [(t, w) for t, w, _ in valid_vals]
        else:
            weighted = [(t, w) for t, w, _ in valid_vals]

        if not weighted:
            avg_target = None
            if not notes and not filtered_out:
                notes.append("계산불가(데이터부족)")
            if filtered_out:
                notes.extend(filtered_out)
            reason = " / ".join(notes) if notes else "알수없음"
        else:
            tw = sum(w for _, w in weighted)
            avg_target = round(sum(t * w for t, w in weighted) / tw)
            note_parts = []
            if notes: note_parts.extend(notes)
            if outlier_removed: note_parts.extend(outlier_removed)
            reason = "정상" if not note_parts else ("(" + ",".join(note_parts) + ")")

        # 사용된 PER/PBR 원본값을 사유에 항상 부기 (검증용)
        per_str = f"PER{per:.1f}" if per else "PER-"
        pbr_str = f"PBR{pbr:.2f}" if pbr else "PBR-"
        reason = f"[{per_str}/{pbr_str}] {reason}"

        # ── 괴리율 ──
        upside = None
        if avg_target and cur > 0:
            upside = round((avg_target - cur) / cur * 100, 1)
            # 저PBR 가치주는 괴리율 300%까지 정상, 일반 300% → 200%
            upside_max = 300 if (pbr and pbr < 0.5) else 200
            if upside < -80 or upside > upside_max:
                reason = f"[{per_str}/{pbr_str}] 괴리율극단({upside:.0f}%)→제외"
                upside = None
                avg_target = None

        return pd.Series({
            "적정가_PER":    per_t_f,
            "적정가_PBR":    pbr_t_f,
            "적정가_Graham": graham_t_f,
            "적정가_평균":   avg_target,
            "괴리율(%)":     upside,
            "적정가_사유":   reason,
        })

    val_df = filtered.apply(_valuation, axis=1)
    for col in val_df.columns:
        filtered[col] = val_df[col]

    # ── ⑩ 100점 만점 기업 종합 평가 ──
    print("  [100점] 기업 종합 평가 계산 중...")
    pt100_rows = filtered.apply(
        lambda row: pd.Series(compute_100pt_score(row.to_dict())), axis=1
    )
    for col in pt100_rows.columns:
        filtered[col] = pt100_rows[col]

    # ── ⑪ 매매 시그널 ──
    print("  [시그널] 매매 시그널 및 가격 밴드 계산 중...")
    def _sig_row(row):
        sc100 = {c: row.get(c) for c in pt100_rows.columns}
        return pd.Series(compute_trading_signal(row.to_dict(), sc100))
    sig_rows = filtered.apply(_sig_row, axis=1)
    for col in sig_rows.columns:
        filtered[col] = sig_rows[col]

    return filtered.sort_values("복합점수", ascending=False)

# ══════════════════════════════════════════════════════════
# 10. 엑셀 출력
# ══════════════════════════════════════════════════════════

def fetch_stock_news_claude(ticker: str, name: str, sector: str, top_n: int = 3) -> str:
    """
    Claude API를 이용해 종목별 주간 핵심 이슈/뉴스 요약 생성 (v30 신규)
    - 실제 웹 검색(web_search tool)을 사용해 최신 뉴스를 찾아 요약
    - 엑셀 한 셀에 들어갈 수 있는 간결한 형태로 반환
    """
    try:
        import requests as _req
        today_str = datetime.today().strftime("%Y년 %m월 %d일")
        prompt = (
            f"오늘은 {today_str}입니다.\n"
            f"한국 주식 종목 '{name}'({ticker}, 업종: {sector})의 "
            f"최근 1주일 이내 핵심 투자 이슈를 {top_n}줄 이내로 간략히 요약해주세요.\n"
            f"형식: '① [날짜] 이슈내용 ② [날짜] 이슈내용 ...' 형태로 작성.\n"
            f"없으면 '특이사항 없음'으로 답하세요. 불필요한 설명 없이 이슈 내용만 작성."
        )
        resp = _req.post(
            "https://api.anthropic.com/v1/messages",
            headers={"Content-Type": "application/json"},
            json={
                "model": "claude-sonnet-4-6",
                "max_tokens": 300,
                "tools": [{"type": "web_search_20250305", "name": "web_search"}],
                "messages": [{"role": "user", "content": prompt}]
            },
            timeout=30
        )
        data = resp.json()
        # 텍스트 블록만 추출
        texts = [b.get("text","") for b in data.get("content",[]) if b.get("type")=="text"]
        result = " ".join(texts).strip()
        return result if result else "뉴스 수집 실패"
    except Exception as e:
        return f"뉴스 수집 오류: {str(e)[:30]}"


def fetch_news_batch(df_top: pd.DataFrame, max_stocks: int = 30) -> dict:
    """
    상위 종목의 뉴스를 병렬로 배치 수집
    max_stocks: API 비용 제한을 위해 상위 N개만 수집
    """
    from concurrent.futures import ThreadPoolExecutor, as_completed
    stocks = list(df_top.iterrows())[:max_stocks]
    news_map = {}

    print(f"  [뉴스] 상위 {len(stocks)}개 종목 뉴스 수집 중 (Claude API)...")

    def _fetch(item):
        code, row = item
        name   = str(row.get("종목명","") or "")
        sector = str(row.get("업종","") or "")
        news = fetch_stock_news_claude(str(code), name, sector)
        return str(code), news

    with ThreadPoolExecutor(max_workers=5) as ex:
        futs = {ex.submit(_fetch, item): item for item in stocks}
        done = 0
        for fut in as_completed(futs):
            code, news = fut.result()
            news_map[code] = news
            done += 1
            if done % 10 == 0:
                print(f"    {done}/{len(stocks)}개 완료...")

    print(f"  [뉴스] 수집 완료: {len(news_map)}개")
    return news_map


def build_excel(df_all, df_top, label, args, output_dir=BASE_DIR):
    if not HAS_OPENPYXL:
        print("  ⚠ openpyxl 미설치 → 엑셀 저장 불가")
        return None

    # v33: 날짜 고정 파일명 → 매일 같은 파일을 덮어씀 (작업스케줄러 자동실행 최적화)
    # 파일명 예: quant_KR_20260617.xlsx  (시각 없음 → 항상 같은 이름)
    fname = datetime.today().strftime(f"quant_{label}_%Y%m%d.xlsx")
    fpath = os.path.join(output_dir, fname)

    # ── 스타일 상수 ──
    KR_FONT    = "맑은 고딕"
    body_font  = Font(name=KR_FONT, size=10)
    _HDR_FILL  = PatternFill("solid", fgColor="1F3864")
    _ALT_FILL  = PatternFill("solid", fgColor="F5F5F5")
    _GRN_FILL  = PatternFill("solid", fgColor="E2EFDA")

    # ── 숫자 포맷 헬퍼 ──
    def _r1(v):
        if v is None: return None
        try: return round(float(v), 1)
        except: return None
    def _r2(v):
        if v is None: return None
        try: return round(float(v), 2)
        except: return None
    def _ri(v):
        if v is None: return None
        try: return int(round(float(v)))
        except: return None
    def _eok(v):
        """원 단위 → 억 단위 변환 (DART 재무 수치용)"""
        if v is None: return None
        try:
            fv = float(v)
            if abs(fv) > 1e8:          # 1억 이상이면 원 단위로 판단 → 억으로 변환
                return int(round(fv / 1e8))
            else:                       # 이미 억 단위이거나 작은 값
                return int(round(fv))
        except: return None

    # ── 다중 키 값 조회 ──
    def _gv(row, *keys):
        for k in keys:
            v = row.get(k)
            if v is not None and not (isinstance(v, float) and np.isnan(v)):
                try:
                    f = float(v)
                    if f != 0: return f
                except:
                    if str(v).strip(): return v
        return None

    # ── 시가총액 조회 ──
    def _marcap(row):
        mc = row.get("시가총액(억)") or row.get("시총")
        if mc: return _safe(mc)
        cur = _safe(row.get("현재가") or row.get("현재가_NV") or row.get("현재가_YF"))
        shr = _safe(row.get("shares"))
        if cur and shr: return round(cur * shr / 1e8, 1)
        return None

    wb = Workbook()
    ws_detail = wb.active          # 시트1: 스크리닝결과상세
    ws_guide   = wb.create_sheet()  # 시트2: 지표해석가이드
    ws_all     = wb.create_sheet()  # 시트3: 전체종목데이터
    ws_signal  = wb.create_sheet()  # 시트4: 매매시그널보드
    ws_news    = wb.create_sheet()  # 시트5: 종목별 주간뉴스 (v30 신규)

    # ── 뉴스 수집 (Claude API, 상위 종목 한정) ──
    news_map = {}
    fetch_news = getattr(args, "fetch_news", True)
    news_top_n = min(getattr(args, "news_top", 30), len(df_top))
    if fetch_news and news_top_n > 0:
        try:
            news_map = fetch_news_batch(df_top, max_stocks=news_top_n)
        except Exception as e:
            print(f"  [뉴스] 수집 실패: {e}")

    # ── is_kr 먼저 정의 ──
    is_kr = label in ("KR","ALL")

    # ── 시트2: 스크리닝 결과 상세 (핵심 컬럼 통합) ──
    ws_detail.title = "스크리닝결과상세"

    # 제목행
    ws_detail.merge_cells("A1:Z1")
    t2 = ws_detail.cell(row=1, column=1,
        value=f"퀀트 스크리닝 상위 {args.top}종목 [{label}]  기준일: {datetime.today().strftime('%Y-%m-%d')}")
    t2.font  = Font(name=KR_FONT, bold=True, size=13, color="FFFFFF")
    t2.fill  = PatternFill("solid", fgColor="1F3864")
    t2.alignment = Alignment(horizontal="center", vertical="center")
    ws_detail.row_dimensions[1].height = 28

    ws_detail.merge_cells("A2:Z2")
    ws_detail.cell(row=2, column=1,
        value=f"가중치: 가치{args.w_value:.0f}% 모멘텀{args.w_mom:.0f}% "
              f"배당{args.w_div:.0f}% 품질{args.w_quality:.0f}% "
              f"성장{args.w_growth:.0f}% 수익성{args.w_prof:.0f}% "
              f"안정{args.w_stability:.0f}%").font = Font(name=KR_FONT, size=10, italic=True)

    # ── 헤더 그룹 정의 ──
    sub_headers = [
        # 기본 식별
        "순위","코드","종목명","시장","업종","현재가","시가총액(억)",
        # 7팩터 점수
        "복합점수","등급","가치","모멘텀","배당","품질","성장","수익성","안정",
        # 핵심 재무
        "PBR","ROE(%)","ROA(%)","DIV(%)","부채비율(%)","이자보상배율",
        "영업이익률(%)","FCF(억)","CFO(억)",
        # 적정가/밸류
        "적정가_PER","적정가_PBR","적정가_Graham","적정가_평균","괴리율(%)","적정가_사유",
        # 100점 평가
        "100점_합계","100점_등급",
        "100점_수익성","100점_성장성","100점_재무안정","100점_현금흐름","100점_밸류","100점_주주친화","100점_정성보정",
        # 5대 체크 (숫자+판정)
        "CFO(판정)","자본잠식(판정)","부채비율(판정)","ROE(판정)","PER(판정)",
        # 매매 시그널
        "매매시그널","현재가위치",
        "적극매수선","매수선","1차목표가","2차목표가","손절선",
        "보유전략","위험신호","보너스근거",
    ]

    # 헤더 색상 그룹
    HDR_GROUPS = {
        range(1,8):   "1F3864",  # 기본
        range(8,17):  "2E75B6",  # 7팩터
        range(17,26): "375623",  # 재무
        range(26,31): "7030A0",  # 적정가
        range(31,41): "833C00",  # 100점
        range(41,46): "404040",  # 체크
        range(46,57): "C00000",  # 시그널
    }
    def _hdr_color(j):
        for rng, color in HDR_GROUPS.items():
            if j in rng: return color
        return "1F3864"

    for j, h in enumerate(sub_headers, 1):
        c = ws_detail.cell(row=3, column=j, value=h)
        c.font      = Font(name=KR_FONT, bold=True, color="FFFFFF", size=9)
        c.fill      = PatternFill("solid", fgColor=_hdr_color(j))
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = _THIN
    ws_detail.row_dimensions[3].height = 30

    for ri, (code, row) in enumerate(df_top.iterrows(), 1):
        score     = _safe(row.get("복합점수"), 0)
        grade_txt, grade_color = grade_label(score)
        cur       = _gv(row, "현재가","현재가_NV","현재가_YF")
        pt100     = _safe(row.get("100점_합계"))
        sig_txt   = str(row.get("매매시그널","") or "")
        sig_color = str(row.get("_signal_color","FFCC00") or "FFCC00")

        row_vals = [
            # 기본
            ri, str(code),
            str(_gv(row,"종목명") or ""),
            str(row.get("시장","") or ""),
            str(row.get("업종","") or ""),
            _ri(cur),
            _ri(_marcap(row)),
            # 7팩터
            _r1(score), str(grade_txt),
            _r1(row.get("가치점수")),   _r1(row.get("모멘텀점수")),
            _r1(row.get("배당점수")),   _r1(row.get("품질점수")),
            _r1(row.get("성장점수")),   _r1(row.get("수익성점수")),
            _r1(row.get("안정점수")),
            # 재무
            _r2(row.get("PBR") or row.get("pbr")),
            _r1(row.get("ROE") or row.get("roe")),
            _r1(row.get("ROA")),
            _r1(row.get("DIV") or 0),
            _r1(row.get("부채비율") or row.get("debt_ratio")),
            _r1(row.get("interest_coverage")),
            _r1(row.get("영업이익률")),
            _eok(row.get("fcf")),
            _eok(row.get("cfo")),
            # 적정가
            _ri(row.get("적정가_PER")),
            _ri(row.get("적정가_PBR")),
            _ri(row.get("적정가_Graham")),
            _ri(row.get("적정가_평균")),
            _r1(row.get("괴리율(%)")),
            str(row.get("적정가_사유","") or ""),
            # 100점
            _r1(pt100),
            str(row.get("100점_등급","") or ""),
            _safe(row.get("100점_수익성")),
            _safe(row.get("100점_성장성")),
            _safe(row.get("100점_재무안정")),
            _safe(row.get("100점_현금흐름")),
            _safe(row.get("100점_밸류")),
            _safe(row.get("100점_주주친화")),
            _safe(row.get("100점_정성보정")),
            # 5대 체크 (숫자+판정)
            str(row.get("체크_CFO","") or ""),
            str(row.get("체크_자본잠식","") or ""),
            str(row.get("체크_부채비율","") or ""),
            str(row.get("체크_ROE","") or ""),
            str(row.get("체크_PER","") or ""),
            # 시그널
            sig_txt,
            str(row.get("현재가위치","") or ""),
            _ri(row.get("적극매수선")),
            _ri(row.get("매수선")),
            _ri(row.get("1차목표가")),
            _ri(row.get("2차목표가")),
            _ri(row.get("손절선")),
            str(row.get("보유전략","") or ""),
            str(row.get("위험신호","없음") or "없음"),
            str(row.get("보너스근거","") or ""),
        ]

        rf = _rank_fill(ri)
        for j, v in enumerate(row_vals, 1):
            c = ws_detail.cell(row=ri+3, column=j, value=v)
            c.fill   = rf
            c.border = _THIN
            c.font   = body_font
            c.alignment = Alignment(horizontal="center", vertical="center",
                                    wrap_text=(j >= 53))
            # 숫자 포맷
            fmt_map = {
                6:"#,##0", 7:"#,##0",
                8:"#,##0.0", 10:"#,##0.0", 11:"#,##0.0", 12:"#,##0.0",
                13:"#,##0.0", 14:"#,##0.0", 15:"#,##0.0", 16:"#,##0.0",
                17:"#,##0.00", 18:"#,##0.0", 19:"#,##0.0", 20:"#,##0.0",
                21:"#,##0.0", 22:"#,##0.0", 23:"#,##0.0", 24:"#,##0", 25:"#,##0",
                26:"#,##0", 27:"#,##0", 28:"#,##0", 29:"#,##0", 30:"#,##0.0",
                31:"#,##0.0",
                48:"#,##0", 49:"#,##0", 50:"#,##0", 51:"#,##0", 52:"#,##0",
            }
            if j in fmt_map and v is not None:
                c.number_format = fmt_map[j]
            # 100점 합계 색상
            if j == 31 and v is not None:
                try:
                    fv = float(v)
                    if fv >= 90:   c.fill = PatternFill("solid", fgColor="0070C0"); c.font = Font(name=KR_FONT, size=9, color="FFFFFF", bold=True)
                    elif fv >= 80: c.fill = PatternFill("solid", fgColor="2E75B6"); c.font = Font(name=KR_FONT, size=9, color="FFFFFF", bold=True)
                    elif fv >= 70: c.fill = PatternFill("solid", fgColor="92D050"); c.font = Font(name=KR_FONT, size=9, bold=True)
                    elif fv >= 60: c.fill = PatternFill("solid", fgColor="FFCC00"); c.font = Font(name=KR_FONT, size=9, bold=True)
                    else:          c.fill = PatternFill("solid", fgColor="FF9900"); c.font = Font(name=KR_FONT, size=9, bold=True)
                except Exception: pass
            # 매매시그널 색상
            if j == 46 and v:
                try:
                    c.fill = PatternFill("solid", fgColor=sig_color)
                    is_dark = sig_color in ("C00000","0070C0","2E75B6","FF4444")
                    c.font = Font(name=KR_FONT, size=9, bold=True,
                                  color="FFFFFF" if is_dark else "000000")
                except Exception: pass
            # 괴리율 색상
            if j == 30 and v is not None:
                try:
                    fv = float(v)
                    c.font = Font(name=KR_FONT, size=9,
                                  color="0070C0" if fv >= 0 else "C00000",
                                  bold=(abs(fv) >= 20))
                except Exception: pass

    ws_detail.freeze_panes = "D4"
    auto_col_width(ws_detail)

    # ── 시트3: 지표 해석 가이드 ──
    ws_guide.title = "지표해석가이드"

    # ── 헬퍼: 가이드 셀 쓰기 ──
    def _gw(row, col, value, bold=False, color="000000", bg=None, wrap=True, size=10, align="left"):
        c = ws_guide.cell(row=row, column=col, value=value)
        c.font      = Font(name=KR_FONT, bold=bold, color=color, size=size)
        c.fill      = PatternFill("solid", fgColor=bg) if bg else PatternFill("none")
        c.border    = _THIN
        c.alignment = Alignment(wrap_text=wrap, vertical="top",
                                horizontal="center" if align=="center" else "left")
        return c

    # 컬럼 너비 수동 설정
    col_widths = [6, 18, 38, 22, 22, 22, 30]
    for i, w in enumerate(col_widths, 1):
        ws_guide.column_dimensions[chr(64+i)].width = w

    # ── 색상 정의 ──
    C_TITLE   = "1F3864"   # 진남색 (대제목)
    C_SEC     = "2E75B6"   # 중간파랑 (섹션)
    C_GREEN   = "E2EFDA"   # 연초록 배경 (좋음)
    C_YELLOW  = "FFF2CC"   # 연노랑 배경 (보통/주의)
    C_RED     = "FCE4D6"   # 연빨강 배경 (위험)
    C_BLUE    = "DEEAF1"   # 연파랑 배경 (정보)
    C_GRAY    = "F5F5F5"   # 연회색 배경 (중립)
    C_ORANGE  = "FCE4D6"   # 주황 배경 (경고)

    # ── 대제목 ──
    ws_guide.merge_cells("A1:G1")
    t = ws_guide.cell(row=1, column=1, value="퀀트 스크리너 지표 해석 완전 가이드")
    t.font      = Font(name=KR_FONT, bold=True, size=16, color="FFFFFF")
    t.fill      = PatternFill("solid", fgColor=C_TITLE)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws_guide.row_dimensions[1].height = 36

    ws_guide.merge_cells("A2:G2")
    sub = ws_guide.cell(row=2, column=1,
        value=f"기준일: {datetime.today().strftime('%Y-%m-%d')}  |  각 지표의 의미, 계산식, 기준값, 투자 판단 활용법")
    sub.font      = Font(name=KR_FONT, size=10, color="FFFFFF", italic=True)
    sub.fill      = PatternFill("solid", fgColor="2E75B6")
    sub.alignment = Alignment(horizontal="center", vertical="center")
    ws_guide.row_dimensions[2].height = 22

    # ── 투자 판단 우선순위 박스 ──
    ws_guide.merge_cells("A3:G3")
    _gw(3, 1, "【 투자 판단 5단계 우선순위 】  1단계: 생존필터(CFO·자본잠식·부채·이자보상) → "
        "2단계: 수익성(ROE·영업이익률) → 3단계: 가격(괴리율·PBR·PER) → "
        "4단계: 현금흐름품질(FCF·CFO>순이익) → 5단계: 복합점수·100점 종합 후 매매시그널",
        bold=True, bg="FFF2CC", size=10, align="left")
    ws_guide.row_dimensions[3].height = 30

    cur_row = 4

    # ── 섹션 정의 ──
    SECTIONS = [

        # ─────────────────────────────────────────────
        ("■ 수익성 지표", None, None),
        # (번호, 지표명, 의미, 계산식, 좋은기준, 보통기준, 위험기준/주의사항)
        ("1", "ROE\n(자기자본이익률)",
         "주주가 맡긴 돈으로 얼마나 벌었는지.\n높을수록 자본 효율이 뛰어난 기업.\n버핏이 가장 중시하는 지표.",
         "순이익 ÷ 자기자본 × 100",
         "✅ 20% 이상\n→ 독점적 해자 가능성\n→ 가산점 +5점",
         "⚠ 10~20%\n→ 보유 가능 수준\n→ 업종 평균과 비교 필요",
         "❌ 10% 미만\n→ 자본 비효율\n→ 주의: 부채로 ROE 부풀리기 가능\n   → ROA와 함께 확인"),

        ("2", "ROA\n(총자산이익률)",
         "총자산 대비 수익률.\nROE가 높아도 ROA가 낮으면\n부채로 수익을 부풀린 것일 수 있음.",
         "순이익 ÷ 총자산 × 100",
         "✅ 7% 이상\n→ 자산 활용 우수",
         "⚠ 3~7%\n→ 보통 수준",
         "❌ 3% 미만\n→ 자산 생산성 낮음\n→ ROE와 괴리 크면 고부채 의심"),

        ("3", "영업이익률",
         "본업으로 매출의 몇 %를 버는지.\n업종별 편차가 매우 크므로\n반드시 동종업계와 비교.",
         "영업이익 ÷ 매출액 × 100",
         "✅ 25% 이상\n→ 독점적 해자 추정\n→ 가산점 +5점\n\n✅ 15~25%\n→ 경쟁우위",
         "⚠ 5~15%\n→ 업종 평균 수준",
         "❌ 5% 미만\n→ 수익성 낮음\n❌ 적자\n→ 하드필터 탈락 가능"),

        ("4", "순이익률",
         "최종적으로 남는 이익 비율.\n영업이익률과 큰 차이가 나면\n이자·세금 부담이 크다는 신호.",
         "당기순이익 ÷ 매출액 × 100",
         "✅ 8% 이상\n→ 우수",
         "⚠ 3~8%\n→ 보통",
         "❌ 3% 미만\n→ 위험\n→ 영업이익률 대비 낮으면\n   이자비용 과다 확인"),

        # ─────────────────────────────────────────────
        ("■ 밸류에이션 지표", None, None),

        ("5", "PBR\n(주가순자산비율)",
         "주가가 순자산의 몇 배인지.\n1배 미만 = 이론상 청산가치보다 싸게 거래.\n현재가·결산기준 두 가지로 산출.",
         "시가총액 ÷ 자기자본(순자산)\n\nPBR갭(%) = (현재가기준 - 결산기준)\n           ÷ 결산기준 × 100",
         "✅ 1배 미만\n→ 저평가 관심\n\nPBR갭 음수\n→ 주가가 자산 대비 내려옴",
         "⚠ 1~2배\n→ 적정 수준",
         "❌ 3배 이상\n→ 고평가 주의\n→ 단, 고ROE 기업은\n   높은 PBR이 정당화될 수 있음"),

        ("6", "PER\n(주가수익비율)",
         "이익 대비 주가 수준.\n낮을수록 싸게 거래되는 것.\n업종 평균과의 비교가 핵심.\n적자 기업은 의미 없음.",
         "주가 ÷ EPS(주당순이익)\n\nPER갭(%) = (현재가기준 - 결산기준)\n           ÷ 결산기준 × 100",
         "✅ 10배 미만\n→ 저평가 매수 선호",
         "⚠ 10~20배\n→ 적정 수준",
         "❌ 25배 이상\n→ 고평가 주의\n→ 성장주는 예외 적용\n   (PEG = PER ÷ 성장률 활용)"),

        ("7", "DIV\n(배당수익률)",
         "주가 대비 연간 배당금 비율.\n3% 이상이면 안정적 인컴 자산.\n배당성향 60% 초과 시 지속가능성 확인.",
         "주당배당금 ÷ 주가 × 100\n\n배당성향 = 배당금 ÷ 순이익 × 100",
         "✅ DIV 3% 이상\n→ 안정적 인컴\n✅ 배당성향 20~50%\n→ 지속 가능",
         "⚠ DIV 1~3%\n⚠ 배당성향 50~70%",
         "❌ 배당성향 80% 초과\n→ 삭감 위험\n❌ DIV 높은데 PER 낮으면\n→ 실적 악화 가능성 확인"),

        ("8", "EV/EBITDA",
         "부채까지 포함한 기업 전체 가치가\n영업이익+감가상각의 몇 배인지.\n부채 많은 기업 비교에 유용.",
         "EV = 시가총액 + 총부채 - 현금\nEBITDA = 영업이익 + 감가상각\n\nEV/EBITDA = EV ÷ EBITDA",
         "✅ 6배 미만\n→ 저평가",
         "⚠ 6~12배\n→ 적정",
         "❌ 15배 이상\n→ 고평가\n→ 장치산업(제조·통신)에서\n   PER보다 더 유용"),

        ("9", "적정가 3종 & 괴리율",
         "세 가지 방법으로 내재가치 추정 후\n평균을 현재가와 비교.\n괴리율이 클수록 저평가.",
         "적정가_PER  = EPS × 업종평균PER\n적정가_PBR  = BPS × 업종평균PBR\n적정가_Graham = √(22.5×EPS×BPS)\n\n괴리율(%) = (적정가평균 - 현재가)\n            ÷ 현재가 × 100",
         "✅ 괴리율 20% 이상\n→ 저평가 매수 기회\n→ 파란색 표시",
         "⚠ 괴리율 0~20%\n→ 소폭 저평가",
         "❌ 괴리율 음수\n→ 고평가\n→ 빨간색 표시\n→ -20% 이하면 매도 검토"),

        # ─────────────────────────────────────────────
        ("■ 재무안정성 지표", None, None),

        ("10", "부채비율",
         "자기자본 대비 총부채 비율.\n업종별 차이 매우 큼.\n금융·지주·건설은 구조적으로 높음.",
         "총부채 ÷ 자기자본 × 100",
         "✅ 100% 미만\n→ 안전\n→ 체크 ✅",
         "⚠ 100~200%\n→ 모니터링\n→ 체크 ⚠",
         "❌ 300% 이상\n→ 위험\n→ 체크 ❌\n→ FCF음수와 겹치면 즉시 경고"),

        ("11", "순부채비율",
         "현금을 제외한 실질 부채 수준.\n음수면 순현금 기업으로 매우 우량.\n부채비율보다 실질 위험을 잘 반영.",
         "(총부채 - 현금및현금성자산)\n÷ 자기자본 × 100",
         "✅ 음수 (순현금 기업)\n→ 최우량 등급",
         "⚠ 0~50%\n→ 안전권",
         "❌ 100% 이상\n→ 실질 부채 과다\n→ 300% 이상이면 감점 -2점"),

        ("12", "유동비율",
         "1년 내 갚아야 할 부채를\n1년 내 현금화 가능한 자산으로\n감당할 수 있는지.",
         "유동자산 ÷ 유동부채 × 100",
         "✅ 200% 이상\n→ 단기 안전",
         "⚠ 150~200%\n→ 양호",
         "❌ 100% 미만\n→ 단기 유동성 위기 가능"),

        ("13", "이자보상배율",
         "영업이익으로 이자를 몇 번 갚을 수 있는지.\n1배 미만 = 이자도 못 버는 한계기업.\nAltman Z보다 더 직관적인 위험 지표.",
         "영업이익 ÷ 이자비용",
         "✅ 5배 이상\n→ 안전",
         "⚠ 2~5배\n→ 양호",
         "❌ 1배 미만\n→ 한계기업\n→ 하드필터 탈락\n→ 실위험 판단 핵심 지표"),

        ("14", "Altman Z-Score",
         "5개 재무비율로 부도 가능성 예측.\n★ 단독 판단 금지 ★\n이자보상배율·FCF와 함께 봐야 함.\n금융·지주·부동산은 구조적으로 낮게 나옴.",
         "Z = 1.2×(운전자본/자산)\n  + 1.4×(이익잉여금/자산)\n  + 3.3×(EBIT/자산)\n  + 0.6×(자기자본/총부채)\n  + 1.0×(매출/자산)",
         "✅ 2.99 이상\n→ 안전지대",
         "⚠ 1.81~2.99\n→ 회색지대\n→ 이자보상 확인 필수",
         "❌ 1.81 미만 + 이자보상<2배\n→ 실제 위험 → 감점\n\n✅ 1.81 미만 + 이자보상≥3배\n→ 업종 구조적 문제\n→ 감점 최소화"),

        # ─────────────────────────────────────────────
        ("■ 현금흐름 지표", None, None),

        ("15", "CFO\n(영업현금흐름)",
         "본업에서 실제로 들어온 현금.\n순이익보다 조작하기 어려워\n이익 품질 판단에 핵심.\n반드시 양수여야 함 (하드필터).",
         "현금흐름표 → 영업활동현금흐름\n(손익계산서 순이익과 다름)",
         "✅ CFO > 0\n→ 생존 필터 통과\n✅ CFO > 순이익\n→ 이익 품질 최상",
         "⚠ CFO ≈ 순이익\n→ 보통 수준",
         "❌ CFO < 0\n→ 하드필터 탈락\n→ 순이익이 플러스여도 의미 없음\n→ '가짜 이익' 가능성"),

        ("16", "FCF\n(잉여현금흐름)",
         "영업현금에서 설비투자(CAPEX)를 뺀\n진짜 자유현금.\n배당·자사주·부채상환 가능한 돈.",
         "FCF = CFO - CAPEX(자본적지출)",
         "✅ FCF > 0\n→ 현금창출 우수\n→ 배당·성장 여력",
         "⚠ FCF ≈ 0\n→ 투자 확대기\n→ 성장주는 일시적 허용",
         "❌ FCF < 0 + 부채비율 > 200%\n→ 감점 -4점\n→ 자금조달(CB·유증) 위험\n\n❌ FCF < 0 + 이자보상 < 2배\n→ 감점 -3점"),

        ("17", "FCF Yield\n(잉여현금수익률)",
         "시가총액 대비 FCF 비율.\n5% 이상이면 현금창출력 우수.\n배당수익률보다 실질 주주환원을 잘 반영.",
         "FCF ÷ 시가총액 × 100",
         "✅ 5% 이상\n→ 매력적\n→ 핵심투자포인트 표시",
         "⚠ 2~5%\n→ 보통",
         "❌ 2% 미만 또는 음수\n→ 현금창출 부족"),

        ("18", "FCF마진",
         "매출액 대비 FCF 비율.\n사업모델의 현금 효율성을 측정.",
         "FCF ÷ 매출액 × 100",
         "✅ 10% 이상\n→ 우수한 사업모델",
         "⚠ 5~10%\n→ 보통",
         "❌ 5% 미만 또는 음수\n→ 자본집약도 높음"),

        # ─────────────────────────────────────────────
        ("■ 성장성 · 모멘텀 지표", None, None),

        ("19", "매출성장률",
         "전년 대비 매출 증가율.\n지속적인 성장이 핵심.\n3년 연속 감소는 사업 경쟁력 약화 신호.",
         "( 당기매출 - 전기매출 ) ÷ 전기매출 × 100",
         "✅ 20% 이상\n→ 고성장 (가산점 +3점)\n✅ 10~20%\n→ 성장 (가산점 +1점)",
         "⚠ 0~10%\n→ 안정성장",
         "❌ -5% 이하\n→ 매출감소 (감점 -1점)\n❌ -10% 이하\n→ 매출급감 (감점 -3점)"),

        ("20", "52주위치(%)",
         "52주 최저~최고가 구간에서\n현재가가 어디에 위치하는지.\n0%=52주 최저가, 100%=52주 최고가.",
         "( 현재가 - 52주최저 )\n÷ ( 52주최고 - 52주최저 ) × 100",
         "✅ 0~30%\n→ 52주 저점 근처\n→ 역발상 매수 기회\n→ 핵심투자포인트 표시",
         "⚠ 30~70%\n→ 중간 구간",
         "❌ 80% 이상\n→ 52주 고점 근처\n→ 추격매수 주의"),

        ("21", "6개월수익률",
         "최근 6개월 주가 상승률.\n모멘텀 팩터의 핵심 지표.\n강한 모멘텀은 추세 지속 경향.",
         "( 현재가 - 6개월전가 ) ÷ 6개월전가 × 100",
         "✅ 15% 이상\n→ 강한 모멘텀\n→ 핵심투자포인트 표시",
         "⚠ 0~15%\n→ 중립",
         "❌ -10% 이하\n→ 하락 추세"),

        # ─────────────────────────────────────────────
        ("■ 100점 만점 종합 평가", None, None),

        ("22", "100점_합계\n& 등급",
         "6개 카테고리 합산 + 정성 보정.\n복합점수(7팩터)와 함께 투자 판단.",
         "수익성20 + 성장성20 + 재무안정20\n+ 현금흐름20 + 밸류10 + 주주친화10\n± 정성보정(최대 ±10점)",
         "✅ 90점 이상 → ★★★★★ 최상급\n✅ 80~89점 → ★★★★  우수",
         "⚠ 70~79점 → ★★★ 양호\n⚠ 60~69점 → ★★  보통",
         "❌ 50~59점 → ★ 주의\n❌ 50점 미만 → 투자재검토"),

        ("23", "정성보정\n(±10점)",
         "데이터로 잡기 어려운 질적 요소를\n가감점으로 반영.",
         "가산: 독점해자+5, 성장산업+5, 고성장+3\n감점: 부도위험-5, FCF음수+고부채-4,\n      매출급감-3, 순부채과다-2",
         "✅ +5점 이상\n→ 강한 경쟁우위 보유",
         "⚠ 0점\n→ 중립",
         "❌ -5점 이하\n→ 실질 위험 존재\n→ 감점 근거 컬럼 반드시 확인"),

        # ─────────────────────────────────────────────
        ("■ 매매시그널 & 가격 밴드", None, None),

        ("24", "매매시그널\n(7단계)",
         "복합점수·100점·괴리율·위험신호를\n종합해 자동 산출하는 투자 의견.\n주의: 참고용이며 최종 판단은 직접.",
         "■■■ 강력매수: 복합≥85 & 100점≥80\n■■  매수:    복합≥75 & 100점≥70\n■   관심:    복합≥65 & 100점≥60\n─   보유:    특이사항 없음\n▽   비중축소: 괴리율 < -10%\n▼▼  매도검토: 실위험 1개 이상\n▼▼▼ 즉시매도: 실위험 2개 이상",
         "✅ ■■■ 강력매수\n→ 분할매수 2~3회 진입",
         "⚠ ─ 보유\n→ 1차목표가까지 보유",
         "❌ ▼▼ 매도검토\n→ 손절선 이탈 시 매도\n❌ ▼▼▼ 즉시매도\n→ 신규 진입 금지"),

        ("25", "가격 밴드\n(6단계)",
         "현재가·적정가 기준으로\n매수·매도 가격선 자동 계산.\n손절선은 반드시 설정 후 진입.",
         "적극매수선 = 현재가 × 0.95\n매수선     = 현재가\n1차목표가  = 적정가 평균\n2차목표가  = 적정가 × 1.15\n추격매수경보 = 적정가 × 1.20\n손절선     = 현재가 × 0.85",
         "✅ 현재가 < 적극매수선\n→ 적극 진입 구간",
         "⚠ 현재가 = 매수선 근처\n→ 현재 매수 가능",
         "❌ 현재가 > 추격매수경보\n→ 고평가, 신규 진입 자제\n❌ 현재가 < 손절선\n→ 즉시 손절"),

        ("26", "위험신호\n(3단계)",
         "개별 지표가 아닌 복합 조건으로\n실질 위험을 판단.\n오탐(False Positive) 최소화.",
         "🔴 실위험: 복합 조건 충족\n   (Altman<1.8 AND IC<2배 등)\n🟡 주의: 단일 조건 (실위험 낮음)\n⬜ 없음: 모두 정상",
         "⬜ 없음\n→ 위험 요소 없음",
         "🟡 주의\n→ 해당 지표 추가 확인\n→ 매수시그널 1단계 하향",
         "🔴 실위험 1개 → 매도검토\n🔴 실위험 2개↑ → 즉시매도\n→ 감점근거 컬럼 반드시 확인"),

        # ─────────────────────────────────────────────
        ("■ 7팩터 복합점수 구성", None, None),

        ("27", "가치점수\n(Value)",
         "PBR·ROE·FCF수익률·EV/EBITDA\n조합으로 계산한 가치 점수.",
         "PBR 30% + ROE 30%\n+ FCF Yield 20% + EV/EBITDA 20%",
         "✅ PBR≤0.5, FCFYield≥8%\n→ 80점 이상 목표",
         "⚠ PBR 1~2배 수준",
         "❌ PBR≥3.0\n❌ EV/EBITDA≥20"),

        ("28", "모멘텀점수\n(Momentum)",
         "52주 위치와 6개월 수익률 기반.\n추세 추종 전략의 핵심 요소.",
         "52주위치 70% + 6개월수익률 30%",
         "✅ 52주 40~75% 구간\n+ 6개월 수익률 양호",
         "⚠ 52주 위치 중립 구간",
         "❌ 52주 0% 또는 100% 극단\n→ 과매도/과매수 판단 필요"),

        ("29", "품질점수\n(Quality)",
         "재무 건전성 종합.\n부채·ROA·영업이익률 기반.",
         "부채비율 + ROA + 영업이익률\n3개 지표 가중 평균",
         "✅ 부채비율≤50%\n✅ ROA≥7%",
         "⚠ 부채비율 100~200%",
         "❌ 부채비율≥300%\n❌ ROA 음수"),

        ("30", "안정점수\n(Stability)",
         "재무 안정성 종합.\nAltman Z·유동비율·ROE·배당 포함.",
         "순부채비율35% + 유동비율15%\n+ AltmanZ20% + ROE15% + 배당15%",
         "✅ Altman≥3.0\n✅ 유동비율≥150%",
         "⚠ Altman 회색지대",
         "❌ Altman≤1.8 (복합 확인)\n❌ 유동비율≤100%"),
    ]

    # ── 헤더 행 출력 ──
    HDR_ROW = cur_row
    hdr_vals = ["No.", "지표명", "의미 및 설명", "계산식", "✅ 좋음 (기준값)", "⚠ 보통 (주의)", "❌ 위험 / 주의사항"]
    for ci, v in enumerate(hdr_vals, 1):
        _gw(HDR_ROW, ci, v, bold=True, color="FFFFFF", bg=C_TITLE, size=10, align="center")
    ws_guide.row_dimensions[HDR_ROW].height = 22
    cur_row += 1

    # ── 섹션 & 데이터 행 출력 ──
    for item in SECTIONS:
        if item[1] is None:
            # 섹션 구분 행
            ws_guide.merge_cells(f"A{cur_row}:G{cur_row}")
            c = ws_guide.cell(row=cur_row, column=1, value=item[0])
            c.font      = Font(name=KR_FONT, bold=True, size=11, color="FFFFFF")
            c.fill      = PatternFill("solid", fgColor=C_SEC)
            c.border    = _THIN
            c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            ws_guide.row_dimensions[cur_row].height = 22
            cur_row += 1
        else:
            no, name, desc, formula, good, mid, bad = item
            row_data = [no, name, desc, formula, good, mid, bad]
            # 행 높이 = 내용에 따라 조정
            ws_guide.row_dimensions[cur_row].height = 75

            for ci, v in enumerate(row_data, 1):
                if ci == 1:   # No.
                    _gw(cur_row, ci, v, bold=True, bg=C_GRAY, size=10, align="center")
                elif ci == 2: # 지표명
                    _gw(cur_row, ci, v, bold=True, bg=C_BLUE, size=10)
                elif ci == 3: # 의미
                    _gw(cur_row, ci, v, size=10)
                elif ci == 4: # 계산식
                    c = ws_guide.cell(row=cur_row, column=ci, value=v)
                    c.font      = Font(name="Courier New", size=9, color="2E4057")
                    c.fill      = PatternFill("solid", fgColor="F0F4F8")
                    c.border    = _THIN
                    c.alignment = Alignment(wrap_text=True, vertical="top")
                elif ci == 5: # 좋음
                    _gw(cur_row, ci, v, size=10, bg=C_GREEN)
                elif ci == 6: # 보통
                    _gw(cur_row, ci, v, size=10, bg=C_YELLOW)
                elif ci == 7: # 위험
                    _gw(cur_row, ci, v, size=10, bg=C_RED)
            cur_row += 1

    # ── 하단 요약 박스 ──
    ws_guide.merge_cells(f"A{cur_row}:G{cur_row}")
    _gw(cur_row, 1,
        "【 실전 투자 체크리스트 】  "
        "① CFO > 0  "
        "② 자본잠식 없음  "
        "③ 부채비율 < 150%  "
        "④ ROE ≥ 10%  "
        "⑤ PER 업종 평균 이하  "
        "⑥ FCF > 0  "
        "⑦ 이자보상배율 ≥ 3배  "
        "→ 7개 모두 통과하면 부실기업 대부분 걸러짐",
        bold=True, bg="D9EAD3", size=10)
    ws_guide.row_dimensions[cur_row].height = 28

    ws_guide.freeze_panes = "A5"


    # ── 시트4: 전체 종목 데이터 ──
    ws_all.title = "전체종목데이터"
    ws_all_hdr_font = Font(name=KR_FONT, bold=True, color="FFFFFF", size=9)
    ws_all_body_font = Font(name=KR_FONT, size=9)

    # 제목행
    ws_all.merge_cells("A1:BW1")
    t_all = ws_all.cell(row=1, column=1,
        value=f"전체 종목 데이터 [{label}]  기준일: {datetime.today().strftime('%Y-%m-%d')}  (복합점수 내림차순)")
    t_all.font  = Font(name=KR_FONT, bold=True, size=12, color="FFFFFF")
    t_all.fill  = PatternFill("solid", fgColor="1F3864")
    t_all.alignment = Alignment(horizontal="center", vertical="center")
    ws_all.row_dimensions[1].height = 26

    # 전체종목 헤더 (기술지표 추가)
    all_headers = ["순위","코드","종목명","시장","업종","현재가",
                   "시가총액(억)" if is_kr else "시총",
                   # 거래량
                   "거래대금5일(억)","거래대금20일(억)","거래량비율VR(%)",
                   # 기술지표 (v32)
                   "MACD선","MACD시그널","MACD히스토","MACD신호",
                   "RSI14","RSI신호",
                   "OBV추세","OBV신호",
                   "BB상단","BB하단","BB%B","BB밴드폭(%)","BB신호",
                   "기술매수신호수","기술매도신호수","기술신호","기술지표요약",
                   # 밸류에이션
                   "PBR(현재가)","PBR(결산기준)","PBR갭(%)",
                   "ROE(%)","ROA(%)",
                   "PER(현재가)","PER(결산기준)","PER갭(%)",
                   "DIV(%)","배당성향(%)",
                   "영업이익률(%)","순이익률(%)","부채비율(%)","순부채비율(%)",
                   "유동비율(%)","Altman-Z",
                   "FCF(억)" if is_kr else "FCF","FCF마진(%)","FCF Yield(%)",
                   "EBITDA(억)" if is_kr else "EBITDA","EV/EBITDA",
                   "매출성장률(%)",
                   "CFO(억)" if is_kr else "CFO",
                   "이자보상배율",
                   "52주위치(%)","52주수익률(%)","6개월수익률(%)",
                   "가치점수","모멘텀점수","배당점수",
                   "품질점수","성장점수","수익성점수","안정점수",
                   "복합점수",
                   "적정가_PER","적정가_PBR","적정가_Graham","적정가_평균","괴리율(%)","적정가_사유",
                   "등급",
                   "100점_합계","100점_등급",
                   "100점_수익성(20)","100점_성장성(20)","100점_재무안정(20)",
                   "100점_현금흐름(20)","100점_밸류(10)","100점_주주친화(10)","100점_정성보정",
                   "CFO(판정)","자본잠식(판정)","부채비율(판정)","ROE(판정)","PER(판정)",
                   "보너스근거","감점근거",
                   "매매시그널","현재가위치",
                   "적극매수선","매수선","1차목표가","2차목표가","추격매수경보","손절선",
                   "보유전략","핵심투자포인트","위험신호"]

    # 헤더 색상 그룹
    _ALL_HDR_GROUPS = {
        range(1,8):   "1F3864",
        range(8,18):  "2E75B6",
        range(18,35): "375623",
        range(35,43): "7030A0",
        range(43,49): "4A4A4A",
        range(49,66): "833C00",
        range(66,76): "C00000",
    }
    def _all_hdr_color(j):
        for rng, color in _ALL_HDR_GROUPS.items():
            if j in rng: return color
        return "1F3864"

    for j, h in enumerate(all_headers, 1):
        c = ws_all.cell(row=2, column=j, value=h)
        c.font      = ws_all_hdr_font
        c.fill      = PatternFill("solid", fgColor=_all_hdr_color(j))
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = _THIN
    ws_all.row_dimensions[2].height = 28

    _all_fmt = {
        6:"#,##0", 7:"#,##0",
        8:"#,##0.00", 9:"#,##0.00", 10:"#,##0.0",
        11:"#,##0.0", 12:"#,##0.0",
        13:"#,##0.0", 14:"#,##0.0", 15:"#,##0.0",
        16:"#,##0.0", 17:"#,##0.0", 18:"#,##0.0", 19:"#,##0.0",
        20:"#,##0.0", 21:"#,##0.0", 22:"#,##0.0", 23:"#,##0.00",
        24:"#,##0", 25:"#,##0.0", 26:"#,##0.0",
        27:"#,##0", 28:"#,##0.0", 29:"#,##0.0",
        30:"#,##0", 31:"#,##0.0",
        32:"#,##0.0", 33:"#,##0.0", 34:"#,##0.0",
        35:"#,##0.0", 36:"#,##0.0", 37:"#,##0.0",
        38:"#,##0.0", 39:"#,##0.0", 40:"#,##0.0", 41:"#,##0.0",
        42:"#,##0.0",
        43:"#,##0", 44:"#,##0", 45:"#,##0", 46:"#,##0", 47:"#,##0.0",
        49:"#,##0.0",
        51:"#,##0.0", 52:"#,##0.0", 53:"#,##0.0",
        54:"#,##0.0", 55:"#,##0.0", 56:"#,##0.0", 57:"#,##0.0",
        68:"#,##0", 69:"#,##0", 70:"#,##0", 71:"#,##0", 72:"#,##0", 73:"#,##0",
    }

    for ri_all, (code_a, row_a) in enumerate(df_all.iterrows(), 1):
        sc_a  = _safe(row_a.get("복합점수"), 0)
        cur_a = _gv(row_a, "현재가","현재가_NV","현재가_YF")
        mc_a  = _marcap(row_a)
        fcf_a = _safe(row_a.get("fcf"))
        eb_a  = _safe(row_a.get("ebitda"))
        # 원 단위 → 억 단위 변환
        fcf_a_eok = _eok(fcf_a)
        eb_a_eok  = _eok(eb_a)
        cfo_a_eok = _eok(row_a.get("cfo"))
        # FCF Yield: FCF(원) / 시총(원) × 100
        fy_a = _r1(fcf_a / (mc_a * 1e8) * 100) if (fcf_a and mc_a and mc_a > 0) else None
        ev_a = None
        if eb_a and eb_a > 0 and mc_a:
            tl_a = _safe(row_a.get("total_liabilities")) or 0
            ca_a = _safe(row_a.get("cash")) or 0
            ev_a = _r1((mc_a * 1e8 + tl_a - ca_a) / eb_a)
        grd_a, grd_color_a = grade_label(sc_a)

        # ── PBR/PER 현재가 기준 재계산 (v28: 주식수 기반) ──
        # 우선순위: ① shares×현재가÷재무수치  ② 시총÷재무수치  ③ 수집값
        _shr_raw = row_a.get("shares")
        shr_a = _safe(_shr_raw) if (_shr_raw is not None and not (hasattr(_shr_raw, '__class__') and _shr_raw.__class__.__name__ == 'NAType')) else _safe(row_a.get("shares_dart"))
        if shr_a is not None and not (isinstance(shr_a, float) and np.isnan(shr_a)):
            shr_a = float(shr_a)
        else:
            shr_a = None
        equity_a = _safe(row_a.get("equity"))
        ni_a     = _safe(row_a.get("net_income"))

        mktcap_won_a = None
        if shr_a and shr_a > 0 and cur_a and cur_a > 0:
            mktcap_won_a = shr_a * cur_a                      # 주식수 × 현재가(원)
        elif mc_a and mc_a > 0:
            mktcap_won_a = mc_a * 1e8                         # 시총(억) → 원

        if mktcap_won_a and equity_a and equity_a > 0:
            pbr_na = round(mktcap_won_a / equity_a, 2)
        else:
            pbr_na = _r2(row_a.get("PBR") or row_a.get("pbr"))  # fallback

        if mktcap_won_a and ni_a and ni_a > 0:
            per_na = round(mktcap_won_a / ni_a, 1)
        else:
            per_na = _r1(row_a.get("PER") or row_a.get("per"))  # fallback

        pbr_ra = _r2(row_a.get("PBR_결산"))
        per_ra = _r1(row_a.get("PER_결산"))
        pbr_ga = _r1((pbr_na-pbr_ra)/abs(pbr_ra)*100) if (pbr_na and pbr_ra and pbr_ra!=0) else None
        per_ga = _r1((per_na-per_ra)/abs(per_ra)*100) if (per_na and per_ra and per_ra>0) else None
        pt100_a   = _safe(row_a.get("100점_합계"))
        sig_txt_a = str(row_a.get("매매시그널","") or "")
        sig_clr_a = str(row_a.get("_signal_color","FFCC00") or "FFCC00")

        all_vals = [
            ri_all,
            str(code_a),
            str(row_a.get("종목명","") or ""),
            str(row_a.get("시장","") or ""),
            str(row_a.get("업종","") or ""),
            _ri(cur_a), _ri(mc_a),
            # 거래량
            _r1(row_a.get("거래대금5일(억)")),
            _r1(row_a.get("거래대금20일(억)")),
            _r1(row_a.get("거래량비율VR(%)")),
            # 기술지표 (v32)
            _r2(row_a.get("MACD선")),
            _r2(row_a.get("MACD시그널")),
            _r2(row_a.get("MACD히스토")),
            str(row_a.get("MACD신호","") or ""),
            _r1(row_a.get("RSI14")),
            str(row_a.get("RSI신호","") or ""),
            str(row_a.get("OBV추세","") or ""),
            str(row_a.get("OBV신호","") or ""),
            _ri(row_a.get("BB상단")),
            _ri(row_a.get("BB하단")),
            _r1(row_a.get("BB%B")),
            _r1(row_a.get("BB밴드폭(%)")),
            str(row_a.get("BB신호","") or ""),
            row_a.get("기술매수신호수"),
            row_a.get("기술매도신호수"),
            str(row_a.get("기술신호","") or ""),
            str(row_a.get("기술지표요약","") or ""),
            # 밸류에이션
            pbr_na, pbr_ra, pbr_ga,
            _r1(row_a.get("ROE") or row_a.get("roe")),
            _r1(row_a.get("ROA")),
            per_na, per_ra, per_ga,
            _r1(row_a.get("DIV") or 0),
            _r1(row_a.get("dividend_payout")),
            _r1(row_a.get("영업이익률")),
            _r1(row_a.get("순이익률")),
            _r1(row_a.get("부채비율") or row_a.get("debt_ratio")),
            _r1(row_a.get("net_debt_ratio")),
            _r1(row_a.get("current_ratio")),
            _r2(row_a.get("altman_z")),
            fcf_a_eok, _r1(row_a.get("fcf_margin")), fy_a,
            eb_a_eok, ev_a,
            _r1(row_a.get("매출성장률")),
            cfo_a_eok,
            _r1(row_a.get("interest_coverage")),
            _r1(row_a.get("52주위치")),
            _r1(row_a.get("52주수익률")),
            _r1(row_a.get("6개월수익률")),
            _r1(row_a.get("가치점수")), _r1(row_a.get("모멘텀점수")),
            _r1(row_a.get("배당점수")), _r1(row_a.get("품질점수")),
            _r1(row_a.get("성장점수")), _r1(row_a.get("수익성점수")),
            _r1(row_a.get("안정점수")),
            _r1(sc_a),
            _ri(row_a.get("적정가_PER")), _ri(row_a.get("적정가_PBR")),
            _ri(row_a.get("적정가_Graham")), _ri(row_a.get("적정가_평균")),
            _r1(row_a.get("괴리율(%)")),
            str(row_a.get("적정가_사유","") or ""),
            str(grd_a),
            # 100점 평가
            _r1(pt100_a),
            str(row_a.get("100점_등급","") or ""),
            _safe(row_a.get("100점_수익성")), _safe(row_a.get("100점_성장성")),
            _safe(row_a.get("100점_재무안정")), _safe(row_a.get("100점_현금흐름")),
            _safe(row_a.get("100점_밸류")), _safe(row_a.get("100점_주주친화")),
            _safe(row_a.get("100점_정성보정")),
            str(row_a.get("체크_CFO","") or ""),
            str(row_a.get("체크_자본잠식","") or ""),
            str(row_a.get("체크_부채비율","") or ""),
            str(row_a.get("체크_ROE","") or ""),
            str(row_a.get("체크_PER","") or ""),
            str(row_a.get("보너스근거","") or ""),
            str(row_a.get("감점근거","") or ""),
            # 매매 시그널
            sig_txt_a,
            str(row_a.get("현재가위치","") or ""),
            _ri(row_a.get("적극매수선")),
            _ri(row_a.get("매수선")),
            _ri(row_a.get("1차목표가")),
            _ri(row_a.get("2차목표가")),
            _ri(row_a.get("추격매수경보")),
            _ri(row_a.get("손절선")),
            str(row_a.get("보유전략","") or ""),
            str(row_a.get("핵심투자포인트","") or ""),
            str(row_a.get("위험신호","없음") or "없음"),
        ]

        # 상위 N종목은 강조색, 나머지는 교번색
        if ri_all <= args.top:
            row_fill = _rank_fill(ri_all)
        elif ri_all % 2 == 0:
            row_fill = _ALT_FILL
        else:
            row_fill = PatternFill("solid", fgColor="FFFFFF")

        for j, v in enumerate(all_vals, 1):
            c = ws_all.cell(row=ri_all+2, column=j, value=v)
            c.font   = ws_all_body_font
            c.fill   = row_fill
            c.border = _THIN
            c.alignment = Alignment(horizontal="center", vertical="center",
                                    wrap_text=(j >= 63))
            fmt_a = _all_fmt.get(j)
            if fmt_a and v is not None:
                c.number_format = fmt_a
            # 괴리율(47열) 색상
            if j == 47 and v is not None:
                try:
                    fv = float(v)
                    c.font = Font(name=KR_FONT, size=9,
                                  color="0070C0" if fv >= 0 else "C00000",
                                  bold=(abs(fv) >= 20))
                except Exception: pass
            # 100점 합계(49열) 색상
            if j == 49 and v is not None:
                try:
                    fv = float(v)
                    if fv >= 90:   c.fill = PatternFill("solid", fgColor="0070C0"); c.font = Font(name=KR_FONT, size=9, color="FFFFFF", bold=True)
                    elif fv >= 80: c.fill = PatternFill("solid", fgColor="2E75B6"); c.font = Font(name=KR_FONT, size=9, color="FFFFFF", bold=True)
                    elif fv >= 70: c.fill = PatternFill("solid", fgColor="92D050"); c.font = Font(name=KR_FONT, size=9, bold=True)
                    elif fv >= 60: c.fill = PatternFill("solid", fgColor="FFCC00"); c.font = Font(name=KR_FONT, size=9, bold=True)
                    else:          c.fill = PatternFill("solid", fgColor="FF9900"); c.font = Font(name=KR_FONT, size=9, bold=True)
                except Exception: pass
            # 매매시그널(66열) 색상
            if j == 66 and v:
                try:
                    c.fill = PatternFill("solid", fgColor=sig_clr_a)
                    is_dark = sig_clr_a in ("C00000","0070C0","2E75B6","FF4444")
                    c.font = Font(name=KR_FONT, size=9, bold=True,
                                  color="FFFFFF" if is_dark else "000000")
                except Exception: pass

    ws_all.freeze_panes = "C3"
    auto_col_width(ws_all)
    print(f"  📋 전체 종목 시트: {len(df_all)}개 종목 × {len(all_headers)}컬럼")

    # ── 시트5: 매매시그널 보드 ──
    ws_signal.title = "매매시그널보드"
    ws_signal_hdr_font  = Font(name=KR_FONT, bold=True, color="FFFFFF", size=11)
    ws_signal_body_font = Font(name=KR_FONT, size=10)

    # 제목
    ws_signal.merge_cells("A1:O1")
    title_cell = ws_signal.cell(row=1, column=1,
        value=f"📊 매매시그널 보드  [{label}]  기준일: {datetime.today().strftime('%Y-%m-%d')}")
    title_cell.font = Font(name=KR_FONT, bold=True, size=14, color="FFFFFF")
    title_cell.fill = PatternFill("solid", fgColor="1F3864")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws_signal.row_dimensions[1].height = 32

    sig_headers = [
        "순위","코드","종목명","현재가",
        "매매시그널","현재가위치",
        "100점","복합점수",
        "적극매수선","매수선","1차목표가","2차목표가","손절선",
        "보유전략","위험신호"
    ]
    SIG_FILLS = {
        "강력매수": PatternFill("solid", fgColor="0070C0"),
        "■■ 매수":  PatternFill("solid", fgColor="2E75B6"),
        "관심":     PatternFill("solid", fgColor="92D050"),
        "보유":     PatternFill("solid", fgColor="FFCC00"),
        "비중축소": PatternFill("solid", fgColor="FF9900"),
        "매도":     PatternFill("solid", fgColor="FF4444"),
        "즉시매도": PatternFill("solid", fgColor="C00000"),
    }

    for j, h in enumerate(sig_headers, 1):
        c = ws_signal.cell(row=2, column=j, value=h)
        c.font = ws_signal_hdr_font
        c.fill = PatternFill("solid", fgColor="2E75B6")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _THIN
    ws_signal.row_dimensions[2].height = 24

    for ri_s, (code_s, row_s) in enumerate(df_top.iterrows(), 1):
        sig_txt  = str(row_s.get("매매시그널","") or "")
        pos_txt  = str(row_s.get("현재가위치","") or "")
        cur_s    = _gv(row_s, "현재가","현재가_NV","현재가_YF")
        pt100_s  = _r1(row_s.get("100점_합계"))
        comp_s   = _r1(row_s.get("복합점수", 0))

        # 시그널별 배경색
        sig_fill = None
        for kw, fill in SIG_FILLS.items():
            if kw in sig_txt:
                sig_fill = fill
                break
        if sig_fill is None:
            sig_fill = PatternFill("solid", fgColor="F5F5F5")

        sig_row_vals = [
            ri_s,
            str(code_s),
            str(row_s.get("종목명","") or ""),
            _ri(cur_s),
            sig_txt,
            pos_txt,
            pt100_s,
            comp_s,
            _ri(row_s.get("적극매수선")),
            _ri(row_s.get("매수선")),
            _ri(row_s.get("1차목표가")),
            _ri(row_s.get("2차목표가")),
            _ri(row_s.get("손절선")),
            str(row_s.get("보유전략","") or ""),
            str(row_s.get("위험신호","없음") or "없음"),
        ]

        is_danger = "위험" in str(row_s.get("위험신호","")) or "즉시" in sig_txt

        for j, v in enumerate(sig_row_vals, 1):
            c = ws_signal.cell(row=ri_s+2, column=j, value=v)
            c.border = _THIN
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=(j >= 14))
            if j == 5:  # 시그널 컬럼 색상
                c.fill = sig_fill
                is_dark = any(kw in sig_txt for kw in ("강력매수","매도","즉시"))
                c.font = Font(name=KR_FONT, size=10, bold=True,
                              color="FFFFFF" if is_dark else "000000")
            elif is_danger and j == 15:
                c.fill = PatternFill("solid", fgColor="FFE7E7")
                c.font = Font(name=KR_FONT, size=10, color="C00000", bold=True)
            else:
                c.fill = _GRN_FILL if ri_s % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
                c.font = ws_signal_body_font
            if j in (4, 9, 10, 11, 12, 13) and v is not None:
                c.number_format = "#,##0"

    ws_signal.freeze_panes = "A3"
    auto_col_width(ws_signal)
    print(f"  📋 매매시그널 보드: {len(df_top)}개 종목")

    # ── 시트5: 종목별 주간 뉴스/이슈 (v30 신규) ──
    ws_news.title = "종목별주간이슈"
    ws_news.merge_cells("A1:F1")
    _nc = ws_news.cell(row=1, column=1,
        value=f"📰 종목별 주간 핵심 이슈  [{label}]  기준일: {datetime.today().strftime('%Y-%m-%d')}"
              f"  (Claude AI 웹검색 기반 자동 요약)")
    _nc.font  = Font(name=KR_FONT, bold=True, size=12, color="FFFFFF")
    _nc.fill  = PatternFill("solid", fgColor="1F3864")
    _nc.alignment = Alignment(horizontal="center", vertical="center")
    ws_news.row_dimensions[1].height = 26

    _news_hdrs = ["순위","코드","종목명","업종","복합점수","주간 핵심 이슈 (Claude AI 자동 요약)"]
    _news_hdr_colors = ["1F3864","1F3864","1F3864","1F3864","2E75B6","7030A0"]
    for j, (h, col) in enumerate(zip(_news_hdrs, _news_hdr_colors), 1):
        c = ws_news.cell(row=2, column=j, value=h)
        c.font      = Font(name=KR_FONT, bold=True, color="FFFFFF", size=10)
        c.fill      = PatternFill("solid", fgColor=col)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = _THIN
    ws_news.row_dimensions[2].height = 24
    ws_news.column_dimensions["F"].width = 80  # 뉴스 컬럼 넓게

    for ri_n, (code_n, row_n) in enumerate(df_top.iterrows(), 1):
        score_n = _safe(row_n.get("복합점수"), 0)
        news_txt = news_map.get(str(code_n), "뉴스 미수집")
        _alt = _GRN_FILL if ri_n % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
        row_news = [
            ri_n,
            str(code_n),
            str(row_n.get("종목명","") or ""),
            str(row_n.get("업종","") or ""),
            _r1(score_n),
            news_txt,
        ]
        for j, v in enumerate(row_news, 1):
            c = ws_news.cell(row=ri_n+2, column=j, value=v)
            c.border    = _THIN
            c.fill      = _alt
            c.font      = Font(name=KR_FONT, size=9)
            c.alignment = Alignment(
                horizontal="center" if j < 6 else "left",
                vertical="top",
                wrap_text=(j == 6)
            )
        ws_news.row_dimensions[ri_n+2].height = max(
            30, min(15 * (len(news_txt)//60 + 1), 90)
        )

    ws_news.freeze_panes = "A3"
    for j in range(1, 6):
        auto_col_width(ws_news)
    print(f"  📰 종목별 주간이슈 시트: {len(news_map)}개 종목 수록")

    wb.save(fpath)
    print(f"\n  ✅ 엑셀 저장: {fpath}")
    return fpath

def save_signal_json(df_top, label, args, output_dir=BASE_DIR):
    try:
        stocks = []
        for rank, (code, row) in enumerate(df_top.iterrows(), 1):
            score = float(_safe(row.get("복합점수"), 0))
            cur   = _safe(row.get("현재가") or row.get("현재가_NV") or row.get("현재가_YF"), 0)
            pt100 = float(_safe(row.get("100점_합계"), 0))
            if score < 50 or not cur:
                continue
            stocks.append({
                "rank":         rank,
                "code":         str(code),
                "name":         str(row.get("종목명","")),
                "market":       str(row.get("시장","")),
                "sector":       str(row.get("업종","")),
                "price":        int(cur) if cur else 0,
                # ── 복합점수 (7팩터) ──
                "composite_score": round(score, 1),
                "grade":        str(row.get("등급","")),
                # ── 100점 종합 평가 ──
                "score_100": {
                    "total":        round(pt100, 1),
                    "grade":        str(row.get("100점_등급","")),
                    "수익성":       _safe(row.get("100점_수익성")),
                    "성장성":       _safe(row.get("100점_성장성")),
                    "재무안정":     _safe(row.get("100점_재무안정")),
                    "현금흐름":     _safe(row.get("100점_현금흐름")),
                    "밸류에이션":   _safe(row.get("100점_밸류")),
                    "주주친화":     _safe(row.get("100점_주주친화")),
                    "정성보정":     _safe(row.get("100점_정성보정")),
                },
                # ── 5대 필수 체크 ──
                "health_check": {
                    "cfo_positive":    str(row.get("체크_CFO","")),
                    "no_impairment":   str(row.get("체크_자본잠식","")),
                    "debt_safe":       str(row.get("체크_부채비율","")),
                    "roe_ok":          str(row.get("체크_ROE","")),
                    "per_ok":          str(row.get("체크_PER","")),
                },
                # ── 매매 시그널 ──
                "trading": {
                    "signal":          str(row.get("매매시그널","")),
                    "position_diag":   str(row.get("현재가위치","")),
                    "price_band": {
                        "strong_buy":  _safe(row.get("적극매수선")),
                        "buy":         _safe(row.get("매수선")),
                        "target_1":    _safe(row.get("1차목표가")),
                        "target_2":    _safe(row.get("2차목표가")),
                        "overbought":  _safe(row.get("추격매수경보")),
                        "stop_loss":   _safe(row.get("손절선")),
                    },
                    "strategy":        str(row.get("보유전략","")),
                    "key_points":      str(row.get("핵심투자포인트","")),
                    "risk_flags":      str(row.get("위험신호","없음")),
                    "bonus_reason":    str(row.get("보너스근거","")),
                    "malus_reason":    str(row.get("감점근거","")),
                },
                # ── 기본 재무 ──
                "fundamentals": {
                    "pbr":          _safe(row.get("PBR") or row.get("pbr")),
                    "roe":          _safe(row.get("ROE") or row.get("roe")),
                    "div":          _safe(row.get("DIV"), 0),
                    "debt_ratio":   _safe(row.get("부채비율") or row.get("debt_ratio")),
                    "interest_coverage": _safe(row.get("interest_coverage")),
                    "cfo":          _safe(row.get("cfo")),
                    "fcf":          _safe(row.get("fcf")),
                },
                # ── 밸류에이션 ──
                "valuation": {
                    "per_target":    _safe(row.get("적정가_PER")),
                    "pbr_target":    _safe(row.get("적정가_PBR")),
                    "graham_target": _safe(row.get("적정가_Graham")),
                    "avg_target":    _safe(row.get("적정가_평균")),
                    "upside_pct":    _safe(row.get("괴리율(%)")),
                },
            })

        data = {
            "scan_date":  datetime.today().strftime("%Y-%m-%d"),
            "scan_time":  datetime.now().strftime("%H:%M"),
            "screener":   f"QuantScreener_{VERSION}_{label}",
            "top_stocks": stocks,
            "filter_conditions": {
                "min_roe": args.min_roe, "max_pbr": args.max_pbr,
                "max_debt": args.max_debt, "min_ic": args.min_ic, "top_n": args.top,
            }
        }

        fname  = datetime.today().strftime(f"매매신호_{label}_%Y%m%d.json")
        fpath  = os.path.join(output_dir, fname)
        with open(fpath, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        print(f"  ✅ 매매신호 JSON: {fpath} ({len(stocks)}개)")
    except Exception as e:
        print(f"  ⚠ JSON 저장 실패: {e}")

# ══════════════════════════════════════════════════════════
# 11. 터미널 요약 출력
# ══════════════════════════════════════════════════════════

def print_summary(df_top, label, args):
    print()
    print("=" * 110)
    print(f"  ★ 퀀트 스크리닝 상위 {args.top}종목 [{label}]  {VERSION}")
    print(f"  기준: PBR≤{args.max_pbr} / ROE≥{args.min_roe}% / 부채비율<{args.max_debt}% / 이자보상>={args.min_ic}배")
    print("=" * 110)
    print(f"  {'순위':<4} {'코드':<8} {'종목명':<14} {'현재가':>8} {'PBR':>5} "
          f"{'ROE%':>7} {'DIV%':>6} {'복합':>6} {'100점':>5} {'시그널':<15} "
          f"{'1차목표가':>10} {'괴리율%':>8} {'손절선':>8}")
    print("  " + "-" * 106)

    for rank, (code, row) in enumerate(df_top.iterrows(), 1):
        name   = str(row.get("종목명",""))[:12]
        cur    = _safe(row.get("현재가") or row.get("현재가_NV") or row.get("현재가_YF"), 0)
        pbr    = _safe(row.get("PBR") or row.get("pbr"), 0)
        roe    = _safe(row.get("ROE") or row.get("roe"), 0)
        div    = _safe(row.get("DIV"), 0.0)
        score  = _safe(row.get("복합점수"), 0)
        pt100  = _safe(row.get("100점_합계"), 0)
        signal = str(row.get("매매시그널","─ 보유") or "─ 보유")[:14]
        tgt1   = _safe(row.get("1차목표가"))
        upside = _safe(row.get("괴리율(%)"))
        stop   = _safe(row.get("손절선"))
        try:
            cur_str  = f"{int(cur):,}"
        except Exception:
            cur_str  = "N/A"
        try:
            tgt_str  = f"{int(tgt1):,}" if tgt1 else "N/A"
        except Exception:
            tgt_str  = "N/A"
        try:
            stop_str = f"{int(stop):,}" if stop else "N/A"
        except Exception:
            stop_str = "N/A"
        upside_str = f"{upside:+.1f}%" if upside is not None else "N/A"
        print(f"  {rank:<4} {code:<8} {name:<14} {cur_str:>8} {pbr:>5.2f} "
              f"{roe:>6.1f}% {div:>5.2f}% {score:>6.1f} {pt100:>5.0f} {signal:<15} "
              f"{tgt_str:>10} {upside_str:>8} {stop_str:>8}")
    print()

# ══════════════════════════════════════════════════════════
# 12. Google Drive 업로드 + 메인 실행
#     (v33: GitHub Actions + Google Drive 자동화)
# ══════════════════════════════════════════════════════════

def upload_to_gdrive(xlsx_files: list, output_dir: str, folder_id: str) -> None:
    """
    생성된 엑셀 파일을 Google Drive 지정 폴더에 업로드합니다.

    인증 방식: Service Account JSON (GitHub Secrets → GDRIVE_SA_JSON 환경변수)
    필요 패키지: google-auth google-api-python-client

    환경변수:
      GDRIVE_SA_JSON    — Service Account JSON 문자열 (GitHub Secret)
      GDRIVE_FOLDER_ID  — 업로드 대상 Drive 폴더 ID
    """
    if not xlsx_files:
        print("  ⚠ 업로드할 엑셀 파일 없음")
        return

    sa_json_str = os.environ.get("GDRIVE_SA_JSON", "").strip()
    if not sa_json_str:
        print("  ⚠ GDRIVE_SA_JSON 환경변수 없음 → Drive 업로드 스킵")
        return

    try:
        from google.oauth2 import service_account
        from googleapiclient.discovery import build
        from googleapiclient.http import MediaFileUpload
    except ImportError:
        print("  ⚠ google-auth / google-api-python-client 미설치")
        print("    pip install google-auth google-api-python-client")
        return

    try:
        sa_info = json.loads(sa_json_str)
        creds = service_account.Credentials.from_service_account_info(
            sa_info,
            scopes=["https://www.googleapis.com/auth/drive"]
        )
        service = build("drive", "v3", credentials=creds, cache_discovery=False)
    except Exception as e:
        print(f"  ⚠ Google Drive 인증 실패: {e}")
        return

    print(f"\n  📤 Google Drive 업로드 시작 ({len(xlsx_files)}개 파일) ...")

    for fname in xlsx_files:
        fpath = os.path.join(output_dir, fname)
        if not os.path.exists(fpath):
            continue
        try:
            # 같은 이름 파일이 폴더에 이미 있으면 덮어쓰기 (업데이트)
            existing = service.files().list(
                q=f"name='{fname}' and '{folder_id}' in parents and trashed=false",
                fields="files(id, name)",
                spaces="drive"
            ).execute().get("files", [])

            media = MediaFileUpload(
                fpath,
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                resumable=True
            )

            if existing:
                # 기존 파일 업데이트
                file_id = existing[0]["id"]
                service.files().update(
                    fileId=file_id,
                    media_body=media
                ).execute()
                print(f"  ✅ 업데이트: {fname}  (Drive 파일 ID: {file_id})")
            else:
                # 신규 업로드
                meta = {"name": fname, "parents": [folder_id]}
                result = service.files().create(
                    body=meta,
                    media_body=media,
                    fields="id"
                ).execute()
                print(f"  ✅ 업로드: {fname}  (Drive 파일 ID: {result.get('id')})")

        except Exception as e:
            print(f"  ⚠ {fname} 업로드 실패: {e}")

    print("  📤 Google Drive 업로드 완료\n")


def _run_scan(dart, args, market_choice, kr_size, us_size):
    all_dfs = []

    # 한국
    if market_choice in ("1","3"):
        print("\n" + "─"*45)
        print("  [한국 시장 스크리닝 시작]")
        print("─"*45)
        tickers, names, markets = fetch_kr_universe()
        df_kr_raw = fetch_kr_all_data(dart, tickers, names, markets, kr_size,
                                       no_cache=getattr(args, "no_cache", False))
        df_kr_raw = compute_scores(df_kr_raw, args)
        df_kr     = apply_hard_filter(df_kr_raw, args)
        if not df_kr.empty:
            df_top_kr = df_kr.head(args.top)
            print(f"  📊 스크리닝 통과: {len(df_kr)}개 / 전체 {len(df_kr_raw)}개 → 상위 {args.top}개 강조")
            print_summary(df_top_kr, "한국 KOSPI/KOSDAQ", args)
            build_excel(df_kr_raw, df_top_kr, "KR", args, args.output_dir)
            save_signal_json(df_top_kr, "KR", args, args.output_dir)
            all_dfs.append(df_kr_raw)

    # 미국
    if market_choice in ("2","3"):
        print("\n" + "─"*45)
        print("  [미국 시장 스크리닝 시작]")
        print("─"*45)
        df_us_raw = fetch_us_all_data(us_size)
        df_us_raw = compute_scores(df_us_raw, args)
        df_us     = apply_hard_filter(df_us_raw, args)
        if not df_us.empty:
            df_top_us = df_us.head(args.top)
            print(f"  📊 스크리닝 통과: {len(df_us)}개 / 전체 {len(df_us_raw)}개 → 상위 {args.top}개 강조")
            print_summary(df_top_us, "미국 S&P500", args)
            build_excel(df_us_raw, df_top_us, "US", args, args.output_dir)
            save_signal_json(df_top_us, "US", args, args.output_dir)
            all_dfs.append(df_us_raw)

    # 통합
    if market_choice == "3" and len(all_dfs) == 2:
        print("\n" + "─"*45)
        print("  [한국+미국 통합 랭킹]")
        combined = pd.concat(all_dfs)
        combined = compute_scores(combined, args)
        if not combined.empty:
            df_top_all = combined.head(args.top)
            print(f"  📊 통합 전체: {len(combined)}개 종목 → 상위 {args.top}개 강조")
            print_summary(df_top_all, "통합 한국+미국", args)
            build_excel(combined, df_top_all, "ALL", args, args.output_dir)
            save_signal_json(df_top_all, "ALL", args, args.output_dir)

    print("\n  ✅ 완료! 저장 파일 확인: quant_*.xlsx / 매매신호_*.json")
    print("  📌 엑셀 시트:")
    print("     ① 스크리닝결과상세  — 상위N종목 핵심 컬럼 요약")
    print("     ② 지표해석가이드    — 지표별 해석 기준")
    print("     ③ 전체종목데이터    — 전 종목 75컬럼 완전 데이터 (복합점수순)")
    print("     ④ 매매시그널보드    — 시그널별 컬러코딩 한눈에 보기")
    print("  📌 JSON: 100점 구조화 + 가격밴드 + 위험신호 포함")
    print("=" * 75 + "\n")

    # ── Google Drive 업로드 (환경변수 GDRIVE_FOLDER_ID 설정 시 자동 실행) ──
    folder_id = os.environ.get("GDRIVE_FOLDER_ID", "").strip()
    if folder_id:
        xlsx_files = [
            f for f in os.listdir(args.output_dir)
            if f.startswith("quant_") and f.endswith(".xlsx")
               and datetime.today().strftime("%Y%m%d") in f
        ]
        upload_to_gdrive(xlsx_files, args.output_dir, folder_id)
    else:
        print("  ℹ Google Drive 업로드 스킵 (GDRIVE_FOLDER_ID 미설정)")


def _make_task_bat():
    """
    Windows 작업 스케줄러용 .bat 파일 자동 생성
    생성 파일: run_quant_daily.bat  (스크립트와 같은 폴더)

    사용법:
      1) 이 함수 실행 → run_quant_daily.bat 생성
      2) Windows 키 → '작업 스케줄러' 검색 → 열기
      3) [작업 만들기] 클릭
         - 일반: '퀀트 스크리너 매일 실행'
         - 트리거: [새로 만들기] → 매일 / 오전 08:50
         - 동작:  [새로 만들기] → 프로그램/스크립트에 run_quant_daily.bat 경로 입력
         - 조건:  '컴퓨터가 AC 전원에 연결된 경우에만 실행' 체크 해제 (노트북)
      4) [확인] → 완료

    추가 팁:
      - 로그는 quant_daily.log 파일에 쌓임
      - 수동 테스트: run_quant_daily.bat 더블클릭
    """
    py_exe  = sys.executable                        # 현재 Python 인터프리터 경로
    script  = os.path.abspath(__file__)             # 이 스크립트의 절대경로
    log_f   = os.path.join(BASE_DIR, "quant_daily.log")
    bat_path = os.path.join(BASE_DIR, "run_quant_daily.bat")

    bat_content = f"""@echo off
REM ============================================================
REM  퀀트 스크리너 v33 — 매일 자동 실행 배치파일
REM  Windows 작업 스케줄러에 등록하여 매일 08:50 자동 실행
REM  생성일: {datetime.now().strftime("%Y-%m-%d %H:%M")}
REM ============================================================

echo [%date% %time%] 퀀트 스크리너 시작 >> "{log_f}"

"{py_exe}" "{script}" --market 1 --scope 2 --auto >> "{log_f}" 2>&1

if %errorlevel% equ 0 (
    echo [%date% %time%] 완료 (성공) >> "{log_f}"
) else (
    echo [%date% %time%] 완료 (오류 코드: %errorlevel%) >> "{log_f}"
)
"""
    try:
        with open(bat_path, "w", encoding="cp949") as f:
            f.write(bat_content)
        print(f"\n  ✅ 배치파일 생성 완료: {bat_path}")
        print(f"  📋 Windows 작업 스케줄러 등록 방법:")
        print(f"     1. Windows키 → '작업 스케줄러' 검색 → 실행")
        print(f"     2. [작업 만들기] → 트리거: 매일 08:50")
        print(f"     3. 동작 → 프로그램: {bat_path}")
        print(f"     4. 실행 로그: {log_f}")
    except Exception as e:
        print(f"  ⚠ 배치파일 생성 실패: {e}")
    return bat_path


def select_mode_interactive():
    """수동 실행 시 대화형 메뉴 (--auto 없을 때)"""
    print()
    print("  ┌────────────────────────────────────────────────────────┐")
    print(f"  │  퀀트 주식 스크리너 {VERSION}  (100점 만점 종합분석)  │")
    print("  │  DART×네이버×yfinance │ 7팩터+100점 │ 한국+미국       │")
    print("  │  3계층 캐시 (DART 7일/보조지표 1일/현재가 매일 갱신)  │")
    print("  └────────────────────────────────────────────────────────┘")
    print()

    # 캐시 상태 표시
    cache_show_status()
    cache_clear(days_old=8)
    print()

    print("  [시장 선택]")
    print("  1. 한국 (KOSPI+KOSDAQ)  ← DART 재무 + 네이버 보조 + yfinance 가격")
    print("  2. 미국 (S&P500)        ← yfinance")
    print("  3. 둘 다 (통합 랭킹 포함)")
    while True:
        mkt = input("\n  선택 (1/2/3): ").strip()
        if mkt in ("1","2","3"): break
        print("  1, 2, 3 중 하나를 입력하세요.")

    print()
    print("  [스크리닝 범위]")
    print("  1. 빠른 테스트  (한국 100개 / 미국 50개,   약 10~15분)")
    print("  2. 중간 범위    (한국 300개 / 미국 200개,  약 30~40분)")
    print("  3. 전체 스크리닝 (전 종목,                  수 시간 소요)")
    print("  4. 작업스케줄러 .bat 파일 생성 (최초 1회)")
    while True:
        scope = input("\n  선택 (1/2/3/4): ").strip()
        if scope in ("1","2","3","4"): break
        print("  1, 2, 3, 4 중 하나를 입력하세요.")

    if scope == "4":
        _make_task_bat()
        sys.exit(0)

    return mkt, scope


def main():
    _print_pkg_status()

    # ── argparse에 --auto / --market / --scope 추가 ──
    # (기존 parse_args는 그대로 유지하고 여기서 추가 인자만 처리)
    import argparse as _ap
    pre = _ap.ArgumentParser(add_help=False)
    pre.add_argument("--auto",   action="store_true",
                     help="비대화형 자동 실행 (작업 스케줄러 전용)")
    pre.add_argument("--market", type=str, default="1",
                     choices=["1","2","3"],
                     help="시장 선택: 1=한국 2=미국 3=둘다 (기본: 1)")
    pre.add_argument("--scope",  type=str, default="2",
                     choices=["1","2","3"],
                     help="범위: 1=100개 2=300개 3=전체 (기본: 2)")
    pre_args, _ = pre.parse_known_args()

    args = parse_args()           # 기존 전체 인자 파싱

    if pre_args.auto:
        # ── 자동 실행 모드 (작업 스케줄러) ──
        market_choice = pre_args.market
        scope         = pre_args.scope
        print(f"\n  🤖 자동 실행 모드  [{datetime.now().strftime('%Y-%m-%d %H:%M')}]")
        print(f"     시장: {'한국' if market_choice=='1' else '미국' if market_choice=='2' else '한국+미국'}")
        print(f"     범위: {'100개' if scope=='1' else '300개' if scope=='2' else '전체'}")
        cache_show_status()
        cache_clear(days_old=8)
    else:
        # ── 수동 실행 모드 (대화형 메뉴) ──
        market_choice, scope = select_mode_interactive()

    kr_size = {"1": 100, "2": 300, "3": None}[scope]
    us_size = {"1":  50, "2": 200, "3": None}[scope]

    # DART 키 로드 (한국 선택 시)
    dart = None
    if market_choice in ("1","3"):
        dart_key = load_dart_key()
        dart = DartClient(dart_key)

    _run_scan(dart, args, market_choice, kr_size, us_size)


if __name__ == "__main__":
    main()

